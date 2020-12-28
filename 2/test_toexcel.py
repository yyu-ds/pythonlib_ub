# -*- coding: utf-8 -*-
"""
Created on Fri Jan 29 15:54:08 2016

@author: ub71894 (4e8e6d0b), CSG
"""


import os
import pandas as pd 
import numpy as np
os.chdir(r"C:\Users\ub71894\Documents\code\Python\testcode")


# make up 2 DataFrame
dat = pd.read_csv(r"C:\Users\ub71894\Documents\data\RA\ra_perm_csv.csv")
a=dat.iloc[1:100,1:5].copy()
b=dat.iloc[100:200,6:10].copy()
c=a


#%% save df to an excel file:
writer = pd.ExcelWriter('temp.xlsx', engine='xlsxwriter')
a.to_excel(writer, sheet_name='DataFrame_a')
b.to_excel(writer, sheet_name='DataFrame_b')
writer.save()


#%% save df to an existing excel file:

from openpyxl import load_workbook

book = load_workbook('temp.xlsx')
writer = pd.ExcelWriter('temp.xlsx', engine='openpyxl') 
writer.book = book
writer.sheets = dict((ws.title, ws) for ws in book.worksheets)

c.to_excel(writer, sheet_name='DataFrame_c')

writer.save()



#%% write number in specific range of workbook:
from openpyxl import Workbook
from openpyxl.compat import range

wb = Workbook()

dest_filename = 'empty_book.xlsx'

ws1 = wb.active
ws1.title = "range names"

for row in range(1, 40):
    ws1.append(range(600))

ws2 = wb.create_sheet(title="Pi")

ws2['F5'] = df2

ws3 = wb.create_sheet(title="Data")
for row in range(10, 20):
    for col in range(27, 54):
        _ = ws3.cell(column=col, row=row, value="%s" % get_column_letter(col))
print(ws3['AA10'].value)

wb.save(filename = dest_filename)


#%% write df in specific range in each setting:
top = 5
writer = pd.ExcelWriter('top5.xlsx')
sheetname = ['dev','prod','combo']
ab= ['/before','/after']
filename = ['output\mfa_before.xlsx','output\mfa_after_existing.xlsx']

for i,file in enumerate(filename):
    for j,sheet in enumerate(sheetname):
        df = pd.read_excel(file,sheetname=sheet)
        df.sort_values(by='AR',ascending=False, inplace=True)
        df2 = df.iloc[:top,:]
        df2.to_excel(writer, sheet_name='Sheet1', startrow=(8*j+1), startcol=(10**i))
        worksheet = writer.sheets['Sheet1']
        worksheet.write((8*j+1), (10**i), sheet+ab[i])

writer.save()from twilio.rest import TwilioRestClient 
