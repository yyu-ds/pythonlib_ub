In this function, you have:
-   Turning format into percent
-   Changing police (bold, size…)
-   Defining vertical/horizontal alignment
-   Coloring
-   Row/ciolumn dimensions
-   Adding borders
'''

import os, numpy as np, pandas as pd
os.chdir(r"C:\Users\ub71894\Documents\code\python\testcode")

#%% use openpyxl to create new xlsx:

from openpyxl import Workbook
from openpyxl.compat import range
from openpyxl.cell import get_column_letter

wb = Workbook()
dest_filename = 'empty_book.xlsx'

ws1 = wb.active
ws1.title = "range names"

for row in range(1, 40):
    ws1.append(range(600))

ws2 = wb.create_sheet(title="Pi")
ws2['F5'] = 3.14

ws3 = wb.create_sheet(title="Data")
for row in range(10, 20):
    for col in range(27, 54):
        _ = ws3.cell(column=col, row=row, value="%s" % get_column_letter(col))

wb.save(filename = dest_filename)







#%% read specific range frpm existing file:

bf_existing = pd.read_excel('top5_AR_5%.xlsx',0,skiprows=1, parse_cols = 'B:H')
af_existing = pd.read_excel('top5_AR_5%.xlsx',0,skiprows=1, parse_cols = 'K:P')
af_new = pd.read_excel('top5_AR_5%.xlsx',0,skiprows=1, parse_cols = 'T:Y')























#%% condition formatting:

import numpy as np
import pandas
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Border, Side, Alignment, Font
from openpyxl.formatting.rule import IconSetRule
from openpyxl.styles import PatternFill

        
#Open output
wb =  openpyxl.load_workbook(output_file_path)
sheet = wb.create_sheet() 
sheet.title = u'RLA-Override'
    sheet.cell(row=3,column=col).value=RLA


def CreateStyleSheet_RLA_Override(sheet):
    
    for col in range(3,7):
        sheet.cell(row=3,column=col).number_format='.0%'
        sheet.cell(row=6,column=col).number_format='.0%'
        sheet.cell(row=4,column=col).number_format='.0%'    
        sheet.cell(row=5,column=col).number_format='.0%'
        
    #Work on style
    for _row in sheet.iter_rows('B1:F7'):
        for _cell in _row:
            _cell.alignment=Alignment(horizontal='center',vertical='center')
            _cell.font=Font(size=9)
            fill=PatternFill(start_color='FFFFFF', end_color='FFFFFF',fill_type='solid')
            _cell.fill=fill 
    for _row in sheet.iter_rows('C3:F3'): #Adjust color for cells
        for _cell in _row:
            color=Color(_cell.value, 'RLA')
            fill=PatternFill(start_color=color, end_color=color,fill_type='solid')
            _cell.fill=fill 
    for _row in sheet.iter_rows('C4:F4'): #Adjust color for cells
        for _cell in _row:
            color=Color(_cell.value, 'Override')
            fill=PatternFill(start_color=color, end_color=color,fill_type='solid')
            _cell.fill=fill
            
                    
    sheet.row_dimensions[2].height=17
    sheet.row_dimensions[3].height=17
    sheet.row_dimensions[4].height=17   
    sheet.row_dimensions[5].height=17    
    sheet.row_dimensions[6].height=17    
    sheet.row_dimensions[7].height=17
    sheet.column_dimensions['B'].width=14
    sheet.column_dimensions['C'].width=9
    sheet.column_dimensions['D'].width=9
    sheet.column_dimensions['E'].width=9   
    sheet.column_dimensions['F'].width=9
    
    for i in ['B1', 'B2','B3','B4', 'B5','B6', 'B7']:
        a = sheet[i]
        a.font=Font(bold=True, size=9)
        a.alignment=Alignment(vertical='center', horizontal='left')
    for i in ['B2', 'C2','D2','E2', 'F2']:
        a = sheet[i]
        a.font=Font(bold=True, size=9)
        a.alignment=Alignment(vertical='center', horizontal='center')
    sheet.merge_cells('B1:F1')
    sheet['B1'].alignment=Alignment(horizontal='center')
    
    #### Add borders ###
    side=Side(style='thin', color="FF000000")
    
    for i in ['C1','D1','E1']: #top border
        a = sheet[i]
        a.border=Border(bottom=side, top=side)
    sheet['F1'].border=Border(bottom=side, right=side, top=side)
    sheet['B1'].border=Border(bottom=side, left=side, top=side)
    
    
    for i in ['C2','D2','E2']: #top border
        a = sheet[i]
        a.border=Border(bottom=side, top=side)
    sheet['F2'].border=Border(bottom=side, right=side, top=side)
    sheet['B2'].border=Border(bottom=side, left=side, top=side, right=side)
    
    sheet['B3'].border=Border(left=side, top=side, right=side)
    sheet['F3'].border=Border(right=side, top=side)
    
    sheet['B4'].border=Border(left=side, right=side)
    sheet['F4'].border=Border(right=side)
    
    for i in ['C5','D5','E5']: #top border
        a = sheet[i]
        a.border=Border(top=side)
   sheet['B5'].border=Border(top=side, left=side, right=side)
    sheet['F5'].border=Border(right=side, top=side)    
    
    for i in ['C6','D6','E6']: #top border
        a = sheet[i]
        a.border=Border(bottom=side)
    sheet['B6'].border=Border(bottom=side, left=side, right=side)
    sheet['F6'].border=Border(bottom=side, right=side)      
    
    sheet['B7'].border=Border(bottom=side, right=side, left=side, top=side)      
    sheet['C7'].border=Border(bottom=side, left=side, top=side) 
    sheet['D7'].border=Border(bottom=side, top=side)       
    sheet['E7'].border=Border(bottom=side, top=side)  
    sheet['F7'].border=Border(bottom=side, top=side, right=side)
