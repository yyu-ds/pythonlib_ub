# -*- coding: utf-8 -*-
"""
Created on Tue Feb  7 14:46:19 2017

Version: 1.0 Initial build
Version: 1.1, 20170313, Remove all irrelevant cols in default data

@author: ub71894 (4e8e6d0b), CSG
"""


import pandas as pd
import numpy as np
from dateutil.relativedelta import relativedelta as rd
import collections
import warnings

#%%
class Attachdefaults(object):
    
    '''    
    This class is used to attach default data to financial statement and then 
    apply default logic to clean the dataset. 
    
    Attachdefaults(findata, findata_id, defdata, defdata_id, **kwargs):

        findata:    financial statement data. User need to clean and drop duplicates
                    in 'findata' before constructing the class
        findata_id: The column name for id in 'findata' and it will be used to link 
                    with default data  
        defdata:    default data. Usually coming from MasterDefault data. User 
                    need to make sure the column name for default time is 
                    'L_DATE_OF_DEFAULT' and the column name for default ending 
                    time is 'L_DEFAULT_END_DT' if it has.
        defdata_id: The column name for id in 'defdata' and it will be used to link 
                    with financial statement.  

        Parameters: 

        isthere_def_enddate:    boolean, default False
                                Whether the defaults have ending time in 'defdata'
        months_default_atleast_last:    int, default 6
                                        the floor of the default period. The unit
                                        is month.
        findata_timestamp:  str, default 'archive_date'
                            the column name of the timestamp for financial statement 
                            in 'findata'
        blackout_mo_bf_def_begins:  int, default 6
                                    When the financial statement is too close to 
                                    the actual default, it should be ignore since 
                                    it's most likely influenced. This number is 
                                    how many months before the default are considered
                                    as 'too close'. 
        blackout_mo_af_def_ends:    int, default 6
                                    When the financial statement is too close to 
                                    the actual default ending date, it should be
                                    ignore since it's most likely still under the 
                                    impact of default. This number is how many months
                                    after the default ends are considered as 'too close'. 
        valid_time_window_mo:   tuple of 2 integers, default (6,18)
                                The valid time window (months) for financial statements 
                                that happens before actual default. Usually we 
                                set 6-18 months as the valid time period. which 
                                means only financial statements occurs 6 to 18 months 
                                before the actual default will be tagged as a 
                                'default' or 'bad' observation.                   

    The procedure detail:

    Part One: Clean 'defdata':
        Step_1: Drop Obs. if its 'L_DATE_OF_DEFAULT' is NAs or duplicated 
        Step_2: Modify 'L_DEFAULT_END_DT' by the rule below:
                L_DEFAULT_END_DT = L_DEFAULT_END_DT,  if the default last longer than 6 months
                L_DEFAULT_END_DT = L_DATE_OF_DEFAULT + 6 months,  if others
        Step_3: Combine defaults for the same obligor if the second one happened within 1 year of the
                previous one or within previous default's end date.

    Part Two: Merge 'findata' and 'defdata' by the input identifiers.

    Part Three: Apply default logic to the merged data:
        Step_1: Blackout all financial statments which are in the time window 
                ['blackout_mo_bf_def_begins' months before default begins,
                blackout_mo_af_def_ends' months after default ends]    
        Step_2: Pick the financial statments that occur in the 'valid time window'. 
                Usually it's 6-18 months before actual default.
        Step_3: Only keep the latest financial statement if multiple lie in valid 
                time window. In other word, we only select the closest financial 
                statement to the actual default and remove others.

    '''

    def __init__(self, findata, findata_id, defdata, defdata_id, **kwargs):
        self.findata = findata.copy()
        self.findata.reset_index(drop=True, inplace=True)
        self.defdata = defdata.copy()
        self.defdata.reset_index(drop=True, inplace=True)
        # add the line below in Ver 1.1 to remove all irrelevant cols in def data:
        self.defdata = self.defdata[[defdata_id]+['L_DATE_OF_DEFAULT','L_DEFAULT_END_DT']]
        self.orisize_findata = len(self.findata)
        self.orisize_defdata = len(self.defdata)
        #self.orinofdef = defdata.def_flag.sum()
        self.isthere_def_enddate = kwargs.pop('isthere_def_enddate',False)
        self.months_default_atleast_last = kwargs.pop('months_default_atleast_last',6)
        self.data_on = findata_id
        self.def_on = defdata_id
        self.findata_timestamp = kwargs.pop('findata_timestamp','archive_date')
        self.blackout_mo_bf_def_begins = kwargs.pop('blackout_mo_bf_def_begins',6)
        self.blackout_mo_af_def_ends = kwargs.pop('blackout_mo_af_def_ends',6)
        self.valid_time_window_mo = kwargs.pop('valid_time_window_mo',(6,18))

    def __cleandef(self):

        if not self.isthere_def_enddate: # then create a column for end date
            self.defdata['L_DEFAULT_END_DT'] = self.defdata['L_DATE_OF_DEFAULT']
        # 1. drop duplicated defaults and keep the one which has the latest L_DEFAULT_END_DT
        previousdefdata = self.defdata.copy()
        self.defdata.dropna(subset=['L_DATE_OF_DEFAULT'], inplace=True)
        self.defdata.sort_values(by=[self.def_on,'L_DATE_OF_DEFAULT','L_DEFAULT_END_DT'], ascending=False, inplace=True)
        self.defdata.drop_duplicates(subset=[self.def_on,'L_DATE_OF_DEFAULT'], keep='first', inplace=True)
        self.defdata.sort_values(by=[self.def_on,'L_DATE_OF_DEFAULT'], inplace=True)
        # To get the dataframe that contains the obs which have been kicked out
        self.def_step1_dropped = pd.concat([previousdefdata, self.defdata])
        self.def_step1_dropped.reset_index(drop=False, inplace=True)
        self.def_step1_dropped.drop_duplicates(subset='index',keep=False,inplace=True)

        self.defdata.reset_index(drop=True, inplace=True)
        self.def_step1_left=self.defdata.copy()

        # 2. modify L_DEFAULT_END_DT by the rule below:
        # L_DEFAULT_END_DT = L_DEFAULT_END_DT            ,  if the default last longer than 6 months
        # L_DEFAULT_END_DT = L_DATE_OF_DEFAULT + 6 months,  if others
        s=0
        for row in self.defdata.iterrows():
            if pd.isnull(row[1]['L_DEFAULT_END_DT']):
                self.defdata.loc[row[0],'L_DEFAULT_END_DT'] = self.defdata.loc[row[0],'L_DATE_OF_DEFAULT'] + rd(months=self.months_default_atleast_last)       
                s+=1
            elif (row[1]['L_DEFAULT_END_DT']-row[1]['L_DATE_OF_DEFAULT']).days <= 30 * self.months_default_atleast_last:
                self.defdata.loc[row[0],'L_DEFAULT_END_DT'] = self.defdata.loc[row[0],'L_DATE_OF_DEFAULT'] + rd(months=self.months_default_atleast_last)
                s+=1
        self.def_step2_dropped = [s,len(self.defdata)]

        self.def_step2_left=self.defdata.copy()

        # 3. combine defaults if the second one happened within one year of the previous one or within previous default's end date
        self.def_step3_dropped=[len(self.defdata),0]

        list_to_remove=[];
        current_obligor = self.defdata.loc[0,self.def_on]
        current_def = self.defdata.loc[0,'L_DATE_OF_DEFAULT']
        current_def_end = self.defdata.loc[0,'L_DEFAULT_END_DT']
        current_index = 0       

        for i in range(1,len(self.defdata)):
            if self.defdata.loc[i,self.def_on] != current_obligor: 
                current_obligor = self.defdata.loc[i,self.def_on]
                current_def = self.defdata.loc[i,'L_DATE_OF_DEFAULT']
                current_def_end = self.defdata.loc[i,'L_DEFAULT_END_DT']
                current_index = i
                continue
            elif (self.defdata.loc[i,'L_DATE_OF_DEFAULT'] < current_def+rd(years=1)) or \
                (self.defdata.loc[i,'L_DATE_OF_DEFAULT'] < current_def_end):
                list_to_remove.append(i)
                current_def_end = self.defdata.loc[i,'L_DEFAULT_END_DT']
                self.defdata.loc[current_index,'L_DEFAULT_END_DT'] = current_def_end
                continue
            else:
                current_obligor = self.defdata.loc[i,self.def_on]
                current_def = self.defdata.loc[i,'L_DATE_OF_DEFAULT']
                current_def_end = self.defdata.loc[i,'L_DEFAULT_END_DT']
                current_index = i
                continue        
        

        self.defdata.drop(self.defdata.index[list_to_remove], inplace=True)
        self.def_step3_dropped[1] = len(self.defdata)
        self.def_step3_left=self.defdata.copy()


    def __run(self):

        self.__cleandef()

        alldata= pd.merge(self.findata, self.defdata, left_on=self.data_on, right_on=self.def_on, how='left')
        good = alldata.query('L_DATE_OF_DEFAULT != L_DATE_OF_DEFAULT') # nan in default date means good
        good['def_flag'] = 0
        temp = self.defdata.dropna(subset=[self.def_on])
        bad = pd.merge(self.findata, temp, left_on=self.data_on, right_on=self.def_on, how='inner')
        bad['def_flag'] = 0 # set 0 first and change to 1 after cleaning

        self.merge_step0_good = good.copy()
        self.merge_step0_bad = bad.copy()


        # 1 blackout all financial statments which are in the time window 
        # ['blackout_mo_bf_def_begins' months before default begins,
        # 'blackout_mo_af_def_ends' months after default ends]      
        previousbad = bad.copy()
        list_to_remove=[]
        for row in bad.iterrows():
            if (row[1]['L_DATE_OF_DEFAULT']-rd(months=self.blackout_mo_bf_def_begins)) \
                <= row[1][self.findata_timestamp] \
                <= (row[1]['L_DEFAULT_END_DT']+rd(months=self.blackout_mo_af_def_ends)):
                list_to_remove.append(row[0])
            else:
                continue        

        bad.drop(bad.index[list_to_remove], inplace=True)

        # To get the dataframe that contains the obs which have been kicked out
        self.merge_step1_dropped = pd.concat([previousbad, bad])
        self.merge_step1_dropped.reset_index(drop=False, inplace=True)
        self.merge_step1_dropped.drop_duplicates(subset='index',keep=False,inplace=True)

        bad.reset_index(drop=True, inplace=True)
        self.merge_step1_left=bad.copy()

        # 2. pick the financial statments that lie if the 'valid time window' 
        #    (usually 6-18 months before default) 
        previousbad = bad.copy()      
        list_to_remove=[]
        for row in bad.iterrows():
            if row[1][self.findata_timestamp] >= \
            (row[1]['L_DATE_OF_DEFAULT']-rd(months=self.valid_time_window_mo[0])) or \
               row[1][self.findata_timestamp] <= \
            (row[1]['L_DATE_OF_DEFAULT']-rd(months=self.valid_time_window_mo[1])):
                list_to_remove.append(row[0])
            else:
                continue        

        good_2 = bad.iloc[list_to_remove].copy()
        good = pd.concat([good, good_2])
        bad.drop(bad.index[list_to_remove], inplace=True)       
        self.merge_step2_left=bad.copy()
        # To get the dataframe that contains the obs which have been kicked out
        self.merge_step2_dropped = pd.concat([previousbad, bad])
        self.merge_step2_dropped.reset_index(drop=False, inplace=True)
        self.merge_step2_dropped.drop_duplicates(subset='index',keep=False,inplace=True)

        # 3.  Only keep the latest financial statement in multiple lie in valid time window:        
             
        bad.sort_values(by=[self.data_on,self.findata_timestamp], inplace=True)
        bad.reset_index(drop=True, inplace=True)
        previousbad = bad.copy() 
        bad.drop_duplicates(subset=[self.data_on], keep='last', inplace=True)

        # To get the dataframe that contains the obs which have been kicked out
        self.merge_step3_dropped = pd.concat([previousbad, bad])
        self.merge_step3_dropped.reset_index(drop=False, inplace=True)
        self.merge_step3_dropped.drop_duplicates(subset='index',keep=False,inplace=True)

        bad.reset_index(drop=True, inplace=True)
        bad['def_flag'] = 1     
        self.merge_step3_left=bad.copy()
        # 4. combine good and bad to construct complete datset
        self.full_data = pd.concat([good, bad])
        self.full_data.reset_index(drop=True, inplace=True)

    def __printstarline(self):
        print("*"*80)

    def __printdoubleline(self):
        print("="*80)

    def __printblankline(self):
        print(" "*80)


    def __report(self):
        self.__printblankline()
        self.__printblankline()
        print('{:^80}'.format('Report'))
        self.__printstarline()
        self.__printblankline()
        print('The input default dataset has {:d} Obs.'.format(self.orisize_defdata))
        print('The input financial statement dataset has {:d} Obs.'.format(self.orisize_findata))
        self.__printblankline()
        self.__printdoubleline()
        print('{:*^80}'.format('Part One: Cleaing procedure for default data:'))
        self.__printblankline()
        print('{:<80}'.format('In the Step_1, {:d} NAs and duplicates Obs. in default dataset are dropped.'.format(len(self.def_step1_dropped))))
        print('{:<80}'.format('Attribute "def_step1_dropped" is the dropped data.'))   
        print('{:<80}'.format('Attribute "def_step1_left" is the processed data.'))   
        self.__printblankline()
        print('{:<80}'.format('In the Step_2, {:d} out of {:d} Obs.\' "L_DEFAULT_END_DT" are modified.'.format(self.def_step2_dropped[0],self.def_step2_dropped[1])))
        print('{:<80}'.format('Attribute "def_step2_left" is the processed data.'))  
        self.__printblankline()
        print('{:<80}'.format('In the Step_3, {:d} defaults are combined into previous defaults.'.format(self.def_step3_dropped[0]-self.def_step3_dropped[1])))
        print('{:<80}'.format('Attribute "def_step3_left" is the processed data.'))  
        self.__printblankline()
        print('{:<80}'.format('After 3 steps cleaning, {:d} defaults are obtained.'.format(self.def_step3_dropped[1])))
        self.__printblankline()
        self.__printdoubleline()
        print('{:*^80}'.format('Part Two: Merging procedure for default and financial dataset:'))
        self.__printblankline()
        print('{:<80}'.format('There are {:d} "good" Obs. and {:d} "default" Obs. in the merged data before \napplying any default logic.'.\
            format(len(self.merge_step0_good),len(self.merge_step0_bad))))
        print('{:<80}'.format('Attribute "merge_step0_good" is the non-default Obs.'))   
        print('{:<80}'.format('Attribute "merge_step0_bad" is the default Obs.'))  
        self.__printblankline()
        self.__printdoubleline()
        print('{:*^80}'.format('Part Three: Applying default logic to the merged dataset:'))
        self.__printblankline()
        print('{:<80}'.format('In the Step_1, {:d} default Obs. are dropped ({:d} are left).'.\
            format(len(self.merge_step1_dropped),len(self.merge_step0_bad)-len(self.merge_step1_dropped))))
        print('{:<80}'.format('Attribute "merge_step1_dropped" is the dropped data.')) 
        print('{:<80}'.format('Attribute "merge_step1_left" is the processed data.'))   
        self.__printblankline()
        print('{:<80}'.format('In the Step_2, {:d} default Obs. are dropped ({:d} are left).'.\
            format(len(self.merge_step2_dropped),len(self.merge_step0_bad)-len(self.merge_step1_dropped)-len(self.merge_step2_dropped))))
        print('{:<80}'.format('Attribute "merge_step2_dropped" is the dropped data.')) 
        print('{:<80}'.format('Attribute "merge_step2_left" is the processed data.'))   
        self.__printblankline()
        print('{:<80}'.format('In the Step_3, {:d} default Obs. are dropped ({:d} are left).'.\
            format(len(self.merge_step3_dropped),len(self.merge_step0_bad)-len(self.merge_step1_dropped)-len(self.merge_step2_dropped)-len(self.merge_step3_dropped))))
        print('{:<80}'.format('Attribute "merge_step3_dropped" is the dropped data.')) 
        print('{:<80}'.format('Attribute "merge_step3_left" is the processed data.'))   
        self.__printblankline()
        self.__printdoubleline()
        print('{:*^80}'.format('Summary after the procedure:'))
        self.__printblankline()
        print('{:<80}'.format('The final cleaned dataset has {:d} Obs. among which there are {:d} valid default Obs. \nThe empirical PD is {:6.4f}%'.\
            format(len(self.full_data),self.full_data.def_flag.sum(),100*self.full_data.def_flag.sum()/len(self.full_data))))       
        self.__printblankline()
        print('{:<80}'.format('Attribute "full_data" is the cleaned dataset.')) 
        self.__printblankline()
        self.__printstarline()
        print('{:<80}'.format('*print(Attachdefaults.__doc__) to check the detail for each step')) 


    def run_and_report(self):

        warnings.filterwarnings('ignore', category=Warning)
        self.__run()
        self.__report()


# -*- coding: utf-8 -*-
"""
Created on Tue Feb  7 14:46:19 2017

Version: 1.0 Initial build
Version: 1.1, 20170313, Remove all irrelevant cols in default data

@author: ub71894 (4e8e6d0b), CSG
"""


import pandas as pd
import numpy as np
from dateutil.relativedelta import relativedelta as rd
import collections
import warnings

#%%
class Attachdefaults(object):
    
    '''    
    This class is used to attach default data to financial statement and then 
    apply default logic to clean the dataset. 
    
    Attachdefaults(findata, findata_id, defdata, defdata_id, **kwargs):

        findata:    financial statement data. User need to clean and drop duplicates
                    in 'findata' before constructing the class
        findata_id: The column name for id in 'findata' and it will be used to link 
                    with default data  
        defdata:    default data. Usually coming from MasterDefault data. User 
                    need to make sure the column name for default time is 
                    'L_DATE_OF_DEFAULT' and the column name for default ending 
                    time is 'L_DEFAULT_END_DT' if it has.
        defdata_id: The column name for id in 'defdata' and it will be used to link 
                    with financial statement.  

        Parameters: 

        isthere_def_enddate:    boolean, default False
                                Whether the defaults have ending time in 'defdata'
        months_default_atleast_last:    int, default 6
                                        the floor of the default period. The unit
                                        is month.
        findata_timestamp:  str, default 'archive_date'
                            the column name of the timestamp for financial statement 
                            in 'findata'
        blackout_mo_bf_def_begins:  int, default 6
                                    When the financial statement is too close to 
                                    the actual default, it should be ignore since 
                                    it's most likely influenced. This number is 
                                    how many months before the default are considered
                                    as 'too close'. 
        blackout_mo_af_def_ends:    int, default 6
                                    When the financial statement is too close to 
                                    the actual default ending date, it should be
                                    ignore since it's most likely still under the 
                                    impact of default. This number is how many months
                                    after the default ends are considered as 'too close'. 
        valid_time_window_mo:   tuple of 2 integers, default (6,18)
                                The valid time window (months) for financial statements 
                                that happens before actual default. Usually we 
                                set 6-18 months as the valid time period. which 
                                means only financial statements occurs 6 to 18 months 
                                before the actual default will be tagged as a 
                                'default' or 'bad' observation.                   

    The procedure detail:

    Part One: Clean 'defdata':
        Step_1: Drop Obs. if its 'L_DATE_OF_DEFAULT' is NAs or duplicated 
        Step_2: Modify 'L_DEFAULT_END_DT' by the rule below:
                L_DEFAULT_END_DT = L_DEFAULT_END_DT,  if the default last longer than 6 months
                L_DEFAULT_END_DT = L_DATE_OF_DEFAULT + 6 months,  if others
        Step_3: Combine defaults for the same obligor if the second one happened within 1 year of the
                previous one or within previous default's end date.

    Part Two: Merge 'findata' and 'defdata' by the input identifiers.

    Part Three: Apply default logic to the merged data:
        Step_1: Blackout all financial statments which are in the time window 
                ['blackout_mo_bf_def_begins' months before default begins,
                blackout_mo_af_def_ends' months after default ends]    
        Step_2: Pick the financial statments that occur in the 'valid time window'. 
                Usually it's 6-18 months before actual default.
        Step_3: Only keep the latest financial statement if multiple lie in valid 
                time window. In other word, we only select the closest financial 
                statement to the actual default and remove others.

    '''

    def __init__(self, findata, findata_id, defdata, defdata_id, **kwargs):
        self.findata = findata.copy()
        self.findata.reset_index(drop=True, inplace=True)
        self.defdata = defdata.copy()
        self.defdata.reset_index(drop=True, inplace=True)
        # add the line below in Ver 1.1 to remove all irrelevant cols in def data:
        self.defdata = self.defdata[[defdata_id]+['L_DATE_OF_DEFAULT','L_DEFAULT_END_DT']]
        self.orisize_findata = len(self.findata)
        self.orisize_defdata = len(self.defdata)
        #self.orinofdef = defdata.def_flag.sum()
        self.isthere_def_enddate = kwargs.pop('isthere_def_enddate',False)
        self.months_default_atleast_last = kwargs.pop('months_default_atleast_last',6)
        self.data_on = findata_id
        self.def_on = defdata_id
        self.findata_timestamp = kwargs.pop('findata_timestamp','archive_date')
        self.blackout_mo_bf_def_begins = kwargs.pop('blackout_mo_bf_def_begins',6)
        self.blackout_mo_af_def_ends = kwargs.pop('blackout_mo_af_def_ends',6)
        self.valid_time_window_mo = kwargs.pop('valid_time_window_mo',(6,18))

    def __cleandef(self):

        if not self.isthere_def_enddate: # then create a column for end date
            self.defdata['L_DEFAULT_END_DT'] = self.defdata['L_DATE_OF_DEFAULT']
        # 1. drop duplicated defaults and keep the one which has the latest L_DEFAULT_END_DT
        previousdefdata = self.defdata.copy()
        self.defdata.dropna(subset=['L_DATE_OF_DEFAULT'], inplace=True)
        self.defdata.sort_values(by=[self.def_on,'L_DATE_OF_DEFAULT','L_DEFAULT_END_DT'], ascending=False, inplace=True)
        self.defdata.drop_duplicates(subset=[self.def_on,'L_DATE_OF_DEFAULT'], keep='first', inplace=True)
        self.defdata.sort_values(by=[self.def_on,'L_DATE_OF_DEFAULT'], inplace=True)
        # To get the dataframe that contains the obs which have been kicked out
        self.def_step1_dropped = pd.concat([previousdefdata, self.defdata])
        self.def_step1_dropped.reset_index(drop=False, inplace=True)
        self.def_step1_dropped.drop_duplicates(subset='index',keep=False,inplace=True)

        self.defdata.reset_index(drop=True, inplace=True)
        self.def_step1_left=self.defdata.copy()

        # 2. modify L_DEFAULT_END_DT by the rule below:
        # L_DEFAULT_END_DT = L_DEFAULT_END_DT            ,  if the default last longer than 6 months
        # L_DEFAULT_END_DT = L_DATE_OF_DEFAULT + 6 months,  if others
        s=0
        for row in self.defdata.iterrows():
            if pd.isnull(row[1]['L_DEFAULT_END_DT']):
                self.defdata.loc[row[0],'L_DEFAULT_END_DT'] = self.defdata.loc[row[0],'L_DATE_OF_DEFAULT'] + rd(months=self.months_default_atleast_last)       
                s+=1
            elif (row[1]['L_DEFAULT_END_DT']-row[1]['L_DATE_OF_DEFAULT']).days <= 30 * self.months_default_atleast_last:
                self.defdata.loc[row[0],'L_DEFAULT_END_DT'] = self.defdata.loc[row[0],'L_DATE_OF_DEFAULT'] + rd(months=self.months_default_atleast_last)
                s+=1
        self.def_step2_dropped = [s,len(self.defdata)]

        self.def_step2_left=self.defdata.copy()

        # 3. combine defaults if the second one happened within one year of the previous one or within previous default's end date
        self.def_step3_dropped=[len(self.defdata),0]

        list_to_remove=[];
        current_obligor = self.defdata.loc[0,self.def_on]
        current_def = self.defdata.loc[0,'L_DATE_OF_DEFAULT']
        current_def_end = self.defdata.loc[0,'L_DEFAULT_END_DT']
        current_index = 0       

        for i in range(1,len(self.defdata)):
            if self.defdata.loc[i,self.def_on] != current_obligor: 
                current_obligor = self.defdata.loc[i,self.def_on]
                current_def = self.defdata.loc[i,'L_DATE_OF_DEFAULT']
                current_def_end = self.defdata.loc[i,'L_DEFAULT_END_DT']
                current_index = i
                continue
            elif (self.defdata.loc[i,'L_DATE_OF_DEFAULT'] < current_def+rd(years=1)) or \
                (self.defdata.loc[i,'L_DATE_OF_DEFAULT'] < current_def_end):
                list_to_remove.append(i)
                current_def_end = self.defdata.loc[i,'L_DEFAULT_END_DT']
                self.defdata.loc[current_index,'L_DEFAULT_END_DT'] = current_def_end
                continue
            else:
                current_obligor = self.defdata.loc[i,self.def_on]
                current_def = self.defdata.loc[i,'L_DATE_OF_DEFAULT']
                current_def_end = self.defdata.loc[i,'L_DEFAULT_END_DT']
                current_index = i
                continue        
        

        self.defdata.drop(self.defdata.index[list_to_remove], inplace=True)
        self.def_step3_dropped[1] = len(self.defdata)
        self.def_step3_left=self.defdata.copy()


    def __run(self):

        self.__cleandef()

        alldata= pd.merge(self.findata, self.defdata, left_on=self.data_on, right_on=self.def_on, how='left')
        good = alldata.query('L_DATE_OF_DEFAULT != L_DATE_OF_DEFAULT') # nan in default date means good
        good['def_flag'] = 0
        temp = self.defdata.dropna(subset=[self.def_on])
        bad = pd.merge(self.findata, temp, left_on=self.data_on, right_on=self.def_on, how='inner')
        bad['def_flag'] = 0 # set 0 first and change to 1 after cleaning

        self.merge_step0_good = good.copy()
        self.merge_step0_bad = bad.copy()


        # 1 blackout all financial statments which are in the time window 
        # ['blackout_mo_bf_def_begins' months before default begins,
        # 'blackout_mo_af_def_ends' months after default ends]      
        previousbad = bad.copy()
        list_to_remove=[]
        for row in bad.iterrows():
            if (row[1]['L_DATE_OF_DEFAULT']-rd(months=self.blackout_mo_bf_def_begins)) \
                <= row[1][self.findata_timestamp] \
                <= (row[1]['L_DEFAULT_END_DT']+rd(months=self.blackout_mo_af_def_ends)):
                list_to_remove.append(row[0])
            else:
                continue        

        bad.drop(bad.index[list_to_remove], inplace=True)

        # To get the dataframe that contains the obs which have been kicked out
        self.merge_step1_dropped = pd.concat([previousbad, bad])
        self.merge_step1_dropped.reset_index(drop=False, inplace=True)
        self.merge_step1_dropped.drop_duplicates(subset='index',keep=False,inplace=True)

        bad.reset_index(drop=True, inplace=True)
        self.merge_step1_left=bad.copy()

        # 2. pick the financial statments that lie if the 'valid time window' 
        #    (usually 6-18 months before default) 
        previousbad = bad.copy()      
        list_to_remove=[]
        for row in bad.iterrows():
            if row[1][self.findata_timestamp] >= \
            (row[1]['L_DATE_OF_DEFAULT']-rd(months=self.valid_time_window_mo[0])) or \
               row[1][self.findata_timestamp] <= \
            (row[1]['L_DATE_OF_DEFAULT']-rd(months=self.valid_time_window_mo[1])):
                list_to_remove.append(row[0])
            else:
                continue        

        good_2 = bad.iloc[list_to_remove].copy()
        good = pd.concat([good, good_2])
        bad.drop(bad.index[list_to_remove], inplace=True)       
        self.merge_step2_left=bad.copy()
        # To get the dataframe that contains the obs which have been kicked out
        self.merge_step2_dropped = pd.concat([previousbad, bad])
        self.merge_step2_dropped.reset_index(drop=False, inplace=True)
        self.merge_step2_dropped.drop_duplicates(subset='index',keep=False,inplace=True)

        # 3.  Only keep the latest financial statement in multiple lie in valid time window:        
             
        bad.sort_values(by=[self.data_on,self.findata_timestamp], inplace=True)
        bad.reset_index(drop=True, inplace=True)
        previousbad = bad.copy() 
        bad.drop_duplicates(subset=[self.data_on], keep='last', inplace=True)

        # To get the dataframe that contains the obs which have been kicked out
        self.merge_step3_dropped = pd.concat([previousbad, bad])
        self.merge_step3_dropped.reset_index(drop=False, inplace=True)
        self.merge_step3_dropped.drop_duplicates(subset='index',keep=False,inplace=True)

        bad.reset_index(drop=True, inplace=True)
        bad['def_flag'] = 1     
        self.merge_step3_left=bad.copy()
        # 4. combine good and bad to construct complete datset
        self.full_data = pd.concat([good, bad])
        self.full_data.reset_index(drop=True, inplace=True)

    def __printstarline(self):
        print("*"*80)

    def __printdoubleline(self):
        print("="*80)

    def __printblankline(self):
        print(" "*80)


    def __report(self):
        self.__printblankline()
        self.__printblankline()
        print('{:^80}'.format('Report'))
        self.__printstarline()
        self.__printblankline()
        print('The input default dataset has {:d} Obs.'.format(self.orisize_defdata))
        print('The input financial statement dataset has {:d} Obs.'.format(self.orisize_findata))
        self.__printblankline()
        self.__printdoubleline()
        print('{:*^80}'.format('Part One: Cleaing procedure for default data:'))
        self.__printblankline()
        print('{:<80}'.format('In the Step_1, {:d} NAs and duplicates Obs. in default dataset are dropped.'.format(len(self.def_step1_dropped))))
        print('{:<80}'.format('Attribute "def_step1_dropped" is the dropped data.'))   
        print('{:<80}'.format('Attribute "def_step1_left" is the processed data.'))   
        self.__printblankline()
        print('{:<80}'.format('In the Step_2, {:d} out of {:d} Obs.\' "L_DEFAULT_END_DT" are modified.'.format(self.def_step2_dropped[0],self.def_step2_dropped[1])))
        print('{:<80}'.format('Attribute "def_step2_left" is the processed data.'))  
        self.__printblankline()
        print('{:<80}'.format('In the Step_3, {:d} defaults are combined into previous defaults.'.format(self.def_step3_dropped[0]-self.def_step3_dropped[1])))
        print('{:<80}'.format('Attribute "def_step3_left" is the processed data.'))  
        self.__printblankline()
        print('{:<80}'.format('After 3 steps cleaning, {:d} defaults are obtained.'.format(self.def_step3_dropped[1])))
        self.__printblankline()
        self.__printdoubleline()
        print('{:*^80}'.format('Part Two: Merging procedure for default and financial dataset:'))
        self.__printblankline()
        print('{:<80}'.format('There are {:d} "good" Obs. and {:d} "default" Obs. in the merged data before \napplying any default logic.'.\
            format(len(self.merge_step0_good),len(self.merge_step0_bad))))
        print('{:<80}'.format('Attribute "merge_step0_good" is the non-default Obs.'))   
        print('{:<80}'.format('Attribute "merge_step0_bad" is the default Obs.'))  
        self.__printblankline()
        self.__printdoubleline()
        print('{:*^80}'.format('Part Three: Applying default logic to the merged dataset:'))
        self.__printblankline()
        print('{:<80}'.format('In the Step_1, {:d} default Obs. are dropped ({:d} are left).'.\
            format(len(self.merge_step1_dropped),len(self.merge_step0_bad)-len(self.merge_step1_dropped))))
        print('{:<80}'.format('Attribute "merge_step1_dropped" is the dropped data.')) 
        print('{:<80}'.format('Attribute "merge_step1_left" is the processed data.'))   
        self.__printblankline()
        print('{:<80}'.format('In the Step_2, {:d} default Obs. are dropped ({:d} are left).'.\
            format(len(self.merge_step2_dropped),len(self.merge_step0_bad)-len(self.merge_step1_dropped)-len(self.merge_step2_dropped))))
        print('{:<80}'.format('Attribute "merge_step2_dropped" is the dropped data.')) 
        print('{:<80}'.format('Attribute "merge_step2_left" is the processed data.'))   
        self.__printblankline()
        print('{:<80}'.format('In the Step_3, {:d} default Obs. are dropped ({:d} are left).'.\
            format(len(self.merge_step3_dropped),len(self.merge_step0_bad)-len(self.merge_step1_dropped)-len(self.merge_step2_dropped)-len(self.merge_step3_dropped))))
        print('{:<80}'.format('Attribute "merge_step3_dropped" is the dropped data.')) 
        print('{:<80}'.format('Attribute "merge_step3_left" is the processed data.'))   
        self.__printblankline()
        self.__printdoubleline()
        print('{:*^80}'.format('Summary after the procedure:'))
        self.__printblankline()
        print('{:<80}'.format('The final cleaned dataset has {:d} Obs. among which there are {:d} valid default Obs. \nThe empirical PD is {:6.4f}%'.\
            format(len(self.full_data),self.full_data.def_flag.sum(),100*self.full_data.def_flag.sum()/len(self.full_data))))       
        self.__printblankline()
        print('{:<80}'.format('Attribute "full_data" is the cleaned dataset.')) 
        self.__printblankline()
        self.__printstarline()
        print('{:<80}'.format('*print(Attachdefaults.__doc__) to check the detail for each step')) 


    def run_and_report(self):

        warnings.filterwarnings('ignore', category=Warning)
        self.__run()
        self.__report()


# -*- coding: utf-8 -*-
"""
Created on Tue Apr  4 10:21:43 2017

@author: ub64283 Liwen Zhang (original author: XU79799)
@author: ub71894 (4e8e6d0b), Add TMnotches(), withknotchesrate() and TMstats()

Version: 1.1: 20191101, modify code to fit new openpyxl function

"""
import pandas as pd
import numpy as np
import openpyxl
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.styles import Border, Side, Alignment, Font
from openpyxl.formatting.rule import IconSetRule
from openpyxl import  Workbook, load_workbook
import scipy.stats as stats
from openpyxl.utils.cell import coordinate_from_string, column_index_from_string


def range_to_index(range_in_str): # New Ver. 1.1. The best patch I have ever developed...

    delimiter = range_in_str.find(':')
    min_cell = range_in_str[:delimiter]
    max_cell = range_in_str[(delimiter+1):]

    xy = coordinate_from_string(min_cell) 
    min_col = column_index_from_string(xy[0]) 
    min_row = xy[1]
    xy = coordinate_from_string(max_cell) 
    max_col = column_index_from_string(xy[0]) 
    max_row = xy[1]

    return({'min_col':min_col, 'max_col':max_col, 'min_row':min_row, 'max_row':max_row})




#Count the number of obligors which have same rating with a margin of y    
def within(MM,y):
    """
    Summary:
    Counts the number of obligors which have the same rating within a margin of y

    Variables:
    MM: type = DataFrame. A dataframe which contains the transition matrix of two
    ratings systems.

    y: type = int. The margin of ratings that will contribute to the final sum.

    Returns:
    An integer counting the number of obligors with the same rating within a margin of y
    """
    x=np.array(MM)
    N=len(x)
    sum1=0
    if y == 0:
        return np.sum(np.diag(x))
    else:
        for i in range(N):
            ll=max(0,i-y) #checks the lower limit of the margin
            uu=min(N+1,i+y+1) #checks the upper limit of the margin
            sum1=sum1+np.sum(x[ll:uu,i]) #for each cell, adds the number that falls within the margin
        return(sum1)

def down_notch(MM):
    N=MM.shape[0]
    sum1 = 0
    for notch in range(1,N):
        temp = 0
        for i in range(notch):
            temp += MM[i,(N-notch+i)]
        sum1 += temp*(N-notch)
    return(sum1)
        
def worse_ratings(MM):
    x=np.array(MM)
    N=len(x)
    sum1=0
    for i in range(N):
        ll=0
        sum1=sum1+np.sum(x[ll:i,i])
    return(sum1)

def up_notch(MM):
    N=MM.shape[0]
    sum1=0
    for notch in range(1,N):
        temp=0
        for i in range(notch):
            temp += MM[(N-notch+i),i]
        sum1 += temp*(N-notch)
    return(sum1)
    
def better_ratings(MM):
    x=np.array(MM)
    N=len(x)
    sum1=0
    for i in range(N):
        uu=N+1
        sum1=sum1+np.sum(x[(i+1):uu,i])
    return(sum1)

def ComputeListOutstandingObligors(df):
    res=[]
    
#    for i in range(df.shape[0]):
#       if (abs(df['Old_Ratings'].iloc[i]-df['New_Ratings_oldms'].iloc[i])>=4):
#           res.append(df['CUSTOMERID'].iloc[i])
    return res

def CreateStyleSheet_TransitionMatrix(sheet, rowshift, colshift, PDRR):
    
    ### assume len(PDRR) <= 21 ###
    
    len_pdrr = len(PDRR)
    
    #### Work on style ###
    for i in range(rowshift-3,rowshift+len_pdrr):
        sheet.row_dimensions[i].height=12
    for j in ['AB','AC','AD']:
        sheet.column_dimensions[j].width=7
    sheet.column_dimensions['B'].width=3
    sheet.column_dimensions['AA'].width=9
    sheet.column_dimensions['AG'].width=9
    sheet.column_dimensions['AH'].width=5
    sheet.column_dimensions['AE'].width=8
    sheet.column_dimensions['AF'].width=8

    for _row in sheet.iter_rows(**range_to_index('B'+ str(rowshift-2) + ':' + chr(68+len_pdrr) + str(rowshift-1))):
        for _cell in _row:
            fill=PatternFill(start_color='FFFFFF', end_color='FFFFFF',fill_type='solid')
            _cell.fill=fill 
    for _row in sheet.iter_rows(**range_to_index('B'+ str(rowshift-2) + ':C' + str(rowshift+len_pdrr))):
        for _cell in _row:
            fill=PatternFill(start_color='FFFFFF', end_color='FFFFFF',fill_type='solid')
            _cell.fill=fill 
    for _row in sheet.iter_rows(**range_to_index(chr(68+len_pdrr) + str(rowshift-2)+ ':' + chr(68+len_pdrr) + str(rowshift+len_pdrr))):
        for _cell in _row:
            fill=PatternFill(start_color='FFFFFF', end_color='FFFFFF',fill_type='solid')
            _cell.fill=fill 
    
    for _row in sheet.iter_rows(**range_to_index('C' + str(rowshift-1)+ ':' + chr(68+len_pdrr) + str(rowshift+len_pdrr))):
        for _cell in _row:
            _cell.alignment=Alignment(horizontal='center',vertical='center')
            _cell.font=Font(size=9)
 
    #for colNum in range(15):
    for colNum in range(len_pdrr):
         col1=colshift+colNum
         sheet.column_dimensions[get_column_letter(col1)].width = 5   
   
   #Format of percentage/notch table
    for _row in sheet.iter_rows(**range_to_index('V'+ str(rowshift-1) + ':A' + chr(65+2) + str(rowshift+1))): 
       for _cell in _row:
           _cell.font=Font(size=9, bold=True)
           _cell.alignment=Alignment(vertical='center', horizontal='center')
           fill=PatternFill(start_color='FFFFFF', end_color='FFFFFF',fill_type='solid')
           _cell.fill=fill 
           
    #Add borders for percentage/notch table:
    side=Side(style='thin', color="FF000000")
    for _row in sheet.iter_rows(**range_to_index('V' + str(rowshift-1) + ':A' + chr(65+2) + str(rowshift-1))): #top border
       for _cell in _row:
           _cell.border=Border(top=side) 
    for _row in sheet.iter_rows(**range_to_index('V' + str(rowshift) + ':A' + chr(65+2) + str(rowshift+1))): #bottom border
       for _cell in _row:
           _cell.border=Border(bottom=side)
    for _row in sheet.iter_rows(**range_to_index('V' + str(rowshift+1) + ':A' + chr(65+2) + str(rowshift+1))): #another bottom border
       for _cell in _row:
           _cell.border=Border(bottom=side)
    sheet['V'+str(rowshift-1)].border=Border(left=side, right=side, top=side)
    sheet['V'+str(rowshift)].border=Border(left=side, right=side)         
    sheet['V'+str(rowshift+1)].border=Border(left=side, right=side, top=side, bottom=side)
    sheet['Z'+str(rowshift-1)].border=Border(right=side, top=side)
    sheet['Z'+str(rowshift)].border=Border(right=side, bottom=side)       
    sheet['Z'+str(rowshift+1)].border=Border(right=side, bottom=side, top=side)   
    sheet['A'+chr(65+2)+str(rowshift-1)].border=Border(right=side, left=side, top=side)
    sheet['A'+chr(65+2)+str(rowshift)].border=Border(right=side, left=side) 
    sheet['A'+chr(65+2)+str(rowshift+1)].border=Border(right=side, left=side, top=side, bottom=side)
    
    ### Do formatting ###
    for rowNum in PDRR:
        rowNum=rowNum-1
        for colNum in range(len(PDRR)): 
            row1=rowshift+rowNum
            col1=colshift+colNum
            if rowNum==colNum:
                fill=PatternFill(start_color='FFFF00', end_color='FFFF00',fill_type='solid')
                sheet.cell(row=row1,column=col1).fill=fill
            elif ((abs(rowNum-colNum) == 1) | (abs(rowNum-colNum)== 2)):
                fill=PatternFill(start_color='FFFFCC', end_color='FFFFCC',fill_type='solid')
                sheet.cell(row=row1,column=col1).fill=fill
            elif abs(rowNum-colNum) == 3:
                fill=PatternFill(start_color='FCD5B4', end_color='FCD5B4',fill_type='solid')
                sheet.cell(row=row1,column=col1).fill=fill
            elif ((abs(rowNum-colNum) > 3) & (not(sheet.cell(row=row1,column=col1).value==None))): 
                fill=PatternFill(start_color='E4DFEC', end_color='E4DFEC',fill_type='solid')
                sheet.cell(row=row1,column=col1).fill=fill 
            else:
                fill=PatternFill(start_color='FFFFFF', end_color='FFFFFF',fill_type='solid')
                sheet.cell(row=row1,column=col1).fill=fill     
        
    #Title format        
    sheet.merge_cells('B'+str(rowshift-3)+':'+chr(67+len_pdrr)+str(rowshift-3))
    sheet.merge_cells('D'+str(rowshift-2)+ ':'+chr(67+len_pdrr)+str(rowshift-2))
    sheet.merge_cells('B'+str(rowshift)+ ':B'+str(rowshift+len_pdrr-1))
    sheet['B'+str(rowshift)].alignment=Alignment(horizontal='center',text_rotation=90, vertical='center')
    sheet['B'+str(rowshift-3)].alignment=Alignment(horizontal='center', vertical='center')
    sheet['D'+str(rowshift-2)].alignment=Alignment(horizontal='center', vertical='center')
    sheet['B'+str(rowshift-3)].font=Font(bold=True)
    sheet['D'+str(rowshift-2)].font=Font(italic=True)
    sheet['B'+str(rowshift)].font=Font(italic=True)
    
    #Add borders
    for _row in sheet.iter_rows(**range_to_index('B'+ str(rowshift-3) + ':' + chr(67+len_pdrr) + str(rowshift-3))): #top border
       for _cell in _row:
           _cell.border=Border(top=side)
    for _row in sheet.iter_rows(**range_to_index('B'+ str(rowshift) + ':' +chr(67+len_pdrr) + str(rowshift))): #top border 2
       for _cell in _row:
           _cell.border=Border(top=side)
    for i in ['B'+ str(rowshift-2),'B' + str(rowshift-1)]: #left border
        a = sheet[i]
        a.border=Border(left=side)
    for _row in sheet.iter_rows(**range_to_index('B' + str(rowshift) + ':B' + str(rowshift+len_pdrr-1))): #left border
       for _cell in _row:
           _cell.border=Border(left=side, right=side)
    for _row in sheet.iter_rows(**range_to_index('C'+ str(rowshift+len_pdrr-1) + ':' + chr(67+len_pdrr) + str(rowshift+len_pdrr-1))): #bottom border
       for _cell in _row:
           _cell.border=Border(bottom=side)
    for _row in sheet.iter_rows(**range_to_index(chr(67+len_pdrr) + str(rowshift) + ':' + chr(67+len_pdrr) + str(rowshift+len_pdrr-1))): #right border
       for _cell in _row:
           _cell.border=Border(right=side)
    sheet['B'+str(rowshift-3)].border=Border(left=side, top=side)
    sheet['B'+str(rowshift+len_pdrr-1)].border=Border(left=side, bottom=side, right=side)
    sheet[chr(67+len_pdrr)+str(rowshift+len_pdrr-1)].border=Border(bottom=side, right=side)
    sheet[chr(67+len_pdrr)+str(rowshift-1)].border=Border(right=side, bottom=side)
    sheet[chr(67+len_pdrr)+str(rowshift-2)].border=Border(right=side)
    sheet[chr(67+len_pdrr)+str(rowshift-3)].border=Border(right=side, top=side)
    
#    for j in ['D','E','F','G','H','I','J','K','L','M','N','O','P','Q','R']:
#        sheet.column_dimensions[j].width=4   
                               
    for j in range(68, 67+len_pdrr):
        if j <= 90:
            sheet.column_dimensions[chr(j)].width=5
        if j >= 91:
            sheet.column_dimensions['A' + chr(j-26)].width=5
    
    sheet.column_dimensions[chr(68+len_pdrr)].width=6
    sheet.column_dimensions['C'].width=6


##### ##### ##### ##### ##### ##### #####
##### #####   Main Function   ##### #####
##### ##### ##### ##### ##### ##### #####

# This function Creates a sheet to compare the two different rating
# Output: will add a sheet in the output file 
def CreateBenchmarkMatrix(Data, output_filename, sheetname, xaxis, yaxis, PDRR):
     
    #add 'try' on 08/23/2017 to make it work on existing file
    try:
        wb = load_workbook(filename = output_filename)
        sheet = wb.create_sheet() 
        sheet.title = sheetname

    except FileNotFoundError:
        wb =  Workbook()
        sheet = wb.create_sheet() 
        sheet.title = sheetname
        wb.remove_sheet(wb.get_sheet_by_name('Sheet')) 

   
    colshift = 4 #not changeable
    rowshift = 5 #changeable
    len_pdrr = len(PDRR)
    
    #### Create table ### 
    
    #label the vertical axis
    for rowNum in range(len_pdrr):
        row1 = rowNum+rowshift
        sheet.cell(row=row1,column=3).value=PDRR[rowNum]
    
    #label the horizontal axis
    for colNum in range(len_pdrr):
        col1 = colNum+colshift
        sheet.cell(row=rowshift-1,column=col1).value=PDRR[colNum]
    
    sheet.cell(row=rowshift-3,column=colshift-2).value = 'Benchmarking Matrix'
    sheet.cell(row=rowshift-2,column=colshift).value = xaxis
    sheet.cell(row=rowshift,column=colshift-2).value = yaxis

    ### Fill up table ###
    aa = np.zeros((len_pdrr,len_pdrr))
    for rowNum in range(len_pdrr):
        for colNum in range(len_pdrr):
            row1=rowshift+rowNum
            col1=colshift+colNum
            temp = sum((Data[yaxis]==PDRR[rowNum]) & (Data[xaxis]==PDRR[colNum]))
            temp=int(temp)            
            if ((not(temp==0))):        
                sheet.cell(row=row1,column=col1).value= temp
                aa[rowNum,colNum]=temp                  
    
    ### Add list of oustanding obligors
#    sheet.cell(row=rowshift+6,column=27).value='Old Prelim PD'
#    sheet.cell(row=rowshift+5,column=27).value='New Prelim PD'        
#    sheet.cell(row=rowshift+4,column=27).value='Outstanding Obligors'
#    sheet.cell(row=rowshift+4,column=colshift+len_pdrr+2).font=Font(bold=True)
#    list_outstanding=ComputeListOutstandingObligors(Data)
#    for j in range(len(list_outstanding)):
#        sheet.cell(row=rowshift+5+(j % 7),column=colshift+len_pdrr+2+j//7).value=int(list_outstanding[j])  
    
    ### Create percentage Table ###
    sheet.cell(row=rowshift-1,column=len_pdrr+8).value='Match'
    sheet.cell(row=rowshift-1,column=len_pdrr+9).value='Within 1'
    sheet.cell(row=rowshift-1,column=len_pdrr+10).value='Within 2'
    sheet.cell(row=rowshift-1,column=len_pdrr+11).value='Outside 5'
    sheet.cell(row=rowshift-1,column=len_pdrr+12).value='Downgrade'
    sheet.cell(row=rowshift-1,column=len_pdrr+13).value='Upgrade'
    sheet.cell(row=rowshift-1,column=len_pdrr+14).value='Total'
    sheet.cell(row=rowshift,column=len_pdrr+7).value='Percentage'
    sheet.cell(row=rowshift+1,column=len_pdrr+7).value='Notch'
    sheet.cell(row=rowshift,column=len_pdrr+8).value=within(aa,0)/np.sum(aa) # Match
    sheet.cell(row=rowshift,column=len_pdrr+8).number_format='.0%'
    sheet.cell(row=rowshift,column=len_pdrr+9).value=within(aa,1)/np.sum(aa) # Within 1
    sheet.cell(row=rowshift,column=len_pdrr+9).number_format='.0%'
    sheet.cell(row=rowshift,column=len_pdrr+10).value=within(aa,2)/np.sum(aa) # Within 2
    sheet.cell(row=rowshift,column=len_pdrr+10).number_format='.0%'
    sheet.cell(row=rowshift,column=len_pdrr+11).value=1-within(aa,4)/np.sum(aa) # Outside 5
    sheet.cell(row=rowshift,column=len_pdrr+11).number_format='.0%'
    sheet.cell(row=rowshift,column=len_pdrr+12).value=worse_ratings(aa)/np.sum(aa) # Downgrade
    sheet.cell(row=rowshift,column=len_pdrr+12).number_format='.0%'
    down_measure = down_notch(aa)
    sheet.cell(row=rowshift+1,column=len_pdrr+12).value=down_measure
    sheet.cell(row=rowshift,column=len_pdrr+13).value=better_ratings(aa)/np.sum(aa) # Upgrade
    sheet.cell(row=rowshift,column=len_pdrr+13).number_format='.0%'
    up_measure = up_notch(aa)
    sheet.cell(row=rowshift+1,column=len_pdrr+13).value=up_measure
    sheet.cell(row=rowshift,column=len_pdrr+14).value=np.sum(aa)
    sheet.cell(row=rowshift+1,column=len_pdrr+14).value=down_measure + up_measure
    
    CreateStyleSheet_TransitionMatrix(sheet, rowshift, colshift, PDRR)
    
    ### Add total per rows and columns
    sheet.cell(row=rowshift-1,column=colshift+len_pdrr).value = 'Total'
    for i in range(len_pdrr):
        row1=rowshift+i
        sheet.cell(row=row1,column=colshift+len_pdrr).value = sum(aa[i,:])
    sheet.cell(row=rowshift+len_pdrr,column=colshift-1).value='Total'
    for i in range(len_pdrr):
        col1=colshift+i
        sheet.cell(row=rowshift+len_pdrr,column=col1).value = sum(aa[:,i])
    sheet.cell(row=rowshift+len_pdrr,column=colshift+len_pdrr).value = np.sum(aa)    
        
    wb.save(output_filename)


def TMnotches(Data, xaxis, yaxis, PDRR):
     
    len_pdrr = len(PDRR)
    
    ### Fill up table ###
    aa = np.zeros((len_pdrr,len_pdrr))
    for rowNum in range(len_pdrr):
        for colNum in range(len_pdrr):
            temp = sum((Data[yaxis]==PDRR[rowNum]) & (Data[xaxis]==PDRR[colNum]))
            temp=int(temp)            
            if ((not(temp==0))):        
                aa[rowNum,colNum]=temp                  
    
    return((down_notch(aa),up_notch(aa)))


def withknotchesrate(Data, xaxis, yaxis, PDRR, k=2):
     
    len_pdrr = len(PDRR)
    
    ### Fill up table ###
    aa = np.zeros((len_pdrr,len_pdrr))
    for rowNum in range(len_pdrr):
        for colNum in range(len_pdrr):
            temp = sum((Data[yaxis]==PDRR[rowNum]) & (Data[xaxis]==PDRR[colNum]))
            temp=int(temp)            
            if ((not(temp==0))):        
                aa[rowNum,colNum]=temp                  
    
    return(within(aa,k)/np.sum(aa))


def TMstats(Data, xaxis, yaxis, PDRR):
     
    len_pdrr = len(PDRR)
    
    ### Fill up table ###
    aa = np.zeros((len_pdrr,len_pdrr))
    for rowNum in range(len_pdrr):
        for colNum in range(len_pdrr):
            temp = sum((Data[yaxis]==PDRR[rowNum]) & (Data[xaxis]==PDRR[colNum]))
            temp=int(temp)            
            if ((not(temp==0))):        
                aa[rowNum,colNum]=temp                  
    
    stats = {
    'Match':within(aa,0)/np.sum(aa),
    'Within_1':within(aa,1)/np.sum(aa),
    'Within_2':within(aa,2)/np.sum(aa),
    'Outside_5':1-within(aa,4)/np.sum(aa),
    'Downgrade':worse_ratings(aa)/np.sum(aa),
    'Upgrade':better_ratings(aa)/np.sum(aa),
    #'CC_3': aa[2,:].sum(),
    #'CC_3-4': aa[2:4,:].sum(),
    #'CC_3-5': aa[2:5,:].sum(),
    #'CC_3-6': aa[2:6,:].sum(),
    }
    
    return(stats)
    # -*- coding: utf-8 -*-
"""
Created on Tue Apr  4 10:21:43 2017

@author: ub64283 Liwen Zhang (original author: XU79799)
@author: ub71894 (4e8e6d0b), Add TMnotches(), withknotchesrate() and TMstats()
"""
import pandas as pd
import numpy as np
import openpyxl
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.styles import Border, Side, Alignment, Font
from openpyxl.formatting.rule import IconSetRule
from openpyxl import  Workbook, load_workbook
import scipy.stats as stats

#Count the number of obligors which have same rating with a margin of y    
def within(MM,y):
    """
    Summary:
    Counts the number of obligors which have the same rating within a margin of y

    Variables:
    MM: type = DataFrame. A dataframe which contains the transition matrix of two
    ratings systems.

    y: type = int. The margin of ratings that will contribute to the final sum.

    Returns:
    An integer counting the number of obligors with the same rating within a margin of y
    """
    x=np.array(MM)
    N=len(x)
    sum1=0
    if y == 0:
        return np.sum(np.diag(x))
    else:
        for i in range(N):
            ll=max(0,i-y) #checks the lower limit of the margin
            uu=min(N+1,i+y+1) #checks the upper limit of the margin
            sum1=sum1+np.sum(x[ll:uu,i]) #for each cell, adds the number that falls within the margin
        return(sum1)

def down_notch(MM):
    N=MM.shape[0]
    sum1 = 0
    for notch in range(1,N):
        temp = 0
        for i in range(notch):
            temp += MM[i,(N-notch+i)]
        sum1 += temp*(N-notch)
    return(sum1)
        
def worse_ratings(MM):
    x=np.array(MM)
    N=len(x)
    sum1=0
    for i in range(N):
        ll=0
        sum1=sum1+np.sum(x[ll:i,i])
    return(sum1)

def up_notch(MM):
    N=MM.shape[0]
    sum1=0
    for notch in range(1,N):
        temp=0
        for i in range(notch):
            temp += MM[(N-notch+i),i]
        sum1 += temp*(N-notch)
    return(sum1)
    
def better_ratings(MM):
    x=np.array(MM)
    N=len(x)
    sum1=0
    for i in range(N):
        uu=N+1
        sum1=sum1+np.sum(x[(i+1):uu,i])
    return(sum1)

def ComputeListOutstandingObligors(df):
    res=[]
    
#    for i in range(df.shape[0]):
#       if (abs(df['Old_Ratings'].iloc[i]-df['New_Ratings_oldms'].iloc[i])>=4):
#           res.append(df['CUSTOMERID'].iloc[i])
    return res

def CreateStyleSheet_TransitionMatrix(sheet, rowshift, colshift, PDRR):
    
    ### assume len(PDRR) <= 21 ###
    
    len_pdrr = len(PDRR)
    
    #### Work on style ###
    for i in range(rowshift-3,rowshift+len_pdrr):
        sheet.row_dimensions[i].height=12
    for j in ['AB','AC','AD']:
        sheet.column_dimensions[j].width=7
    sheet.column_dimensions['B'].width=3
    sheet.column_dimensions['AA'].width=9
    sheet.column_dimensions['AG'].width=9
    sheet.column_dimensions['AH'].width=5
    sheet.column_dimensions['AE'].width=8
    sheet.column_dimensions['AF'].width=8

    for _row in sheet.iter_rows('B'+ str(rowshift-2) + ':' + chr(68+len_pdrr) + str(rowshift-1)):
        for _cell in _row:
            fill=PatternFill(start_color='FFFFFF', end_color='FFFFFF',fill_type='solid')
            _cell.fill=fill 
    for _row in sheet.iter_rows('B'+ str(rowshift-2) + ':C' + str(rowshift+len_pdrr)):
        for _cell in _row:
            fill=PatternFill(start_color='FFFFFF', end_color='FFFFFF',fill_type='solid')
            _cell.fill=fill 
    for _row in sheet.iter_rows(chr(68+len_pdrr) + str(rowshift-2)+ ':' + chr(68+len_pdrr) + str(rowshift+len_pdrr)):
        for _cell in _row:
            fill=PatternFill(start_color='FFFFFF', end_color='FFFFFF',fill_type='solid')
            _cell.fill=fill 
    
    for _row in sheet.iter_rows('C' + str(rowshift-1)+ ':' + chr(68+len_pdrr) + str(rowshift+len_pdrr)):
        for _cell in _row:
            _cell.alignment=Alignment(horizontal='center',vertical='center')
            _cell.font=Font(size=9)
 
    #for colNum in range(15):
    for colNum in range(len_pdrr):
         col1=colshift+colNum
         sheet.column_dimensions[get_column_letter(col1)].width = 5   
   
   #Format of percentage/notch table
    for _row in sheet.iter_rows('V'+ str(rowshift-1) + ':A' + chr(65+2) + str(rowshift+1)): 
       for _cell in _row:
           _cell.font=Font(size=9, bold=True)
           _cell.alignment=Alignment(vertical='center', horizontal='center')
           fill=PatternFill(start_color='FFFFFF', end_color='FFFFFF',fill_type='solid')
           _cell.fill=fill 
           
    #Add borders for percentage/notch table:
    side=Side(style='thin', color="FF000000")
    for _row in sheet.iter_rows('V' + str(rowshift-1) + ':A' + chr(65+2) + str(rowshift-1)): #top border
       for _cell in _row:
           _cell.border=Border(top=side) 
    for _row in sheet.iter_rows('V' + str(rowshift) + ':A' + chr(65+2) + str(rowshift+1)): #bottom border
       for _cell in _row:
           _cell.border=Border(bottom=side)
    for _row in sheet.iter_rows('V' + str(rowshift+1) + ':A' + chr(65+2) + str(rowshift+1)): #another bottom border
       for _cell in _row:
           _cell.border=Border(bottom=side)
    sheet['V'+str(rowshift-1)].border=Border(left=side, right=side, top=side)
    sheet['V'+str(rowshift)].border=Border(left=side, right=side)         
    sheet['V'+str(rowshift+1)].border=Border(left=side, right=side, top=side, bottom=side)
    sheet['Z'+str(rowshift-1)].border=Border(right=side, top=side)
    sheet['Z'+str(rowshift)].border=Border(right=side, bottom=side)       
    sheet['Z'+str(rowshift+1)].border=Border(right=side, bottom=side, top=side)   
    sheet['A'+chr(65+2)+str(rowshift-1)].border=Border(right=side, left=side, top=side)
    sheet['A'+chr(65+2)+str(rowshift)].border=Border(right=side, left=side) 
    sheet['A'+chr(65+2)+str(rowshift+1)].border=Border(right=side, left=side, top=side, bottom=side)
    
    ### Do formatting ###
    for rowNum in PDRR:
        rowNum=rowNum-1
        for colNum in range(len(PDRR)): 
            row1=rowshift+rowNum
            col1=colshift+colNum
            if rowNum==colNum:
                fill=PatternFill(start_color='FFFF00', end_color='FFFF00',fill_type='solid')
                sheet.cell(row=row1,column=col1).fill=fill
            elif ((abs(rowNum-colNum) == 1) | (abs(rowNum-colNum)== 2)):
                fill=PatternFill(start_color='FFFFCC', end_color='FFFFCC',fill_type='solid')
                sheet.cell(row=row1,column=col1).fill=fill
            elif abs(rowNum-colNum) == 3:
                fill=PatternFill(start_color='FCD5B4', end_color='FCD5B4',fill_type='solid')
                sheet.cell(row=row1,column=col1).fill=fill
            elif ((abs(rowNum-colNum) > 3) & (not(sheet.cell(row=row1,column=col1).value==None))): 
                fill=PatternFill(start_color='E4DFEC', end_color='E4DFEC',fill_type='solid')
                sheet.cell(row=row1,column=col1).fill=fill 
            else:
                fill=PatternFill(start_color='FFFFFF', end_color='FFFFFF',fill_type='solid')
                sheet.cell(row=row1,column=col1).fill=fill     
        
    #Title format        
    sheet.merge_cells('B'+str(rowshift-3)+':'+chr(67+len_pdrr)+str(rowshift-3))
    sheet.merge_cells('D'+str(rowshift-2)+ ':'+chr(67+len_pdrr)+str(rowshift-2))
    sheet.merge_cells('B'+str(rowshift)+ ':B'+str(rowshift+len_pdrr-1))
    sheet['B'+str(rowshift)].alignment=Alignment(horizontal='center',text_rotation=90, vertical='center')
    sheet['B'+str(rowshift-3)].alignment=Alignment(horizontal='center', vertical='center')
    sheet['D'+str(rowshift-2)].alignment=Alignment(horizontal='center', vertical='center')
    sheet['B'+str(rowshift-3)].font=Font(bold=True)
    sheet['D'+str(rowshift-2)].font=Font(italic=True)
    sheet['B'+str(rowshift)].font=Font(italic=True)
    
    #Add borders
    for _row in sheet.iter_rows('B'+ str(rowshift-3) + ':' + chr(67+len_pdrr) + str(rowshift-3)): #top border
       for _cell in _row:
           _cell.border=Border(top=side)
    for _row in sheet.iter_rows('B'+ str(rowshift) + ':' +chr(67+len_pdrr) + str(rowshift)): #top border 2
       for _cell in _row:
           _cell.border=Border(top=side)
    for i in ['B'+ str(rowshift-2),'B' + str(rowshift-1)]: #left border
        a = sheet[i]
        a.border=Border(left=side)
    for _row in sheet.iter_rows('B' + str(rowshift) + ':B' + str(rowshift+len_pdrr-1)): #left border
       for _cell in _row:
           _cell.border=Border(left=side, right=side)
    for _row in sheet.iter_rows('C'+ str(rowshift+len_pdrr-1) + ':' + chr(67+len_pdrr) + str(rowshift+len_pdrr-1)): #bottom border
       for _cell in _row:
           _cell.border=Border(bottom=side)
    for _row in sheet.iter_rows(chr(67+len_pdrr) + str(rowshift) + ':' + chr(67+len_pdrr) + str(rowshift+len_pdrr-1)): #right border
       for _cell in _row:
           _cell.border=Border(right=side)
    sheet['B'+str(rowshift-3)].border=Border(left=side, top=side)
    sheet['B'+str(rowshift+len_pdrr-1)].border=Border(left=side, bottom=side, right=side)
    sheet[chr(67+len_pdrr)+str(rowshift+len_pdrr-1)].border=Border(bottom=side, right=side)
    sheet[chr(67+len_pdrr)+str(rowshift-1)].border=Border(right=side, bottom=side)
    sheet[chr(67+len_pdrr)+str(rowshift-2)].border=Border(right=side)
    sheet[chr(67+len_pdrr)+str(rowshift-3)].border=Border(right=side, top=side)
    
#    for j in ['D','E','F','G','H','I','J','K','L','M','N','O','P','Q','R']:
#        sheet.column_dimensions[j].width=4   
                               
    for j in range(68, 67+len_pdrr):
        if j <= 90:
            sheet.column_dimensions[chr(j)].width=5
        if j >= 91:
            sheet.column_dimensions['A' + chr(j-26)].width=5
    
    sheet.column_dimensions[chr(68+len_pdrr)].width=6
    sheet.column_dimensions['C'].width=6


##### ##### ##### ##### ##### ##### #####
##### #####   Main Function   ##### #####
##### ##### ##### ##### ##### ##### #####

# This function Creates a sheet to compare the two different rating
# Output: will add a sheet in the output file 
def CreateBenchmarkMatrix(Data, output_filename, sheetname, xaxis, yaxis, PDRR):
     
    #add 'try' on 08/23/2017 to make it work on existing file
    try:
        wb = load_workbook(filename = output_filename)
        sheet = wb.create_sheet() 
        sheet.title = sheetname

    except FileNotFoundError:
        wb =  Workbook()
        sheet = wb.create_sheet() 
        sheet.title = sheetname
        wb.remove_sheet(wb.get_sheet_by_name('Sheet')) 

   
    colshift = 4 #not changeable
    rowshift = 5 #changeable
    len_pdrr = len(PDRR)
    
    #### Create table ### 
    
    #label the vertical axis
    for rowNum in range(len_pdrr):
        row1 = rowNum+rowshift
        sheet.cell(row=row1,column=3).value=PDRR[rowNum]
    
    #label the horizontal axis
    for colNum in range(len_pdrr):
        col1 = colNum+colshift
        sheet.cell(row=rowshift-1,column=col1).value=PDRR[colNum]
    
    sheet.cell(row=rowshift-3,column=colshift-2).value = 'Benchmarking Matrix'
    sheet.cell(row=rowshift-2,column=colshift).value = xaxis
    sheet.cell(row=rowshift,column=colshift-2).value = yaxis

    ### Fill up table ###
    aa = np.zeros((len_pdrr,len_pdrr))
    for rowNum in range(len_pdrr):
        for colNum in range(len_pdrr):
            row1=rowshift+rowNum
            col1=colshift+colNum
            temp = sum((Data[yaxis]==PDRR[rowNum]) & (Data[xaxis]==PDRR[colNum]))
            temp=int(temp)            
            if ((not(temp==0))):        
                sheet.cell(row=row1,column=col1).value= temp
                aa[rowNum,colNum]=temp                  
    
    ### Add list of oustanding obligors
#    sheet.cell(row=rowshift+6,column=27).value='Old Prelim PD'
#    sheet.cell(row=rowshift+5,column=27).value='New Prelim PD'        
#    sheet.cell(row=rowshift+4,column=27).value='Outstanding Obligors'
#    sheet.cell(row=rowshift+4,column=colshift+len_pdrr+2).font=Font(bold=True)
#    list_outstanding=ComputeListOutstandingObligors(Data)
#    for j in range(len(list_outstanding)):
#        sheet.cell(row=rowshift+5+(j % 7),column=colshift+len_pdrr+2+j//7).value=int(list_outstanding[j])  
    
    ### Create percentage Table ###
    sheet.cell(row=rowshift-1,column=len_pdrr+8).value='Match'
    sheet.cell(row=rowshift-1,column=len_pdrr+9).value='Within 1'
    sheet.cell(row=rowshift-1,column=len_pdrr+10).value='Within 2'
    sheet.cell(row=rowshift-1,column=len_pdrr+11).value='Outside 5'
    sheet.cell(row=rowshift-1,column=len_pdrr+12).value='Downgrade'
    sheet.cell(row=rowshift-1,column=len_pdrr+13).value='Upgrade'
    sheet.cell(row=rowshift-1,column=len_pdrr+14).value='Total'
    sheet.cell(row=rowshift,column=len_pdrr+7).value='Percentage'
    sheet.cell(row=rowshift+1,column=len_pdrr+7).value='Notch'
    sheet.cell(row=rowshift,column=len_pdrr+8).value=within(aa,0)/np.sum(aa) # Match
    sheet.cell(row=rowshift,column=len_pdrr+8).number_format='.0%'
    sheet.cell(row=rowshift,column=len_pdrr+9).value=within(aa,1)/np.sum(aa) # Within 1
    sheet.cell(row=rowshift,column=len_pdrr+9).number_format='.0%'
    sheet.cell(row=rowshift,column=len_pdrr+10).value=within(aa,2)/np.sum(aa) # Within 2
    sheet.cell(row=rowshift,column=len_pdrr+10).number_format='.0%'
    sheet.cell(row=rowshift,column=len_pdrr+11).value=1-within(aa,4)/np.sum(aa) # Outside 5
    sheet.cell(row=rowshift,column=len_pdrr+11).number_format='.0%'
    sheet.cell(row=rowshift,column=len_pdrr+12).value=worse_ratings(aa)/np.sum(aa) # Downgrade
    sheet.cell(row=rowshift,column=len_pdrr+12).number_format='.0%'
    down_measure = down_notch(aa)
    sheet.cell(row=rowshift+1,column=len_pdrr+12).value=down_measure
    sheet.cell(row=rowshift,column=len_pdrr+13).value=better_ratings(aa)/np.sum(aa) # Upgrade
    sheet.cell(row=rowshift,column=len_pdrr+13).number_format='.0%'
    up_measure = up_notch(aa)
    sheet.cell(row=rowshift+1,column=len_pdrr+13).value=up_measure
    sheet.cell(row=rowshift,column=len_pdrr+14).value=np.sum(aa)
    sheet.cell(row=rowshift+1,column=len_pdrr+14).value=down_measure + up_measure
    
    CreateStyleSheet_TransitionMatrix(sheet, rowshift, colshift, PDRR)
    
    ### Add total per rows and columns
    sheet.cell(row=rowshift-1,column=colshift+len_pdrr).value = 'Total'
    for i in range(len_pdrr):
        row1=rowshift+i
        sheet.cell(row=row1,column=colshift+len_pdrr).value = sum(aa[i,:])
    sheet.cell(row=rowshift+len_pdrr,column=colshift-1).value='Total'
    for i in range(len_pdrr):
        col1=colshift+i
        sheet.cell(row=rowshift+len_pdrr,column=col1).value = sum(aa[:,i])
    sheet.cell(row=rowshift+len_pdrr,column=colshift+len_pdrr).value = np.sum(aa)    
        
    wb.save(output_filename)


def TMnotches(Data, xaxis, yaxis, PDRR):
     
    len_pdrr = len(PDRR)
    
    ### Fill up table ###
    aa = np.zeros((len_pdrr,len_pdrr))
    for rowNum in range(len_pdrr):
        for colNum in range(len_pdrr):
            temp = sum((Data[yaxis]==PDRR[rowNum]) & (Data[xaxis]==PDRR[colNum]))
            temp=int(temp)            
            if ((not(temp==0))):        
                aa[rowNum,colNum]=temp                  
    
    return((down_notch(aa),up_notch(aa)))


def withknotchesrate(Data, xaxis, yaxis, PDRR, k=2):
     
    len_pdrr = len(PDRR)
    
    ### Fill up table ###
    aa = np.zeros((len_pdrr,len_pdrr))
    for rowNum in range(len_pdrr):
        for colNum in range(len_pdrr):
            temp = sum((Data[yaxis]==PDRR[rowNum]) & (Data[xaxis]==PDRR[colNum]))
            temp=int(temp)            
            if ((not(temp==0))):        
                aa[rowNum,colNum]=temp                  
    
    return(within(aa,k)/np.sum(aa))


def TMstats(Data, xaxis, yaxis, PDRR):
     
    len_pdrr = len(PDRR)
    
    ### Fill up table ###
    aa = np.zeros((len_pdrr,len_pdrr))
    for rowNum in range(len_pdrr):
        for colNum in range(len_pdrr):
            temp = sum((Data[yaxis]==PDRR[rowNum]) & (Data[xaxis]==PDRR[colNum]))
            temp=int(temp)            
            if ((not(temp==0))):        
                aa[rowNum,colNum]=temp                  
    
    stats = {
    'Match':within(aa,0)/np.sum(aa),
    'Within_1':within(aa,1)/np.sum(aa),
    'Within_2':within(aa,2)/np.sum(aa),
    'Outside_5':1-within(aa,4)/np.sum(aa),
    'Downgrade':worse_ratings(aa)/np.sum(aa),
    'Upgrade':better_ratings(aa)/np.sum(aa),
    'CC_3': aa[2,:].sum(),
    'CC_3-4': aa[2:4,:].sum(),
    'CC_3-5': aa[2:5,:].sum(),
    'CC_3-6': aa[2:6,:].sum(),
    }
    
    return(stats)

    # -*- coding: utf-8 -*-
"""
Created on Mon Mar 20 14:11:42 2017

Version: 1.0: Initial build
Version: 1.1: 20170512, Modify the class to handle missing data. And add new method LinearReg(); Minor change in __SomersD()
Version: 1.2: 20180511, Add new private method __LinReg() and modify method modelselection(): add linear regression as one option.
@author: ub71894 (4e8e6d0b), CSG
"""
import os
import pandas as pd
import numpy as np
import warnings
import matplotlib.pyplot as plt
import statsmodels.api as sm
from sklearn.metrics import roc_auc_score
from itertools import combinations, product
from numba import jit
from PDScorecardTool.Process import SomersD

#%%
class MFA(object):
    
    '''    
    This class is used to realize all calculations in MFA procedure.
    
    MFA(data, model):

        data:       data for development. Make sure it has column 'def_flag' as default
                    indicator and the factors' names are the same as the ones in 'model'

        model:      PDModle class. It saves all parameters for the model
    

        quant_only: boolean, default 'False'. Added in Ver.1.6
                    set as 'True' if only normalize on quant factors.

        missing:    str, default 'median'. Added in Ver.1.1
                    statistics of the data used to fill missing data (NA data)



    Methods:

        describe()

        corr(factors='all', raw=False, output=True)

        VIF(factors='all', fullmodel='all')

        reweight(on='good', multiplier=1, random_state=None)

        recover()

        LogReg(factors='all', summary=False)

        LinearReg(factors='all', dependentvar='Final_PD_Risk_Rating', summary=False)

        modelAR(quant_weight=0, quali_weight=0, quantweight=0.5)
        
        plotAR(quant_weight=0, quali_weight=0, quantweight_range=[0,1], savefig=True)
        
        modelselection(factors, atleast_p=3, best_k=5, by='SomersD')
        
        ARgridsearch(**kw):  
        
    '''

    def __init__(self, data, model, quant_only=False, missing='median'):

        self.model = model
        self.rawdata = data.copy()
        self.rawdata.reset_index(drop=True, inplace=True)
        self.stats = self.__describe()
        self.normdata = self.__normalization(quant_only, missing)
        # save original normdata for later recover
        self.oridata = self.normdata.copy()


    def __describe(self):
        '''

        Privata method that calculate key statistics for quant factors on raw data

        '''
        warnings.filterwarnings('ignore', category=Warning)
        dat = self.rawdata[self.model.quant_factor]
        stats = dat.describe(percentiles=[0.01,0.05,0.1,0.25,0.5,0.75,0.9,0.95,0.99])  
        temp={}
        temp['Skew'] = pd.Series(dat.skew())
        temp['Kurtosis'] = dat.kurtosis()
        stats = pd.concat([stats, pd.DataFrame(temp).T])
        return stats


    def __normalization(self, quant_only, missing):
        '''

        Private method that apply normalization on financial factors.

        '''
        warnings.filterwarnings('ignore', category=Warning)
        # Invalid_Neg
        normdata = self.rawdata.copy()
        for i,neg_source in enumerate(self.model.Invalid_Neg):
            if neg_source:
                col=self.model.quant_factor[i]
                normdata[col][ (normdata[col]<0) & (normdata[neg_source]<0) ] = self.model.cap[i]
                # treat NA in 'neg_source' as negative value
                normdata[col][ (normdata[col]<0) & pd.isnull(normdata[neg_source]) ] = self.model.cap[i]

        # cap/floor for quant factors:
        for i, col in enumerate(self.model.quant_factor):
            normdata[col] = np.clip(normdata[col], self.model.floor[i], self.model.cap[i])        

        # quant factors transformation:
        for i, col in enumerate(self.model.quant_factor):
            if self.model.quant_log[i]:
                normdata[col] = np.log(normdata[col])
            
        # quant factors normalization:
        for i, col in enumerate(self.model.quant_factor):
            normdata[col]  = 50*(normdata[col] - self.model.doc_mean[i]) / self.model.doc_std[i]      
        
        # quant factors flip sign:  
        for i, col in enumerate(self.model.quant_factor):
            normdata[col] = normdata[col] * self.model.quant_multiplier[i]     


        if quant_only:
            exec('normdata[self.model.quant_factor] = normdata[self.model.quant_factor].\
            fillna(normdata[self.model.quant_factor].{missing}())'.format(missing=missing))
        else:
            # calibration for quali factors:
            for i, col in enumerate(self.model.quali_factor):
                normdata[col].replace(to_replace=self.model.qualimapping[i],inplace=True) 
                 
            # fill missing data. added in Ver. 1.1
            exec('normdata[self.model.quant_factor+self.model.quali_factor] = normdata[self.model.quant_factor+self.model.quali_factor].\
                fillna(normdata[self.model.quant_factor+self.model.quali_factor].{missing}())'.format(missing=missing))

        return normdata       


    def __AUROC(self, score, dependentvar, isthere_def=True) :
        '''

        Private method that calculate area under ROC.

        '''
        if isthere_def:
            score = np.array(score)
            y_true  = [int(x) for x in list(self.normdata[dependentvar])] 
            return roc_auc_score(y_true, score) 
        else:
            return 'NA'


    def __SomersD_byAUROC(self, score, dependentvar, isthere_def=True) :
        '''

        Private method that calculate SomersD by 2*AUROC-1.

        '''
        if isthere_def:
            score = np.array(score)
            y_true  = [int(x) for x in list(self.normdata[dependentvar])] 
            return 2*roc_auc_score(y_true, score)-1
        else:
            return 'NA'


    @jit
    def __SomersD(self, score, dependentvar, sign=1, unit_concordance=True): 
        '''
        Private method to calculate SomersD
        New version of SomersD function which leverages numba.jit to accelerate
        the calculation. 

        '''
        x = np.array(score)
        # added in Ver 1.1 to deal with missing PDRR sometimes.
        y_true = self.normdata[dependentvar].copy()
        y_true.dropna(inplace=True)
        y = np.array(y_true)
        n = len(x)
        C = 0
        D = 0
        if unit_concordance==True:
            for i in range(n):
                for j in range(i):
                    dx = x[i] - x[j]
                    dy = y[i] - y[j]
                    if (dx*dy) > 0:
                        cij = 1
                        dij = 1
                    elif  (dx*dy) < 0:
                        cij = -1
                        dij = 1
                    else:
                        cij = 0
                        if dy==0: 
                            dij = 0
                        else: 
                            dij=1
                    C = C + cij
                    D = D + dij
        else:
            for i in range(n):
                for j in range(i):
                    dx = x[i] - x[j]
                    dy = y[i] - y[j]
                    if  (dx*dy) > 0:
                        cij = abs(dy)
                        dij = abs(dy)
                    elif (dx*dy) < 0:
                        cij = -abs(dy)
                        dij = abs(dy)
                    else:
                        cij = 0
                        if dy==0: 
                            dij = 0
                        else: 
                            dij=1
                    C = C + cij
                    D = D + dij    
        
        return sign*C/D


    def describe(self):
        '''

        Public method that present key statistics for quant factors on raw data
        
        '''
        return self.stats


    def corr(self, factors='all', raw=False, output=True):
        '''

        Public method that calculate correlation matrix for multifactors

        Parameters: 

        factors:    str {'quant', 'quali', 'all'} or None, default 'all'
                    Multifactors for matrix calculations. 

        raw:        boolean, default False
                    whether using raw data or normalized data. Default setting is
                    False since usually we care about the corr of normalized factors.

        output:     boolean, default True
                    whether save the corr matrix to excel file.

        '''
        warnings.filterwarnings('ignore', category=Warning)
        if factors=='quant':
            names= self.model.quant_factor
        elif factors=='quali':
            names= self.model.quali_factor
        elif factors=='all':
            names = self.model.quant_factor + self.model.quali_factor
        else: # user input a customized list
            names = factors
            
        
        if raw:
            suffix = '_raw'
            corr = self.rawdata[names].corr()
            if output:
                corr.to_excel('CorrMat_'+factors+suffix+'.xlsx')
            return corr

        else:
            suffix = '_norm'
            corr = self.normdata[names].corr()
            if output:
                corr.to_excel('CorrMat_'+factors+suffix+'.xlsx')
            return corr    


    def VIF(self, factors='all', fullmodel='all'):
        '''

        Public method that calculate VIF for factors

        Parameters: 

        factors:    str {'quant','quali', 'all', factorname}, default 'all'
                    factors for VIF calculation. 

        fullmodel:  str {'quant','quali', 'all'}, default 'all'
                    factors busket for VIF calculation. 

        '''        

        if factors=='quant':
            names= self.model.quant_factor
        elif factors=='quali':
            names= self.model.quali_factor
        elif factors=='all':
            names= self.model.quant_factor + self.model.quali_factor 
        else: # user input a single factor's name
            names = [factors]   
        
        if fullmodel=='quant':
            fullmodelnames= self.model.quant_factor
        elif fullmodel=='quali':
            fullmodelnames= self.model.quali_factor
        else:
            fullmodelnames = self.model.quant_factor + self.model.quali_factor 

        result=[]
        for factor in names:
            ols_model = sm.OLS(endog=np.array(self.normdata[factor]),exog=\
                sm.add_constant(self.normdata[fullmodelnames].drop(factor,axis=1)), missing='drop')
            ols_result = ols_model.fit()
            result.append(1/(1-ols_result.rsquared))

        return pd.DataFrame({'VIF':result}, index=names)


    def reweight(self, on='good', multiplier=1, random_state=None):
        '''

        Public method that performs data 'sampling' for reweighting purpose.
        Parameters: 

        on:             str 'good' or 'bad', default 'good'
                        which event for reweight

        multiplier:     int or float, default 1

        random_state:   int or numpy.random.RandomState
                        Seed for the random number generator (if int), or 
                        numpy RandomState object.
    
        ''' 
        enum   = {'good':0,'bad':1}
        data   = self.normdata.query('def_flag=={on}'.format(on=enum[on])).copy()
        data_c = self.normdata.query('def_flag=={not_on}'.format(not_on=1-enum[on]))
        
        if multiplier<=1: # sample without replacement
            sample = int(multiplier*len(data))
            new = pd.concat([data.sample(n=sample, replace=False, random_state=random_state), data_c], ignore_index=True)   

        else:       # sample without replacement for fraction part, times sample for int part
            int_part = int(multiplier)
            frac_part = multiplier-int_part
            sample = int(frac_part*len(data))
            new = pd.concat([data]*int_part, ignore_index=True)
            if sample !=0:
                new = pd.concat([new, data.sample(n=sample, replace=False, random_state=random_state)], ignore_index=True)
            new = pd.concat([new, data_c], ignore_index=True)
        
        self.normdata = new.copy()


    def recover(self):
        '''

        Public method that recover reweighted normdata back to original normdata
    

        '''
        print('Current "normdata" has {n} rows.\n'.format(n=len(self.normdata)))
        self.normdata = self.oridata.copy()
        print('After recovery, "normdata" is back to {n} rows.\n'.format(n=len(self.normdata)))


    def LogReg(self, factors='all', summary=False):
        '''

        Public method that perform Logistic Regression with intercept

        Parameters: 

        factors:    str {'quant','quali', all'} or a list of strs, default 'all'
                    factors for LR. 

        summary:    boolean, default False
                    whether to print out the summary for LR.
        

        Return:     a tuple has parameters implied weights and corresponding AR


        '''      
        if factors=='quant':
            names = self.model.quant_factor
        elif factors=='quali':
            names = self.model.quali_factor
        elif factors=='all':
            names = self.model.quant_factor + self.model.quali_factor 
        else:
            names = factors


        x = sm.add_constant(self.normdata[names],prepend = True)
        logit = sm.Logit(self.normdata.def_flag, x)
        result = logit.fit(disp=0)
        if summary:
            print(result.summary())

        if factors!='all':
            coeff = result.params[1:]
            weight_coeff = [x/coeff.sum() for x in coeff]
            score = weight_coeff*self.normdata[names] 
            AUROC = roc_auc_score(self.normdata.def_flag,score.sum(axis=1))
            SomersD = 2*AUROC-1

            result1 = pd.DataFrame({'weights':weight_coeff}, index=names)
            result2 = pd.DataFrame({'AR':[AUROC, SomersD]}, index=['AUROC', 'SomersD'])
            return (result1, result2)

        else: # both quant and quali factors
            idx =len(self.model.quant_factor)+1
            coeff_quant = result.params[1:idx]
            coeff_quali = result.params[idx:]
            weight_coeff_quant = [x/coeff_quant.sum() for x in coeff_quant]
            weight_coeff_quali = [x/coeff_quali.sum() for x in coeff_quali]
            weight_quantmodule = coeff_quant.sum() / (coeff_quant.sum() + coeff_quali.sum())
            AUROC = roc_auc_score(self.normdata.def_flag, result.fittedvalues)
            SomersD = 2*AUROC-1

            result1 = pd.DataFrame({'weights':weight_coeff_quant+weight_coeff_quali}, index=names)
            result2 = pd.DataFrame({'AR':[AUROC, SomersD]}, index=['AUROC', 'SomersD'])
            result3 = pd.DataFrame({'quantmoduleweight':weight_quantmodule}, index=['byLogisticRegresion'])
            return (result1, result2, result3)


    def LinearReg(self, factors='all', dependentvar='Final_PD_Risk_Rating', summary=False):
        '''

        Public method that perform Logistic Regression with intercept

        Parameters: 

        factors:        str {'quant','quali', all'} or a list of strs, default 'all'
                        factors for LR. 

        dependentvar:   str, default 'Final_PD_Risk_Rating' 
                        column name of y which is dependent variable in the regression

        summary:        boolean, default False
                        whether to print out the summary for LR.
        

        Return:     a tuple has parameters implied weights and corresponding AR


        ''' 
        
        if factors=='quant':
            names = self.model.quant_factor
        elif factors=='quali':
            names = self.model.quali_factor
        elif factors=='all':
            names = self.model.quant_factor + self.model.quali_factor 
        else:
            names = factors

        #dataforLR = self.normdata.copy()
        #dataforLR.dropna(subset=[names+[dependentvar]], how='any', inplace=True)

        #x = sm.add_constant(dataforLR[names],prepend = True)
        #OLS = sm.OLS(dataforLR[dependentvar], x)
        x = sm.add_constant(self.normdata[names],prepend = True)
        OLS = sm.OLS(self.normdata[dependentvar], x, missing='drop')
        result = OLS.fit(disp=0)
        if summary:
            print(result.summary())

        if factors!='all':
            coeff = result.params[1:]
            weight_coeff = [x/coeff.sum() for x in coeff]
            SomersD = self.__SomersD(result.fittedvalues, dependentvar)

            result1 = pd.DataFrame({'weights':weight_coeff}, index=names)
            result2 = pd.DataFrame({'AR':[SomersD]}, index=['SomersD'])
            return (result1, result2)

        else: # both quant and quali factors
            idx =len(self.model.quant_factor)+1
            coeff_quant = result.params[1:idx]
            coeff_quali = result.params[idx:]
            weight_coeff_quant = [x/coeff_quant.sum() for x in coeff_quant]
            weight_coeff_quali = [x/coeff_quali.sum() for x in coeff_quali]
            weight_quantmodule = coeff_quant.sum() / (coeff_quant.sum() + coeff_quali.sum())
            SomersD = self.__SomersD(result.fittedvalues, dependentvar)

            result1 = pd.DataFrame({'weights':weight_coeff_quant+weight_coeff_quali}, index=names)
            result2 = pd.DataFrame({'AR':[SomersD]}, index=['SomersD'])
            result3 = pd.DataFrame({'quantmoduleweight':weight_quantmodule}, index=['byLogisticRegresion'])
            return (result1, result2, result3)

    
    def modelAR(self, quant_weight=0, quali_weight=0, quantweight=0.5, isthere_def=True, dependentvar='def_flag', use_msms=False, update_msms=False):
        '''

        Public method that calculate model's AR according to input model setting

        Parameters: 

        quant_weight:   list, default 0 
                        list of quant factors' weights

        quali_weight:   list, default 0 
                        list of quali factors' weights

        quantweight:    float, default 0.5
                        the weight of quant module in total score.

        isthere_def:    boolean, default True
                        whether the analysis is on the default event. If not, for 
                        example we want to check the accuracy of factor on PDRR,
                        then we should input  
                        isthere_def=False, dependentvar='PDRR'

        dependentvar:   str, default 'def_flag'
                        column name for dependent variable.          

        use_msms:       boolean, default False
                        'msms' means module statistics which saved in model setting
                        whether use module mean and std in model setting.
                        To use the mean and std of quant and quali module in class 'model' if True.
                        To use the mean and std of quant and quali module that calibrated from the data if False.
        update_msms:    boolean, default False
                        update model.quantmean, model.qualimean, model.quantstd and model.qualistd with calibrated value from data
        Return:   AR in a Dataframe 
        '''      
        if quant_weight and quali_weight: # means all factors

            temp_quant = quant_weight*self.normdata[self.model.quant_factor]
            temp_quali = quali_weight*self.normdata[self.model.quali_factor]

            if use_msms:
                quantmean =  self.model.quantmean
                quantstd =  self.model.quantstd
                qualimean =  self.model.qualimean
                qualistd =  self.model.qualistd
            else:
                quantmean =  temp_quant.sum(axis=1).mean()
                quantstd =  temp_quant.sum(axis=1).std()
                qualimean =  temp_quali.sum(axis=1).mean()
                qualistd =  temp_quali.sum(axis=1).std()

            if update_msms:
                self.model.quantmean = quantmean
                self.model.quantstd = quantstd
                self.model.qualimean = qualimean
                self.model.qualistd = qualistd
                self.model.quantweight = quantweight
                self.model.qualiweight = 1 - quantweight


            score = quantweight*50*((temp_quant).sum(axis=1)-quantmean) / quantstd+\
                    (1-quantweight)*50*((temp_quali).sum(axis=1)-qualimean) / qualistd

            AUROC = self.__AUROC(score, dependentvar, isthere_def)
            if isthere_def:
                SomersD = 2*AUROC-1
            else:
                SomersD = self.__SomersD(score, dependentvar)
            return(pd.DataFrame({'AR':[AUROC, SomersD]}, index=['AUROC', 'SomersD']))

        elif quant_weight:
            score =  (quant_weight*self.normdata[self.model.quant_factor]).sum(axis=1)
            AUROC = self.__AUROC(score, dependentvar, isthere_def)
            if isthere_def:
                SomersD = 2*AUROC-1
            else:
                SomersD = self.__SomersD(score, dependentvar)
            return(pd.DataFrame({'AR':[AUROC, SomersD]}, index=['AUROC', 'SomersD'])) 

        elif quali_weight:
            score =  (quali_weight*self.normdata[self.model.quali_factor]).sum(axis=1)
            AUROC = self.__AUROC(score, dependentvar, isthere_def)
            if isthere_def:
                SomersD = 2*AUROC-1
            else:
                SomersD = self.__SomersD(score, dependentvar)
            return(pd.DataFrame({'AR':[AUROC, SomersD]}, index=['AUROC', 'SomersD']))   
        
        else:
            return 0


    def plotAR(self, quant_weight=0, quali_weight=0, quantweight_range=[0,1], isthere_def=True, dependentvar='def_flag', savefig=True): 
        '''

        Public method that draws the relationship between model's AR and quantweight

        Parameters: 

        quant_weight:       list, default 0 
                            list of quant factors' weights

        quali_weight:       list, default 0 
                            list of quali factors' weights

        quantweight_range:  list, default [0,1]
                            starting and ending points of quantweight in this plot  

        isthere_def:        boolean, default True
                            whether the analysis is on the default event. If not, for 
                            example we want to check the accuracy of factor on PDRR,
                            then we should input  
                            isthere_def=False, dependentvar='PDRR'

        dependentvar:       str, default 'def_flag'
                            column name for dependent variable. 

        savefig:            boolean, default True
                            whether save png file in current working directory. 
        
        '''      
        weightlist=[];  ARlist=[]
        weight =  quantweight_range[0]  
        while weight<=quantweight_range[1]:
            weightlist.append(weight)
            ARlist.append(self.modelAR(quant_weight, quali_weight, quantweight=weight,\
                isthere_def=isthere_def, dependentvar=dependentvar).loc['SomersD','AR'])
            weight += 0.01  
        AR_max = max(ARlist)
        weigh_max = weightlist[ARlist.index(max(ARlist))]
        
        # plot
        f,ax = plt.subplots(1,1,figsize=(8,8))  
        plt.plot(weightlist, ARlist, 'b')
        ax.annotate('Max={ar:.3f} @ weight={wt:.2f}'.format(ar=AR_max, wt=weigh_max), \
            xy=(weigh_max, AR_max), xytext=(weigh_max-0.1, AR_max+0.002),\
            arrowprops=dict(facecolor='green', shrink=0.05))
        ax.set_ylim([min(ARlist),max(ARlist)+0.01])
        plt.xlabel('Weight of Quant')
        plt.ylabel('AccuracyRate by SomersD')

        if savefig:
            s=1;
            for file in os.listdir('.'):
                if file.startswith("ARplot"):
                    s+=1
            f.savefig('ARplot_{suffix}.png'.format(suffix=s))

        return(f)


    def __LogReg(self, names):
        '''
        Privata method that run LR with fixed setting in factors.
        '''
        x = sm.add_constant(self.normdata[names],prepend = True)
        logit = sm.Logit(self.normdata.def_flag, x)
        result = logit.fit(disp=0)
        BIC = result.bic
        AUROC = roc_auc_score(self.normdata.def_flag, result.fittedvalues)
        SomersD = 2*AUROC-1
        return (pd.DataFrame({'Measure':[SomersD, BIC]}, index=['SomersD','BIC']))

    def __LinReg(self, names, dependentvar):
        '''
        Privata method that run LR with fixed setting in factors.
        '''
        x = sm.add_constant(self.normdata[names],prepend = True)
        linear = sm.OLS(self.normdata[dependentvar], x, missing='drop')
        result = linear.fit(disp=0)
        R2 = result.rsquared
        SD = SomersD(self.normdata[dependentvar], result.fittedvalues)
        
        return (pd.DataFrame({'Measure':[SD, R2]}, index=['SomersD','R-squared']))

    def modelselection(self, factors, atleast_p=3, best_k=5, by='SomersD', regressor='logistic', dependentvar=None):
        '''

        Public method that performs model selection by Logistic Regression. 
        According to 'SomersD' or 'BIC', it returns the best k model settings.

        Parameters: 

        factors:        list of strings
                        full list of factor candidates

        atleast_p:      int, default 3
                        minimal requirement for the number of factors in the model.

        best_k:         int, default 5
                        how many best settings are kept.

        by:             str {'SomersD' or 'BIC'}, default 'SomersD'
                        the measure used as criterion in model selection.

        regressor:      str {'logistic' or 'linear'}. Add in Ver. 1.2
                        the regression model that user chooses

        dependentvar:   str, the name for dependent variable in linear regression.
                        Add in Ver. 1.2

        Return:         best_k settings and measure in a Dataframe 
        ''' 
        if regressor=='logistic':

            enum = {'SomersD':0, 'BIC':1}
            idx = enum.pop(by, 0)   

            setting=[];    measure=[]
            for p in range(atleast_p, len(factors)+1):
                for names in combinations(factors, p):
                    setting.append(names)
                    measure.append(self.__LogReg(list(names)).iloc[idx,0])
            result = pd.DataFrame({'Setting':setting,'Measure':measure}, columns=['Setting','Measure'])
            result.sort_values(by='Measure', ascending=idx, inplace=True)
            return (result.reset_index(drop=True).iloc[:best_k,])

        elif regressor=='linear':

            enum = {'SomersD':0, 'R-squared':1}
            idx = enum.pop(by, 0)   

            setting=[];    measure=[]
            for p in range(atleast_p, len(factors)+1):
                for names in combinations(factors, p):
                    setting.append(names)
                    measure.append(self.__LinReg(list(names), dependentvar).iloc[idx,0])
            result = pd.DataFrame({'Setting':setting,'Measure':measure}, columns=['Setting','Measure'])
            result.sort_values(by='Measure', ascending=idx, inplace=True)
            return (result.reset_index(drop=True).iloc[:best_k,])
        else:
            return (0)


    def ARgridsearch(self, **kw):   
        '''

        Public method that performs Grid Search for factors' best weights according to
        Somers'D.

        Parameters: 

        quant_factor:       list of strings(quant factors's names)
                            ex. model.quant_factor or ['quant1','quant2']

        quant_weight_range: list of tuples that indicate the beginning and ending 
                            points in searching for quant factors. 
                            ex. [(0.25,0.35),(0.1,0.2),(0.2,0.3),(0.1,0.2),(0.25,0.35)]

        quali_factor:       list of strings(quali factors's names)
                            ex. model.quali_factor or ['quali1','quali2']

        quali_weight_range: list of tuples that indicate the beginning and ending 
                            points in searching for quali factors. 
                            ex. [(0.25,0.35),(0.4,0.5),(0.2,0.3),(0.05,0.15)]

        delta_factor:       float, default 0.05
                            grid increment for factors in searching      

        quantweight_range:  tuple that indicates the beginning and ending 
                            points in searching for quantweight.    

        delta_quantweight:  float, default 0.05
                            grid increment for quantweight in searching   

        best_k:             int, default 20
                            how many best settings are kept.

        isthere_def:        boolean, default True
                            whether the analysis is on the default event. If not, for 
                            example we want to check the accuracy of factor on PDRR,
                            then we should input  
                            isthere_def=False, dependentvar='PDRR'

        dependentvar:       str, default 'def_flag'
                            column name for dependent variable.      

        use_msms:           boolean, default False
                            'msms' means module statistics which saved in model setting
                            whether use module mean and std in model setting.
                            To use the mean and std of quant and quali module in class 'model' if True.
                            To use the mean and std of quant and quali module that calibrated from the data if False.

        Return:   best_k weights settings and AR(SomersD) in a Dataframe 
        '''     
        quant_factor        = kw.pop('quant_factor',False)
        quant_weight_range  = kw.pop('quant_weight_range',False)    
        quali_factor        = kw.pop('quali_factor',False)
        quali_weight_range  = kw.pop('quali_weight_range',False)
        delta_factor        = kw.pop('delta_factor',0.05)
        quantweight_range   = kw.pop('quantweight_range',False)
        delta_quantweight   = kw.pop('delta_quantweight',0.05)
        best_k              = kw.pop('best_k',20)   
        isthere_def         = kw.pop('isthere_def',True)   
        dependentvar        = kw.pop('dependentvar','def_flag')   
        use_msms            = kw.pop('use_msms',False)

        if isthere_def:
            arfunc = self.__SomersD_byAUROC
        else:
            arfunc = self.__SomersD
            
        rows_list = [];     args = []       

        # case 1: both quant and quali:
        if quant_factor and quali_factor:
            # prepare data
            quant_weight_range  = dict(zip(quant_factor,quant_weight_range))
            quali_weight_range  = dict(zip(quali_factor,quali_weight_range))    

            # prepare grid for each quant and quali factors:
            for i in quant_factor:
                args.append(np.linspace(quant_weight_range[i][0],quant_weight_range[i][1], \
                    num=int(round((quant_weight_range[i][1]-quant_weight_range[i][0])/delta_factor+1))))
            for i in quali_factor:
                args.append(np.linspace(quali_weight_range[i][0],quali_weight_range[i][1], \
                    num=int(round((quali_weight_range[i][1]-quali_weight_range[i][0])/delta_factor+1))))        
            if len(quantweight_range)==1:
                args.append(np.array(quantweight_range))
            else:
                args.append(np.linspace(quantweight_range[0],quantweight_range[1],num=int(round((quantweight_range[1]-quantweight_range[0])/delta_quantweight+1))))
            
            lens_qq = len(quant_factor)
            weights = product(*args)   

            if use_msms:
                quantmean =  self.model.quantmean
                quantstd =  self.model.quantstd
                qualimean =  self.model.qualimean
                qualistd =  self.model.qualistd

            for all in weights:
                # check the sum of weights:
                if sum(all[:lens_qq])==1 and sum(all[lens_qq:-1])==1:
                    quant_weights = list(all)[:lens_qq]
                    quali_weights = list(all)[lens_qq:-1]

                    temp_quant = quant_weights*self.normdata[quant_factor]
                    temp_quali = quali_weights*self.normdata[quali_factor]
                    
                    if not use_msms:
                        quantmean =  temp_quant.sum(axis=1).mean()
                        quantstd =  temp_quant.sum(axis=1).std()
                        qualimean =  temp_quali.sum(axis=1).mean()
                        qualistd =  temp_quali.sum(axis=1).std()

                    score =  all[-1]*50*((temp_quant).sum(axis=1)-quantmean) / quantstd + \
                    (1-all[-1])*50*((temp_quali).sum(axis=1)-qualimean) / qualistd
                
                    accrat = arfunc(score, dependentvar)

                    dict1 = {}
                    dict1.update({'AR':accrat})
                    dict1.update(dict(zip(quant_factor,quant_weights)))
                    dict1.update(dict(zip(quali_factor,quali_weights)))
                    dict1.update({'quantweight':all[-1]})
                    rows_list.append(dict1) 
    

            result = pd.DataFrame(rows_list)
            result = result[quant_factor+quali_factor+['quantweight']+['AR']]
            result.sort_values(by='AR',ascending=False, inplace=True)
            result.reset_index(drop=True, inplace=True)
            return(result.head(best_k)) 
    
        elif bool(quant_factor) != bool(quali_factor):# means quant_factor xor quali_factor
            
            if quant_factor:
                quant_weight_range  = dict(zip(quant_factor,quant_weight_range))
                names = quant_factor
                for i in names:
                    args.append(np.linspace(quant_weight_range[i][0],quant_weight_range[i][1], \
                    num=int(round((quant_weight_range[i][1]-quant_weight_range[i][0])/delta_factor+1))))
            else:
                quali_weight_range  = dict(zip(quali_factor,quali_weight_range))
                names = quali_factor
                for i in names:
                    args.append(np.linspace(quali_weight_range[i][0],quali_weight_range[i][1], \
                    num=int(round((quali_weight_range[i][1]-quali_weight_range[i][0])/delta_factor+1))))    

            weights = product(*args)    
            for all in weights:    
                # check the sum of weights:   
                if sum(all)==1:
                    names_weights = list(all)                   
                    score = (self.normdata[names]*names_weights).sum(axis=1)        
                    accrat = arfunc(score, dependentvar)

                    dict1 = {}
                    dict1.update({'AR':accrat})
                    dict1.update(dict(zip(names,names_weights)))
                    rows_list.append(dict1) 

            result = pd.DataFrame(rows_list)
            result = result[names+['AR']]
            result.sort_values(by='AR',ascending=False, inplace=True)
            result.reset_index(drop=True, inplace=True)
            return(result.head(best_k))
           
        else:
            return(0)



# -*- coding: utf-8 -*-
"""
Created on Fri Oct  7 10:47:28 2016

Version: 1.0: Initial build
Version: 1.1, 20170327, Add method update()
Version: 1.2, 20170330, Add masterscale as new class attribute.
Version: 1.3, 20170502, Change masterscale as optional attribute.
Version: 1.4, 20180508, Add method reset()
Version: 1.5, 20180810, Add method save()

@author: ub71894 (4e8e6d0b), CSG

"""

import collections
import pandas as pd
import numpy as np
import warnings
import pickle

class PDModel(object):

    def __init__(self,  PDInfo_file, model_name, version, quant_factor, quali_factor, masterscale_file=False):

        self.model_name = model_name
        self.version = version
        self.quant_factor = quant_factor
        self.quali_factor = quali_factor
        self.masterscale_file = masterscale_file

        PDInfo = pd.read_excel(PDInfo_file,sheet_name=r'PD Models')

        Parameters=PDInfo.loc[(PDInfo['Model Name.1']==self.model_name) & (PDInfo['Version Number']==self.version)]
        # Quant parameters df
        quant_para=pd.DataFrame(Parameters.loc[:,'Quant_Name_1':'Quant_Wt_1'].values, 
            columns=['Factor', 'Lower', 'Upper', 'Invalid_Neg', 'Take_log','Mean', 'Std', 'Multiplier','Weight'])     
        for i in range((len(self.quant_factor)-1)):
            quant_para=pd.concat([quant_para, 
                pd.DataFrame(Parameters.loc[:,'Quant_Name_'+str(i+2):'Quant_Wt_'+str(i+2)].values,
                columns=['Factor', 'Lower', 'Upper', 'Invalid_Neg', 'Take_log','Mean', 'Std', 'Multiplier','Weight'])],  ignore_index=True)              
        # Quali parameters df
        quali_para=pd.DataFrame(Parameters.loc[:,'Qual_1_Question':'Qual_1_Wt'].values, 
            columns=['Factor', 'A', 'B', 'C', 'D','E', 'F', 'G','Weight'])
        for i in range((len(self.quali_factor)-1)):
            quali_para=pd.concat([quali_para, 
                pd.DataFrame(Parameters.loc[:,'Qual_'+str(i+2)+'_Question':'Qual_'+str(i+2)+'_Wt'].values, 
                columns=['Factor', 'A', 'B', 'C', 'D','E', 'F', 'G','Weight']) ], ignore_index=True)
            
        self.quant_weight=quant_para['Weight'].tolist()  
        self.floor=quant_para['Lower'].tolist()
        self.cap=quant_para['Upper'].tolist()
        self.doc_mean=quant_para['Mean'].tolist()
        self.doc_std=quant_para['Std'].tolist()
        self.quant_log = quant_para['Take_log'].fillna(value = False).tolist()
        self.quant_multiplier = quant_para['Multiplier'].tolist()        

        self.quali_weight=quali_para['Weight'].tolist()    
        qualimapping=[]
        for i in range(len(quali_para)):
           qualimapping.append(quali_para.loc[i,'A':'G'].dropna().to_dict()) 
        self.qualimapping = qualimapping           
           
        self.quantmean=Parameters['Quant_Mean_Module'].iat[0]
        self.quantstd=Parameters['Quant_StdDev_Module'].iat[0]
        self.quantweight=Parameters['Quant_Weight'].iat[0] 
        self.qualimean=Parameters['Qual_Mean_Module'].iat[0]
        self.qualistd=Parameters['Qual_StdDev_Module'].iat[0]
        self.qualiweight=1-self.quantweight
        self.Invalid_Neg=quant_para['Invalid_Neg'].tolist()  

        if Parameters['Less than or Equal to'].iat[0]>=999999:
            self.slope1=self.slope2=Parameters['Slope'].iat[0]
            self.intercept1=self.intercept2=Parameters['Intercept'].iat[0]
            self.cutoff=1e7
        else:
            self.slope1=Parameters['Slope'].iat[0]
            self.intercept1=Parameters['Intercept'].iat[0]
            self.slope2=Parameters['Slope.1'].iat[0]
            self.intercept2=Parameters['Intercept.1'].iat[0]
            self.cutoff=Parameters['Less than or Equal to'].iat[0]
        # new in Ver. 1.3
        if self.masterscale_file:
            self.MS = pd.read_excel(self.masterscale_file) # add in Ver. 1.2
        else:
            print("Waring: MasterScale file is not assigned to PDModel instance.\n")

    
    def reset(self):
       for key, value in self.__dict__.items():
            if key!='MS':
                exec('self.{key} = 0'.format(key=key)) 

    def update(self, kw):
        for key, value in kw.items():
            exec('self.{key} = value'.format(key=key))

    def __repr__(self):
        od = collections.OrderedDict(sorted(self.__dict__.items()))
        inner_lines = '\n'.join('%s = %s' % (k, v) for k, v in od.items())
        return '<\n%s\n>' % inner_lines

    def save(self, filename):
        filehandler = open(filename, 'wb')
        pickle.dump(self, filehandler)

    __str__=__repr__

# -*- coding: utf-8 -*-
"""
Created on Thu Mar 16 12:50:12 2017

Version: 1.0:   Initial build
Version: 1.1:   20170327, Add function quanttrans() which calculate cap/floor, 
                mean and std of quant factors.
Version: 1.2:   20170405, Add function normalization(), PD_frPDRR(), 
                logitPD_frPDRR(), getPDRR(), getTM(), and getTMnotches().
                Jack debugs method qualitrans() to make it work correctly in the
                model whose has no true default.
Version: 1.3:   20170424, Add function getwithknotchesrate(), SomersD() and 
                buildpseudodef()
Version: 1.4:   20170512, Modify function qualitrans() and quanttrans(), 
                normalization()
Version: 1.5:   20170802, Modify function getPDRR(). Add new function 
                RAreplica()
Version: 1.6:   20180508, Modify function normalization(), PD_frPDRR() and 
                logitPD_frPDRR()
Version: 1.7:   20180713, Add function buildpseudodef2()
Version: 1.8:   20180808, Add function MAUG_mapping(), NAICS_mapping()
Version: 1.9:   20180820, Add function ExtRating_mapping()
Version: 2.0:   20180906, Modify function getPDRR(), getTM(), getTMnotches(), 
                getwithknotchesrate() to make them have option for the version 
                of Masterscale
Version: 2.1:   20190604, Add function PD_frPDRR_autoMS() and 
                logitPD_frPDRR_autoMS()
Version: 2.2:   20191105, Add function gcar_converter()
Version: 2.3:   20200121, Modify a typo in MAUG_mapping()
Version: 2.4:   20200318, Add a tuple for americas office code: gcar_americas_office_code
                gcar_cif_cleaner()
Version: 2.5:   20200330, Modify MAUG_mapping(), Major Update for NAICS_mapping()           
Version: 2.6:   20200331, Add 3 functions: func_sd, func_rla and func_ovd 
                          All of them are designed for GroupBy.apply.
Version: 2.7:   20200605, Modify function SomersD()      
Version: 2.8:   20200930, Modify function gcar_converter()   
Version: 2.9:   20201113, Add function GICS_mapping()  
Version: 3.0:   20201119, Add function correlation_dist(), reorder_cols(), 
                          cluster_corr(), cluster_corr_old(), plot_cluster_corr(),
                          pca_n_comp()        
@author: ub71894 (4e8e6d0b), CSG
"""

import pandas as pd
import numpy as np
import warnings
from numba import jit
from PDScorecardTool.CreateBenchmarkMatrix import CreateBenchmarkMatrix
from PDScorecardTool.CreateBenchmarkMatrix import TMnotches, withknotchesrate
from PDScorecardTool._info_data import naics_code, gics_code
from scipy.cluster import hierarchy
from scipy.stats import pearsonr
from sklearn import decomposition
from sklearn.preprocessing import StandardScaler
import matplotlib.pyplot as plt
import seaborn as sns
import locale
locale.setlocale(locale.LC_NUMERIC, '')


# for qualitative factors recalibration

def qualitrans(data, model, isthere_def=True, dependentvar='def_flag', output=True):
    """
    This function transforms raw qualitative factors (letters) into numeric value
    based on logit mean PD of each bucket.
    
    Parameters:

        data:           the input dataset in DataFrame. Make sure no NA in quali
                        factors

        model:          PDModel class
                        
        isthere_def:    boolean, default True
                        whether the calibration is based on true default event.

        dependentvar:   str, default 'def_flag'
                        column name for dependent variable. If 'isthere_def=False',
                        'dependentvar' should be some PD.  

        output:         boolean, default True
                        whether save the calibration to excel file.                           

    Return:
        a list of dictionaries that saves all mapping between letters to values
    
    """    

    calibration = {}
    for quali_col in model.quali_factor: #for each quali column
        dat = data[[quali_col,dependentvar]].copy() #takes quali_col and dependentvar columns from data (preserves index)
        dat.dropna(how='any', inplace=True)  # added in Ver. 1.4
        temp = dat.groupby(by=[quali_col]).mean() #for quali_col, find mean dependentvar
        temp.rename(columns={dependentvar:'mean_PD'},inplace=True) #renames dependentvar to mean_PD
        temp['count'] = dat.groupby(by=[quali_col]).count() #checks number of items with quali_col
        if isthere_def: #if there is a default
            temp['defaults'] = data[[quali_col,dependentvar]].groupby(by=[quali_col]).sum()#add to default column
            temp = temp[['count','defaults','mean_PD']] #sort into count, defaults, mean_PD
            # to set 1 as the minimal #def. Move here in Ver. 1.2
            temp.defaults.replace({0:1}, inplace=True)
            temp.loc[temp.defaults==1, 'mean_PD'] = 1/temp.loc[temp.defaults==1, 'count']
        else:
            temp = temp[['count','mean_PD']] #just sort into count, mean_pd
        
        temp['Rawscore'] = np.log(temp.mean_PD/(1-temp.mean_PD)) #calculates raw score        
        #new mean
        temp['mean'] = sum(temp['Rawscore'] * temp['count'])/temp['count'].sum() #takes the mean of raw score                
        #new std
        temp['std'] = np.sqrt(sum((temp['count'] * ((temp['Rawscore'] - temp['mean'])**2))/(temp['count'].sum()-1))) #takes the std of the raw score
        #new score
        temp['score'] = 50* (temp['Rawscore']-temp['mean'])/temp['std']#calculated new score
        calibration.update({quali_col:temp}) #adds tempscore to each qualitative column
    if output:
        writer_Def = pd.ExcelWriter('qualitrans_output.xlsx', engine='xlsxwriter')#creates a Pandas Excel Writer
        s=0
        for quali_col in model.quali_factor:
            calibration[quali_col].to_excel(writer_Def,startrow=(8*s+1), startcol=2)
            #Writes to excel
            s+=1
    #convert format to PDModel class's attribute
    result=[]
    for quali_col in model.quali_factor:
        result.append(dict(dict(calibration[quali_col])['score']))

    return({'qualimapping':result})


def quanttrans(data, model, floor=0.05, cap=0.95):
    """
    This function calculates the floor, cap, mean and std of the data. The output
    can be used as the updated parameters for quantitative factors normailization.
    Modified in Ver. 1.4

    Parameters:

        data:   the input dataset in DataFrame. Make sure no NA in quant
                factors

        model:  PDModel class

        floor:  float, default 0.05
                quantile for floor 

        cap:    float, default 0.95
                quantile for cap 

        
    Return:
        a dictionary that saves floor, cap, doc_mean, doc_std
    
    """    
    normdata = data.copy()
    warnings.filterwarnings('ignore', category=Warning)

    # get new cap and floor from valid observations:
    floor_list=[];  cap_list=[]

    for i, factor in enumerate(model.quant_factor):
        if not model.Invalid_Neg[i]:
            floor_list.append(normdata[factor].quantile(floor))
            cap_list.append(normdata[factor].quantile(cap))
        else:
            negative = model.Invalid_Neg[i]
            floor_list.append(normdata[factor][normdata[negative]>0].quantile(floor))
            cap_list.append(normdata[factor][normdata[negative]>0].quantile(cap)) 


    # Invalid_Neg
    for i,neg_source in enumerate(model.Invalid_Neg):
        if neg_source:
            col=model.quant_factor[i]
            normdata[col][ (normdata[col]<0) & (normdata[neg_source]<0) ] = cap_list[i]
            normdata[col][ (normdata[col]<0) & pd.isnull(normdata[neg_source]) ] = cap_list[i]

    # cap/floor for quant factors:
    for i, col in enumerate(model.quant_factor):
        normdata[col] = np.clip(normdata[col], floor_list[i], cap_list[i])        

    # quant factors transformation:
    for i, col in enumerate(model.quant_factor):
        if model.quant_log[i]:
            normdata[col] = np.log(normdata[col])

    # get new mean and std:
    doc_mean=[];    doc_std=[]

    for i, col in enumerate(model.quant_factor):
        doc_mean.append(normdata[col].mean())     
        doc_std.append(normdata[col].std())     

    dictionary = {'floor':floor_list, 'cap':cap_list, 'doc_mean':doc_mean, 'doc_std':doc_std}
    
    return(dictionary)


def normalization(data, model, quant_only=False, missing='median'):
    """
    This function performs normalization on rawdata with the setting in class 'model'
    
    Parameters:
    
        data:       the input dataset in DataFrame. Make sure no NA in quant
                    factors

        model:      PDModel class


        quant_only: boolean, default 'False'. Added in Ver.1.6
                    set as 'True' if only normalize on quant factors.


        missing:    str, default 'median'. Added in Ver.1.4
                    statistics of the data used to fill missing data (NA data)
       


    Return:
        a DataFrame that has normalized factors.
    
    """    
    warnings.filterwarnings('ignore', category=Warning)
    # Invalid_Neg
    normdata = data.copy()
    # modify the loop below in Ver 1.4:
    for i,neg_source in enumerate(model.Invalid_Neg):
        if neg_source:
            col=model.quant_factor[i]
            normdata[col][ (normdata[col]<0) & (normdata[neg_source]<0) ] = model.cap[i]
            # treat NA in 'neg_source' as negative value
            normdata[col][ (normdata[col]<0) & pd.isnull(normdata[neg_source]) ] = model.cap[i]

    # cap/floor for quant factors:
    for i, col in enumerate(model.quant_factor):
        normdata[col] = np.clip(normdata[col], model.floor[i], model.cap[i])        

    # quant factors transformation:
    for i, col in enumerate(model.quant_factor):
        if model.quant_log[i]:
            normdata[col] = np.log(normdata[col])
            
    # quant factors normalization:
    for i, col in enumerate(model.quant_factor):
        normdata[col]  = 50*(normdata[col] - model.doc_mean[i]) / model.doc_std[i]      
        
    # quant factors flip sign:  
    for i, col in enumerate(model.quant_factor):
        normdata[col] = normdata[col] * model.quant_multiplier[i]     

    if quant_only:
        exec('normdata[model.quant_factor] = normdata[model.quant_factor].\
            fillna(normdata[model.quant_factor].{missing}())'.format(missing=missing))
    else:
        # calibration for quali factors:
        for i, col in enumerate(model.quali_factor):
            normdata[col].replace(to_replace=model.qualimapping[i],inplace=True) 
        # fill missing data. added in Ver. 1.4
        exec('normdata[model.quant_factor+model.quali_factor] = normdata[model.quant_factor+model.quali_factor].\
            fillna(normdata[model.quant_factor+model.quali_factor].{missing}())'.format(missing=missing))

    return normdata


def PD_frPDRR(data, model, PDRR, ms_ver='new'):
    """
    This function calculates implied PD according to the PDRR which user assigned.
    The underling MasterScale is the new version one. 
    
    Parameters:

        data:   the input dataset in DataFrame. Make sure no NA in quant
                factors

        model:  PDModel class

        PDRR:   str
                column name for PDRR. ex. PDRR='Final_PD_Risk_Rating'
        
        ms_ver: str, default 'new'
                version of master default 'old' or 'new'
    Return:
        a DataFrame which is a copy of input data but has one more column called
        'PD_frPDRR'.
    
    """    
    dat = data.copy()
    if ms_ver=='old':
        mid_pd = model.MS['old_mid']
    else:
        mid_pd = model.MS['new_mid']

    MS_dict = dict(zip(model.MS['PDRR'], mid_pd))

    dat['PD_frPDRR'] = data[PDRR].transform(lambda x: np.nan if pd.isnull(x) else MS_dict[x]) 
    return (dat)


def logitPD_frPDRR(data, model, PDRR, ms_ver='new'):
    """
    This function calculates implied logitPD according to the PDRR which user assigned.
    The underling MasterScale is the new version one.
    
    Parameters:

        data:   the input dataset in DataFrame. Make sure no NA in quant
                factors

        model:  PDModel class

        PDRR:   str
                column name for PDRR. ex. PDRR='Final_PD_Risk_Rating'

        ms_ver: str, default 'new'
                version of master default 'old' or 'new'
    Return:
        a DataFrame which is a copy of input data but has one more column called
        'logitPD_frPDRR'.
    
    """        
    dat = data.copy()
    if ms_ver=='old':
        mid_pd = model.MS['old_mid']
    else:
        mid_pd = model.MS['new_mid']

    MS_dict = dict(zip(model.MS['PDRR'], mid_pd))
    pl_pd = data[PDRR].transform(lambda x: np.nan if pd.isnull(x) else MS_dict[x]).tolist()

    # convert PD to logit PD
    dat['logitPD_frPDRR'] = [np.log(x/(1-x)) for x in pl_pd]
    return (dat)


def PD_frPDRR_autoMS(data, model, PDRR, timestamp='archive_date'):  # add in ver. 2.1     

    """
    This function calculates implied PD according to the PDRR which user assigned.
    The underling MasterScale is automatically assigned by archive_date
    (New MS is in production from 05/01/2016)
    
    Parameters:

        data:       the input dataset in DataFrame. Make sure no NA in quant
                    factors
    
        model:      PDModel class
    
        PDRR:       str
                    column name for PDRR. ex. PDRR='Final_PD_Risk_Rating'
        
        timestamp:  str, default 'archive_date'
                    new MS apply after 20160501.
    Return:
        a DataFrame which is a copy of input data but has one more column called
        'PD_frPDRR'.
    
    """    

    dat_old = data.query('{}<20160501'.format(timestamp))
    dat_new = data.query('{}>=20160501'.format(timestamp))
    old_mid = model.MS['old_mid']
    new_mid = model.MS['new_mid']
    old_MS_dict = dict(zip(model.MS['PDRR'], old_mid))
    new_MS_dict = dict(zip(model.MS['PDRR'], new_mid))

    dat_old['PD_frPDRR'] = dat_old[PDRR].transform(lambda x: np.nan if pd.isnull(x) else old_MS_dict[x]) 
    dat_new['PD_frPDRR'] = dat_new[PDRR].transform(lambda x: np.nan if pd.isnull(x) else new_MS_dict[x]) 

    return (pd.concat([dat_old, dat_new], axis=0).sort_index())


def logitPD_frPDRR_autoMS(data, model, PDRR, timestamp='archive_date'): # add in ver. 2.1

    """
    This function calculates implied logitPD according to the PDRR which user assigned.
    The underling MasterScale is automatically assigned by archive_date
    (New MS is in production from 05/01/2016)
    
    Parameters:

        data:       the input dataset in DataFrame. Make sure no NA in quant
                    factors
    
        model:      PDModel class
    
        PDRR:       str
                    column name for PDRR. ex. PDRR='Final_PD_Risk_Rating'
        
        timestamp:  str, default 'archive_date'
                    new MS apply after 20160501.
    Return:
        a DataFrame which is a copy of input data but has one more column called
        'logitPD_frPDRR'.
    
    """    

    dat_old = data.query('{}<20160501'.format(timestamp))
    dat_new = data.query('{}>=20160501'.format(timestamp))
    old_mid = model.MS['old_mid']
    new_mid = model.MS['new_mid']
    old_MS_dict = dict(zip(model.MS['PDRR'], old_mid))
    new_MS_dict = dict(zip(model.MS['PDRR'], new_mid))
    pl_pd_old = dat_old[PDRR].transform(lambda x: np.nan if pd.isnull(x) else old_MS_dict[x]) 
    pl_pd_new = dat_new[PDRR].transform(lambda x: np.nan if pd.isnull(x) else new_MS_dict[x]) 

    # convert PD to logit PD
    dat_old['logitPD_frPDRR'] = [np.log(x/(1-x)) for x in pl_pd_old]
    dat_new['logitPD_frPDRR'] = [np.log(x/(1-x)) for x in pl_pd_new]

    return (pd.concat([dat_old, dat_new], axis=0).sort_index())


def getTotalscore(data, model):
    """
    This function applies 'model' on the 'data': Goes through normalization, weighted
    summation to get model's total score for each obligor.
    
    Parameters:

        data:   the input dataset in DataFrame. Make sure no NA in quant
                factors

        model:  PDModel class


    Return:
        a DataFrame which is a copy of input data but has more columns. The result
        is in column 'Ratings'.
    
    """    
    dat = normalization(data,model)

    if (model.quant_factor and model.quali_factor): # That's mean the model have both quant and quali factors
        dat['quantscore'] = (model.quant_weight * dat[model.quant_factor].values).sum(axis=1)
        dat['quantscore'] = 50*( dat['quantscore'] - model.quantmean) / model.quantstd
        dat['qualiscore'] = (model.quali_weight * dat[model.quali_factor].values).sum(axis=1)
        dat['qualiscore'] = 50*( dat['qualiscore'] - model.qualimean) / model.qualistd
        dat['Totalscore'] = dat['quantscore']*model.quantweight + dat['qualiscore'] *model.qualiweight
    else: # quant factors only. New in Version 1.5
        dat['Totalscore'] =  (model.quant_weight * dat[model.quant_factor].values).sum(axis=1)

    return (dat)


def getPDRR(data, model, ms_ver='new'):
    """
    This function applies 'model' on the 'data': Goes through normalization, weighted
    summation, calibration to get model PD for each obligor and uses the 'MasterScale'
    in 'model' to get the model PDRR (or 'Prelim_PD_Risk_Rating_Uncap')
    
    Parameters:

        data:   the input dataset in DataFrame. Make sure no NA in quant
                factors

        model:  PDModel class

        ms_ver: str, default 'new' New in Version 2.0
                version of master default 'old' or 'new'
    Return:
        a DataFrame which is a copy of input data but has more columns. The result
        is in column 'Ratings'.
    
    """    
    dat = normalization(data,model)

    if (model.quant_factor and model.quali_factor): # That's mean the model have both quant and quali factors
        dat['quantscore'] = (model.quant_weight * dat[model.quant_factor].values).sum(axis=1)
        dat['quantscore'] = 50*( dat['quantscore'] - model.quantmean) / model.quantstd
        dat['qualiscore'] = (model.quali_weight * dat[model.quali_factor].values).sum(axis=1)
        dat['qualiscore'] = 50*( dat['qualiscore'] - model.qualimean) / model.qualistd
        dat['score'] = dat['quantscore']*model.quantweight + dat['qualiscore'] *model.qualiweight
    else: # quant factors only. New in Version 1.5
        dat['score'] =  (model.quant_weight * dat[model.quant_factor].values).sum(axis=1)

    logitPD = []
    for obs in dat.iterrows():
        if obs[1].score < model.cutoff:
            logitPD.append(model.intercept1 + model.slope1*obs[1]['score'])
        else:
            logitPD.append(model.intercept2 + model.slope2*obs[1]['score'])
    dat['logitPD'] = logitPD

    dat['PD'] = dat['logitPD'].apply(lambda x: 100*np.exp(x)/(1+np.exp(x)))
    Ratings = []
    if ms_ver=='old':
        low_pd = model.MS['old_low']
    else:
        low_pd = model.MS['new_low']

    for i in dat.iterrows():
        Ratings.append(sum(low_pd<=(i[1].PD/100)))
    dat['Ratings'] = Ratings
    return (dat)


def getTM(data, model1, model2, ms_ver='new', PDRR_range=(1,20)):
    """
    This function generates Transition Matrix in an Excel file. The TM is coming
    from the same data but different model settings.
    
    Parameters:

        data:       the input dataset in DataFrame. Make sure no NA in quant
                    factors

        model1:     PDModel class
                    Model 1's setting

        model2:     PDModel class
                    Model 2's setting  

        ms_ver:     str, default 'new' New in Version 2.0
                    version of master default 'old' or 'new'

        PDRR_range: tuple, default (1,20)
                    the range of PDRR, (min rating, max rating)
        
    Return:
        An excel file called 'Matrix_Output.xlsx' 
    
    """    

    dat1 = getPDRR(data, model1, ms_ver)
    dat2 = getPDRR(data, model2, ms_ver)
    dat = pd.concat([dat1.Ratings, dat2.Ratings], axis=1)
    dat.columns = pd.Index(['PDRR_model1','PDRR_model2'])
    CreateBenchmarkMatrix(dat, 'Matrix_Output.xlsx', 'Benchmarking Matrix', 'PDRR_model1', 'PDRR_model2', PDRR=range(PDRR_range[0],PDRR_range[1]+1))

    
def getTMnotches(data, model1, model2, ms_ver='new', PDRR_range=(1,20)):
    """
    This function calculates two statistics from function getTM(): 
    'down_notches' and 'up_notches'. They're used to measure the mismatch.
    
    Parameters:

        data:       the input dataset in DataFrame. Make sure no NA in quant
                    factors

        model1:     PDModel class
                    Model 1's setting

        model2:     PDModel class
                    Model 2's setting  

        ms_ver:     str, default 'new' New in Version 2.0
                    version of master default 'old' or 'new'
                          
        PDRR_range: tuple, default (1,20)
                    the range of PDRR, (min rating, max rating)
        
    Return:
        tuple: (down_notches, up_notches)
    
    """
    dat1 = getPDRR(data, model1, ms_ver)
    dat2 = getPDRR(data, model2, ms_ver)
    dat = pd.concat([dat1.Ratings, dat2.Ratings], axis=1)
    dat.columns = pd.Index(['PDRR_model1','PDRR_model2'])
    return(TMnotches(dat,'PDRR_model1', 'PDRR_model2', PDRR=range(PDRR_range[0],PDRR_range[1]+1)))


def getwithknotchesrate(data, model1, model2, ms_ver='new', PDRR_range=(1,20), k=2):
    """
    This function calculates the rate that with k notches matches.
    
    Parameters:

        data:       the input dataset in DataFrame. Make sure no NA in quant
                    factors

        model1:     PDModel class
                    Model 1's setting

        model2:     PDModel class
                    Model 2's setting  
        
        ms_ver:     str, default 'new' New in Version 2.0
                    version of master default 'old' or 'new'
                                              
        PDRR_range: tuple, default (1,20)
                    the range of PDRR, (min rating, max rating)

        k:          int, default 2
                    the rate is within k notches
    Return:
        tuple: (down_notches, up_notches)
    
    """
    dat1 = getPDRR(data, model1, ms_ver)
    dat2 = getPDRR(data, model2, ms_ver)
    dat = pd.concat([dat1.Ratings, dat2.Ratings], axis=1)
    dat.columns = pd.Index(['PDRR_model1','PDRR_model2'])
    return(withknotchesrate(dat,'PDRR_model1', 'PDRR_model2', PDRR=range(PDRR_range[0],PDRR_range[1]+1),k=k))
 

def buildpseudodef(data, def_flag='def_flag', PDRR='Final_PD_Risk_Rating', pseudodef_PDRR=[13,14,15]):
    """
    This function constructs pseudo defaults according the raings
    
    Parameters:

        data:           the input dataset in DataFrame. Make sure no NA in quant
                        factors

        def_flag:       str, default 'def_flag'
                        column name for default flag

        PDRR:           str, default 'Final_PD_Risk_Rating'
                        column name for PDRR
                                  
        pseudodef_PDRR: list, default [13,14,15]
                        the range of PDRR that treated as defaulted
        
    Return:
        DataFrame: Data with pseudo defaults.

    """
    warnings.filterwarnings('ignore', category=Warning)
    dat = data.copy()
    n_truedef = int(dat[def_flag].sum())
    print ("There're {n} true defaults in the input dateset.".format(n=n_truedef))
    dat2 = dat[dat[PDRR].isin(pseudodef_PDRR)]
    dat2[def_flag]=1
    dat[dat[PDRR].isin(pseudodef_PDRR)] = dat2
    n_totaldef = int(dat[def_flag].sum())
    n_pseudodef = n_totaldef - n_truedef
    print ("There're {n} defaults in the output dateset. Among them, {m} are pseudo defaults.".\
        format(n=n_totaldef, m=n_pseudodef))

    return(dat)



def buildpseudodef2(data, **kw):  # add in ver. 1.7
    """
    This function constructs pseudo defaults based on the multiple conditions:
    1. current (at time t) PDRR is at 'pseudodef_PDRR'
    2. the downgrade gap between time t and time t-1 is equal or larger than 'gap_PDRR'
    
    Parameters:

        data:           the input dataset in DataFrame. Make sure no NA in quant
                        factors

        def_flag:       str, default 'def_flag'
                        column name for default flag

        PDRR:           str, default 'Final_PD_Risk_Rating'
                        column name for PDRR

        pseudodef_PDRR: list, default [13,14]
                        the range of PDRR that treated as defaulted

        gap_PDRR:       int, default 2
                        the gap between PDRR at t-1 and PDRR at t 

        timestamp:      str, default 'archive_date'
                        column name for timestamp
        
        idn:            str, default 'CUSTOMERID'
                        column name for customer identifier

    Return:
        DataFrame: Data with pseudo defaults.

    """
    def_flag =          kw.pop('def_flag','def_flag')
    PDRR =              kw.pop('PDRR','Final_PD_Risk_Rating')
    pseudodef_PDRR =    kw.pop('pseudodef_PDRR',[13,14])
    gap_PDRR =          kw.pop('gap_PDRR',2 )
    timestamp =         kw.pop('timestamp','archive_date')
    idn =               kw.pop('idn','CUSTOMERID')


    #warnings.filterwarnings('ignore', category=Warning)
    dat = data.copy()
    n_truedef = int(dat[def_flag].sum())
    print ("There're {n} true defaults in the input dateset.".format(n=n_truedef))

    dat.sort_values(by=[idn,timestamp], inplace=True)
    dat.reset_index(drop=True, inplace=True)
    idlist = dat[idn].unique().tolist()

    id_pseudo=[] # stores all indices of 'dat' that match our criterion
    for id in idlist:
        temp = dat[dat[idn]==id]
        for i in range(len(temp)-1):      
            diff = temp[PDRR].iloc[i+1]-temp[PDRR].iloc[i]
            if (diff>=gap_PDRR) & (temp[PDRR].iloc[i+1] in pseudodef_PDRR):
                id_pseudo.append(temp.index[i+1])
            else:
                continue

    for id in id_pseudo:
        temp = dat.loc[id]
        temp.def_flag = True
        dat.loc[id] = temp

    n_totaldef = int(dat[def_flag].sum())
    n_pseudodef = n_totaldef - n_truedef
    print ("There're {n} defaults in the output dateset. Among them, {m} are pseudo defaults.".\
        format(n=n_totaldef, m=n_pseudodef))

    return(dat)


@jit
def SomersD(y_true, y_score, sign=1, unit_concordance=True): 
    '''

    New version of SomersD function which leverages numba.jit to accelerate
    the calculation. In production from Ver. 1.6.

    '''
    
    #x = np.array(y_score)
    #y = np.array(y_true)
    # mofified in Ver 2.7. To make it auto drop missing value before calculation
    dat=pd.DataFrame()                        #new lines in Ver 2.7                 
    dat['true'] = y_true                      #new lines in Ver 2.7                 
    dat['score'] = y_score                    #new lines in Ver 2.7                
    dat.dropna(how='any', inplace=True)       #new lines in Ver 2.7                   
    x = np.array(dat.score)                   #new lines in Ver 2.7                 
    y = np.array(dat.true)                    #new lines in Ver 2.7              

    n = len(x)
    C = 0
    D = 0
    if unit_concordance==True:
        for i in range(n):
            for j in range(i):
                dx = x[i] - x[j]
                dy = y[i] - y[j]
                if (dx*dy) > 0:
                    cij = 1
                    dij = 1
                elif  (dx*dy) < 0:
                    cij = -1
                    dij = 1
                else:
                    cij = 0
                    if dy==0: 
                        dij = 0
                    else: 
                        dij=1
                C = C + cij
                D = D + dij
    else:
        for i in range(n):
            for j in range(i):
                dx = x[i] - x[j]
                dy = y[i] - y[j]
                if  (dx*dy) > 0:
                    cij = abs(dy)
                    dij = abs(dy)
                elif (dx*dy) < 0:
                    cij = -abs(dy)
                    dij = abs(dy)
                else:
                    cij = 0
                    if dy==0: 
                        dij = 0
                    else: 
                        dij=1
                C = C + cij
                D = D + dij            
    return sign*C/D


def RAreplica(data, model, ms_ver='new'):
    """
    This function applies 'model' on the 'data': Goes through normalization, weighted
    summation, calibration to get model PD for each obligor and uses the 'MasterScale'
    in 'model' to get the PDRR and all intermediate results. New in Version 1.5.
    
    Parameters:

        data:   the input dataset in DataFrame. Make sure no NA in quant
                factors

        model:  PDModel class

        ms_ver: str, default 'new'
                the masterscale version to use.  {‘old’, ‘new’, timecutoff}
                'old':      old version ms which was used before April,2016
                'new':      new version ms which was used after April,2016
                timecutoff: e.x. '2016/05/01'
                            the function will apply old version ms for all observations 
                            if its 'archive_date' is before this date; and apply new 
                            version ms for all observations if its 'archive_date' is on 
                            or after this date. Note that: user MUST make sure all observations
                            have valid 'archive_date'. 

    Return:
        a DataFrame which is a copy of input data but has more columns. And the function will
        export the DataFrame into an Excel File 'RAreplica.xlsx' into current working directory. 
    
    """    
    dat = normalization(data,model)

    if (model.quant_factor and model.quali_factor): # That's mean the model have both quant and quali factors
        dat['quantscore'] = (model.quant_weight * dat[model.quant_factor].values).sum(axis=1)
        dat['quantscore'] = 50*( dat['quantscore'] - model.quantmean) / model.quantstd
        dat['qualiscore'] = (model.quali_weight * dat[model.quali_factor].values).sum(axis=1)
        dat['qualiscore'] = 50*( dat['qualiscore'] - model.qualimean) / model.qualistd
        dat['finalscore'] = dat['quantscore']*model.quantweight + dat['qualiscore'] *model.qualiweight
    else: # quant factors only
        dat['finalscore'] =  (model.quant_weight * dat[model.quant_factor].values).sum(axis=1)

    logitPD = []
    for obs in dat.iterrows():
        if obs[1].finalscore < model.cutoff:
            logitPD.append(model.intercept1 + model.slope1*obs[1]['finalscore'])
        else:
            logitPD.append(model.intercept2 + model.slope2*obs[1]['finalscore'])
    dat['logitPD'] = logitPD
    dat['PD'] = dat['logitPD'].apply(lambda x: 100*np.exp(x)/(1+np.exp(x)))

    Ratings = []
    if ms_ver=='new':
        ms_used = model.MS['new_low']
        for i in dat.iterrows():
            Ratings.append(sum(ms_used<=(i[1].PD/100)))
        dat['Ratings'] = Ratings
    elif ms_ver=='old':
        ms_used = model.MS['old_low']
        for i in dat.iterrows():
            Ratings.append(sum(ms_used<=(i[1].PD/100)))
        dat['Ratings'] = Ratings
    else:
        for i in dat.iterrows():
            if i[1].archive_date >= pd.to_datetime(ms_ver): #use new masterscale
                Ratings.append(sum(model.MS['new_low']<=(i[1].PD/100)))
            else: # use old masterscale
                Ratings.append(sum(model.MS['old_low']<=(i[1].PD/100)))

    dat['Ratings'] = Ratings
    dat.to_excel('RAreplica.xlsx')

    return (dat)


def MAUG_mapping(data, col_guidance='Underwriter_Guideline'):
    """
    This function is to get MAUG code based on 'col_guidance'.(usually, its name is 'Underwriter_Guideline'
    in RA) And then mapping MAUG to industry based on doc 'MAUG-0100 UG to MAUG Mapping.pdf'. New in Version 1.8

    Parameters:

        data:           the input dataset in DataFrame. 

        col_guidance:   str, default 'Underwriter_Guideline'
                        column name

    Return:
        a DataFrame which is a copy of input data but has 2 more columns. 'MAUG' and 'Industry_by_MAUG'
    
    """    
    processed_data = data.copy()
    processed_data['MAUG'] = ['MAUG_'+str(x) for x in processed_data[col_guidance]] # modified on Ver.2.5
        
    # 440 is old UG code and it still exist after 20160723
    processed_data['MAUG'] = processed_data['MAUG'].replace(to_replace={'MAUG_440':'MAUG_125'})
    # 490 is old UG code and it still exist after 20160723
    processed_data['MAUG'] = processed_data['MAUG'].replace(to_replace={'MAUG_490':'MAUG_305'})
    # 705 is old UG code and it still exist after 20160723
    processed_data['MAUG'] = processed_data['MAUG'].replace(to_replace={'MAUG_705':'MAUG_340'})
    # 720 is old UG code and it still exist after 20160723
    processed_data['MAUG'] = processed_data['MAUG'].replace(to_replace={'MAUG_720':'MAUG_350'})
    # 835 is old UG code and it still exist after 20160723
    processed_data['MAUG'] = processed_data['MAUG'].replace(to_replace={'MAUG_835':'MAUG_325'})
    # 150-1 is old UG code and it still exist after 20160723
    processed_data['MAUG'] = processed_data['MAUG'].replace(to_replace={'MAUG_150-1':'MAUG_105'})

    processed_data['MAUG'] = processed_data['MAUG'].replace(to_replace={'MAUG_227':'MAUG_115'})

    processed_data['MAUG'] = processed_data['MAUG'].replace(to_replace={'MAUG_240':'MAUG_115'})

    processed_data['MAUG'] = processed_data['MAUG'].replace(to_replace={'MAUG_245':'MAUG_115'})

    processed_data['MAUG'] = processed_data['MAUG'].replace(to_replace={'MAUG_340-1':'MAUG_105'})

    processed_data['MAUG'] = processed_data['MAUG'].replace(to_replace={'MAUG_408':'MAUG_160'})

    processed_data['MAUG'] = processed_data['MAUG'].replace(to_replace={'MAUG_450':'MAUG_105'})

    processed_data['MAUG'] = processed_data['MAUG'].replace(to_replace={'MAUG_455':'MAUG_205'})

    processed_data['MAUG'] = processed_data['MAUG'].replace(to_replace={'MAUG_470':'MAUG_180'})

    processed_data['MAUG'] = processed_data['MAUG'].replace(to_replace={'MAUG_480':'MAUG_165'})

    processed_data['MAUG'] = processed_data['MAUG'].replace(to_replace={'MAUG_810':np.nan})

    processed_data['MAUG'] = processed_data['MAUG'].replace(to_replace={'MAUG_820':np.nan})

    processed_data['MAUG'] = processed_data['MAUG'].replace(to_replace={'MAUG_840':np.nan})

    processed_data['MAUG'] = processed_data['MAUG'].replace(to_replace={'MAUG_230':np.nan})

    processed_data['MAUG'] = processed_data['MAUG'].replace(to_replace={'MAUG_235':np.nan})

    processed_data['MAUG'] = processed_data['MAUG'].replace(to_replace={'MAUG_750-01':np.nan})

    processed_data['MAUG'] = processed_data['MAUG'].replace(to_replace={'MAUG_465':np.nan})

    processed_data['MAUG'] = processed_data['MAUG'].replace(to_replace={'MAUG_CBUG-185':'MAUG_180'})

    processed_data['MAUG'] = processed_data['MAUG'].replace(to_replace={'MAUG_481':'MAUG_105'})

    processed_data['MAUG'] = processed_data['MAUG'].replace(to_replace={'MAUG_806':'MAUG_395'})

    processed_data['MAUG'] = processed_data['MAUG'].replace(to_replace={'MAUG_730':'MAUG_350'})

    processed_data['MAUG'] = processed_data['MAUG'].replace(to_replace={'MAUG_445':'MAUG_105'})

    processed_data['MAUG'] = processed_data['MAUG'].replace(to_replace={'MAUG_405-01':'MAUG_155'})

    processed_data['MAUG'] = processed_data['MAUG'].replace(to_replace={'MAUG_405-02':'MAUG_155'})

    processed_data['MAUG'] = processed_data['MAUG'].replace(to_replace={'MAUG_405-03':'MAUG_155'})

    processed_data['MAUG'] = processed_data['MAUG'].replace(to_replace={'MAUG_910':'MAUG_435'})

    processed_data['MAUG'] = processed_data['MAUG'].replace(to_replace={'MAUG_905':'MAUG_105'})

    processed_data['MAUG'] = processed_data['MAUG'].replace(to_replace={'MAUG_850':'MAUG_105'})

    processed_data['MAUG'] = processed_data['MAUG'].replace(to_replace={'MAUG_225':'MAUG_115'})

    processed_data['MAUG'] = processed_data['MAUG'].replace(to_replace={'MAUG_837':'MAUG_320'})

    processed_data['MAUG'] = processed_data['MAUG'].replace(to_replace={'MAUG_855':'MAUG_105'})

    processed_data['MAUG'] = processed_data['MAUG'].replace(to_replace={'MAUG_626':'MAUG_355'})

    processed_data['MAUG'] = processed_data['MAUG'].replace(to_replace={'MAUG_250':'MAUG_115'})

    processed_data['MAUG'] = processed_data['MAUG'].replace(to_replace={'MAUG_805':'MAUG_415'})

    processed_data['MAUG'] = processed_data['MAUG'].replace(to_replace={'MAUG_845':'MAUG_390'})

    processed_data['MAUG'] = processed_data['MAUG'].replace(to_replace={'MAUG_485':'MAUG_310'})

    processed_data['MAUG'] = processed_data['MAUG'].replace(to_replace={'MAUG_CBUG-135':'MAUG_105'})

    processed_data['MAUG'] = processed_data['MAUG'].replace(to_replace={'MAUG_425-1':'MAUG_110'})

    processed_data['MAUG'] = processed_data['MAUG'].replace(to_replace={'MAUG_CBUG-190':'MAUG_165'})

    processed_data['MAUG'] = processed_data['MAUG'].replace(to_replace={'MAUG_CBUG-125':'MAUG_175'})

    processed_data['MAUG'] = processed_data['MAUG'].replace(to_replace={'MAUG_CBUG-130':'MAUG_120'})

    processed_data['MAUG'] = processed_data['MAUG'].replace(to_replace={'MAUG_340-2':'MAUG_330'})


    sofar_mapping = {
    'MAUG_105': 'General Industries',
    'MAUG_110': 'Entertainment',
    'MAUG_115': 'Media and Telecom',
    'MAUG_120': 'Food and Beverage',
    'MAUG_125': 'Health Care',
    'MAUG_130': 'Insurance',
    'MAUG_135': 'Power & Utilities',
    'MAUG_140': 'Independent Exploration and Production',
    'MAUG_143': 'Midstream Energy',
    'MAUG_145': 'Independent Refining',
    'MAUG_147': 'Drilling and Oilfield Services',
    'MAUG_150': 'Auto and Auto Parts',
    'MAUG_155': 'Agribusiness',
    'MAUG_160': 'Wine Industry',
    'MAUG_165': 'Technology',
    'MAUG_170': 'Leasing Companies',
    'MAUG_175': 'Engineering and Construction',
    'MAUG_180': 'Retail',
    'MAUG_205': 'Asian Corporate Banking',
    'MAUG_210': 'Japanese Corporate Banking in Canada',
    'MAUG_215': 'Asian Corporate Leasing & Finance',
    'MAUG_305': 'Business Banking',
    'MAUG_307': 'Small Business Banking',
    'MAUG_310': 'Business Diversity Lending',
    'MAUG_320': 'Corporate Leasing Transactions',
    'MAUG_325': 'Leasing and Asset Finance Division',
    'MAUG_330': 'Project Finance',
    'MAUG_333': 'Commodity Finance',
    'MAUG_335': 'Mezzanine Finance',
    'MAUG_340': 'Commercial Finance Loans',
    'MAUG_345': 'Commodity and Structured Trade Finance',
    'MAUG_350': 'Trading Asset Reliant',
    'MAUG_355': 'Capital Call Bridge / Subscription Lending',
    'MAUG_360': 'Securities Broker / Dealers',
    'MAUG_365': 'Asset Managers',
    'MAUG_367': 'Investment Funds (Mutual Funds)',
    'MAUG_370': 'Clearinghouses and Exchanges',
    'MAUG_375': 'Industrial Development Bonds',
    'MAUG_380': 'Insured Domestic Institutions',
    'MAUG_381': 'Foreign Banks',
    'MAUG_383': 'Global Financial Solutions',
    'MAUG_385': 'Securitization',
    'MAUG_390': 'Debt Obligation and Passive Equity Investments',
    'MAUG_395': 'Individuals',
    'MAUG_405': 'Real Estate Industries',
    'MAUG_410': 'Institutional Real Estate Lending',
    'MAUG_415': 'Homeowner Association Loans',
    'MAUG_420': 'Retail Equity Equivalent Loan Program',
    'MAUG_425': 'Debtor-in-Possession Loans',
    'MAUG_430': 'Finance and Mortgage Companies',
    'MAUG_435': 'Investor Perm Loans',
    'MAUG_505': 'Community Development Financing',
    'MAUG_510': 'Public Finance',
    'MAUG_515': 'Non-Profit Organizations',
    'MAUG_605': 'Brazil Agribusiness',
    'MAUG_610': 'Placeholder for pending General Corporate Latin America',
    'MAUG_615': 'Placeholder for pending Retail Latin America',
    'MAUG_620': 'Placeholder for pending Utilities Latin America',
    'MAUG_625': 'Placeholder for pending Oil & Gas Latin America',
    'MAUG_630': 'Placeholder for pending Metals and Mining Latin America',
    'MAUG_635': 'Placeholder for pending Food and Beverage Latin America'
    }

    processed_data['Industry_by_MAUG'] = processed_data['MAUG'].replace(sofar_mapping)

    return (processed_data)


def NAICS_mapping(data, col_guidance='NAICS_Cd', first_digits=2):
    """
    This function is to convert NAICS code to industry. New in Version 1.8. Major
    update in Version 2.5 to make it work for different granular level mapping

    Parameters:

        data:           the input dataset in DataFrame. 

        col_guidance:   str, default 'NAICS_Cd'
                        column name

        first_digits:   integer from 2 to 6, default 2. 
                        First digits used for the industry mapping

    Return:
        a DataFrame which is a copy of input data but has one more column: 'Industry_by_NAICS'
    
    """    
    processed_data = data.copy()
    processed_data['Industry_by_NAICS'] = processed_data[col_guidance].fillna(999999)
    processed_data['Industry_by_NAICS'] = [str(int(x))[:first_digits] for x in processed_data['Industry_by_NAICS']]
    processed_data['Industry_by_NAICS'] = processed_data['Industry_by_NAICS'].replace(naics_code)
    
    # to deal with 'Others':
    pl_code = [str(x) for x in range(0,10**first_digits) if str(x) not in list(naics_code.keys()) ]
    pd_others = dict(zip(pl_code,['Unknow_NAICS_Code',]*len(pl_code)))
    processed_data['Industry_by_NAICS'] = processed_data['Industry_by_NAICS'].replace(pd_others)

    return (processed_data)


def ExtRating_mapping(data, col_ext_rating='SPRating'):
    """
    This function is to convert external S&P rating to internal PDRR. New in Version 1.9

    Parameters:

        data:               the input dataset in DataFrame. 

        col_ext_rating:     str, default 'SPRating'
                            column name for S&P rating

    Return:
        a DataFrame which is a copy of input data but has one more column: 'ExternalRating'
    
    """   
    processed_data = data.copy()

    ext_ratings_mapping={
    'AAA':2,
    'AA+':2,
    'AA':2,
    'AA-':2, # modified on 1/31/2019, previous is '3'
    'A+':3,
    'A':3,
    'A-':4,
    'BBB+':5,
    'BBB':6,
    'BBB-':7,
    'BB+':8,
    'BB':9,  
    'BB-':10,
    'B+':11,
    'B':12,
    'B-':13,
    'CCC+':14, # modified on 10/314/2020, previous is '15'
    'CCC':15,
    'CCC-':15,
    'CC':15,


    }

    processed_data['ExternalRating'] = processed_data[col_ext_rating]
    processed_data['ExternalRating'] = processed_data['ExternalRating'].replace(ext_ratings_mapping)
    return (processed_data)


def gcar_converter(data, cols, inplace=False, ignore_pct=True): # Added in Ver. 2.2
    '''
    This function is to convert GCAR data's financial statement into float. New in Version 2.2

    Parameters:

        data:           the input dataset in DataFrame. 
    
        cols:           str, or list of str
                        column name for the conversion
    
        inplace:        bool, default False
                        If True, do operation inplace and return None.

        ignore_pct:     bool, default True. New in Version 2.8
                        If True, convert '123.45%' to 123.45        
                        If False, convert '123.45%' to 1.2345                

    Return:
        DataFrame with converted columns.
    '''
    def _mycast(s):

        if pd.isnull(s):
            return(np.nan)

        elif isinstance(s, float) or isinstance(s, int):
            return (s)

        elif s[-1]=='%' and ignore_pct:  # to convert '123.45%' to 123.45
            if ',' in s:
                return (locale.atof(s[:-1]))
            else:
                return (float(s[:-1]))

        elif s[-1]=='%' and not ignore_pct:  # to convert '123.45%' to 1.2345
            if ',' in s:
                return (locale.atof(s[:-1])/100)
            else:
                return (float(s[:-1])/100)

        elif ',' in s:  # to convert '123,345.000' to 123456.0
            return (locale.atof(s))

        else:
            try: 
                return (float(s))
            except ValueError:
                return(np.nan)

    if isinstance(cols, str):
        cols=[cols,]

    if inplace:
        data[cols] = data[cols].applymap(_mycast)
    else:
        dat = data.copy()
        dat[cols] = dat[cols].applymap(_mycast)
        return (dat)



def gcar_cif_cleaner(data, col='Borrower CIF', inplace=False): # Added in Ver. 2.4
    '''
    This function is to convert GCAR data's CIF into correct string format. New in Version 2.4

    Parameters:

        data:       the input dataset in DataFrame. 

        col:        str, default 'Borrower CIF'
                    column name for the CIF

        inplace:    bool, default False
                    If True, do operation inplace and return None.

    Return:
        DataFrame with converted columns.
    '''
    def _my_cast(cif):
        if cif[0]!='R':
            return (f'{int(cif):08d}')
        else:
            return (cif)

    if inplace:
        data[col] = data[col].astype(str)
        data[col] = data[col].apply(_my_cast)
    else:
        tmp_series = data[col].astype(str)
        return (tmp_series.apply(_my_cast))


# Added in Ver. 2.4
gcar_Americas_office_code = (
3103,3104,3108,3110,3116,3120,3134,3138,3147,3149,3153,3157,3158,3160,
3161,3165,3170,3191,3220,3500,3770,
3281,3282,3286,
3250,3260) # add on 06/10/2020


gcar_EMEA_office_code = (
3330, 3332, 3461, 6277, 3591, 6276, 3337, 3340, 3341, 3342, 3343, 33422, 
3345, 6412, 3219, 3348, 3349, 6415, 3351, 6418, 3225, 3226, 3354, 3356, 
3357, 6430, 3231, 3232, 3361, 3234, 3233, 3364, 3237, 3238, 3239, 3235, 
3236, 3240, 3371, 6444, 6447, 3377, 6457, 6423, 6334, 6335, 6207, 6337, 
6338, 3523, 3524, 3525, 3526, 7232, 7233, 6990, 6227, 6359, 6360, 6362, 
6363, 3550, 3424, 6431, 3301, 6246, 3304, 3561, 3305, 6249, 6762, 3309, 
3310, 6434, 3321)  # add on 09/01/2020


def func_sd(data, target_col='Final_PD_Risk_Rating', score_col='Prelim_PD_Risk_Rating', min_num=6):
    '''
    This function is to calculate SomersD between target_col and score_col.
    And it's designed for 'GroupBy.apply' and applied to group-wise object.
    New in Version 2.6

    Parameters:

        data:           the input dataset in DataFrame. 

        target_col:     str, default 'Final_PD_Risk_Rating'
                        column name for the targeted column.

        score_col:      str, default 'Prelim_PD_Risk_Rating'
                        column name for the score column.

        min_num:        int, default 6
                        the threshold. If the number of obs is less than it, the result is NA 
                        when the number of obs less than 

    Return:
        combined results is used in GroupBy.apply
    '''
    if len(data) < min_num:
        return np.nan
    else:
        return SomersD(data[target_col], data[score_col])




def func_rla(data, rla_col='RLA_Notches'):
    '''
    This function is to calculate RLA rate.
    And it's designed for 'GroupBy.apply' and applied to group-wise object.
    New in Version 2.6

    Parameters:

        data:       the input dataset in DataFrame. 

        rla_col:    str, default 'RLA_Notches'
                    column name for RLA column.

    Return:
        combined results is used in GroupBy.apply
    '''
    N = len(data)
    n_RLA = len(data.query(f'{rla_col}!=0'))    
    return (n_RLA/N)


def func_ovd(data, ovd_col='Override_Action'):
    '''
    This function is to calculate Override rate.
    And it's designed for 'GroupBy.apply' and applied to group-wise object.
    New in Version 2.6

    Parameters:

        data:       the input dataset in DataFrame. 

        ovd_col:    str, default 'Override_Action'
                    column name for Override.

    Return:
        combined results is used in GroupBy.apply
    '''

    N = len(data)   
    n_Override = len(data.query(f'{ovd_col}!=0'))
    return (n_Override/N)



#  Blueberry, Cinnabar, Selective Yellow, Sea Green
google_color = ['#4285F4','#EA4335','#FBBC05','#34A853']



def GICS_mapping(data, col_guidance, first_digits=2):
    """
    This function is to convert GICS code to industry. New in Version 2.9. 


    Parameters:

        data:           the input dataset in DataFrame. 

        col_guidance:   str
                        column name for GICS code

        first_digits:   integer [2,4,6,8], default 2. 
                        First digits used for the industry mapping

    Return:
        a DataFrame which is a copy of input data but has one more column: 'Industry_by_GICS'
    
    """    
    processed_data = data.copy()
    processed_data['Industry_by_GICS'] = processed_data[col_guidance].fillna(99999999)
    processed_data['Industry_by_GICS'] = [str(int(x))[:first_digits] for x in processed_data['Industry_by_GICS']]
    processed_data['Industry_by_GICS'] = processed_data['Industry_by_GICS'].replace(gics_code)
    
    return (processed_data)


def cluster_corr_old(dat, pl_cols):

    pl_cat1 = [pl_cols[0],]
    pl_cat2 = []

    for name in pl_cols[1:]:
        if np.abs(dat[[name]+pl_cat1].corr()).iloc[0,1:].mean() > 0.75:
            pl_cat1.append(name)
        else:
            pl_cat2.append(name)
    if len(pl_cat2)==0:
        return((pl_cat1,))
    else:
        return((pl_cat1,) + cluster_corr(dat, pl_cat2))


def correlation_dist(m, n):
    """
    This function is to calculate the distance of two variables based on correlation
    New in Version 3.0

    Parameters:

        m,n:   2 variables

    Return:
        distance, scalar
    
    """   

    df = pd.DataFrame({'m':m,'n':n})
    df.dropna(inplace=True)
    return(1-np.abs(pearsonr(df.m, df.n)[0]))


def reorder_cols(dat, cols):
    """
    This function is to re-rank the columns based on their correlation distance
    New in Version 3.0

    Parameters:

        dat:    the input dataset in DataFrame. 

        cols:   list of str
                column names


    Return:

        a list which is a re-ranked column names
    
    """    

    m = len(cols)
    X = dat[cols].T.to_numpy()
    k = 0
    dm = np.empty((m * (m - 1)) // 2, dtype=np.double)
    for i in range(0, m - 1):
        for j in range(i + 1, m):
            dm[k] = correlation_dist(X[i], X[j])
            k = k + 1

    yy = hierarchy.linkage(dm, method='average', optimal_ordering=False)
    Z = hierarchy.dendrogram(yy, no_plot=True, color_threshold=-np.inf)
    list_reorder = [cols[i] for i in Z['leaves']]

    return(list_reorder)


def cluster_corr(dat, list_cols):
    """
    This function is to group columns based on re-ranked columns names.
    New in Version 3.0
    
    Parameters:

        dat:        the input dataset in DataFrame. 

        list_cols:  list of str
                    column names


    Return:

        a tuple (grouped_names , re-ranked names)
    
    """    


    pl_cols = reorder_cols(dat, list_cols)
    pl_cat = []
    X = np.abs(dat[pl_cols].corr()).to_numpy()
    m = len(pl_cols)
    start=0
    for i in range(m-1):
        tmp = X[i+1,i]
        if tmp>=0.7:
            continue
        else:
            end=i
            pl_cat.append(pl_cols[start:(end+1)])
            start=end+1
            continue
    pl_cat.append(pl_cols[start:])
    return((pl_cat, pl_cols))


def plot_cluster_corr(dat, cat_list):
    """
    This function is to generate heatmap plot for the re-ranked cols
    
    Parameters:

        dat:        the input dataset in DataFrame. 

        cat_list:   tuple, the first output of function cluster_corr()

    """        
    cols = []
    for cat in cat_list:
        cols = cols+cat

    f, ax = plt.subplots(figsize=(10*len(cols)/25, 6*len(cols)/25))
    sns.heatmap(np.abs(dat[cols].corr()), linewidths=.3, cmap='Blues', ax=ax)


def pca_n_comp(dat, cols, threshold=0.9):
    df = dat[cols].dropna(subset=cols)
    scaler = StandardScaler()
    scaler.fit(df)
    df_pca = scaler.transform(df)

    for i in range(1,len(cols)+1):
        pca = decomposition.PCA(n_components=i)
        pca.fit(df_pca)
        if pca.explained_variance_ratio_.sum()>threshold:
            n = i
            explained_variance = pca.explained_variance_ratio_.sum()
            break
    return((n, explained_variance))


# Added in Ver. 3.0
compustat_Americas_country_code = ('USA','CAN','BMU','BRA','MEX','CHL','ARG',
    'CYM','COL','PER','BHS','BRB','DOM','HND','PAN')

# Added in Ver. 3.0
compustat_EMEA_country_code = ('GBR','FRA','NLD','DEU','CHE','IRL','SWE','LUX',
    'ESP','RUS','GRC','ITA','FIN','NOR','MCO','BEL','DNK','PRT','ZAF','HUN',
    'ISR','AUT','CYP','POL','TUR','UKR')

# -*- coding: utf-8 -*-
"""
Created on Fri Feb 24 10:57:30 2017

Version: 1.0: Initial build
Version: 1.1, 20170317, Modify method barplot() to avoid negative y tick for mean PD
Version: 1.2, 20170320, Debug. Delete 'plt.show()' in qqplot() to avoid saving blank png.
Version: 1.3, 20170321, Mofify method barplot() by Jack to make it compatible to 
                        grade combinations, such as 'B/C'.
Version: 1.4, 20170327, Modify method __normalization() to correctly treat 'Invalid Negative'
                        in some quant factors in normalization procedure.
Version: 1.5, 20170329, Modify method barplot() to correctly plot mean PD point when its default
                        comes from floor(at least 1 default for each category)
Version: 1.6, 20170404, Modify method __Somersd() to accelerate the calculation and add one new 
                        method 'logitPDplot()' to plot the relationship between quant factor and 
                        obligor's logitPD.
Version: 1.7, 20170404, Modify method 'logitPDplot()' and add new method 'PDplot()'
Version: 1.8: 20170510, Modify the class to handle missing data. Minor change in __SomersD().

@author: ub71894 (4e8e6d0b), CSG
"""

import pandas as pd
import numpy as np
import warnings
import seaborn as sns
import matplotlib.pyplot as plt
import statsmodels.api as sm
from sklearn.metrics import roc_auc_score
import itertools
from numba import jit
from PDScorecardTool.Process import logitPD_frPDRR
#%%
class SFA(object):
    
    '''    
    This class is used to generate plots, tables and statistics related to SFA
    
    SFA(data, model):

        data:       data for development. Make sure it has column 'def_flag' as default
                    indicator and the factors' names are the same as the ones in 'model'
        model:      PDModle class. It saves all parameters for the model
    
        missing:    str, default 'median'. Added in Ver.1.8
                    statistics of the data used to fill missing data (NA data)

    Methods:

        describe()

        distplot(factors='quant', raw=True, savefig=True, clip=[0.01,0.99])

        barplot(factors='quali', savefig=True)

        qqplot(factors='quant', raw=True, savefig=True, clip=[0.01,0.99])

        violinplot(factors='quant', raw=True, savefig=True, clip=[0.01,0.99])

        boxplot(factors='quant', raw=True, savefig=True, clip=[0.01,0.99])

        logitPDplot(factors='quant', raw=True, PDsource='def_flag', bin_on='default', numofbins=10, order=1, savefig=True)
        
        corr(factors='all', raw=False, output=True)

        ARanalysis(factors='all', isthere_def=True, dependentvar='def_flag', output=True)
    

    Applicability:
                            |  Only quant   |  Only quali   |     ALL              
    ---------------------------------------------------------------------
    Only raw data:          |  describe()   |  barplot()    |
    ---------------------------------------------------------------------
                            |               |               |
    Only normalized data:   |               |               | ARanalysis()
    ---------------------------------------------------------------------
                            |               |               |
    Both:                   |  distplot()   |               | corr()
                            |  violinplot() |               |
                            |  boxplot()    |               |
                            |  qqplot()     |               | 
                            |  logitPDplot()|               |
                            |  PDplot()     |               |
    =====================================================================
    '''

    def __init__(self, data, model, missing='median'):

        self.model = model
        self.rawdata = data.copy()
        self.rawdata.reset_index(drop=True, inplace=True)
        self.stats = self.__describe()
        self.normdata = self.__normalization(missing)
        

    def __describe(self):
        '''

        Privata method that calculate key statistics for quant factors on raw data

        '''
        warnings.filterwarnings('ignore', category=Warning)
        dat = self.rawdata[self.model.quant_factor]
        stats = dat.describe(percentiles=[0.01,0.05,0.1,0.25,0.5,0.75,0.9,0.95,0.99])  
        temp={}
        temp['Skew'] = pd.Series(dat.skew())
        temp['Kurtosis'] = dat.kurtosis()
        stats = pd.concat([stats, pd.DataFrame(temp).T])
        return stats


    def __normalization(self, missing):
        '''

        Private method that apply normalization on financial factors.

        '''
        warnings.filterwarnings('ignore', category=Warning)
        # Invalid_Neg
        normdata = self.rawdata.copy()
        # modify the loop below in Ver 1.4:
        for i,neg_source in enumerate(self.model.Invalid_Neg):
            if neg_source:
                col=self.model.quant_factor[i]
                normdata[col][ (normdata[col]<0) & (normdata[neg_source]<0) ] = self.model.cap[i]
                # treat NA in 'neg_source' as negative value
                normdata[col][ (normdata[col]<0) & pd.isnull(normdata[neg_source]) ] = self.model.cap[i]

        # cap/floor for quant factors:
        for i, col in enumerate(self.model.quant_factor):
            normdata[col] = np.clip(normdata[col], self.model.floor[i], self.model.cap[i])        

        # calibration for quali factors:
        for i, col in enumerate(self.model.quali_factor):
            normdata[col].replace(to_replace=self.model.qualimapping[i],inplace=True) 

        # quant factors transformation:
        for i, col in enumerate(self.model.quant_factor):
            if self.model.quant_log[i]:
                normdata[col] = np.log(normdata[col])
            
        # quant factors normalization:
        for i, col in enumerate(self.model.quant_factor):
            normdata[col]  = 50*(normdata[col] - self.model.doc_mean[i]) / self.model.doc_std[i]      
        
        # quant factors flip sign:  
        for i, col in enumerate(self.model.quant_factor):
            normdata[col] = normdata[col] * self.model.quant_multiplier[i]     
        
        # fill missing data. added in Ver. 1.8
        exec('normdata[self.model.quant_factor+self.model.quali_factor] = normdata[self.model.quant_factor+self.model.quali_factor].\
            fillna(normdata[self.model.quant_factor+self.model.quali_factor].{missing}(), inplace=True)'.format(missing=missing))
        
        return normdata
    

    def __Spearman(self, factor, dependentvar) :
        '''

        Private method that calculate Spearman correlation.

        '''
        rho = self.ardata.corr(method='spearman')[factor][dependentvar]
        return rho


    def __Pearson(self, factor, dependentvar) :
        '''

        Private method that calculate Pearson correlation.

        '''       
        rho = self.ardata.corr(method='pearson')[factor][dependentvar]
        return rho        


    def __AUROC(self, factor, dependentvar, isthere_def) :
        '''

        Private method that calculate area under ROC.

        '''
        if isthere_def:
            y_score = self.ardata[factor].tolist()
            y_true  = [int(x) for x in list(self.ardata[dependentvar])] 
            return roc_auc_score(y_true, y_score) 
        else:
            return 'NA'


    def __SomersD_old(self, factor, dependentvar, sign=1, unit_concordance=True): # debugged on 20160817
        '''

        Private method that calculate SomersD using Josh's code. It's debugged 
        by me on 20160817. when dependentvar='def_flag', SomersD = 2*AUROC-1

        '''
        x = np.array(self.ardata[factor])
        y = np.array(self.ardata[dependentvar])
        n = len(x)
        C = 0
        D = 0
        if unit_concordance==True:
            for i, j in itertools.combinations(np.arange(n), 2):
                dx = x[i] - x[j]
                dy = y[i] - y[j]
                if (((dx<0) & (dy<0)) | ((dx>0) & (dy>0))):
                    cij = 1
                    dij = 1
                elif (((dx<0) & (dy>0)) | ((dx>0) & (dy<0))):
                    cij = -1
                    dij = 1
                else:
                    cij = 0
                    if dy==0: 
                        dij = 0
                    else: 
                        dij=1
                C = C + cij
                D = D + dij
        else:
            for i, j in itertools.combinations(np.arange(n), 2):
                dx = x[i] - x[j]
                dy = y[i] - y[j]
                if (((dx<0) & (dy<0)) | ((dx>0) & (dy>0))):
                    cij = abs(dy)
                    dij = abs(dy)
                elif (((dx<0) & (dy>0)) | ((dx>0) & (dy<0))):
                    cij = -abs(dy)
                    dij = abs(dy)
                else:
                    cij = 0
                    if dy==0: 
                        dij = 0
                    else: 
                        dij=1
                C = C + cij
                D = D + dij    
        
        return sign*C/D



    @jit
    def __SomersD(self, factor, dependentvar, sign=1, unit_concordance=True): 
        '''

        New version of SomersD function which leverages numba.jit to accelerate
        the calculation. In production from Ver. 1.6.

        '''
        x = np.array(self.ardata[factor])
        # added in Ver 1.8 to deal with missing PDRR sometimes.
        y_true = self.normdata[dependentvar].copy()
        y_true.dropna(inplace=True)
        y = np.array(y_true)
        n = len(x)
        C = 0
        D = 0
        if unit_concordance==True:
            for i in range(n):
                for j in range(i):
                    dx = x[i] - x[j]
                    dy = y[i] - y[j]
                    if (dx*dy) > 0:
                        cij = 1
                        dij = 1
                    elif  (dx*dy) < 0:
                        cij = -1
                        dij = 1
                    else:
                        cij = 0
                        if dy==0: 
                            dij = 0
                        else: 
                            dij=1
                    C = C + cij
                    D = D + dij
        else:
            for i in range(n):
                for j in range(i):
                    dx = x[i] - x[j]
                    dy = y[i] - y[j]
                    if  (dx*dy) > 0:
                        cij = abs(dy)
                        dij = abs(dy)
                    elif (dx*dy) < 0:
                        cij = -abs(dy)
                        dij = abs(dy)
                    else:
                        cij = 0
                        if dy==0: 
                            dij = 0
                        else: 
                            dij=1
                    C = C + cij
                    D = D + dij    
        
        return sign*C/D


    def describe(self):
        '''

        Public method that present key statistics for quant factors on raw data
        
        '''
        return self.stats
        

    def distplot(self, factors='quant', raw=True, savefig=True, clip=[0.01,0.99]):
        '''

        Public method that present distribution plot for quant factors
        It will ignore +-inf and NA data
        Note that, for raw quant factors, it makes plot based on clipped data 
        and exclude outliers.
        
        Parameters: 

        factors:    str {'quant','all', factorname} or None, default 'quant'
                    quant factors for plotting. 

        raw:        boolean, default True
                    whether using raw data or normalized data

        savefig:    boolean, default True
                    whether save png file in current working directory.  

        clip:       list of 2 scalars. default [0.01,0.99]
                    Lower and upper bounds percentiles for datapoints used to fit
                    ONLY work for raw data
        '''
        warnings.filterwarnings('ignore', category=Warning)
        if factors=='all' or factors=='quant':
            factors= self.model.quant_factor 
        else: # user input a single factor's name
            factors = [factors]
        if raw:
            suffix = '_raw' + str(int(clip[0]*100))+'-'+str(int(clip[1]*100))
            for factor in factors:
                # only use clipped data and exclude outliers:
                floor, cap = self.rawdata[factor].quantile(clip)
                dataforplot = self.rawdata.query('{floor}<{factor}<{cap}'.format(factor=factor,floor=floor,cap=cap))            
                # 3 plots: hist, pdf and cdf
                fig,ax=plt.subplots(1,3,figsize=(18,6))
                fig.suptitle(factor+suffix)
                sns.distplot(dataforplot[factor], kde=False , ax=ax[0], color='g')
                ax[0].set_xlabel('Histogram')
                sns.distplot(dataforplot[factor], hist=False, kde_kws=dict(shade=True), ax=ax[1], color='g')
                ax[1].set_xlabel('Probability density function')
                sns.distplot(dataforplot[factor], hist=False, kde_kws=dict(cumulative=True,shade=True), ax=ax[2], color='g')
                ax[2].set_xlabel('Cumulative density function')
                if savefig:
                    fig.savefig(factor+suffix+'_quantdist.png')
        else:
            suffix = '_norm'
            for factor in factors:     
                # 3 plots: hist, pdf and cdf    
                fig,ax=plt.subplots(1,3,figsize=(18,6))
                fig.suptitle(factor+suffix)
                sns.distplot(self.normdata[factor], kde=False , ax=ax[0], color='g')
                ax[0].set_xlabel('Histogram')
                sns.distplot(self.normdata[factor], hist=False, kde_kws=dict(shade=True), ax=ax[1], color='g')
                ax[1].set_xlabel('Probability density function')
                sns.distplot(self.normdata[factor], hist=False, kde_kws=dict(cumulative=True,shade=True), ax=ax[2], color='g')
                ax[2].set_xlabel('Cumulative density function')
                if savefig:
                    fig.savefig(factor+suffix+'_quantdist.png')


    def barplot(self, factors='quali', savefig=True):
        '''

        Public method that present distribution plot for quali factors
        It will ignore NA data

        Parameters: 

        factors:    str {'quali','all', factorname} or None, default 'quali'
                    quali factors for plotting. 

        savefig:    boolean, default True
                    whether save png file in current working directory.  
        
        '''

        
        # add this adjusted mean function in Ver 1.5 for plotting
        def _adjmean(x):
            if x.sum()==0:
                return (1/x.size)
            else:
                return x.mean()

        warnings.filterwarnings('ignore', category=Warning)
        if factors=='all' or factors=='quali':
            factors= self.model.quali_factor
        else: # user input a single factor's name
            factors = [factors]

        for factor in factors:
            letter = list(set(self.rawdata[factor])) # add in Ver. 1.3
            letter.sort()
            fig = plt.figure()
            fig.suptitle(factor)
            ax1 = fig.add_subplot(111)
            sns.countplot(x=factor, data=self.rawdata, ax=ax1,order=letter, alpha=0.7, palette='Blues_d')
            ax1.set_xlabel('Answers')
            ax2 = ax1.twinx()
            sns.pointplot(x=factor, y='def_flag', data=self.rawdata, ax=ax2, order=letter, ci=0, estimator=_adjmean) # modified in Ver 1.5
            ax2.grid(None)
            ax2.set_ylabel('mean of PD')
            ax2.set_ylim(bottom=0) # add this line in Ver 1.1
            if savefig:
                fig.savefig(factor+'_qualibar.png')


    def qqplot(self, factors='quant', raw=True, savefig=True, clip=[0.01,0.99]):
        '''

        Public method that make qq plot on quant factors
        It uses statsmodels.api.qqplot to create graph.
        Note that:  1. for raw quant factors, it makes plot based on clipped data 
                      and exclude outliers.
                    2. Usually this plot is trival for our modeling work since we 
                        don't expect the quant factor has normal distribution.
        Parameters: 

        factors:    str {'quant','all', factorname} or None, default 'quant'
                    quant factors for plotting. 

        raw:        boolean, default True
                    whether using raw data or normalized data.

        savefig:    boolean, default True
                    whether save png file in current working directory. 

        clip:       list of 2 scalars. default [0.01,0.99]
                    Lower and upper bounds percentiles for datapoints used to fit
                    ONLY work for raw data
        '''
        warnings.filterwarnings('ignore', category=Warning)
        if factors=='quant' or factors=='all':
            factors= self.model.quant_factor
        else: # user input a single factor's name
            factors = [factors]
        if raw:
            suffix = '_raw' + str(int(clip[0]*100))+'-'+str(int(clip[1]*100))
            for factor in factors:
                # only use clipped data and exclude outliers:
                floor, cap = self.rawdata[factor].quantile(clip)
                dataforplot = self.rawdata.query('{floor}<{factor}<{cap}'.format(factor=factor,floor=floor,cap=cap))  
                dataforplot = dataforplot[factor]
                sm.qqplot(dataforplot,fit=True,line='45')
                plt.title(dataforplot.name)
                if savefig:
                    plt.savefig(factor+suffix+'_quantqqplot.png')
        else:
            suffix = '_norm'
            for factor in factors:
                dataforplot= self.normdata[factor]
                sm.qqplot(dataforplot,fit=True,line='45')
                plt.title(dataforplot.name)
                if savefig:
                    plt.savefig(factor+suffix+'_quantqqplot.png')


    def violinplot(self, factors='quant', raw=True, savefig=True, clip=[0.01,0.99]):
        '''

        Public method that make violin plot to show distribution and quartile of
        the quant factors. It splits data into default(bad) and non-default(good)
        and seaborn.violinplot() to create the graph.

        Note that, for raw quant factors, it makes plot based on clipped data 
        and exclude outliers.

        Parameters: 

        factors:    str {'quant','all', factorname} or None, default 'quant'
                    quant factors for plotting. 

        raw:        boolean, default True
                    whether using raw data or normalized data.

        savefig:    boolean, default True
                    whether save png file in current working directory. 

        clip:       list of 2 scalars. default [0.01,0.99]
                    Lower and upper bounds percentiles for datapoints used to fit
                    ONLY work for raw data
        '''
        warnings.filterwarnings('ignore', category=Warning)
        if factors=='quant' or factors=='all':
            factors= self.model.quant_factor
        else: # user input a single factor's name
            factors = [factors]
        if raw:
            suffix = '_raw' + str(int(clip[0]*100))+'-'+str(int(clip[1]*100))
            for factor in factors:
                # only use clipped data and exclude outliers:
                floor, cap = self.rawdata[factor].quantile(clip)
                dataforplot = self.rawdata.query('{floor}<{factor}<{cap}'.format(factor=factor,floor=floor,cap=cap))  
                dataforplot["all"] = ""
                dataforplot['event'] = dataforplot['def_flag'].replace(to_replace={1:'Default',0:'Non-Default'})
                fig,ax=plt.subplots(1,1,figsize=(8,8))
                ax = sns.violinplot(x='all', y=factor, hue='event', inner="quartile", data=dataforplot, split=True,  cut=0, palette="Paired")
                if savefig:
                    fig.savefig(factor+suffix+'_quantviolinplot.png')
        else:
            suffix = '_norm'
            dataforplot= self.normdata.copy()
            for factor in factors:
                dataforplot["all"] = ""
                dataforplot['event'] = dataforplot['def_flag'].replace(to_replace={1:'Default',0:'Non-Default'})
                fig,ax=plt.subplots(1,1,figsize=(8,8))
                ax = sns.violinplot(x='all', y=factor, hue='event', inner="quartile", data=dataforplot, split=True,  cut=0, palette="Paired")
                if savefig:
                    fig.savefig(factor+suffix+'_quantviolinplot.png')  


    def boxplot(self, factors='quant', raw=True, savefig=True, clip=[0.01,0.99]):
        '''

        Public method that make box plot to show distribution and quartile of
        the quant factors. It splits data into default(bad) and non-default(good)
        and seaborn.boxplot() to create the graph.

        Note that, for raw quant factors, it makes plot based on clipped data 
        and exclude outliers.

        Parameters: 

        factors:    str {'quant','all', factorname} or None, default 'quant'
                    quant factors for plotting. 

        raw:        boolean, default True
                    whether using raw data or normalized data.

        savefig:    boolean, default True
                    whether save png file in current working directory. 

        clip:       list of 2 scalars. default [0.01,0.99]
                    Lower and upper bounds percentiles for datapoints used to fit
                    ONLY work for raw data
        '''
        warnings.filterwarnings('ignore', category=Warning)
        if factors=='quant' or factors=='all':
            factors= self.model.quant_factor
        else: # user input a single factor's name
            factors = [factors]
        if raw:
            suffix = '_raw_clip' + str(int(clip[0]*100))+'-'+str(int(clip[1]*100))
            for factor in factors:
                # only use clipped data and exclude outliers:
                floor, cap = self.rawdata[factor].quantile(clip)
                dataforplot = self.rawdata.query('{floor}<{factor}<{cap}'.format(factor=factor,floor=floor,cap=cap))  
                dataforplot["all"] = ""
                dataforplot['event'] = dataforplot['def_flag'].replace(to_replace={1:'Default',0:'Non-Default'})
                fig,ax=plt.subplots(1,1,figsize=(8,8))
                ax = sns.boxplot(x="all", y=factor, hue='event', data=dataforplot, palette="Paired")
                if savefig:
                    fig.savefig(factor+suffix+'_quantboxplot.png')
        else:
            suffix = '_norm'
            dataforplot = self.normdata.copy()
            for factor in factors:
                dataforplot["all"] = ""
                dataforplot['event'] = dataforplot['def_flag'].replace(to_replace={1:'Default',0:'Non-Default'})
                fig,ax=plt.subplots(1,1,figsize=(8,8))
                ax = sns.boxplot(x="all", y=factor, hue='event', data=dataforplot, palette="Paired")
                if savefig:
                    fig.savefig(factor+suffix+'_quantboxplot.png')       

    
    def logitPDplot(self, factors='quant', raw=True, PDsource='def_flag', bin_on='default', numofbins=10, \
        plot_on_binnum=False, order=1,  lowess=False, savefig=True):
        '''

        Public method that plots the relationship between logitPD to factor value 


        Parameters: 

        factors:        str {'quant','all', factorname} or None, default 'quant'
                        quant factors for plotting. 

        raw:            boolean, default True
                        whether using raw data or normalized data

        PDsource:       str, default 'def_flag'
                        the name of the column which is the source of PD. For true default
                        empirical PD, PDsource='def_flag'; For non-true default PD, user can
                        assign a PDRR name. ex. PDsource='Final_PD_Risk_Rating'

        bin_on:         {'default','obligor'}, default 'default'
                        when PDsource='def_flag', user needs to choose the even binning is on
                        the number of default or the number of obligor. In other words, user 
                        needs to choose "even obligors" or "even defaults" in each bin. 

        numofbins:      int, default 10
                        the number of bins in binning procedure.

        plot_on_binnum: boolean, default False, Added in Ver. 1.7
                        plot logitPD vs. factor mean in each bin, if False
                        plot logitPD vs. the sequence number of each bin, if True

        order:          int, default 1
                        the plot will estimate a linear regression in default setting.
                        if order>1, the plot will estimate a polynomial regression.

        lowess:         boolean, default False, Added in Ver. 1.7
                        if True, estimate a nonparametric lowess model (locally weighted linear regression)
        
        savefig:        boolean, default True
                        whether save png file in current working directory. 

        '''

        if factors=='quant' or factors=='all':
            factors= self.model.quant_factor
        else: # user input a single factor's name
            factors = [factors]

        if raw:
            data = self.rawdata.copy()
            suffix = '_raw'
        else:
            data = self.normdata.copy()
            suffix = '_norm'


        if (PDsource=='def_flag') & (bin_on=='default'):
            suffix += '_evendef_(#bins={nbins}_order={order})'.format(nbins=numofbins,order=order)
            for factor in factors:
                data_plot = data[[factor]+['def_flag']].copy()
                # get the bins from default data:
                temp = data_plot.query("def_flag>0")                
                try:
                    _,bins = pd.qcut(temp[factor], numofbins, retbins=True)
                    bins[0] = -np.inf; bins[-1] = np.inf    # Added in Ver. 1.7                
                except ValueError:
                    print('Not enought defaults for the binning. Try smaller number of bins or Even-Obligors binning method')
                    return 0
                # apply the bins to all data
                data_plot['bin'] = pd.cut(data_plot[factor], bins, labels=list(range(1,numofbins+1))) # Added in Ver. 1.7
                bucket = data_plot.groupby(by='bin').mean()
                bucket['count'] = data_plot.groupby(by='bin').count().def_flag
                bucket['Ndef'] = data_plot.groupby(by='bin').sum().def_flag
                bucket['logit_PD'] = np.array([np.log(x/(1-x)) for x in bucket.def_flag.tolist()])
                bucket.reset_index(drop=False, inplace=True) # Added in Ver. 1.7
                # kick out -inf in logit PD due to 0 default in some bin
                bucket = bucket.query('logit_PD > {invalid}'.format(invalid=-np.inf))                
                # scatter plot with regression 
                if plot_on_binnum: # Added in Ver. 1.7
                    if lowess:
                        g = sns.lmplot(x='bin', y="logit_PD", data=bucket, lowess=lowess, ci=None, scatter_kws={"s": 20})
                    else:
                        g = sns.lmplot(x='bin', y="logit_PD", data=bucket, order=order, ci=None, scatter_kws={"s": 20})
                else:
                    if lowess:
                        g = sns.lmplot(x=factor, y="logit_PD", data=bucket, lowess=lowess, ci=None, scatter_kws={"s": 20})
                    else:
                        g = sns.lmplot(x=factor, y="logit_PD", data=bucket, order=order, ci=None, scatter_kws={"s": 20})
                
                if savefig:
                    g.savefig(factor+suffix+'_logitPDplot.png')


        elif (PDsource=='def_flag') & (bin_on=='obligor'):
            suffix += '_evenobligor_(#bins={nbins}_order={order})'.format(nbins=numofbins,order=order)
            for factor in factors:
                data_plot = data[[factor]+['def_flag']].copy()  
                try:
                    data_plot['bin'] = pd.qcut(data_plot[factor], numofbins, labels=list(range(1,numofbins+1))) # Added in Ver. 1.7
                except ValueError:
                    print('Too many obligors have the same value due to cap/floor. Try smaller number of bins.')
                    return 0
                bucket = data_plot.groupby(by='bin').mean()
                bucket['count'] = data_plot.groupby(by='bin').count().def_flag
                bucket['Ndef'] = data_plot.groupby(by='bin').sum().def_flag
                bucket['logit_PD'] = np.array([np.log(x/(1-x)) for x in bucket.def_flag.tolist()])
                bucket.reset_index(drop=False, inplace=True) # Added in Ver. 1.7
                # kick out -inf in logit PD due to 0 default in some bin
                bucket = bucket.query('logit_PD > {invalid}'.format(invalid=-np.inf))
                # scatter plot with regression 
                if plot_on_binnum: # Added in Ver. 1.7
                    if lowess:
                        g = sns.lmplot(x='bin', y="logit_PD", data=bucket, lowess=lowess, ci=None, scatter_kws={"s": 20})
                    else:
                        g = sns.lmplot(x='bin', y="logit_PD", data=bucket, order=order, ci=None, scatter_kws={"s": 20})
                else:
                    if lowess:
                        g = sns.lmplot(x=factor, y="logit_PD", data=bucket, lowess=lowess, ci=None, scatter_kws={"s": 20})
                    else:
                        g = sns.lmplot(x=factor, y="logit_PD", data=bucket, order=order, ci=None, scatter_kws={"s": 20})

                if savefig:
                    g.savefig(factor+suffix+'_logitPDplot.png')
        

        elif PDsource: # means PDsource has been assigned
            try:
                data[PDsource]
            except KeyError:
                print('No column in dataset is called "{PDsource}".'.format(PDsource=PDsource))
                return 0
            suffix += '_{PDRR}_(order={order})'.format(PDRR=PDsource,order=order)
            # prepare PDRR implied logitPD of each obligor:
            data = logitPD_frPDRR(data, self.model, PDsource)
            for factor in factors:
                # scatter plot with regression 
                g = sns.lmplot(x=factor, y="logitPD_frPDRR", data=data, order=order, ci=None, scatter_kws={"s": 20})
                if savefig:    
                    g.savefig(factor+suffix+'_logitPDplot.png')    


        else:
            return 0


    def PDplot(self, factors='quant', raw=True, PDsource='def_flag', bin_on='default', numofbins=10, \
        plot_on_binnum=False, order=1,  lowess=False, savefig=True):
        '''

        Public method that plots the relationship between PD to factor value 


        Parameters: 

        factors:        str {'quant','all', factorname} or None, default 'quant'
                        quant factors for plotting. 

        raw:            boolean, default True
                        whether using raw data or normalized data

        PDsource:       str, default 'def_flag'
                        the name of the column which is the source of PD. For true default
                        empirical PD, PDsource='def_flag'; For non-true default PD, user can
                        assign a PDRR name. ex. PDsource='Final_PD_Risk_Rating'

        bin_on:         {'default','obligor'}, default 'default'
                        when PDsource='def_flag', user needs to choose the even binning is on
                        the number of default or the number of obligor. In other words, user 
                        needs to choose "even obligors" or "even defaults" in each bin. 

        numofbins:      int, default 10
                        the number of bins in binning procedure.

        plot_on_binnum: boolean, default False, Added in Ver. 1.7
                        plot logitPD vs. factor mean in each bin, if False
                        plot logitPD vs. the sequence number of each bin, if True

        order:          int, default 1
                        the plot will estimate a linear regression in default setting.
                        if order>1, the plot will estimate a polynomial regression.
        
        lowess:         boolean, default False, Added in Ver. 1.7
                        if True, estimate a nonparametric lowess model (locally weighted linear regression)
        
        savefig:        boolean, default True
                        whether save png file in current working directory. 

        '''

        if factors=='quant' or factors=='all':
            factors= self.model.quant_factor
        else: # user input a single factor's name
            factors = [factors]

        if raw:
            data = self.rawdata.copy()
            suffix = '_raw'
        else:
            data = self.normdata.copy()
            suffix = '_norm'


        if (PDsource=='def_flag') & (bin_on=='default'):
            suffix += '_evendef_(#bins={nbins}_order={order})'.format(nbins=numofbins,order=order)
            for factor in factors:
                data_plot = data[[factor]+['def_flag']].copy()
                # get the bins from default data:
                temp = data_plot.query("def_flag>0")                
                try:
                    _,bins = pd.qcut(temp[factor], numofbins, retbins=True)
                    bins[0] = -np.inf; bins[-1] = np.inf    # Added in Ver. 1.7                
                except ValueError:
                    print('Not enought defaults for the binning. Try smaller number of bins or Even-Obligors binning method')
                    return 0
                # apply the bins to all data
                data_plot['bin'] = pd.cut(data_plot[factor], bins, labels=list(range(1,numofbins+1))) # Added in Ver. 1.7
                bucket = data_plot.groupby(by='bin').mean()
                bucket['count'] = data_plot.groupby(by='bin').count().def_flag
                bucket['Ndef'] = data_plot.groupby(by='bin').sum().def_flag
                bucket['logit_PD'] = np.array([np.log(x/(1-x)) for x in bucket.def_flag.tolist()])
                bucket.reset_index(drop=False, inplace=True) # Added in Ver. 1.7
                # kick out -inf in logit PD due to 0 default in some bin
                #bucket = bucket.query('logit_PD > {invalid}'.format(invalid=-np.inf))                
                # scatter plot with regression 
                if plot_on_binnum: # Added in Ver. 1.7
                    if lowess:
                        g = sns.lmplot(x='bin', y="def_flag", data=bucket, lowess=lowess, ci=None, scatter_kws={"s": 20})
                    else:
                        g = sns.lmplot(x='bin', y="def_flag", data=bucket, order=order, ci=None, scatter_kws={"s": 20})
                else:
                    if lowess:
                        g = sns.lmplot(x=factor, y="def_flag", data=bucket, lowess=lowess, ci=None, scatter_kws={"s": 20})
                    else:
                        g = sns.lmplot(x=factor, y="def_flag", data=bucket, order=order, ci=None, scatter_kws={"s": 20})

                if savefig:
                    g.savefig(factor+suffix+'_PDplot.png')


        elif (PDsource=='def_flag') & (bin_on=='obligor'):
            suffix += '_evenobligor_(#bins={nbins}_order={order})'.format(nbins=numofbins,order=order)
            for factor in factors:
                data_plot = data[[factor]+['def_flag']].copy()  
                try:
                    data_plot['bin'] = pd.qcut(data_plot[factor], numofbins, labels=list(range(1,numofbins+1))) # Added in Ver. 1.7
                except ValueError:
                    print('Too many obligors have the same value due to cap/floor. Try smaller number of bins.')
                    return 0
                bucket = data_plot.groupby(by='bin').mean()
                bucket['count'] = data_plot.groupby(by='bin').count().def_flag
                bucket['Ndef'] = data_plot.groupby(by='bin').sum().def_flag
                bucket['logit_PD'] = np.array([np.log(x/(1-x)) for x in bucket.def_flag.tolist()])
                bucket.reset_index(drop=False, inplace=True) # Added in Ver. 1.7
                # kick out -inf in logit PD due to 0 default in some bin
                #bucket = bucket.query('logit_PD > {invalid}'.format(invalid=-np.inf))
                # scatter plot with regression 
                if plot_on_binnum: # Added in Ver. 1.7
                    if lowess:
                        g = sns.lmplot(x='bin', y="def_flag", data=bucket, lowess=lowess, ci=None, scatter_kws={"s": 20})
                    else:
                        g = sns.lmplot(x='bin', y="def_flag", data=bucket, order=order, ci=None, scatter_kws={"s": 20})
                else:
                    if lowess:
                        g = sns.lmplot(x=factor, y="def_flag", data=bucket, lowess=lowess, ci=None, scatter_kws={"s": 20})
                    else:
                        g = sns.lmplot(x=factor, y="def_flag", data=bucket, order=order, ci=None, scatter_kws={"s": 20})
                    
                if savefig:
                    g.savefig(factor+suffix+'_PDplot.png')
        

        elif PDsource: # means PDsource has been assigned
            try:
                data[PDsource]
            except KeyError:
                print('No column in dataset is called "{PDsource}".'.format(PDsource=PDsource))
                return 0
            suffix += '_{PDRR}_(order={order})'.format(PDRR=PDsource,order=order)
            # prepare PDRR implied logitPD of each obligor:
            data = PD_frPDRR(data, self.model, PDsource)
            for factor in factors:
                # scatter plot with regression 
                g = sns.lmplot(x=factor, y="PD_frPDRR", data=data, order=order, ci=None, scatter_kws={"s": 20})
                if savefig:    
                    g.savefig(factor+suffix+'_PDplot.png')    


        else:
            return 0


    def corr(self, factors='all', raw=False, output=True):
        '''

        Public method that calculate correlation matrix for multifactors

        Parameters: 

        factors:    str {'quant', 'quali', 'all'} or None, default 'all'
                    Multifactors for matrix calculations. 

        raw:        boolean, default False
                    whether using raw data or normalized data. Default setting is
                    False since usually we care about the corr of normalized factors. 

        output:     boolean, default True
                    whether save the corr matrix to excel file.

        '''
        warnings.filterwarnings('ignore', category=Warning)
        if factors=='quant':
            names= self.model.quant_factor
        elif factors=='quali':
            names= self.model.quali_factor
        else: # all factors
            names = self.model.quant_factor + self.model.quali_factor
        
        if raw:
            suffix = '_raw'
            corr = self.rawdata[names].corr()
            if output:
                corr.to_excel('CorrMat_'+factors+suffix+'.xlsx')
            return corr

        else:
            suffix = '_norm'
            corr = self.normdata[names].corr()
            if output:
                corr.to_excel('CorrMat_'+factors+suffix+'.xlsx')
            return corr    


    def ARanalysis(self, factors='all', isthere_def=True, dependentvar='def_flag', output=True):
        '''
        
        Public method that calculate correlation matrix for multifactors

        Parameters: 

        factors:        str {'quant','quali', all', factorname}, default 'all'
                        Multifactors for matrix calculations. 

        isthere_def:    boolean, default True
                        whether the analysis is on the default event. If not, for 
                        example we want to check the accuracy of factor on PDRR,
                        then we should input  
                        isthere_def=False, dependentvar='PDRR'

        dependentvar:   str, default 'def_flag'
                        column name for dependent variable.  
                                    
        output:         boolean, default True
                        whether save the corr matrix to excel file.

        '''
        warnings.filterwarnings('ignore', category=Warning)
        if factors=='quant':
            names= self.model.quant_factor
        elif factors=='quali':
            names= self.model.quali_factor
        elif factors=='all':
            names= self.model.quant_factor + self.model.quali_factor 
        else: # user input a single factor's name
            names = [factors]   

        self.ardata = self.normdata[names+[dependentvar]].copy()
        
        d={}
        for factor in names:
            rho = self.__Pearson(factor, dependentvar)
            spearmanrho = self.__Spearman(factor, dependentvar)
            AUROC = self.__AUROC(factor, dependentvar, isthere_def)
            if isthere_def:
                SomersD = 2*AUROC-1
            else:
                SomersD = self.__SomersD(factor, dependentvar)
            d.update({factor:pd.Series([rho, spearmanrho, AUROC, SomersD], index= ['Correlation','Spearman','AUROC','SomersD'])})   
        result = pd.DataFrame(d, columns=[names])
        if output:
            result.to_excel('AR_'+factors+'.xlsx')
        return result            
        
naics_code ={
"11":"Agriculture, Forestry, Fishing and Hunting",
"111":"Crop Production",
"1111":"Oilseed and Grain Farming",
"11111":"Soybean Farming",
"111110":"Soybean Farming",
"11112":"Oilseed (except Soybean) Farming",
"111120":"Oilseed (except Soybean) Farming ",
"11113":"Dry Pea and Bean Farming",
"111130":"Dry Pea and Bean Farming ",
"11114":"Wheat Farming",
"111140":"Wheat Farming",
"11115":"Corn Farming",
"111150":"Corn Farming ",
"11116":"Rice Farming",
"111160":"Rice Farming",
"11119":"Other Grain Farming",
"111191":"Oilseed and Grain Combination Farming ",
"111199":"All Other Grain Farming ",
"1112":"Vegetable and Melon Farming",
"11121":"Vegetable and Melon Farming",
"111211":"Potato Farming ",
"111219":"Other Vegetable (except Potato) and Melon Farming ",
"1113":"Fruit and Tree Nut Farming",
"11131":"Orange Groves",
"111310":"Orange Groves",
"11132":"Citrus (except Orange) Groves",
"111320":"Citrus (except Orange) Groves ",
"11133":"Noncitrus Fruit and Tree Nut Farming",
"111331":"Apple Orchards ",
"111332":"Grape Vineyards ",
"111333":"Strawberry Farming ",
"111334":"Berry (except Strawberry) Farming ",
"111335":"Tree Nut Farming ",
"111336":"Fruit and Tree Nut Combination Farming ",
"111339":"Other Noncitrus Fruit Farming ",
"1114":"Greenhouse, Nursery, and Floriculture Production",
"11141":"Food Crops Grown Under Cover",
"111411":"Mushroom Production ",
"111419":"Other Food Crops Grown Under Cover ",
"11142":"Nursery and Floriculture Production",
"111421":"Nursery and Tree Production ",
"111422":"Floriculture Production ",
"1119":"Other Crop Farming",
"11191":"Tobacco Farming",
"111910":"Tobacco Farming",
"11192":"Cotton Farming",
"111920":"Cotton Farming",
"11193":"Sugarcane Farming",
"111930":"Sugarcane Farming",
"11194":"Hay Farming",
"111940":"Hay Farming ",
"11199":"All Other Crop Farming",
"111991":"Sugar Beet Farming ",
"111992":"Peanut Farming ",
"111998":"All Other Miscellaneous Crop Farming ",
"112":"Animal Production and Aquaculture",
"1121":"Cattle Ranching and Farming",
"11211":"Beef Cattle Ranching and Farming, including Feedlots",
"112111":"Beef Cattle Ranching and Farming ",
"112112":"Cattle Feedlots ",
"11212":"Dairy Cattle and Milk Production",
"112120":"Dairy Cattle and Milk Production",
"11213":"Dual-Purpose Cattle Ranching and Farming",
"112130":"Dual-Purpose Cattle Ranching and Farming ",
"1122":"Hog and Pig Farming",
"11221":"Hog and Pig Farming",
"112210":"Hog and Pig Farming ",
"1123":"Poultry and Egg Production",
"11231":"Chicken Egg Production",
"112310":"Chicken Egg Production ",
"11232":"Broilers and Other Meat Type Chicken Production",
"112320":"Broilers and Other Meat Type Chicken Production ",
"11233":"Turkey Production",
"112330":"Turkey Production",
"11234":"Poultry Hatcheries",
"112340":"Poultry Hatcheries",
"11239":"Other Poultry Production",
"112390":"Other Poultry Production ",
"1124":"Sheep and Goat Farming",
"11241":"Sheep Farming",
"112410":"Sheep Farming",
"11242":"Goat Farming",
"112420":"Goat Farming",
"1125":"Aquaculture",
"11251":"Aquaculture",
"112511":"Finfish Farming and Fish Hatcheries ",
"112512":"Shellfish Farming ",
"112519":"Other Aquaculture ",
"1129":"Other Animal Production",
"11291":"Apiculture",
"112910":"Apiculture",
"11292":"Horses and Other Equine Production",
"112920":"Horses and Other Equine Production",
"11293":"Fur-Bearing Animal and Rabbit Production",
"112930":"Fur-Bearing Animal and Rabbit Production",
"11299":"All Other Animal Production",
"112990":"All Other Animal Production ",
"113":"Forestry and Logging",
"1131":"Timber Tract Operations",
"11311":"Timber Tract Operations",
"113110":"Timber Tract Operations",
"1132":"Forest Nurseries and Gathering of Forest Products",
"11321":"Forest Nurseries and Gathering of Forest Products",
"113210":"Forest Nurseries and Gathering of Forest Products ",
"1133":"Logging",
"11331":"Logging",
"113310":"Logging ",
"114":"Fishing, Hunting and Trapping",
"1141":"Fishing",
"11411":"Fishing",
"114111":"Finfish Fishing ",
"114112":"Shellfish Fishing ",
"114119":"Other Marine Fishing ",
"1142":"Hunting and Trapping",
"11421":"Hunting and Trapping",
"114210":"Hunting and Trapping",
"115":"Support Activities for Agriculture and Forestry",
"1151":"Support Activities for Crop Production",
"11511":"Support Activities for Crop Production",
"115111":"Cotton Ginning ",
"115112":"Soil Preparation, Planting, and Cultivating ",
"115113":"Crop Harvesting, Primarily by Machine ",
"115114":"Postharvest Crop Activities (except Cotton Ginning) ",
"115115":"Farm Labor Contractors and Crew Leaders ",
"115116":"Farm Management Services ",
"1152":"Support Activities for Animal Production",
"11521":"Support Activities for Animal Production",
"115210":"Support Activities for Animal Production",
"1153":"Support Activities for Forestry",
"11531":"Support Activities for Forestry",
"115310":"Support Activities for Forestry",
"21":"Mining, Quarrying, and Oil and Gas Extraction",
"211":"Oil and Gas Extraction",
"2111":"Oil and Gas Extraction",
"21112":"Crude Petroleum Extraction ",
"211120":"Crude Petroleum Extraction ",
"21113":"Natural Gas Extraction ",
"211130":"Natural Gas Extraction ",
"212":"Mining (except Oil and Gas)",
"2121":"Coal Mining",
"21211":"Coal Mining",
"212111":"Bituminous Coal and Lignite Surface Mining ",
"212112":"Bituminous Coal Underground Mining ",
"212113":"Anthracite Mining ",
"2122":"Metal Ore Mining",
"21221":"Iron Ore Mining",
"212210":"Iron Ore Mining",
"21222":"Gold Ore and Silver Ore Mining",
"212221":"Gold Ore Mining ",
"212222":"Silver Ore Mining ",
"21223":"Copper, Nickel, Lead, and Zinc Mining",
"212230":"Copper, Nickel, Lead, and Zinc Mining ",
"21229":"Other Metal Ore Mining",
"212291":"Uranium-Radium-Vanadium Ore Mining ",
"212299":"All Other Metal Ore Mining ",
"2123":"Nonmetallic Mineral Mining and Quarrying",
"21231":"Stone Mining and Quarrying",
"212311":"Dimension Stone Mining and Quarrying ",
"212312":"Crushed and Broken Limestone Mining and Quarrying ",
"212313":"Crushed and Broken Granite Mining and Quarrying ",
"212319":"Other Crushed and Broken Stone Mining and Quarrying ",
"21232":"Sand, Gravel, Clay, and Ceramic and Refractory Minerals Mining and Quarrying",
"212321":"Construction Sand and Gravel Mining ",
"212322":"Industrial Sand Mining ",
"212324":"Kaolin and Ball Clay Mining ",
"212325":"Clay and Ceramic and Refractory Minerals Mining ",
"21239":"Other Nonmetallic Mineral Mining and Quarrying",
"212391":"Potash, Soda, and Borate Mineral Mining ",
"212392":"Phosphate Rock Mining ",
"212393":"Other Chemical and Fertilizer Mineral Mining ",
"212399":"All Other Nonmetallic Mineral Mining ",
"213":"Support Activities for Mining",
"2131":"Support Activities for Mining",
"21311":"Support Activities for Mining",
"213111":"Drilling Oil and Gas Wells",
"213112":"Support Activities for Oil and Gas Operations ",
"213113":"Support Activities for Coal Mining ",
"213114":"Support Activities for Metal Mining ",
"213115":"Support Activities for Nonmetallic Minerals (except Fuels) Mining ",
"22":"Utilities",
"221":"Utilities ",
"2211":"Electric Power Generation, Transmission and Distribution",
"22111":"Electric Power Generation ",
"221111":"Hydroelectric Power Generation ",
"221112":"Fossil Fuel Electric Power Generation ",
"221113":"Nuclear Electric Power Generation ",
"221114":"Solar Electric Power Generation ",
"221115":"Wind Electric Power Generation ",
"221116":"Geothermal Electric Power Generation ",
"221117":"Biomass Electric Power Generation ",
"221118":"Other Electric Power Generation ",
"22112":"Electric Power Transmission, Control, and Distribution ",
"221121":"Electric Bulk Power Transmission and Control ",
"221122":"Electric Power Distribution ",
"2212":"Natural Gas Distribution ",
"22121":"Natural Gas Distribution ",
"221210":"Natural Gas Distribution ",
"2213":"Water, Sewage and Other Systems ",
"22131":"Water Supply and Irrigation Systems ",
"221310":"Water Supply and Irrigation Systems ",
"22132":"Sewage Treatment Facilities ",
"221320":"Sewage Treatment Facilities ",
"22133":"Steam and Air-Conditioning Supply ",
"221330":"Steam and Air-Conditioning Supply ",
"23":"Construction",
"236":"Construction of Buildings",
"2361":"Residential Building Construction",
"23611":"Residential Building Construction",
"236115":"New Single-Family Housing Construction (except For-Sale Builders) ",
"236116":"New Multifamily Housing Construction (except For-Sale Builders) ",
"236117":"New Housing For-Sale Builders ",
"236118":"Residential Remodelers ",
"2362":"Nonresidential Building Construction",
"23621":"Industrial Building Construction",
"236210":"Industrial Building Construction ",
"23622":"Commercial and Institutional Building Construction",
"236220":"Commercial and Institutional Building Construction ",
"237":"Heavy and Civil Engineering Construction",
"2371":"Utility System Construction",
"23711":"Water and Sewer Line and Related Structures Construction",
"237110":"Water and Sewer Line and Related Structures Construction ",
"23712":"Oil and Gas Pipeline and Related Structures Construction",
"237120":"Oil and Gas Pipeline and Related Structures Construction ",
"23713":"Power and Communication Line and Related Structures Construction",
"237130":"Power and Communication Line and Related Structures Construction ",
"2372":"Land Subdivision",
"23721":"Land Subdivision",
"237210":"Land Subdivision ",
"2373":"Highway, Street, and Bridge Construction",
"23731":"Highway, Street, and Bridge Construction",
"237310":"Highway, Street, and Bridge Construction ",
"2379":"Other Heavy and Civil Engineering Construction",
"23799":"Other Heavy and Civil Engineering Construction",
"237990":"Other Heavy and Civil Engineering Construction ",
"238":"Specialty Trade Contractors",
"2381":"Foundation, Structure, and Building Exterior Contractors",
"23811":"Poured Concrete Foundation and Structure Contractors ",
"238110":"Poured Concrete Foundation and Structure Contractors ",
"23812":"Structural Steel and Precast Concrete Contractors ",
"238120":"Structural Steel and Precast Concrete Contractors ",
"23813":"Framing Contractors ",
"238130":"Framing Contractors ",
"23814":"Masonry Contractors ",
"238140":"Masonry Contractors ",
"23815":"Glass and Glazing Contractors ",
"238150":"Glass and Glazing Contractors ",
"23816":"Roofing Contractors ",
"238160":"Roofing Contractors ",
"23817":"Siding Contractors ",
"238170":"Siding Contractors ",
"23819":"Other Foundation, Structure, and Building Exterior Contractors ",
"238190":"Other Foundation, Structure, and Building Exterior Contractors ",
"2382":"Building Equipment Contractors",
"23821":"Electrical Contractors and Other Wiring Installation Contractors",
"238210":"Electrical Contractors and Other Wiring Installation Contractors",
"23822":"Plumbing, Heating, and Air-Conditioning Contractors",
"238220":"Plumbing, Heating, and Air-Conditioning Contractors ",
"23829":"Other Building Equipment Contractors",
"238290":"Other Building Equipment Contractors ",
"2383":"Building Finishing Contractors",
"23831":"Drywall and Insulation Contractors",
"238310":"Drywall and Insulation Contractors ",
"23832":"Painting and Wall Covering Contractors",
"238320":"Painting and Wall Covering Contractors",
"23833":"Flooring Contractors",
"238330":"Flooring Contractors",
"23834":"Tile and Terrazzo Contractors",
"238340":"Tile and Terrazzo Contractors",
"23835":"Finish Carpentry Contractors",
"238350":"Finish Carpentry Contractors",
"23839":"Other Building Finishing Contractors",
"238390":"Other Building Finishing Contractors",
"2389":"Other Specialty Trade Contractors",
"23891":"Site Preparation Contractors",
"238910":"Site Preparation Contractors",
"23899":"All Other Specialty Trade Contractors",
"238990":"All Other Specialty Trade Contractors",
"31":"Manufacturing",
"32":"Manufacturing",
"33":"Manufacturing",
"311":"Food Manufacturing",
"3111":"Animal Food Manufacturing",
"31111":"Animal Food Manufacturing",
"311111":"Dog and Cat Food Manufacturing ",
"311119":"Other Animal Food Manufacturing ",
"3112":"Grain and Oilseed Milling",
"31121":"Flour Milling and Malt Manufacturing",
"311211":"Flour Milling ",
"311212":"Rice Milling ",
"311213":"Malt Manufacturing ",
"31122":"Starch and Vegetable Fats and Oils Manufacturing",
"311221":"Wet Corn Milling ",
"311224":"Soybean and Other Oilseed Processing ",
"311225":"Fats and Oils Refining and Blending ",
"31123":"Breakfast Cereal Manufacturing",
"311230":"Breakfast Cereal Manufacturing",
"3113":"Sugar and Confectionery Product Manufacturing",
"31131":"Sugar Manufacturing",
"311313":"Beet Sugar Manufacturing ",
"311314":"Cane Sugar Manufacturing ",
"31134":"Nonchocolate Confectionery Manufacturing",
"311340":"Nonchocolate Confectionery Manufacturing",
"31135":"Chocolate and Confectionery Manufacturing",
"311351":"Chocolate and Confectionery Manufacturing from Cacao Beans ",
"311352":"Confectionery Manufacturing from Purchased Chocolate ",
"3114":"Fruit and Vegetable Preserving and Specialty Food Manufacturing",
"31141":"Frozen Food Manufacturing",
"311411":"Frozen Fruit, Juice, and Vegetable Manufacturing ",
"311412":"Frozen Specialty Food Manufacturing ",
"31142":"Fruit and Vegetable Canning, Pickling, and Drying",
"311421":"Fruit and Vegetable Canning ",
"311422":"Specialty Canning ",
"311423":"Dried and Dehydrated Food Manufacturing ",
"3115":"Dairy Product Manufacturing",
"31151":"Dairy Product (except Frozen) Manufacturing",
"311511":"Fluid Milk Manufacturing ",
"311512":"Creamery Butter Manufacturing ",
"311513":"Cheese Manufacturing ",
"311514":"Dry, Condensed, and Evaporated Dairy Product Manufacturing ",
"31152":"Ice Cream and Frozen Dessert Manufacturing",
"311520":"Ice Cream and Frozen Dessert Manufacturing",
"3116":"Animal Slaughtering and Processing",
"31161":"Animal Slaughtering and Processing",
"311611":"Animal (except Poultry) Slaughtering ",
"311612":"Meat Processed from Carcasses ",
"311613":"Rendering and Meat Byproduct Processing ",
"311615":"Poultry Processing ",
"3117":"Seafood Product Preparation and Packaging",
"31171":"Seafood Product Preparation and Packaging",
"311710":"Seafood Product Preparation and Packaging",
"3118":"Bakeries and Tortilla Manufacturing",
"31181":"Bread and Bakery Product Manufacturing",
"311811":"Retail Bakeries ",
"311812":"Commercial Bakeries ",
"311813":"Frozen Cakes, Pies, and Other Pastries Manufacturing ",
"31182":"Cookie, Cracker, and Pasta Manufacturing",
"311821":"Cookie and Cracker Manufacturing ",
"311824":"Dry Pasta, Dough, and Flour Mixes Manufacturing from Purchased Flour ",
"31183":"Tortilla Manufacturing",
"311830":"Tortilla Manufacturing",
"3119":"Other Food Manufacturing",
"31191":"Snack Food Manufacturing",
"311911":"Roasted Nuts and Peanut Butter Manufacturing ",
"311919":"Other Snack Food Manufacturing ",
"31192":"Coffee and Tea Manufacturing",
"311920":"Coffee and Tea Manufacturing ",
"31193":"Flavoring Syrup and Concentrate Manufacturing",
"311930":"Flavoring Syrup and Concentrate Manufacturing",
"31194":"Seasoning and Dressing Manufacturing",
"311941":"Mayonnaise, Dressing, and Other Prepared Sauce Manufacturing ",
"311942":"Spice and Extract Manufacturing ",
"31199":"All Other Food Manufacturing",
"311991":"Perishable Prepared Food Manufacturing ",
"311999":"All Other Miscellaneous Food Manufacturing ",
"312":"Beverage and Tobacco Product Manufacturing",
"3121":"Beverage Manufacturing",
"31211":"Soft Drink and Ice Manufacturing",
"312111":"Soft Drink Manufacturing ",
"312112":"Bottled Water Manufacturing ",
"312113":"Ice Manufacturing ",
"31212":"Breweries",
"312120":"Breweries",
"31213":"Wineries",
"312130":"Wineries ",
"31214":"Distilleries",
"312140":"Distilleries ",
"3122":"Tobacco Manufacturing",
"31223":"Tobacco Manufacturing",
"312230":"Tobacco Manufacturing ",
"313":"Textile Mills",
"3131":"Fiber, Yarn, and Thread Mills",
"31311":"Fiber, Yarn, and Thread Mills",
"313110":"Fiber, Yarn, and Thread Mills ",
"3132":"Fabric Mills",
"31321":"Broadwoven Fabric Mills",
"313210":"Broadwoven Fabric Mills",
"31322":"Narrow Fabric Mills and Schiffli Machine Embroidery",
"313220":"Narrow Fabric Mills and Schiffli Machine Embroidery",
"31323":"Nonwoven Fabric Mills",
"313230":"Nonwoven Fabric Mills",
"31324":"Knit Fabric Mills",
"313240":"Knit Fabric Mills",
"3133":"Textile and Fabric Finishing and Fabric Coating Mills",
"31331":"Textile and Fabric Finishing Mills",
"313310":"Textile and Fabric Finishing Mills ",
"31332":"Fabric Coating Mills",
"313320":"Fabric Coating Mills",
"314":"Textile Product Mills",
"3141":"Textile Furnishings Mills",
"31411":"Carpet and Rug Mills",
"314110":"Carpet and Rug Mills",
"31412":"Curtain and Linen Mills",
"314120":"Curtain and Linen Mills",
"3149":"Other Textile Product Mills",
"31491":"Textile Bag and Canvas Mills",
"314910":"Textile Bag and Canvas Mills ",
"31499":"All Other Textile Product Mills",
"314994":"Rope, Cordage, Twine, Tire Cord, and Tire Fabric Mills ",
"314999":"All Other Miscellaneous Textile Product Mills ",
"315":"Apparel Manufacturing",
"3151":"Apparel Knitting Mills",
"31511":"Hosiery and Sock Mills",
"315110":"Hosiery and Sock Mills",
"31519":"Other Apparel Knitting Mills",
"315190":"Other Apparel Knitting Mills ",
"3152":"Cut and Sew Apparel Manufacturing",
"31521":"Cut and Sew Apparel Contractors ",
"315210":"Cut and Sew Apparel Contractors ",
"31522":"Men’s and Boys’ Cut and Sew Apparel Manufacturing ",
"315220":"Men’s and Boys’ Cut and Sew Apparel Manufacturing ",
"31524":"Women’s, Girls’, and Infants’ Cut and Sew Apparel Manufacturing",
"315240":"Women’s, Girls’, and Infants’ Cut and Sew Apparel Manufacturing ",
"31528":"Other Cut and Sew Apparel Manufacturing ",
"315280":"Other Cut and Sew Apparel Manufacturing ",
"3159":"Apparel Accessories and Other Apparel Manufacturing",
"31599":"Apparel Accessories and Other Apparel Manufacturing",
"315990":"Apparel Accessories and Other Apparel Manufacturing ",
"316":"Leather and Allied Product Manufacturing",
"3161":"Leather and Hide Tanning and Finishing",
"31611":"Leather and Hide Tanning and Finishing",
"316110":"Leather and Hide Tanning and Finishing",
"3162":"Footwear Manufacturing",
"31621":"Footwear Manufacturing",
"316210":"Footwear Manufacturing ",
"3169":"Other Leather and Allied Product Manufacturing",
"31699":"Other Leather and Allied Product Manufacturing",
"316992":"Women's Handbag and Purse Manufacturing ",
"316998":"All Other Leather Good and Allied Product Manufacturing ",
"321":"Wood Product Manufacturing",
"3211":"Sawmills and Wood Preservation",
"32111":"Sawmills and Wood Preservation",
"321113":"Sawmills ",
"321114":"Wood Preservation ",
"3212":"Veneer, Plywood, and Engineered Wood Product Manufacturing",
"32121":"Veneer, Plywood, and Engineered Wood Product Manufacturing",
"321211":"Hardwood Veneer and Plywood Manufacturing ",
"321212":"Softwood Veneer and Plywood Manufacturing ",
"321213":"Engineered Wood Member (except Truss) Manufacturing ",
"321214":"Truss Manufacturing ",
"321219":"Reconstituted Wood Product Manufacturing ",
"3219":"Other Wood Product Manufacturing",
"32191":"Millwork",
"321911":"Wood Window and Door Manufacturing ",
"321912":"Cut Stock, Resawing Lumber, and Planing ",
"321918":"Other Millwork (including Flooring) ",
"32192":"Wood Container and Pallet Manufacturing",
"321920":"Wood Container and Pallet Manufacturing",
"32199":"All Other Wood Product Manufacturing",
"321991":"Manufactured Home (Mobile Home) Manufacturing ",
"321992":"Prefabricated Wood Building Manufacturing ",
"321999":"All Other Miscellaneous Wood Product Manufacturing ",
"322":"Paper Manufacturing",
"3221":"Pulp, Paper, and Paperboard Mills",
"32211":"Pulp Mills",
"322110":"Pulp Mills ",
"32212":"Paper Mills",
"322121":"Paper (except Newsprint) Mills ",
"322122":"Newsprint Mills ",
"32213":"Paperboard Mills",
"322130":"Paperboard Mills ",
"3222":"Converted Paper Product Manufacturing",
"32221":"Paperboard Container Manufacturing",
"322211":"Corrugated and Solid Fiber Box Manufacturing ",
"322212":"Folding Paperboard Box Manufacturing ",
"322219":"Other Paperboard Container Manufacturing ",
"32222":"Paper Bag and Coated and Treated Paper Manufacturing",
"322220":"Paper Bag and Coated and Treated Paper Manufacturing",
"32223":"Stationery Product Manufacturing",
"322230":"Stationery Product Manufacturing",
"32229":"Other Converted Paper Product Manufacturing",
"322291":"Sanitary Paper Product Manufacturing ",
"322299":"All Other Converted Paper Product Manufacturing ",
"323":"Printing and Related Support Activities",
"3231":"Printing and Related Support Activities",
"32311":"Printing",
"323111":"Commercial Printing (except Screen and Books) ",
"323113":"Commercial Screen Printing ",
"323117":"Books Printing ",
"32312":"Support Activities for Printing",
"323120":"Support Activities for Printing",
"324":"Petroleum and Coal Products Manufacturing",
"3241":"Petroleum and Coal Products Manufacturing",
"32411":"Petroleum Refineries",
"324110":"Petroleum Refineries",
"32412":"Asphalt Paving, Roofing, and Saturated Materials Manufacturing",
"324121":"Asphalt Paving Mixture and Block Manufacturing ",
"324122":"Asphalt Shingle and Coating Materials Manufacturing ",
"32419":"Other Petroleum and Coal Products Manufacturing",
"324191":"Petroleum Lubricating Oil and Grease Manufacturing ",
"324199":"All Other Petroleum and Coal Products Manufacturing ",
"325":"Chemical Manufacturing",
"3251":"Basic Chemical Manufacturing",
"32511":"Petrochemical Manufacturing",
"325110":"Petrochemical Manufacturing",
"32512":"Industrial Gas Manufacturing",
"325120":"Industrial Gas Manufacturing",
"32513":"Synthetic Dye and Pigment Manufacturing",
"325130":"Synthetic Dye and Pigment Manufacturing",
"32518":"Other Basic Inorganic Chemical Manufacturing",
"325180":"Other Basic Inorganic Chemical Manufacturing ",
"32519":"Other Basic Organic Chemical Manufacturing",
"325193":"Ethyl Alcohol Manufacturing ",
"325194":"Cyclic Crude, Intermediate, and Gum and Wood Chemical Manufacturing ",
"325199":"All Other Basic Organic Chemical Manufacturing ",
"3252":"Resin, Synthetic Rubber, and Artificial and Synthetic Fibers and Filaments Manufacturing",
"32521":"Resin and Synthetic Rubber Manufacturing",
"325211":"Plastics Material and Resin Manufacturing ",
"325212":"Synthetic Rubber Manufacturing ",
"32522":"Artificial and Synthetic Fibers and Filaments Manufacturing",
"325220":"Artificial and Synthetic Fibers and Filaments Manufacturing",
"3253":"Pesticide, Fertilizer, and Other Agricultural Chemical Manufacturing",
"32531":"Fertilizer Manufacturing",
"325311":"Nitrogenous Fertilizer Manufacturing ",
"325312":"Phosphatic Fertilizer Manufacturing ",
"325314":"Fertilizer (Mixing Only) Manufacturing ",
"32532":"Pesticide and Other Agricultural Chemical Manufacturing",
"325320":"Pesticide and Other Agricultural Chemical Manufacturing",
"3254":"Pharmaceutical and Medicine Manufacturing",
"32541":"Pharmaceutical and Medicine Manufacturing",
"325411":"Medicinal and Botanical Manufacturing ",
"325412":"Pharmaceutical Preparation Manufacturing ",
"325413":"In-Vitro Diagnostic Substance Manufacturing ",
"325414":"Biological Product (except Diagnostic) Manufacturing ",
"3255":"Paint, Coating, and Adhesive Manufacturing",
"32551":"Paint and Coating Manufacturing",
"325510":"Paint and Coating Manufacturing",
"32552":"Adhesive Manufacturing",
"325520":"Adhesive Manufacturing",
"3256":"Soap, Cleaning Compound, and Toilet Preparation Manufacturing",
"32561":"Soap and Cleaning Compound Manufacturing",
"325611":"Soap and Other Detergent Manufacturing ",
"325612":"Polish and Other Sanitation Good Manufacturing ",
"325613":"Surface Active Agent Manufacturing ",
"32562":"Toilet Preparation Manufacturing",
"325620":"Toilet Preparation Manufacturing",
"3259":"Other Chemical Product and Preparation Manufacturing",
"32591":"Printing Ink Manufacturing",
"325910":"Printing Ink Manufacturing",
"32592":"Explosives Manufacturing",
"325920":"Explosives Manufacturing",
"32599":"All Other Chemical Product and Preparation Manufacturing",
"325991":"Custom Compounding of Purchased Resins ",
"325992":"Photographic Film, Paper, Plate, and Chemical Manufacturing ",
"325998":"All Other Miscellaneous Chemical Product and Preparation Manufacturing ",
"326":"Plastics and Rubber Products Manufacturing",
"3261":"Plastics Product Manufacturing",
"32611":"Plastics Packaging Materials and Unlaminated Film and Sheet Manufacturing",
"326111":"Plastics Bag and Pouch Manufacturing ",
"326112":"Plastics Packaging Film and Sheet (including Laminated) Manufacturing ",
"326113":"Unlaminated Plastics Film and Sheet (except Packaging) Manufacturing ",
"32612":"Plastics Pipe, Pipe Fitting, and Unlaminated Profile Shape Manufacturing",
"326121":"Unlaminated Plastics Profile Shape Manufacturing ",
"326122":"Plastics Pipe and Pipe Fitting Manufacturing ",
"32613":"Laminated Plastics Plate, Sheet (except Packaging), and Shape Manufacturing",
"326130":"Laminated Plastics Plate, Sheet (except Packaging), and Shape Manufacturing",
"32614":"Polystyrene Foam Product Manufacturing",
"326140":"Polystyrene Foam Product Manufacturing",
"32615":"Urethane and Other Foam Product (except Polystyrene) Manufacturing",
"326150":"Urethane and Other Foam Product (except Polystyrene) Manufacturing",
"32616":"Plastics Bottle Manufacturing",
"326160":"Plastics Bottle Manufacturing",
"32619":"Other Plastics Product Manufacturing",
"326191":"Plastics Plumbing Fixture Manufacturing ",
"326199":"All Other Plastics Product Manufacturing ",
"3262":"Rubber Product Manufacturing",
"32621":"Tire Manufacturing",
"326211":"Tire Manufacturing (except Retreading) ",
"326212":"Tire Retreading ",
"32622":"Rubber and Plastics Hoses and Belting Manufacturing",
"326220":"Rubber and Plastics Hoses and Belting Manufacturing",
"32629":"Other Rubber Product Manufacturing",
"326291":"Rubber Product Manufacturing for Mechanical Use ",
"326299":"All Other Rubber Product Manufacturing ",
"327":"Nonmetallic Mineral Product Manufacturing",
"3271":"Clay Product and Refractory Manufacturing",
"32711":"Pottery, Ceramics, and Plumbing Fixture Manufacturing",
"327110":"Pottery, Ceramics, and Plumbing Fixture Manufacturing ",
"32712":"Clay Building Material and Refractories Manufacturing",
"327120":"Clay Building Material and Refractories Manufacturing ",
"3272":"Glass and Glass Product Manufacturing",
"32721":"Glass and Glass Product Manufacturing",
"327211":"Flat Glass Manufacturing ",
"327212":"Other Pressed and Blown Glass and Glassware Manufacturing ",
"327213":"Glass Container Manufacturing ",
"327215":"Glass Product Manufacturing Made of Purchased Glass ",
"3273":"Cement and Concrete Product Manufacturing",
"32731":"Cement Manufacturing",
"327310":"Cement Manufacturing",
"32732":"Ready-Mix Concrete Manufacturing",
"327320":"Ready-Mix Concrete Manufacturing",
"32733":"Concrete Pipe, Brick, and Block Manufacturing",
"327331":"Concrete Block and Brick Manufacturing ",
"327332":"Concrete Pipe Manufacturing ",
"32739":"Other Concrete Product Manufacturing",
"327390":"Other Concrete Product Manufacturing ",
"3274":"Lime and Gypsum Product Manufacturing",
"32741":"Lime Manufacturing",
"327410":"Lime Manufacturing",
"32742":"Gypsum Product Manufacturing",
"327420":"Gypsum Product Manufacturing",
"3279":"Other Nonmetallic Mineral Product Manufacturing",
"32791":"Abrasive Product Manufacturing",
"327910":"Abrasive Product Manufacturing",
"32799":"All Other Nonmetallic Mineral Product Manufacturing",
"327991":"Cut Stone and Stone Product Manufacturing ",
"327992":"Ground or Treated Mineral and Earth Manufacturing ",
"327993":"Mineral Wool Manufacturing ",
"327999":"All Other Miscellaneous Nonmetallic Mineral Product Manufacturing ",
"331":"Primary Metal Manufacturing",
"3311":"Iron and Steel Mills and Ferroalloy Manufacturing",
"33111":"Iron and Steel Mills and Ferroalloy Manufacturing",
"331110":"Iron and Steel Mills and Ferroalloy Manufacturing ",
"3312":"Steel Product Manufacturing from Purchased Steel",
"33121":"Iron and Steel Pipe and Tube Manufacturing from Purchased Steel",
"331210":"Iron and Steel Pipe and Tube Manufacturing from Purchased Steel",
"33122":"Rolling and Drawing of Purchased Steel",
"331221":"Rolled Steel Shape Manufacturing ",
"331222":"Steel Wire Drawing ",
"3313":"Alumina and Aluminum Production and Processing",
"33131":"Alumina and Aluminum Production and Processing",
"331313":"Alumina Refining and Primary Aluminum Production ",
"331314":"Secondary Smelting and Alloying of Aluminum ",
"331315":"Aluminum Sheet, Plate, and Foil Manufacturing ",
"331318":"Other Aluminum Rolling, Drawing, and Extruding ",
"3314":"Nonferrous Metal (except Aluminum) Production and Processing",
"33141":"Nonferrous Metal (except Aluminum) Smelting and Refining",
"331410":"Nonferrous Metal (except Aluminum) Smelting and Refining ",
"33142":"Copper Rolling, Drawing, Extruding, and Alloying",
"331420":"Copper Rolling, Drawing, Extruding, and Alloying",
"33149":"Nonferrous Metal (except Copper and Aluminum) Rolling, Drawing, Extruding, and Alloying",
"331491":"Nonferrous Metal (except Copper and Aluminum) Rolling, Drawing, and Extruding ",
"331492":"Secondary Smelting, Refining, and Alloying of Nonferrous Metal (except Copper and Aluminum) ",
"3315":"Foundries",
"33151":"Ferrous Metal Foundries",
"331511":"Iron Foundries ",
"331512":"Steel Investment Foundries ",
"331513":"Steel Foundries (except Investment) ",
"33152":"Nonferrous Metal Foundries",
"331523":"Nonferrous Metal Die-Casting Foundries ",
"331524":"Aluminum Foundries (except Die-Casting) ",
"331529":"Other Nonferrous Metal Foundries (except Die-Casting) ",
"332":"Fabricated Metal Product Manufacturing",
"3321":"Forging and Stamping",
"33211":"Forging and Stamping",
"332111":"Iron and Steel Forging ",
"332112":"Nonferrous Forging ",
"332114":"Custom Roll Forming ",
"332117":"Powder Metallurgy Part Manufacturing ",
"332119":"Metal Crown, Closure, and Other Metal Stamping (except Automotive) ",
"3322":"Cutlery and Handtool Manufacturing",
"33221":"Cutlery and Handtool Manufacturing",
"332215":"Metal Kitchen Cookware, Utensil, Cutlery, and Flatware (except Precious) Manufacturing ",
"332216":"Saw Blade and Handtool Manufacturing ",
"3323":"Architectural and Structural Metals Manufacturing",
"33231":"Plate Work and Fabricated Structural Product Manufacturing",
"332311":"Prefabricated Metal Building and Component Manufacturing ",
"332312":"Fabricated Structural Metal Manufacturing ",
"332313":"Plate Work Manufacturing ",
"33232":"Ornamental and Architectural Metal Products Manufacturing",
"332321":"Metal Window and Door Manufacturing ",
"332322":"Sheet Metal Work Manufacturing ",
"332323":"Ornamental and Architectural Metal Work Manufacturing ",
"3324":"Boiler, Tank, and Shipping Container Manufacturing",
"33241":"Power Boiler and Heat Exchanger Manufacturing",
"332410":"Power Boiler and Heat Exchanger Manufacturing",
"33242":"Metal Tank (Heavy Gauge) Manufacturing",
"332420":"Metal Tank (Heavy Gauge) Manufacturing",
"33243":"Metal Can, Box, and Other Metal Container (Light Gauge) Manufacturing",
"332431":"Metal Can Manufacturing ",
"332439":"Other Metal Container Manufacturing ",
"3325":"Hardware Manufacturing",
"33251":"Hardware Manufacturing",
"332510":"Hardware Manufacturing",
"3326":"Spring and Wire Product Manufacturing",
"33261":"Spring and Wire Product Manufacturing",
"332613":"Spring Manufacturing ",
"332618":"Other Fabricated Wire Product Manufacturing ",
"3327":"Machine Shops; Turned Product; and Screw, Nut, and Bolt Manufacturing",
"33271":"Machine Shops",
"332710":"Machine Shops",
"33272":"Turned Product and Screw, Nut, and Bolt Manufacturing",
"332721":"Precision Turned Product Manufacturing ",
"332722":"Bolt, Nut, Screw, Rivet, and Washer Manufacturing ",
"3328":"Coating, Engraving, Heat Treating, and Allied Activities",
"33281":"Coating, Engraving, Heat Treating, and Allied Activities",
"332811":"Metal Heat Treating ",
"332812":"Metal Coating, Engraving (except Jewelry and Silverware), and Allied Services to Manufacturers ",
"332813":"Electroplating, Plating, Polishing, Anodizing, and Coloring ",
"3329":"Other Fabricated Metal Product Manufacturing",
"33291":"Metal Valve Manufacturing",
"332911":"Industrial Valve Manufacturing ",
"332912":"Fluid Power Valve and Hose Fitting Manufacturing ",
"332913":"Plumbing Fixture Fitting and Trim Manufacturing ",
"332919":"Other Metal Valve and Pipe Fitting Manufacturing ",
"33299":"All Other Fabricated Metal Product Manufacturing",
"332991":"Ball and Roller Bearing Manufacturing",
"332992":"Small Arms Ammunition Manufacturing ",
"332993":"Ammunition (except Small Arms) Manufacturing ",
"332994":"Small Arms, Ordnance, and Ordnance Accessories Manufacturing ",
"332996":"Fabricated Pipe and Pipe Fitting Manufacturing ",
"332999":"All Other Miscellaneous Fabricated Metal Product Manufacturing ",
"333":"Machinery Manufacturing",
"3331":"Agriculture, Construction, and Mining Machinery Manufacturing",
"33311":"Agricultural Implement Manufacturing",
"333111":"Farm Machinery and Equipment Manufacturing ",
"333112":"Lawn and Garden Tractor and Home Lawn and Garden Equipment Manufacturing ",
"33312":"Construction Machinery Manufacturing",
"333120":"Construction Machinery Manufacturing",
"33313":"Mining and Oil and Gas Field Machinery Manufacturing",
"333131":"Mining Machinery and Equipment Manufacturing ",
"333132":"Oil and Gas Field Machinery and Equipment Manufacturing ",
"3332":"Industrial Machinery Manufacturing",
"33324":"Industrial Machinery Manufacturing",
"333241":"Food Product Machinery Manufacturing ",
"333242":"Semiconductor Machinery Manufacturing ",
"333243":"Sawmill, Woodworking, and Paper Machinery Manufacturing ",
"333244":"Printing Machinery and Equipment Manufacturing ",
"333249":"Other Industrial Machinery Manufacturing ",
"3333":"Commercial and Service Industry Machinery Manufacturing",
"33331":"Commercial and Service Industry Machinery Manufacturing",
"333314":"Optical Instrument and Lens Manufacturing ",
"333316":"Photographic and Photocopying Equipment Manufacturing ",
"333318":"Other Commercial and Service Industry Machinery Manufacturing ",
"3334":"Ventilation, Heating, Air-Conditioning, and Commercial Refrigeration Equipment Manufacturing",
"33341":"Ventilation, Heating, Air-Conditioning, and Commercial Refrigeration Equipment Manufacturing",
"333413":"Industrial and Commercial Fan and Blower and Air Purification Equipment Manufacturing ",
"333414":"Heating Equipment (except Warm Air Furnaces) Manufacturing ",
"333415":"Air-Conditioning and Warm Air Heating Equipment and Commercial and Industrial Refrigeration Equipment Manufacturing ",
"3335":"Metalworking Machinery Manufacturing",
"33351":"Metalworking Machinery Manufacturing",
"333511":"Industrial Mold Manufacturing ",
"333514":"Special Die and Tool, Die Set, Jig, and Fixture Manufacturing ",
"333515":"Cutting Tool and Machine Tool Accessory Manufacturing ",
"333517":"Machine Tool Manufacturing ",
"333519":"Rolling Mill and Other Metalworking Machinery Manufacturing ",
"3336":"Engine, Turbine, and Power Transmission Equipment Manufacturing",
"33361":"Engine, Turbine, and Power Transmission Equipment Manufacturing",
"333611":"Turbine and Turbine Generator Set Units Manufacturing ",
"333612":"Speed Changer, Industrial High-Speed Drive, and Gear Manufacturing ",
"333613":"Mechanical Power Transmission Equipment Manufacturing ",
"333618":"Other Engine Equipment Manufacturing ",
"3339":"Other General Purpose Machinery Manufacturing",
"33391":"Pump and Compressor Manufacturing",
"333912":"Air and Gas Compressor Manufacturing ",
"333914":"Measuring, Dispensing, and Other Pumping Equipment Manufacturing ",
"33392":"Material Handling Equipment Manufacturing",
"333921":"Elevator and Moving Stairway Manufacturing ",
"333922":"Conveyor and Conveying Equipment Manufacturing ",
"333923":"Overhead Traveling Crane, Hoist, and Monorail System Manufacturing ",
"333924":"Industrial Truck, Tractor, Trailer, and Stacker Machinery Manufacturing ",
"33399":"All Other General Purpose Machinery Manufacturing",
"333991":"Power-Driven Handtool Manufacturing ",
"333992":"Welding and Soldering Equipment Manufacturing ",
"333993":"Packaging Machinery Manufacturing ",
"333994":"Industrial Process Furnace and Oven Manufacturing ",
"333995":"Fluid Power Cylinder and Actuator Manufacturing ",
"333996":"Fluid Power Pump and Motor Manufacturing ",
"333997":"Scale and Balance Manufacturing ",
"333999":"All Other Miscellaneous General Purpose Machinery Manufacturing ",
"334":"Computer and Electronic Product Manufacturing",
"3341":"Computer and Peripheral Equipment Manufacturing",
"33411":"Computer and Peripheral Equipment Manufacturing",
"334111":"Electronic Computer Manufacturing ",
"334112":"Computer Storage Device Manufacturing ",
"334118":"Computer Terminal and Other Computer Peripheral Equipment Manufacturing ",
"3342":"Communications Equipment Manufacturing",
"33421":"Telephone Apparatus Manufacturing",
"334210":"Telephone Apparatus Manufacturing",
"33422":"Radio and Television Broadcasting and Wireless Communications Equipment Manufacturing",
"334220":"Radio and Television Broadcasting and Wireless Communications Equipment Manufacturing",
"33429":"Other Communications Equipment Manufacturing",
"334290":"Other Communications Equipment Manufacturing",
"3343":"Audio and Video Equipment Manufacturing",
"33431":"Audio and Video Equipment Manufacturing",
"334310":"Audio and Video Equipment Manufacturing",
"3344":"Semiconductor and Other Electronic Component Manufacturing",
"33441":"Semiconductor and Other Electronic Component Manufacturing",
"334412":"Bare Printed Circuit Board Manufacturing  ",
"334413":"Semiconductor and Related Device Manufacturing ",
"334416":"Capacitor, Resistor, Coil, Transformer, and Other Inductor Manufacturing ",
"334417":"Electronic Connector Manufacturing ",
"334418":"Printed Circuit Assembly (Electronic Assembly) Manufacturing ",
"334419":"Other Electronic Component Manufacturing ",
"3345":"Navigational, Measuring, Electromedical, and Control Instruments Manufacturing",
"33451":"Navigational, Measuring, Electromedical, and Control Instruments Manufacturing",
"334510":"Electromedical and Electrotherapeutic Apparatus Manufacturing ",
"334511":"Search, Detection, Navigation, Guidance, Aeronautical, and Nautical System and Instrument Manufacturing ",
"334512":"Automatic Environmental Control Manufacturing for Residential, Commercial, and Appliance Use ",
"334513":"Instruments and Related Products Manufacturing for Measuring, Displaying, and Controlling Industrial Process Variables ",
"334514":"Totalizing Fluid Meter and Counting Device Manufacturing ",
"334515":"Instrument Manufacturing for Measuring and Testing Electricity and Electrical Signals ",
"334516":"Analytical Laboratory Instrument Manufacturing ",
"334517":"Irradiation Apparatus Manufacturing ",
"334519":"Other Measuring and Controlling Device Manufacturing ",
"3346":"Manufacturing and Reproducing Magnetic and Optical Media",
"33461":"Manufacturing and Reproducing Magnetic and Optical Media",
"334613":"Blank Magnetic and Optical Recording Media Manufacturing ",
"334614":"Software and Other Prerecorded Compact Disc, Tape, and Record Reproducing ",
"335":"Electrical Equipment, Appliance, and Component Manufacturing",
"3351":"Electric Lighting Equipment Manufacturing",
"33511":"Electric Lamp Bulb and Part Manufacturing",
"335110":"Electric Lamp Bulb and Part Manufacturing",
"33512":"Lighting Fixture Manufacturing",
"335121":"Residential Electric Lighting Fixture Manufacturing ",
"335122":"Commercial, Industrial, and Institutional Electric Lighting Fixture Manufacturing ",
"335129":"Other Lighting Equipment Manufacturing ",
"3352":"Household Appliance Manufacturing",
"33521":"Small Electrical Appliance Manufacturing",
"335210":"Small Electrical Appliance Manufacturing",
"33522":"Major Household Appliance Manufacturing ",
"335220":"Major Household Appliance Manufacturing ",
"3353":"Electrical Equipment Manufacturing",
"33531":"Electrical Equipment Manufacturing",
"335311":"Power, Distribution, and Specialty Transformer Manufacturing ",
"335312":"Motor and Generator Manufacturing ",
"335313":"Switchgear and Switchboard Apparatus Manufacturing ",
"335314":"Relay and Industrial Control Manufacturing ",
"3359":"Other Electrical Equipment and Component Manufacturing",
"33591":"Battery Manufacturing",
"335911":"Storage Battery Manufacturing ",
"335912":"Primary Battery Manufacturing ",
"33592":"Communication and Energy Wire and Cable Manufacturing",
"335921":"Fiber Optic Cable Manufacturing ",
"335929":"Other Communication and Energy Wire Manufacturing ",
"33593":"Wiring Device Manufacturing",
"335931":"Current-Carrying Wiring Device Manufacturing ",
"335932":"Noncurrent-Carrying Wiring Device Manufacturing ",
"33599":"All Other Electrical Equipment and Component Manufacturing",
"335991":"Carbon and Graphite Product Manufacturing ",
"335999":"All Other Miscellaneous Electrical Equipment and Component Manufacturing ",
"336":"Transportation Equipment Manufacturing",
"3361":"Motor Vehicle Manufacturing",
"33611":"Automobile and Light Duty Motor Vehicle Manufacturing",
"336111":"Automobile Manufacturing ",
"336112":"Light Truck and Utility Vehicle Manufacturing ",
"33612":"Heavy Duty Truck Manufacturing",
"336120":"Heavy Duty Truck Manufacturing",
"3362":"Motor Vehicle Body and Trailer Manufacturing",
"33621":"Motor Vehicle Body and Trailer Manufacturing",
"336211":"Motor Vehicle Body Manufacturing ",
"336212":"Truck Trailer Manufacturing ",
"336213":"Motor Home Manufacturing ",
"336214":"Travel Trailer and Camper Manufacturing ",
"3363":"Motor Vehicle Parts Manufacturing",
"33631":"Motor Vehicle Gasoline Engine and Engine Parts Manufacturing",
"336310":"Motor Vehicle Gasoline Engine and Engine Parts Manufacturing",
"33632":"Motor Vehicle Electrical and Electronic Equipment Manufacturing",
"336320":"Motor Vehicle Electrical and Electronic Equipment Manufacturing",
"33633":"Motor Vehicle Steering and Suspension Components (except Spring) Manufacturing",
"336330":"Motor Vehicle Steering and Suspension Components (except Spring) Manufacturing",
"33634":"Motor Vehicle Brake System Manufacturing",
"336340":"Motor Vehicle Brake System Manufacturing",
"33635":"Motor Vehicle Transmission and Power Train Parts Manufacturing",
"336350":"Motor Vehicle Transmission and Power Train Parts Manufacturing",
"33636":"Motor Vehicle Seating and Interior Trim Manufacturing",
"336360":"Motor Vehicle Seating and Interior Trim Manufacturing",
"33637":"Motor Vehicle Metal Stamping",
"336370":"Motor Vehicle Metal Stamping",
"33639":"Other Motor Vehicle Parts Manufacturing",
"336390":"Other Motor Vehicle Parts Manufacturing",
"3364":"Aerospace Product and Parts Manufacturing",
"33641":"Aerospace Product and Parts Manufacturing",
"336411":"Aircraft Manufacturing ",
"336412":"Aircraft Engine and Engine Parts Manufacturing ",
"336413":"Other Aircraft Parts and Auxiliary Equipment Manufacturing ",
"336414":"Guided Missile and Space Vehicle Manufacturing ",
"336415":"Guided Missile and Space Vehicle Propulsion Unit and Propulsion Unit Parts Manufacturing ",
"336419":"Other Guided Missile and Space Vehicle Parts and Auxiliary Equipment Manufacturing ",
"3365":"Railroad Rolling Stock Manufacturing",
"33651":"Railroad Rolling Stock Manufacturing",
"336510":"Railroad Rolling Stock Manufacturing",
"3366":"Ship and Boat Building",
"33661":"Ship and Boat Building",
"336611":"Ship Building and Repairing ",
"336612":"Boat Building ",
"3369":"Other Transportation Equipment Manufacturing",
"33699":"Other Transportation Equipment Manufacturing",
"336991":"Motorcycle, Bicycle, and Parts Manufacturing ",
"336992":"Military Armored Vehicle, Tank, and Tank Component Manufacturing ",
"336999":"All Other Transportation Equipment Manufacturing ",
"337":"Furniture and Related Product Manufacturing",
"3371":"Household and Institutional Furniture and Kitchen Cabinet Manufacturing",
"33711":"Wood Kitchen Cabinet and Countertop Manufacturing",
"337110":"Wood Kitchen Cabinet and Countertop Manufacturing",
"33712":"Household and Institutional Furniture Manufacturing",
"337121":"Upholstered Household Furniture Manufacturing ",
"337122":"Nonupholstered Wood Household Furniture Manufacturing ",
"337124":"Metal Household Furniture Manufacturing ",
"337125":"Household Furniture (except Wood and Metal) Manufacturing ",
"337127":"Institutional Furniture Manufacturing ",
"3372":"Office Furniture (including Fixtures) Manufacturing",
"33721":"Office Furniture (including Fixtures) Manufacturing",
"337211":"Wood Office Furniture Manufacturing ",
"337212":"Custom Architectural Woodwork and Millwork Manufacturing ",
"337214":"Office Furniture (except Wood) Manufacturing ",
"337215":"Showcase, Partition, Shelving, and Locker Manufacturing ",
"3379":"Other Furniture Related Product Manufacturing",
"33791":"Mattress Manufacturing",
"337910":"Mattress Manufacturing",
"33792":"Blind and Shade Manufacturing",
"337920":"Blind and Shade Manufacturing",
"339":"Miscellaneous Manufacturing",
"3391":"Medical Equipment and Supplies Manufacturing",
"33911":"Medical Equipment and Supplies Manufacturing",
"339112":"Surgical and Medical Instrument Manufacturing ",
"339113":"Surgical Appliance and Supplies Manufacturing ",
"339114":"Dental Equipment and Supplies Manufacturing ",
"339115":"Ophthalmic Goods Manufacturing ",
"339116":"Dental Laboratories ",
"3399":"Other Miscellaneous Manufacturing",
"33991":"Jewelry and Silverware Manufacturing",
"339910":"Jewelry and Silverware Manufacturing ",
"33992":"Sporting and Athletic Goods Manufacturing",
"339920":"Sporting and Athletic Goods Manufacturing",
"33993":"Doll, Toy, and Game Manufacturing",
"339930":"Doll, Toy, and Game Manufacturing",
"33994":"Office Supplies (except Paper) Manufacturing",
"339940":"Office Supplies (except Paper) Manufacturing",
"33995":"Sign Manufacturing",
"339950":"Sign Manufacturing",
"33999":"All Other Miscellaneous Manufacturing",
"339991":"Gasket, Packing, and Sealing Device Manufacturing ",
"339992":"Musical Instrument Manufacturing ",
"339993":"Fastener, Button, Needle, and Pin Manufacturing ",
"339994":"Broom, Brush, and Mop Manufacturing ",
"339995":"Burial Casket Manufacturing ",
"339999":"All Other Miscellaneous Manufacturing ",
"42":"Wholesale Trade",
"423":"Merchant Wholesalers, Durable Goods ",
"4231":"Motor Vehicle and Motor Vehicle Parts and Supplies Merchant Wholesalers ",
"42311":"Automobile and Other Motor Vehicle Merchant Wholesalers ",
"423110":"Automobile and Other Motor Vehicle Merchant Wholesalers ",
"42312":"Motor Vehicle Supplies and New Parts Merchant Wholesalers ",
"423120":"Motor Vehicle Supplies and New Parts Merchant Wholesalers ",
"42313":"Tire and Tube Merchant Wholesalers ",
"423130":"Tire and Tube Merchant Wholesalers ",
"42314":"Motor Vehicle Parts (Used) Merchant Wholesalers ",
"423140":"Motor Vehicle Parts (Used) Merchant Wholesalers ",
"4232":"Furniture and Home Furnishing Merchant Wholesalers ",
"42321":"Furniture Merchant Wholesalers ",
"423210":"Furniture Merchant Wholesalers ",
"42322":"Home Furnishing Merchant Wholesalers ",
"423220":"Home Furnishing Merchant Wholesalers ",
"4233":"Lumber and Other Construction Materials Merchant Wholesalers ",
"42331":"Lumber, Plywood, Millwork, and Wood Panel Merchant Wholesalers ",
"423310":"Lumber, Plywood, Millwork, and Wood Panel Merchant Wholesalers ",
"42332":"Brick, Stone, and Related Construction Material Merchant Wholesalers ",
"423320":"Brick, Stone, and Related Construction Material Merchant Wholesalers ",
"42333":"Roofing, Siding, and Insulation Material Merchant Wholesalers ",
"423330":"Roofing, Siding, and Insulation Material Merchant Wholesalers ",
"42339":"Other Construction Material Merchant Wholesalers ",
"423390":"Other Construction Material Merchant Wholesalers ",
"4234":"Professional and Commercial Equipment and Supplies Merchant Wholesalers ",
"42341":"Photographic Equipment and Supplies Merchant Wholesalers ",
"423410":"Photographic Equipment and Supplies Merchant Wholesalers ",
"42342":"Office Equipment Merchant Wholesalers ",
"423420":"Office Equipment Merchant Wholesalers ",
"42343":"Computer and Computer Peripheral Equipment and Software Merchant Wholesalers ",
"423430":"Computer and Computer Peripheral Equipment and Software Merchant Wholesalers ",
"42344":"Other Commercial Equipment Merchant Wholesalers ",
"423440":"Other Commercial Equipment Merchant Wholesalers ",
"42345":"Medical, Dental, and Hospital Equipment and Supplies Merchant Wholesalers ",
"423450":"Medical, Dental, and Hospital Equipment and Supplies Merchant Wholesalers ",
"42346":"Ophthalmic Goods Merchant Wholesalers ",
"423460":"Ophthalmic Goods Merchant Wholesalers ",
"42349":"Other Professional Equipment and Supplies Merchant Wholesalers ",
"423490":"Other Professional Equipment and Supplies Merchant Wholesalers ",
"4235":"Metal and Mineral (except Petroleum) Merchant Wholesalers ",
"42351":"Metal Service Centers and Other Metal Merchant Wholesalers ",
"423510":"Metal Service Centers and Other Metal Merchant Wholesalers ",
"42352":"Coal and Other Mineral and Ore Merchant Wholesalers ",
"423520":"Coal and Other Mineral and Ore Merchant Wholesalers ",
"4236":"Household Appliances and Electrical and Electronic Goods Merchant Wholesalers ",
"42361":"Electrical Apparatus and Equipment, Wiring Supplies, and Related Equipment Merchant Wholesalers ",
"423610":"Electrical Apparatus and Equipment, Wiring Supplies, and Related Equipment Merchant Wholesalers ",
"42362":"Household Appliances, Electric Housewares, and Consumer Electronics Merchant Wholesalers ",
"423620":"Household Appliances, Electric Housewares, and Consumer Electronics Merchant Wholesalers ",
"42369":"Other Electronic Parts and Equipment Merchant Wholesalers ",
"423690":"Other Electronic Parts and Equipment Merchant Wholesalers ",
"4237":"Hardware, and Plumbing and Heating Equipment and Supplies Merchant Wholesalers ",
"42371":"Hardware Merchant Wholesalers ",
"423710":"Hardware Merchant Wholesalers ",
"42372":"Plumbing and Heating Equipment and Supplies (Hydronics) Merchant Wholesalers ",
"423720":"Plumbing and Heating Equipment and Supplies (Hydronics) Merchant Wholesalers ",
"42373":"Warm Air Heating and Air-Conditioning Equipment and Supplies Merchant Wholesalers ",
"423730":"Warm Air Heating and Air-Conditioning Equipment and Supplies Merchant Wholesalers ",
"42374":"Refrigeration Equipment and Supplies Merchant Wholesalers ",
"423740":"Refrigeration Equipment and Supplies Merchant Wholesalers ",
"4238":"Machinery, Equipment, and Supplies Merchant Wholesalers ",
"42381":"Construction and Mining (except Oil Well) Machinery and Equipment Merchant Wholesalers ",
"423810":"Construction and Mining (except Oil Well) Machinery and Equipment Merchant Wholesalers ",
"42382":"Farm and Garden Machinery and Equipment Merchant Wholesalers ",
"423820":"Farm and Garden Machinery and Equipment Merchant Wholesalers ",
"42383":"Industrial Machinery and Equipment Merchant Wholesalers ",
"423830":"Industrial Machinery and Equipment Merchant Wholesalers ",
"42384":"Industrial Supplies Merchant Wholesalers ",
"423840":"Industrial Supplies Merchant Wholesalers",
"42385":"Service Establishment Equipment and Supplies Merchant Wholesalers ",
"423850":"Service Establishment Equipment and Supplies Merchant Wholesalers ",
"42386":"Transportation Equipment and Supplies (except Motor Vehicle) Merchant Wholesalers ",
"423860":"Transportation Equipment and Supplies (except Motor Vehicle) Merchant Wholesalers ",
"4239":"Miscellaneous Durable Goods Merchant Wholesalers ",
"42391":"Sporting and Recreational Goods and Supplies Merchant Wholesalers",
"423910":"Sporting and Recreational Goods and Supplies Merchant Wholesalers ",
"42392":"Toy and Hobby Goods and Supplies Merchant Wholesalers ",
"423920":"Toy and Hobby Goods and Supplies Merchant Wholesalers ",
"42393":"Recyclable Material Merchant Wholesalers ",
"423930":"Recyclable Material Merchant Wholesalers ",
"42394":"Jewelry, Watch, Precious Stone, and Precious Metal Merchant Wholesalers ",
"423940":"Jewelry, Watch, Precious Stone, and Precious Metal Merchant Wholesalers ",
"42399":"Other Miscellaneous Durable Goods Merchant Wholesalers ",
"423990":"Other Miscellaneous Durable Goods Merchant Wholesalers ",
"424":"Merchant Wholesalers, Nondurable Goods ",
"4241":"Paper and Paper Product Merchant Wholesalers ",
"42411":"Printing and Writing Paper Merchant Wholesalers ",
"424110":"Printing and Writing Paper Merchant Wholesalers ",
"42412":"Stationery and Office Supplies Merchant Wholesalers ",
"424120":"Stationery and Office Supplies Merchant Wholesalers ",
"42413":"Industrial and Personal Service Paper Merchant Wholesalers ",
"424130":"Industrial and Personal Service Paper Merchant Wholesalers ",
"4242":"Drugs and Druggists' Sundries Merchant Wholesalers ",
"42421":"Drugs and Druggists' Sundries Merchant Wholesalers ",
"424210":"Drugs and Druggists' Sundries Merchant Wholesalers ",
"4243":"Apparel, Piece Goods, and Notions Merchant Wholesalers ",
"42431":"Piece Goods, Notions, and Other Dry Goods Merchant Wholesalers ",
"424310":"Piece Goods, Notions, and Other Dry Goods Merchant Wholesalers ",
"42432":"Men's and Boys' Clothing and Furnishings Merchant Wholesalers ",
"424320":"Men's and Boys' Clothing and Furnishings Merchant Wholesalers ",
"42433":"Women's, Children's, and Infants' Clothing and Accessories Merchant Wholesalers ",
"424330":"Women's, Children's, and Infants' Clothing and Accessories Merchant Wholesalers ",
"42434":"Footwear Merchant Wholesalers ",
"424340":"Footwear Merchant Wholesalers ",
"4244":"Grocery and Related Product Merchant Wholesalers ",
"42441":"General Line Grocery Merchant Wholesalers ",
"424410":"General Line Grocery Merchant Wholesalers ",
"42442":"Packaged Frozen Food Merchant Wholesalers ",
"424420":"Packaged Frozen Food Merchant Wholesalers ",
"42443":"Dairy Product (except Dried or Canned) Merchant Wholesalers ",
"424430":"Dairy Product (except Dried or Canned) Merchant Wholesalers ",
"42444":"Poultry and Poultry Product Merchant Wholesalers ",
"424440":"Poultry and Poultry Product Merchant Wholesalers ",
"42445":"Confectionery Merchant Wholesalers ",
"424450":"Confectionery Merchant Wholesalers ",
"42446":"Fish and Seafood Merchant Wholesalers ",
"424460":"Fish and Seafood Merchant Wholesalers ",
"42447":"Meat and Meat Product Merchant Wholesalers ",
"424470":"Meat and Meat Product Merchant Wholesalers ",
"42448":"Fresh Fruit and Vegetable Merchant Wholesalers ",
"424480":"Fresh Fruit and Vegetable Merchant Wholesalers ",
"42449":"Other Grocery and Related Products Merchant Wholesalers ",
"424490":"Other Grocery and Related Products Merchant Wholesalers ",
"4245":"Farm Product Raw Material Merchant Wholesalers ",
"42451":"Grain and Field Bean Merchant Wholesalers ",
"424510":"Grain and Field Bean Merchant Wholesalers ",
"42452":"Livestock Merchant Wholesalers ",
"424520":"Livestock Merchant Wholesalers ",
"42459":"Other Farm Product Raw Material Merchant Wholesalers ",
"424590":"Other Farm Product Raw Material Merchant Wholesalers ",
"4246":"Chemical and Allied Products Merchant Wholesalers ",
"42461":"Plastics Materials and Basic Forms and Shapes Merchant Wholesalers ",
"424610":"Plastics Materials and Basic Forms and Shapes Merchant Wholesalers ",
"42469":"Other Chemical and Allied Products Merchant Wholesalers ",
"424690":"Other Chemical and Allied Products Merchant Wholesalers ",
"4247":"Petroleum and Petroleum Products Merchant Wholesalers ",
"42471":"Petroleum Bulk Stations and Terminals ",
"424710":"Petroleum Bulk Stations and Terminals ",
"42472":"Petroleum and Petroleum Products Merchant Wholesalers (except Bulk Stations and Terminals) ",
"424720":"Petroleum and Petroleum Products Merchant Wholesalers (except Bulk Stations and Terminals) ",
"4248":"Beer, Wine, and Distilled Alcoholic Beverage Merchant Wholesalers ",
"42481":"Beer and Ale Merchant Wholesalers ",
"424810":"Beer and Ale Merchant Wholesalers ",
"42482":"Wine and Distilled Alcoholic Beverage Merchant Wholesalers ",
"424820":"Wine and Distilled Alcoholic Beverage Merchant Wholesalers ",
"4249":"Miscellaneous Nondurable Goods Merchant Wholesalers ",
"42491":"Farm Supplies Merchant Wholesalers ",
"424910":"Farm Supplies Merchant Wholesalers ",
"42492":"Book, Periodical, and Newspaper Merchant Wholesalers ",
"424920":"Book, Periodical, and Newspaper Merchant Wholesalers ",
"42493":"Flower, Nursery Stock, and Florists' Supplies Merchant Wholesalers ",
"424930":"Flower, Nursery Stock, and Florists' Supplies Merchant Wholesalers ",
"42494":"Tobacco and Tobacco Product Merchant Wholesalers ",
"424940":"Tobacco and Tobacco Product Merchant Wholesalers ",
"42495":"Paint, Varnish, and Supplies Merchant Wholesalers ",
"424950":"Paint, Varnish, and Supplies Merchant Wholesalers ",
"42499":"Other Miscellaneous Nondurable Goods Merchant Wholesalers ",
"424990":"Other Miscellaneous Nondurable Goods Merchant Wholesalers ",
"425":"Wholesale Electronic Markets and Agents and Brokers ",
"4251":"Wholesale Electronic Markets and Agents and Brokers ",
"42511":"Business to Business Electronic Markets ",
"425110":"Business to Business Electronic Markets ",
"42512":"Wholesale Trade Agents and Brokers ",
"425120":"Wholesale Trade Agents and Brokers ",
"44":"Retail Trade",
"45":"Retail Trade",
"441":"Motor Vehicle and Parts Dealers ",
"4411":"Automobile Dealers ",
"44111":"New Car Dealers ",
"441110":"New Car Dealers ",
"44112":"Used Car Dealers ",
"441120":"Used Car Dealers ",
"4412":"Other Motor Vehicle Dealers ",
"44121":"Recreational Vehicle Dealers ",
"441210":"Recreational Vehicle Dealers ",
"44122":"Motorcycle, Boat, and Other Motor Vehicle Dealers ",
"441222":"Boat Dealers ",
"441228":"Motorcycle, ATV, and All Other Motor Vehicle Dealers ",
"4413":"Automotive Parts, Accessories, and Tire Stores ",
"44131":"Automotive Parts and Accessories Stores ",
"441310":"Automotive Parts and Accessories Stores ",
"44132":"Tire Dealers ",
"441320":"Tire Dealers ",
"442":"Furniture and Home Furnishings Stores ",
"4421":"Furniture Stores ",
"44211":"Furniture Stores ",
"442110":"Furniture Stores ",
"4422":"Home Furnishings Stores ",
"44221":"Floor Covering Stores ",
"442210":"Floor Covering Stores ",
"44229":"Other Home Furnishings Stores ",
"442291":"Window Treatment Stores ",
"442299":"All Other Home Furnishings Stores ",
"443":"Electronics and Appliance Stores ",
"4431":"Electronics and Appliance Stores ",
"44314":"Electronics and Appliance Stores ",
"443141":"Household Appliance Stores ",
"443142":"Electronics Stores ",
"444":"Building Material and Garden Equipment and Supplies Dealers ",
"4441":"Building Material and Supplies Dealers ",
"44411":"Home Centers ",
"444110":"Home Centers ",
"44412":"Paint and Wallpaper Stores ",
"444120":"Paint and Wallpaper Stores ",
"44413":"Hardware Stores ",
"444130":"Hardware Stores ",
"44419":"Other Building Material Dealers ",
"444190":"Other Building Material Dealers ",
"4442":"Lawn and Garden Equipment and Supplies Stores ",
"44421":"Outdoor Power Equipment Stores ",
"444210":"Outdoor Power Equipment Stores ",
"44422":"Nursery, Garden Center, and Farm Supply Stores ",
"444220":"Nursery, Garden Center, and Farm Supply Stores ",
"445":"Food and Beverage Stores ",
"4451":"Grocery Stores ",
"44511":"Supermarkets and Other Grocery (except Convenience) Stores ",
"445110":"Supermarkets and Other Grocery (except Convenience) Stores ",
"44512":"Convenience Stores ",
"445120":"Convenience Stores ",
"4452":"Specialty Food Stores ",
"44521":"Meat Markets ",
"445210":"Meat Markets ",
"44522":"Fish and Seafood Markets ",
"445220":"Fish and Seafood Markets ",
"44523":"Fruit and Vegetable Markets ",
"445230":"Fruit and Vegetable Markets ",
"44529":"Other Specialty Food Stores ",
"445291":"Baked Goods Stores ",
"445292":"Confectionery and Nut Stores ",
"445299":"All Other Specialty Food Stores ",
"4453":"Beer, Wine, and Liquor Stores ",
"44531":"Beer, Wine, and Liquor Stores ",
"445310":"Beer, Wine, and Liquor Stores ",
"446":"Health and Personal Care Stores ",
"4461":"Health and Personal Care Stores ",
"44611":"Pharmacies and Drug Stores ",
"446110":"Pharmacies and Drug Stores ",
"44612":"Cosmetics, Beauty Supplies, and Perfume Stores ",
"446120":"Cosmetics, Beauty Supplies, and Perfume Stores ",
"44613":"Optical Goods Stores ",
"446130":"Optical Goods Stores ",
"44619":"Other Health and Personal Care Stores ",
"446191":"Food (Health) Supplement Stores ",
"446199":"All Other Health and Personal Care Stores ",
"447":"Gasoline Stations ",
"4471":"Gasoline Stations ",
"44711":"Gasoline Stations with Convenience Stores ",
"447110":"Gasoline Stations with Convenience Stores ",
"44719":"Other Gasoline Stations ",
"447190":"Other Gasoline Stations ",
"448":"Clothing and Clothing Accessories Stores ",
"4481":"Clothing Stores ",
"44811":"Men's Clothing Stores ",
"448110":"Men's Clothing Stores ",
"44812":"Women's Clothing Stores ",
"448120":"Women's Clothing Stores ",
"44813":"Children's and Infants' Clothing Stores ",
"448130":"Children's and Infants' Clothing Stores ",
"44814":"Family Clothing Stores ",
"448140":"Family Clothing Stores ",
"44815":"Clothing Accessories Stores ",
"448150":"Clothing Accessories Stores ",
"44819":"Other Clothing Stores ",
"448190":"Other Clothing Stores ",
"4482":"Shoe Stores ",
"44821":"Shoe Stores ",
"448210":"Shoe Stores ",
"4483":"Jewelry, Luggage, and Leather Goods Stores ",
"44831":"Jewelry Stores ",
"448310":"Jewelry Stores ",
"44832":"Luggage and Leather Goods Stores ",
"448320":"Luggage and Leather Goods Stores ",
"451":"Sporting Goods, Hobby, Musical Instrument, and Book Stores ",
"4511":"Sporting Goods, Hobby, and Musical Instrument Stores ",
"45111":"Sporting Goods Stores ",
"451110":"Sporting Goods Stores ",
"45112":"Hobby, Toy, and Game Stores ",
"451120":"Hobby, Toy, and Game Stores ",
"45113":"Sewing, Needlework, and Piece Goods Stores ",
"451130":"Sewing, Needlework, and Piece Goods Stores ",
"45114":"Musical Instrument and Supplies Stores ",
"451140":"Musical Instrument and Supplies Stores ",
"4512":"Book Stores and News Dealers ",
"45121":"Book Stores and News Dealers ",
"451211":"Book Stores ",
"451212":"News Dealers and Newsstands ",
"452":"General Merchandise Stores ",
"4522":"Department Stores ",
"45221":"Department Stores ",
"452210":"Department Stores ",
"4523":"General Merchandise Stores, including Warehouse Clubs and Supercenters ",
"45231":"General Merchandise Stores, including Warehouse Clubs and Supercenters ",
"452311":"Warehouse Clubs and Supercenters ",
"452319":"All Other General Merchandise Stores ",
"453":"Miscellaneous Store Retailers ",
"4531":"Florists ",
"45311":"Florists ",
"453110":"Florists ",
"4532":"Office Supplies, Stationery, and Gift Stores ",
"45321":"Office Supplies and Stationery Stores ",
"453210":"Office Supplies and Stationery Stores ",
"45322":"Gift, Novelty, and Souvenir Stores ",
"453220":"Gift, Novelty, and Souvenir Stores ",
"4533":"Used Merchandise Stores ",
"45331":"Used Merchandise Stores ",
"453310":"Used Merchandise Stores ",
"4539":"Other Miscellaneous Store Retailers ",
"45391":"Pet and Pet Supplies Stores ",
"453910":"Pet and Pet Supplies Stores ",
"45392":"Art Dealers ",
"453920":"Art Dealers ",
"45393":"Manufactured (Mobile) Home Dealers ",
"453930":"Manufactured (Mobile) Home Dealers ",
"45399":"All Other Miscellaneous Store Retailers ",
"453991":"Tobacco Stores ",
"453998":"All Other Miscellaneous Store Retailers (except Tobacco Stores) ",
"454":"Nonstore Retailers ",
"4541":"Electronic Shopping and Mail-Order Houses ",
"45411":"Electronic Shopping and Mail-Order Houses ",
"454110":"Electronic Shopping and Mail-Order Houses ",
"4542":"Vending Machine Operators ",
"45421":"Vending Machine Operators ",
"454210":"Vending Machine Operators ",
"4543":"Direct Selling Establishments ",
"45431":"Fuel Dealers ",
"454310":"Fuel Dealers ",
"45439":"Other Direct Selling Establishments ",
"454390":"Other Direct Selling Establishments ",
"48":"Transportation and Warehousing",
"49":"Transportation and Warehousing",
"481":"Air Transportation",
"4811":"Scheduled Air Transportation",
"48111":"Scheduled Air Transportation",
"481111":"Scheduled Passenger Air Transportation ",
"481112":"Scheduled Freight Air Transportation ",
"4812":"Nonscheduled Air Transportation",
"48121":"Nonscheduled Air Transportation",
"481211":"Nonscheduled Chartered Passenger Air Transportation ",
"481212":"Nonscheduled Chartered Freight Air Transportation ",
"481219":"Other Nonscheduled Air Transportation ",
"482":"Rail Transportation",
"4821":"Rail Transportation",
"48211":"Rail Transportation",
"482111":"Line-Haul Railroads ",
"482112":"Short Line Railroads ",
"483":"Water Transportation",
"4831":"Deep Sea, Coastal, and Great Lakes Water Transportation",
"48311":"Deep Sea, Coastal, and Great Lakes Water Transportation",
"483111":"Deep Sea Freight Transportation ",
"483112":"Deep Sea Passenger Transportation ",
"483113":"Coastal and Great Lakes Freight Transportation ",
"483114":"Coastal and Great Lakes Passenger Transportation ",
"4832":"Inland Water Transportation",
"48321":"Inland Water Transportation",
"483211":"Inland Water Freight Transportation ",
"483212":"Inland Water Passenger Transportation ",
"484":"Truck Transportation",
"4841":"General Freight Trucking",
"48411":"General Freight Trucking, Local",
"484110":"General Freight Trucking, Local ",
"48412":"General Freight Trucking, Long-Distance",
"484121":"General Freight Trucking, Long-Distance, Truckload ",
"484122":"General Freight Trucking, Long-Distance, Less Than Truckload ",
"4842":"Specialized Freight Trucking",
"48421":"Used Household and Office Goods Moving",
"484210":"Used Household and Office Goods Moving",
"48422":"Specialized Freight (except Used Goods) Trucking, Local",
"484220":"Specialized Freight (except Used Goods) Trucking, Local ",
"48423":"Specialized Freight (except Used Goods) Trucking, Long-Distance",
"484230":"Specialized Freight (except Used Goods) Trucking, Long-Distance ",
"485":"Transit and Ground Passenger Transportation",
"4851":"Urban Transit Systems",
"48511":"Urban Transit Systems",
"485111":"Mixed Mode Transit Systems ",
"485112":"Commuter Rail Systems ",
"485113":"Bus and Other Motor Vehicle Transit Systems ",
"485119":"Other Urban Transit Systems ",
"4852":"Interurban and Rural Bus Transportation",
"48521":"Interurban and Rural Bus Transportation",
"485210":"Interurban and Rural Bus Transportation",
"4853":"Taxi and Limousine Service",
"48531":"Taxi Service",
"485310":"Taxi Service ",
"48532":"Limousine Service",
"485320":"Limousine Service",
"4854":"School and Employee Bus Transportation",
"48541":"School and Employee Bus Transportation",
"485410":"School and Employee Bus Transportation",
"4855":"Charter Bus Industry",
"48551":"Charter Bus Industry",
"485510":"Charter Bus Industry",
"4859":"Other Transit and Ground Passenger Transportation",
"48599":"Other Transit and Ground Passenger Transportation",
"485991":"Special Needs Transportation ",
"485999":"All Other Transit and Ground Passenger Transportation ",
"486":"Pipeline Transportation",
"4861":"Pipeline Transportation of Crude Oil",
"48611":"Pipeline Transportation of Crude Oil",
"486110":"Pipeline Transportation of Crude Oil",
"4862":"Pipeline Transportation of Natural Gas",
"48621":"Pipeline Transportation of Natural Gas",
"486210":"Pipeline Transportation of Natural Gas",
"4869":"Other Pipeline Transportation",
"48691":"Pipeline Transportation of Refined Petroleum Products",
"486910":"Pipeline Transportation of Refined Petroleum Products",
"48699":"All Other Pipeline Transportation",
"486990":"All Other Pipeline Transportation",
"487":"Scenic and Sightseeing Transportation",
"4871":"Scenic and Sightseeing Transportation, Land",
"48711":"Scenic and Sightseeing Transportation, Land",
"487110":"Scenic and Sightseeing Transportation, Land",
"4872":"Scenic and Sightseeing Transportation, Water",
"48721":"Scenic and Sightseeing Transportation, Water",
"487210":"Scenic and Sightseeing Transportation, Water",
"4879":"Scenic and Sightseeing Transportation, Other",
"48799":"Scenic and Sightseeing Transportation, Other",
"487990":"Scenic and Sightseeing Transportation, Other",
"488":"Support Activities for Transportation",
"4881":"Support Activities for Air Transportation",
"48811":"Airport Operations",
"488111":"Air Traffic Control",
"488119":"Other Airport Operations ",
"48819":"Other Support Activities for Air Transportation",
"488190":"Other Support Activities for Air Transportation",
"4882":"Support Activities for Rail Transportation",
"48821":"Support Activities for Rail Transportation",
"488210":"Support Activities for Rail Transportation",
"4883":"Support Activities for Water Transportation",
"48831":"Port and Harbor Operations",
"488310":"Port and Harbor Operations",
"48832":"Marine Cargo Handling",
"488320":"Marine Cargo Handling",
"48833":"Navigational Services to Shipping",
"488330":"Navigational Services to Shipping ",
"48839":"Other Support Activities for Water Transportation",
"488390":"Other Support Activities for Water Transportation",
"4884":"Support Activities for Road Transportation",
"48841":"Motor Vehicle Towing",
"488410":"Motor Vehicle Towing",
"48849":"Other Support Activities for Road Transportation",
"488490":"Other Support Activities for Road Transportation ",
"4885":"Freight Transportation Arrangement",
"48851":"Freight Transportation Arrangement",
"488510":"Freight Transportation Arrangement ",
"4889":"Other Support Activities for Transportation",
"48899":"Other Support Activities for Transportation",
"488991":"Packing and Crating ",
"488999":"All Other Support Activities for Transportation ",
"491":"Postal Service",
"4911":"Postal Service",
"49111":"Postal Service",
"491110":"Postal Service",
"492":"Couriers and Messengers",
"4921":"Couriers and Express Delivery Services",
"49211":"Couriers and Express Delivery Services",
"492110":"Couriers and Express Delivery Services",
"4922":"Local Messengers and Local Delivery",
"49221":"Local Messengers and Local Delivery",
"492210":"Local Messengers and Local Delivery",
"493":"Warehousing and Storage",
"4931":"Warehousing and Storage",
"49311":"General Warehousing and Storage",
"493110":"General Warehousing and Storage ",
"49312":"Refrigerated Warehousing and Storage",
"493120":"Refrigerated Warehousing and Storage",
"49313":"Farm Product Warehousing and Storage",
"493130":"Farm Product Warehousing and Storage",
"49319":"Other Warehousing and Storage",
"493190":"Other Warehousing and Storage",
"51":"Information",
"511":"Publishing Industries (except Internet)",
"5111":"Newspaper, Periodical, Book, and Directory Publishers",
"51111":"Newspaper Publishers",
"511110":"Newspaper Publishers ",
"51112":"Periodical Publishers",
"511120":"Periodical Publishers ",
"51113":"Book Publishers",
"511130":"Book Publishers ",
"51114":"Directory and Mailing List Publishers",
"511140":"Directory and Mailing List Publishers ",
"51119":"Other Publishers",
"511191":"Greeting Card Publishers ",
"511199":"All Other Publishers ",
"5112":"Software Publishers",
"51121":"Software Publishers",
"511210":"Software Publishers",
"512":"Motion Picture and Sound Recording Industries",
"5121":"Motion Picture and Video Industries",
"51211":"Motion Picture and Video Production",
"512110":"Motion Picture and Video Production ",
"51212":"Motion Picture and Video Distribution",
"512120":"Motion Picture and Video Distribution",
"51213":"Motion Picture and Video Exhibition",
"512131":"Motion Picture Theaters (except Drive-Ins) ",
"512132":"Drive-In Motion Picture Theaters ",
"51219":"Postproduction Services and Other Motion Picture and Video Industries",
"512191":"Teleproduction and Other Postproduction Services ",
"512199":"Other Motion Picture and Video Industries ",
"5122":"Sound Recording Industries",
"51223":"Music Publishers",
"512230":"Music Publishers",
"51224":"Sound Recording Studios",
"512240":"Sound Recording Studios",
"51225":"Record Production and Distribution",
"512250":"Record Production and Distribution",
"51229":"Other Sound Recording Industries",
"512290":"Other Sound Recording Industries",
"515":"Broadcasting (except Internet)",
"5151":"Radio and Television Broadcasting",
"51511":"Radio Broadcasting",
"515111":"Radio Networks ",
"515112":"Radio Stations ",
"51512":"Television Broadcasting",
"515120":"Television Broadcasting",
"5152":"Cable and Other Subscription Programming",
"51521":"Cable and Other Subscription Programming",
"515210":"Cable and Other Subscription Programming",
"517":"Telecommunications",
"5173":"Wired and Wireless Telecommunications Carriers",
"51731":"Wired and Wireless Telecommunications Carriers",
"517311":"Wired Telecommunications Carriers ",
"517312":"Wireless Telecommunications Carriers (except Satellite)",
"5174":"Satellite Telecommunications",
"51741":"Satellite Telecommunications",
"517410":"Satellite Telecommunications",
"5179":"Other Telecommunications",
"51791":"Other Telecommunications",
"517911":"Telecommunications Resellers ",
"517919":"All Other Telecommunications ",
"518":"Data Processing, Hosting, and Related Services",
"5182":"Data Processing, Hosting, and Related Services",
"51821":"Data Processing, Hosting, and Related Services",
"518210":"Data Processing, Hosting, and Related Services",
"519":"Other Information Services",
"5191":"Other Information Services",
"51911":"News Syndicates",
"519110":"News Syndicates",
"51912":"Libraries and Archives",
"519120":"Libraries and Archives ",
"51913":"Internet Publishing and Broadcasting and Web Search Portals",
"519130":"Internet Publishing and Broadcasting and Web Search Portals",
"51919":"All Other Information Services",
"519190":"All Other Information Services",
"52":"Finance and Insurance",
"521":"Monetary Authorities-Central Bank",
"5211":"Monetary Authorities-Central Bank",
"52111":"Monetary Authorities-Central Bank",
"521110":"Monetary Authorities-Central Bank",
"522":"Credit Intermediation and Related Activities",
"5221":"Depository Credit Intermediation ",
"52211":"Commercial Banking ",
"522110":"Commercial Banking ",
"52212":"Savings Institutions ",
"522120":"Savings Institutions ",
"52213":"Credit Unions ",
"522130":"Credit Unions ",
"52219":"Other Depository Credit Intermediation ",
"522190":"Other Depository Credit Intermediation ",
"5222":"Nondepository Credit Intermediation ",
"52221":"Credit Card Issuing ",
"522210":"Credit Card Issuing ",
"52222":"Sales Financing ",
"522220":"Sales Financing ",
"52229":"Other Nondepository Credit Intermediation ",
"522291":"Consumer Lending ",
"522292":"Real Estate Credit ",
"522293":"International Trade Financing ",
"522294":"Secondary Market Financing ",
"522298":"All Other Nondepository Credit Intermediation ",
"5223":"Activities Related to Credit Intermediation ",
"52231":"Mortgage and Nonmortgage Loan Brokers ",
"522310":"Mortgage and Nonmortgage Loan Brokers ",
"52232":"Financial Transactions Processing, Reserve, and Clearinghouse Activities ",
"522320":"Financial Transactions Processing, Reserve, and Clearinghouse Activities ",
"52239":"Other Activities Related to Credit Intermediation ",
"522390":"Other Activities Related to Credit Intermediation ",
"523":"Securities, Commodity Contracts, and Other Financial Investments and Related Activities",
"5231":"Securities and Commodity Contracts Intermediation and Brokerage",
"52311":"Investment Banking and Securities Dealing ",
"523110":"Investment Banking and Securities Dealing ",
"52312":"Securities Brokerage ",
"523120":"Securities Brokerage ",
"52313":"Commodity Contracts Dealing ",
"523130":"Commodity Contracts Dealing ",
"52314":"Commodity Contracts Brokerage ",
"523140":"Commodity Contracts Brokerage ",
"5232":"Securities and Commodity Exchanges",
"52321":"Securities and Commodity Exchanges",
"523210":"Securities and Commodity Exchanges",
"5239":"Other Financial Investment Activities",
"52391":"Miscellaneous Intermediation ",
"523910":"Miscellaneous Intermediation ",
"52392":"Portfolio Management ",
"523920":"Portfolio Management ",
"52393":"Investment Advice ",
"523930":"Investment Advice ",
"52399":"All Other Financial Investment Activities ",
"523991":"Trust, Fiduciary, and Custody Activities ",
"523999":"Miscellaneous Financial Investment Activities ",
"524":"Insurance Carriers and Related Activities",
"5241":"Insurance Carriers",
"52411":"Direct Life, Health, and Medical Insurance Carriers ",
"524113":"Direct Life Insurance Carriers ",
"524114":"Direct Health and Medical Insurance Carriers ",
"52412":"Direct Insurance (except Life, Health, and Medical) Carriers ",
"524126":"Direct Property and Casualty Insurance Carriers ",
"524127":"Direct Title Insurance Carriers ",
"524128":"Other Direct Insurance (except Life, Health, and Medical) Carriers ",
"52413":"Reinsurance Carriers ",
"524130":"Reinsurance Carriers ",
"5242":"Agencies, Brokerages, and Other Insurance Related Activities",
"52421":"Insurance Agencies and Brokerages ",
"524210":"Insurance Agencies and Brokerages ",
"52429":"Other Insurance Related Activities ",
"524291":"Claims Adjusting ",
"524292":"Third Party Administration of Insurance and Pension Funds ",
"524298":"All Other Insurance Related Activities ",
"525":"Funds, Trusts, and Other Financial Vehicles ",
"5251":"Insurance and Employee Benefit Funds ",
"52511":"Pension Funds ",
"525110":"Pension Funds ",
"52512":"Health and Welfare Funds ",
"525120":"Health and Welfare Funds ",
"52519":"Other Insurance Funds ",
"525190":"Other Insurance Funds ",
"5259":"Other Investment Pools and Funds",
"52591":"Open-End Investment Funds ",
"525910":"Open-End Investment Funds ",
"52592":"Trusts, Estates, and Agency Accounts ",
"525920":"Trusts, Estates, and Agency Accounts ",
"52599":"Other Financial Vehicles ",
"525990":"Other Financial Vehicles ",
"53":"Real Estate and Rental and Leasing",
"531":"Real Estate",
"5311":"Lessors of Real Estate",
"53111":"Lessors of Residential Buildings and Dwellings ",
"531110":"Lessors of Residential Buildings and Dwellings ",
"53112":"Lessors of Nonresidential Buildings (except Miniwarehouses) ",
"531120":"Lessors of Nonresidential Buildings (except Miniwarehouses) ",
"53113":"Lessors of Miniwarehouses and Self-Storage Units ",
"531130":"Lessors of Miniwarehouses and Self-Storage Units ",
"53119":"Lessors of Other Real Estate Property ",
"531190":"Lessors of Other Real Estate Property ",
"5312":"Offices of Real Estate Agents and Brokers",
"53121":"Offices of Real Estate Agents and Brokers",
"531210":"Offices of Real Estate Agents and Brokers",
"5313":"Activities Related to Real Estate",
"53131":"Real Estate Property Managers ",
"531311":"Residential Property Managers ",
"531312":"Nonresidential Property Managers ",
"53132":"Offices of Real Estate Appraisers ",
"531320":"Offices of Real Estate Appraisers ",
"53139":"Other Activities Related to Real Estate ",
"531390":"Other Activities Related to Real Estate ",
"532":"Rental and Leasing Services",
"5321":"Automotive Equipment Rental and Leasing",
"53211":"Passenger Car Rental and Leasing",
"532111":"Passenger Car Rental ",
"532112":"Passenger Car Leasing ",
"53212":"Truck, Utility Trailer, and RV (Recreational Vehicle) Rental and Leasing",
"532120":"Truck, Utility Trailer, and RV (Recreational Vehicle) Rental and Leasing ",
"5322":"Consumer Goods Rental",
"53221":"Consumer Electronics and Appliances Rental",
"532210":"Consumer Electronics and Appliances Rental",
"53228":"Other Consumer Goods Rental ",
"532281":"Formal Wear and Costume Rental",
"532282":"Video Tape and Disc Rental",
"532283":"Home Health Equipment Rental ",
"532284":"Recreational Goods Rental ",
"532289":"All Other Consumer Goods Rental ",
"5323":"General Rental Centers",
"53231":"General Rental Centers",
"532310":"General Rental Centers",
"5324":"Commercial and Industrial Machinery and Equipment Rental and Leasing",
"53241":"Construction, Transportation, Mining, and Forestry Machinery and Equipment Rental and Leasing",
"532411":"Commercial Air, Rail, and Water Transportation Equipment Rental and Leasing ",
"532412":"Construction, Mining, and Forestry Machinery and Equipment Rental and Leasing ",
"53242":"Office Machinery and Equipment Rental and Leasing",
"532420":"Office Machinery and Equipment Rental and Leasing",
"53249":"Other Commercial and Industrial Machinery and Equipment Rental and Leasing",
"532490":"Other Commercial and Industrial Machinery and Equipment Rental and Leasing ",
"533":"Lessors of Nonfinancial Intangible Assets (except Copyrighted Works)",
"5331":"Lessors of Nonfinancial Intangible Assets (except Copyrighted Works)",
"53311":"Lessors of Nonfinancial Intangible Assets (except Copyrighted Works)",
"533110":"Lessors of Nonfinancial Intangible Assets (except Copyrighted Works)",
"54":"Professional, Scientific, and Technical Services",
"541":"Professional, Scientific, and Technical Services",
"5411":"Legal Services",
"54111":"Offices of Lawyers",
"541110":"Offices of Lawyers",
"54112":"Offices of Notaries",
"541120":"Offices of Notaries",
"54119":"Other Legal Services",
"541191":"Title Abstract and Settlement Offices ",
"541199":"All Other Legal Services ",
"5412":"Accounting, Tax Preparation, Bookkeeping, and Payroll Services",
"54121":"Accounting, Tax Preparation, Bookkeeping, and Payroll Services",
"541211":"Offices of Certified Public Accountants ",
"541213":"Tax Preparation Services ",
"541214":"Payroll Services ",
"541219":"Other Accounting Services ",
"5413":"Architectural, Engineering, and Related Services",
"54131":"Architectural Services",
"541310":"Architectural Services",
"54132":"Landscape Architectural Services",
"541320":"Landscape Architectural Services",
"54133":"Engineering Services",
"541330":"Engineering Services",
"54134":"Drafting Services",
"541340":"Drafting Services",
"54135":"Building Inspection Services",
"541350":"Building Inspection Services",
"54136":"Geophysical Surveying and Mapping Services",
"541360":"Geophysical Surveying and Mapping Services",
"54137":"Surveying and Mapping (except Geophysical) Services",
"541370":"Surveying and Mapping (except Geophysical) Services",
"54138":"Testing Laboratories",
"541380":"Testing Laboratories",
"5414":"Specialized Design Services",
"54141":"Interior Design Services",
"541410":"Interior Design Services",
"54142":"Industrial Design Services",
"541420":"Industrial Design Services",
"54143":"Graphic Design Services",
"541430":"Graphic Design Services",
"54149":"Other Specialized Design Services",
"541490":"Other Specialized Design Services",
"5415":"Computer Systems Design and Related Services",
"54151":"Computer Systems Design and Related Services",
"541511":"Custom Computer Programming Services ",
"541512":"Computer Systems Design Services ",
"541513":"Computer Facilities Management Services ",
"541519":"Other Computer Related Services",
"5416":"Management, Scientific, and Technical Consulting Services",
"54161":"Management Consulting Services",
"541611":"Administrative Management and General Management Consulting Services ",
"541612":"Human Resources Consulting Services ",
"541613":"Marketing Consulting Services ",
"541614":"Process, Physical Distribution, and Logistics Consulting Services ",
"541618":"Other Management Consulting Services ",
"54162":"Environmental Consulting Services",
"541620":"Environmental Consulting Services",
"54169":"Other Scientific and Technical Consulting Services",
"541690":"Other Scientific and Technical Consulting Services",
"5417":"Scientific Research and Development Services",
"54171":"Research and Development in the Physical, Engineering, and Life Sciences",
"541713":"Research and Development in Nanotechnology ",
"541714":"Research and Development in Biotechnology (except Nanobiotechnology)",
"541715":"Research and Development in the Physical, Engineering, and Life Sciences (except Nanotechnology and Biotechnology) ",
"54172":"Research and Development in the Social Sciences and Humanities",
"541720":"Research and Development in the Social Sciences and Humanities ",
"5418":"Advertising, Public Relations, and Related Services",
"54181":"Advertising Agencies",
"541810":"Advertising Agencies",
"54182":"Public Relations Agencies",
"541820":"Public Relations Agencies",
"54183":"Media Buying Agencies",
"541830":"Media Buying Agencies",
"54184":"Media Representatives",
"541840":"Media Representatives",
"54185":"Outdoor Advertising",
"541850":"Outdoor Advertising",
"54186":"Direct Mail Advertising",
"541860":"Direct Mail Advertising",
"54187":"Advertising Material Distribution Services",
"541870":"Advertising Material Distribution Services",
"54189":"Other Services Related to Advertising",
"541890":"Other Services Related to Advertising ",
"5419":"Other Professional, Scientific, and Technical Services",
"54191":"Marketing Research and Public Opinion Polling",
"541910":"Marketing Research and Public Opinion Polling",
"54192":"Photographic Services",
"541921":"Photography Studios, Portrait ",
"541922":"Commercial Photography ",
"54193":"Translation and Interpretation Services",
"541930":"Translation and Interpretation Services",
"54194":"Veterinary Services",
"541940":"Veterinary Services ",
"54199":"All Other Professional, Scientific, and Technical Services",
"541990":"All Other Professional, Scientific, and Technical Services",
"55":"Management of Companies and Enterprises",
"551":"Management of Companies and Enterprises",
"5511":"Management of Companies and Enterprises",
"55111":"Management of Companies and Enterprises",
"551111":"Offices of Bank Holding Companies ",
"551112":"Offices of Other Holding Companies ",
"551114":"Corporate, Subsidiary, and Regional Managing Offices ",
"56":"Administrative and Support and Waste Management and Remediation Services",
"561":"Administrative and Support Services",
"5611":"Office Administrative Services",
"56111":"Office Administrative Services",
"561110":"Office Administrative Services",
"5612":"Facilities Support Services",
"56121":"Facilities Support Services",
"561210":"Facilities Support Services",
"5613":"Employment Services",
"56131":"Employment Placement Agencies and Executive Search Services",
"561311":"Employment Placement Agencies ",
"561312":"Executive Search Services ",
"56132":"Temporary Help Services",
"561320":"Temporary Help Services",
"56133":"Professional Employer Organizations",
"561330":"Professional Employer Organizations",
"5614":"Business Support Services",
"56141":"Document Preparation Services",
"561410":"Document Preparation Services",
"56142":"Telephone Call Centers",
"561421":"Telephone Answering Services ",
"561422":"Telemarketing Bureaus and Other Contact Centers ",
"56143":"Business Service Centers",
"561431":"Private Mail Centers ",
"561439":"Other Business Service Centers (including Copy Shops) ",
"56144":"Collection Agencies",
"561440":"Collection Agencies",
"56145":"Credit Bureaus",
"561450":"Credit Bureaus",
"56149":"Other Business Support Services",
"561491":"Repossession Services ",
"561492":"Court Reporting and Stenotype Services ",
"561499":"All Other Business Support Services ",
"5615":"Travel Arrangement and Reservation Services",
"56151":"Travel Agencies",
"561510":"Travel Agencies",
"56152":"Tour Operators",
"561520":"Tour Operators",
"56159":"Other Travel Arrangement and Reservation Services",
"561591":"Convention and Visitors Bureaus ",
"561599":"All Other Travel Arrangement and Reservation Services ",
"5616":"Investigation and Security Services",
"56161":"Investigation, Guard, and Armored Car Services",
"561611":"Investigation Services ",
"561612":"Security Guards and Patrol Services ",
"561613":"Armored Car Services ",
"56162":"Security Systems Services",
"561621":"Security Systems Services (except Locksmiths) ",
"561622":"Locksmiths ",
"5617":"Services to Buildings and Dwellings",
"56171":"Exterminating and Pest Control Services",
"561710":"Exterminating and Pest Control Services",
"56172":"Janitorial Services",
"561720":"Janitorial Services ",
"56173":"Landscaping Services",
"561730":"Landscaping Services",
"56174":"Carpet and Upholstery Cleaning Services",
"561740":"Carpet and Upholstery Cleaning Services",
"56179":"Other Services to Buildings and Dwellings",
"561790":"Other Services to Buildings and Dwellings ",
"5619":"Other Support Services",
"56191":"Packaging and Labeling Services",
"561910":"Packaging and Labeling Services",
"56192":"Convention and Trade Show Organizers",
"561920":"Convention and Trade Show Organizers",
"56199":"All Other Support Services",
"561990":"All Other Support Services",
"562":"Waste Management and Remediation Services",
"5621":"Waste Collection ",
"56211":"Waste Collection ",
"562111":"Solid Waste Collection ",
"562112":"Hazardous Waste Collection ",
"562119":"Other Waste Collection ",
"5622":"Waste Treatment and Disposal ",
"56221":"Waste Treatment and Disposal ",
"562211":"Hazardous Waste Treatment and Disposal ",
"562212":"Solid Waste Landfill ",
"562213":"Solid Waste Combustors and Incinerators ",
"562219":"Other Nonhazardous Waste Treatment and Disposal ",
"5629":"Remediation and Other Waste Management Services ",
"56291":"Remediation Services ",
"562910":"Remediation Services ",
"56292":"Materials Recovery Facilities ",
"562920":"Materials Recovery Facilities ",
"56299":"All Other Waste Management Services ",
"562991":"Septic Tank and Related Services ",
"562998":"All Other Miscellaneous Waste Management Services ",
"61":"Educational Services",
"611":"Educational Services",
"6111":"Elementary and Secondary Schools",
"61111":"Elementary and Secondary Schools ",
"611110":"Elementary and Secondary Schools ",
"6112":"Junior Colleges",
"61121":"Junior Colleges",
"611210":"Junior Colleges ",
"6113":"Colleges, Universities, and Professional Schools",
"61131":"Colleges, Universities, and Professional Schools",
"611310":"Colleges, Universities, and Professional Schools ",
"6114":"Business Schools and Computer and Management Training",
"61141":"Business and Secretarial Schools",
"611410":"Business and Secretarial Schools ",
"61142":"Computer Training",
"611420":"Computer Training ",
"61143":"Professional and Management Development Training",
"611430":"Professional and Management Development Training ",
"6115":"Technical and Trade Schools ",
"61151":"Technical and Trade Schools",
"611511":"Cosmetology and Barber Schools ",
"611512":"Flight Training ",
"611513":"Apprenticeship Training ",
"611519":"Other Technical and Trade Schools ",
"6116":"Other Schools and Instruction",
"61161":"Fine Arts Schools",
"611610":"Fine Arts Schools ",
"61162":"Sports and Recreation Instruction",
"611620":"Sports and Recreation Instruction ",
"61163":"Language Schools",
"611630":"Language Schools ",
"61169":"All Other Schools and Instruction",
"611691":"Exam Preparation and Tutoring ",
"611692":"Automobile Driving Schools ",
"611699":"All Other Miscellaneous Schools and Instruction ",
"6117":"Educational Support Services",
"61171":"Educational Support Services",
"611710":"Educational Support Services",
"62":"Health Care and Social Assistance",
"621":"Ambulatory Health Care Services",
"6211":"Offices of Physicians",
"62111":"Offices of Physicians",
"621111":"Offices of Physicians (except Mental Health Specialists) ",
"621112":"Offices of Physicians, Mental Health Specialists ",
"6212":"Offices of Dentists",
"62121":"Offices of Dentists",
"621210":"Offices of Dentists ",
"6213":"Offices of Other Health Practitioners",
"62131":"Offices of Chiropractors",
"621310":"Offices of Chiropractors ",
"62132":"Offices of Optometrists",
"621320":"Offices of Optometrists",
"62133":"Offices of Mental Health Practitioners (except Physicians)",
"621330":"Offices of Mental Health Practitioners (except Physicians) ",
"62134":"Offices of Physical, Occupational and Speech Therapists, and Audiologists",
"621340":"Offices of Physical, Occupational and Speech Therapists, and Audiologists ",
"62139":"Offices of All Other Health Practitioners",
"621391":"Offices of Podiatrists ",
"621399":"Offices of All Other Miscellaneous Health Practitioners ",
"6214":"Outpatient Care Centers",
"62141":"Family Planning Centers",
"621410":"Family Planning Centers ",
"62142":"Outpatient Mental Health and Substance Abuse Centers",
"621420":"Outpatient Mental Health and Substance Abuse Centers ",
"62149":"Other Outpatient Care Centers",
"621491":"HMO Medical Centers ",
"621492":"Kidney Dialysis Centers ",
"621493":"Freestanding Ambulatory Surgical and Emergency Centers ",
"621498":"All Other Outpatient Care Centers ",
"6215":"Medical and Diagnostic Laboratories",
"62151":"Medical and Diagnostic Laboratories",
"621511":"Medical Laboratories ",
"621512":"Diagnostic Imaging Centers ",
"6216":"Home Health Care Services",
"62161":"Home Health Care Services",
"621610":"Home Health Care Services",
"6219":"Other Ambulatory Health Care Services",
"62191":"Ambulance Services",
"621910":"Ambulance Services ",
"62199":"All Other Ambulatory Health Care Services",
"621991":"Blood and Organ Banks ",
"621999":"All Other Miscellaneous Ambulatory Health Care Services ",
"622":"Hospitals",
"6221":"General Medical and Surgical Hospitals",
"62211":"General Medical and Surgical Hospitals",
"622110":"General Medical and Surgical Hospitals ",
"6222":"Psychiatric and Substance Abuse Hospitals",
"62221":"Psychiatric and Substance Abuse Hospitals",
"622210":"Psychiatric and Substance Abuse Hospitals ",
"6223":"Specialty (except Psychiatric and Substance Abuse) Hospitals",
"62231":"Specialty (except Psychiatric and Substance Abuse) Hospitals",
"622310":"Specialty (except Psychiatric and Substance Abuse) Hospitals ",
"623":"Nursing and Residential Care Facilities",
"6231":"Nursing Care Facilities (Skilled Nursing Facilities)",
"62311":"Nursing Care Facilities (Skilled Nursing Facilities)",
"623110":"Nursing Care Facilities (Skilled Nursing Facilities) ",
"6232":"Residential Intellectual and Developmental Disability, Mental Health, and Substance Abuse Facilities",
"62321":"Residential Intellectual and Developmental Disability Facilities",
"623210":"Residential Intellectual and Developmental Disability Facilities ",
"62322":"Residential Mental Health and Substance Abuse Facilities",
"623220":"Residential Mental Health and Substance Abuse Facilities ",
"6233":"Continuing Care Retirement Communities and Assisted Living Facilities for the Elderly",
"62331":"Continuing Care Retirement Communities and Assisted Living Facilities for the Elderly",
"623311":"Continuing Care Retirement Communities ",
"623312":"Assisted Living Facilities for the Elderly ",
"6239":"Other Residential Care Facilities",
"62399":"Other Residential Care Facilities",
"623990":"Other Residential Care Facilities ",
"624":"Social Assistance",
"6241":"Individual and Family Services",
"62411":"Child and Youth Services",
"624110":"Child and Youth Services ",
"62412":"Services for the Elderly and Persons with Disabilities",
"624120":"Services for the Elderly and Persons with Disabilities ",
"62419":"Other Individual and Family Services",
"624190":"Other Individual and Family Services ",
"6242":"Community Food and Housing, and Emergency and Other Relief Services",
"62421":"Community Food Services",
"624210":"Community Food Services ",
"62422":"Community Housing Services",
"624221":"Temporary Shelters ",
"624229":"Other Community Housing Services ",
"62423":"Emergency and Other Relief Services",
"624230":"Emergency and Other Relief Services ",
"6243":"Vocational Rehabilitation Services",
"62431":"Vocational Rehabilitation Services",
"624310":"Vocational Rehabilitation Services ",
"6244":"Child Day Care Services",
"62441":"Child Day Care Services",
"624410":"Child Day Care Services ",
"71":"Arts, Entertainment, and Recreation",
"711":"Performing Arts, Spectator Sports, and Related Industries",
"7111":"Performing Arts Companies",
"71111":"Theater Companies and Dinner Theaters",
"711110":"Theater Companies and Dinner Theaters ",
"71112":"Dance Companies",
"711120":"Dance Companies ",
"71113":"Musical Groups and Artists",
"711130":"Musical Groups and Artists ",
"71119":"Other Performing Arts Companies",
"711190":"Other Performing Arts Companies ",
"7112":"Spectator Sports",
"71121":"Spectator Sports",
"711211":"Sports Teams and Clubs ",
"711212":"Racetracks ",
"711219":"Other Spectator Sports ",
"7113":"Promoters of Performing Arts, Sports, and Similar Events",
"71131":"Promoters of Performing Arts, Sports, and Similar Events with Facilities",
"711310":"Promoters of Performing Arts, Sports, and Similar Events with Facilities ",
"71132":"Promoters of Performing Arts, Sports, and Similar Events without Facilities",
"711320":"Promoters of Performing Arts, Sports, and Similar Events without Facilities ",
"7114":"Agents and Managers for Artists, Athletes, Entertainers, and Other Public Figures",
"71141":"Agents and Managers for Artists, Athletes, Entertainers, and Other Public Figures",
"711410":"Agents and Managers for Artists, Athletes, Entertainers, and Other Public Figures",
"7115":"Independent Artists, Writers, and Performers",
"71151":"Independent Artists, Writers, and Performers",
"711510":"Independent Artists, Writers, and Performers ",
"712":"Museums, Historical Sites, and Similar Institutions",
"7121":"Museums, Historical Sites, and Similar Institutions",
"71211":"Museums",
"712110":"Museums ",
"71212":"Historical Sites",
"712120":"Historical Sites",
"71213":"Zoos and Botanical Gardens",
"712130":"Zoos and Botanical Gardens ",
"71219":"Nature Parks and Other Similar Institutions",
"712190":"Nature Parks and Other Similar Institutions",
"713":"Amusement, Gambling, and Recreation Industries",
"7131":"Amusement Parks and Arcades",
"71311":"Amusement and Theme Parks",
"713110":"Amusement and Theme Parks ",
"71312":"Amusement Arcades",
"713120":"Amusement Arcades",
"7132":"Gambling Industries",
"71321":"Casinos (except Casino Hotels)",
"713210":"Casinos (except Casino Hotels)",
"71329":"Other Gambling Industries",
"713290":"Other Gambling Industries ",
"7139":"Other Amusement and Recreation Industries",
"71391":"Golf Courses and Country Clubs",
"713910":"Golf Courses and Country Clubs",
"71392":"Skiing Facilities",
"713920":"Skiing Facilities",
"71393":"Marinas",
"713930":"Marinas",
"71394":"Fitness and Recreational Sports Centers",
"713940":"Fitness and Recreational Sports Centers ",
"71395":"Bowling Centers",
"713950":"Bowling Centers",
"71399":"All Other Amusement and Recreation Industries",
"713990":"All Other Amusement and Recreation Industries ",
"72":"Accommodation and Food Services",
"721":"Accommodation",
"7211":"Traveler Accommodation",
"72111":"Hotels (except Casino Hotels) and Motels",
"721110":"Hotels (except Casino Hotels) and Motels ",
"72112":"Casino Hotels",
"721120":"Casino Hotels",
"72119":"Other Traveler Accommodation",
"721191":"Bed-and-Breakfast Inns ",
"721199":"All Other Traveler Accommodation ",
"7212":"RV (Recreational Vehicle) Parks and Recreational Camps",
"72121":"RV (Recreational Vehicle) Parks and Recreational Camps",
"721211":"RV (Recreational Vehicle) Parks and Campgrounds ",
"721214":"Recreational and Vacation Camps (except Campgrounds) ",
"7213":"Rooming and Boarding Houses, Dormitories, and Workers' Camps",
"72131":"Rooming and Boarding Houses, Dormitories, and Workers' Camps",
"721310":"Rooming and Boarding Houses, Dormitories, and Workers' Camps ",
"722":"Food Services and Drinking Places",
"7223":"Special Food Services",
"72231":"Food Service Contractors",
"722310":"Food Service Contractors",
"72232":"Caterers",
"722320":"Caterers",
"72233":"Mobile Food Services",
"722330":"Mobile Food Services",
"7224":"Drinking Places (Alcoholic Beverages)",
"72241":"Drinking Places (Alcoholic Beverages)",
"722410":"Drinking Places (Alcoholic Beverages) ",
"7225":"Restaurants and Other Eating Places",
"72251":"Restaurants and Other Eating Places",
"722511":"Full-Service Restaurants ",
"722513":"Limited-Service Restaurants ",
"722514":"Cafeterias, Grill Buffets, and Buffets ",
"722515":"Snack and Nonalcoholic Beverage Bars ",
"81":"Other Services (except Public Administration)",
"811":"Repair and Maintenance",
"8111":"Automotive Repair and Maintenance",
"81111":"Automotive Mechanical and Electrical Repair and Maintenance",
"811111":"General Automotive Repair ",
"811112":"Automotive Exhaust System Repair ",
"811113":"Automotive Transmission Repair ",
"811118":"Other Automotive Mechanical and Electrical Repair and Maintenance ",
"81112":"Automotive Body, Paint, Interior, and Glass Repair",
"811121":"Automotive Body, Paint, and Interior Repair and Maintenance ",
"811122":"Automotive Glass Replacement Shops ",
"81119":"Other Automotive Repair and Maintenance",
"811191":"Automotive Oil Change and Lubrication Shops ",
"811192":"Car Washes ",
"811198":"All Other Automotive Repair and Maintenance ",
"8112":"Electronic and Precision Equipment Repair and Maintenance",
"81121":"Electronic and Precision Equipment Repair and Maintenance",
"811211":"Consumer Electronics Repair and Maintenance ",
"811212":"Computer and Office Machine Repair and Maintenance ",
"811213":"Communication Equipment Repair and Maintenance ",
"811219":"Other Electronic and Precision Equipment Repair and Maintenance ",
"8113":"Commercial and Industrial Machinery and Equipment (except Automotive and Electronic) Repair and Maintenance",
"81131":"Commercial and Industrial Machinery and Equipment (except Automotive and Electronic) Repair and Maintenance",
"811310":"Commercial and Industrial Machinery and Equipment (except Automotive and Electronic) Repair and Maintenance ",
"8114":"Personal and Household Goods Repair and Maintenance",
"81141":"Home and Garden Equipment and Appliance Repair and Maintenance",
"811411":"Home and Garden Equipment Repair and Maintenance ",
"811412":"Appliance Repair and Maintenance ",
"81142":"Reupholstery and Furniture Repair",
"811420":"Reupholstery and Furniture Repair",
"81143":"Footwear and Leather Goods Repair",
"811430":"Footwear and Leather Goods Repair",
"81149":"Other Personal and Household Goods Repair and Maintenance",
"811490":"Other Personal and Household Goods Repair and Maintenance ",
"812":"Personal and Laundry Services",
"8121":"Personal Care Services ",
"81211":"Hair, Nail, and Skin Care Services ",
"812111":"Barber Shops ",
"812112":"Beauty Salons ",
"812113":"Nail Salons ",
"81219":"Other Personal Care Services ",
"812191":"Diet and Weight Reducing Centers ",
"812199":"Other Personal Care Services ",
"8122":"Death Care Services ",
"81221":"Funeral Homes and Funeral Services ",
"812210":"Funeral Homes and Funeral Services ",
"81222":"Cemeteries and Crematories ",
"812220":"Cemeteries and Crematories ",
"8123":"Drycleaning and Laundry Services ",
"81231":"Coin-Operated Laundries and Drycleaners ",
"812310":"Coin-Operated Laundries and Drycleaners ",
"81232":"Drycleaning and Laundry Services (except Coin-Operated) ",
"812320":"Drycleaning and Laundry Services (except Coin-Operated) ",
"81233":"Linen and Uniform Supply ",
"812331":"Linen Supply ",
"812332":"Industrial Launderers ",
"8129":"Other Personal Services ",
"81291":"Pet Care (except Veterinary) Services ",
"812910":"Pet Care (except Veterinary) Services ",
"81292":"Photofinishing ",
"812921":"Photofinishing Laboratories (except One-Hour) ",
"812922":"One-Hour Photofinishing ",
"81293":"Parking Lots and Garages ",
"812930":"Parking Lots and Garages ",
"81299":"All Other Personal Services ",
"812990":"All Other Personal Services ",
"813":"Religious, Grantmaking, Civic, Professional, and Similar Organizations",
"8131":"Religious Organizations ",
"81311":"Religious Organizations ",
"813110":"Religious Organizations ",
"8132":"Grantmaking and Giving Services ",
"81321":"Grantmaking and Giving Services ",
"813211":"Grantmaking Foundations ",
"813212":"Voluntary Health Organizations ",
"813219":"Other Grantmaking and Giving Services ",
"8133":"Social Advocacy Organizations ",
"81331":"Social Advocacy Organizations ",
"813311":"Human Rights Organizations ",
"813312":"Environment, Conservation and Wildlife Organizations ",
"813319":"Other Social Advocacy Organizations ",
"8134":"Civic and Social Organizations ",
"81341":"Civic and Social Organizations ",
"813410":"Civic and Social Organizations ",
"8139":"Business, Professional, Labor, Political, and Similar Organizations ",
"81391":"Business Associations ",
"813910":"Business Associations ",
"81392":"Professional Organizations ",
"813920":"Professional Organizations ",
"81393":"Labor Unions and Similar Labor Organizations ",
"813930":"Labor Unions and Similar Labor Organizations ",
"81394":"Political Organizations ",
"813940":"Political Organizations ",
"81399":"Other Similar Organizations (except Business, Professional, Labor, and Political Organizations) ",
"813990":"Other Similar Organizations (except Business, Professional, Labor, and Political Organizations) ",
"814":"Private Households",
"8141":"Private Households",
"81411":"Private Households",
"814110":"Private Households",
"92":"Public Administration",
"921":"Executive, Legislative, and Other General Government Support ",
"9211":"Executive, Legislative, and Other General Government Support ",
"92111":"Executive Offices ",
"921110":"Executive Offices ",
"92112":"Legislative Bodies ",
"921120":"Legislative Bodies ",
"92113":"Public Finance Activities ",
"921130":"Public Finance Activities ",
"92114":"Executive and Legislative Offices, Combined ",
"921140":"Executive and Legislative Offices, Combined ",
"92115":"American Indian and Alaska Native Tribal Governments ",
"921150":"American Indian and Alaska Native Tribal Governments ",
"92119":"Other General Government Support ",
"921190":"Other General Government Support ",
"922":"Justice, Public Order, and Safety Activities ",
"9221":"Justice, Public Order, and Safety Activities ",
"92211":"Courts ",
"922110":"Courts ",
"92212":"Police Protection ",
"922120":"Police Protection ",
"92213":"Legal Counsel and Prosecution ",
"922130":"Legal Counsel and Prosecution ",
"92214":"Correctional Institutions ",
"922140":"Correctional Institutions ",
"92215":"Parole Offices and Probation Offices ",
"922150":"Parole Offices and Probation Offices ",
"92216":"Fire Protection ",
"922160":"Fire Protection ",
"92219":"Other Justice, Public Order, and Safety Activities ",
"922190":"Other Justice, Public Order, and Safety Activities ",
"923":"Administration of Human Resource Programs ",
"9231":"Administration of Human Resource Programs ",
"92311":"Administration of Education Programs ",
"923110":"Administration of Education Programs ",
"92312":"Administration of Public Health Programs ",
"923120":"Administration of Public Health Programs ",
"92313":"Administration of Human Resource Programs (except Education, Public Health, and Veterans' Affairs Programs) ",
"923130":"Administration of Human Resource Programs (except Education, Public Health, and Veterans' Affairs Programs) ",
"92314":"Administration of Veterans' Affairs ",
"923140":"Administration of Veterans' Affairs ",
"924":"Administration of Environmental Quality Programs ",
"9241":"Administration of Environmental Quality Programs ",
"92411":"Administration of Air and Water Resource and Solid Waste Management Programs ",
"924110":"Administration of Air and Water Resource and Solid Waste Management Programs ",
"92412":"Administration of Conservation Programs ",
"924120":"Administration of Conservation Programs ",
"925":"Administration of Housing Programs, Urban Planning, and Community Development ",
"9251":"Administration of Housing Programs, Urban Planning, and Community Development ",
"92511":"Administration of Housing Programs ",
"925110":"Administration of Housing Programs ",
"92512":"Administration of Urban Planning and Community and Rural Development ",
"925120":"Administration of Urban Planning and Community and Rural Development ",
"926":"Administration of Economic Programs ",
"9261":"Administration of Economic Programs ",
"92611":"Administration of General Economic Programs ",
"926110":"Administration of General Economic Programs ",
"92612":"Regulation and Administration of Transportation Programs ",
"926120":"Regulation and Administration of Transportation Programs ",
"92613":"Regulation and Administration of Communications, Electric, Gas, and Other Utilities ",
"926130":"Regulation and Administration of Communications, Electric, Gas, and Other Utilities ",
"92614":"Regulation of Agricultural Marketing and Commodities ",
"926140":"Regulation of Agricultural Marketing and Commodities ",
"92615":"Regulation, Licensing, and Inspection of Miscellaneous Commercial Sectors ",
"926150":"Regulation, Licensing, and Inspection of Miscellaneous Commercial Sectors ",
"927":"Space Research and Technology ",
"9271":"Space Research and Technology ",
"92711":"Space Research and Technology ",
"927110":"Space Research and Technology ",
"928":"National Security and International Affairs ",
"9281":"National Security and International Affairs ",
"92811":"National Security ",
"928110":"National Security ",
"92812":"International Affairs ",
"928120":"International Affairs ",
}



gics_code = {
'10': 'Energy', 
'15': 'Materials', 
'20': 'Industrials', 
'25': 'Consumer Discretionary', 
'30': 'Consumer Staples', 
'35': 'Health Care', 
'40': 'Financials', 
'45': 'Information Technology', 
'50': 'Communication Services', 
'55': 'Utilities', 
'60': 'Real Estate', 
'1010': 'Energy', 
'1510': 'Materials', 
'2010': 'Capital Goods', 
'2020': 'Commercial  & Professional Services', 
'2030': 'Transportation', 
'2510': 'Automobiles & Components', 
'2520': 'Consumer Durables & Apparel', 
'2530': 'Consumer Services', 
'2540': 'Media (discontinued effective close of September 28, 2018)', 
'2550': 'Retailing', 
'3010': 'Food & Staples Retailing', 
'3020': 'Food, Beverage & Tobacco', 
'3030': 'Household & Personal Products', 
'3510': 'Health Care Equipment & Services', 
'3520': 'Pharmaceuticals, Biotechnology & Life Sciences', 
'4010': 'Banks', 
'4020': 'Diversified Financials', 
'4030': 'Insurance', 
'4040': 'Real Estate - - discontinued effective close of Aug 31, 2016', 
'4510': 'Software & Services', 
'4520': 'Technology Hardware & Equipment', 
'4530': 'Semiconductors & Semiconductor Equipment', 
'5010': 'Telecommunication Services', 
'5020': 'Media & Entertainment', 
'5510': 'Utilities', 
'6010': 'Real Estate', 
'101010': 'Energy Equipment & Services', 
'101020': 'Oil, Gas & Consumable Fuels', 
'151010': 'Chemicals', 
'151020': 'Construction Materials', 
'151030': 'Containers & Packaging', 
'151040': 'Metals & Mining', 
'151050': 'Paper & Forest Products', 
'201010': 'Aerospace & Defense', 
'201020': 'Building Products', 
'201030': 'Construction & Engineering', 
'201040': 'Electrical Equipment', 
'201050': 'Industrial Conglomerates', 
'201060': 'Machinery', 
'201070': 'Trading Companies & Distributors', 
'202010': 'Commercial Services & Supplies', 
'202020': 'Professional Services', 
'203010': 'Air Freight & Logistics', 
'203020': 'Airlines', 
'203030': 'Marine', 
'203040': 'Road & Rail', 
'203050': 'Transportation Infrastructure', 
'251010': 'Auto Components', 
'251020': 'Automobiles', 
'252010': 'Household Durables', 
'252020': 'Leisure Products', 
'252030': 'Textiles, Apparel & Luxury Goods', 
'253010': 'Hotels, Restaurants & Leisure', 
'253020': 'Diversified Consumer Services', 
'254010': 'Media (discontinued effective close of September 28, 2018)', 
'255010': 'Distributors', 
'255020': 'Internet & Direct Marketing Retail', 
'255030': 'Multiline Retail', 
'255040': 'Specialty Retail', 
'301010': 'Food & Staples Retailing', 
'302010': 'Beverages', 
'302020': 'Food Products', 
'302030': 'Tobacco', 
'303010': 'Household Products', 
'303020': 'Personal Products', 
'351010': 'Health Care Equipment & Supplies', 
'351020': 'Health Care Providers & Services', 
'351030': 'Health Care Technology', 
'352010': 'Biotechnology', 
'352020': 'Pharmaceuticals', 
'352030': 'Life Sciences Tools & Services', 
'401010': 'Banks', 
'401020': 'Thrifts & Mortgage Finance', 
'402010': 'Diversified Financial Services', 
'402020': 'Consumer Finance', 
'402030': 'Capital Markets', 
'402040': 'Mortgage Real Estate Investment \nTrusts (REITs)', 
'403010': 'Insurance', 
'404010': 'Real Estate -- Discontinued effective 04/28/2006', 
'404020': 'Real Estate Investment Trusts (REITs) - discontinued effective close of Aug 31, 2016', 
'404030': 'Real Estate Management & Development (discontinued effective close of August 31, 2016)', 
'451010': 'Internet Software & Services (discontinued effective close of September 28, 2018)', 
'451020': 'IT Services', 
'451030': 'Software', 
'452010': 'Communications Equipment', 
'452020': 'Technology Hardware, Storage & Peripherals', 
'452030': 'Electronic Equipment, Instruments & Components', 
'452040': 'Office Electronics - Discontinued effective 02/28/2014', 
'452050': 'Semiconductor Equipment & Products -- Discontinued effective 04/30/2003.', 
'453010': 'Semiconductors & Semiconductor Equipment', 
'501010': 'Diversified Telecommunication Services', 
'501020': 'Wireless Telecommunication Services', 
'502010': 'Media', 
'502020': 'Entertainment', 
'502030': 'Interactive Media & Services', 
'551010': 'Electric Utilities', 
'551020': 'Gas Utilities', 
'551030': 'Multi-Utilities', 
'551040': 'Water Utilities', 
'551050': 'Independent Power and Renewable Electricity Producers', 
'601010': 'Equity Real Estate \nInvestment Trusts \n(REITs)', 
'601020': 'Real Estate Management & Development', 
'10101010': 'Oil & Gas Drilling', 
'10101020': 'Oil & Gas Equipment & Services', 
'10102010': 'Integrated Oil & Gas', 
'10102020': 'Oil & Gas Exploration & Production', 
'10102030': 'Oil & Gas Refining & Marketing', 
'10102040': 'Oil & Gas Storage & Transportation', 
'10102050': 'Coal & Consumable Fuels', 
'15101010': 'Commodity Chemicals', 
'15101020': 'Diversified Chemicals', 
'15101030': 'Fertilizers & Agricultural Chemicals', 
'15101040': 'Industrial Gases', 
'15101050': 'Specialty Chemicals', 
'15102010': 'Construction Materials', 
'15103010': 'Metal & Glass Containers', 
'15103020': 'Paper Packaging', 
'15104010': 'Aluminum', 
'15104020': 'Diversified Metals & Mining', 
'15104025': 'Copper', 
'15104030': 'Gold', 
'15104040': 'Precious Metals & Minerals', 
'15104045': 'Silver', 
'15104050': 'Steel', 
'15105010': 'Forest Products', 
'15105020': 'Paper Products', 
'20101010': 'Aerospace & Defense', 
'20102010': 'Building Products', 
'20103010': 'Construction & Engineering', 
'20104010': 'Electrical Components & Equipment', 
'20104020': 'Heavy Electrical Equipment', 
'20105010': 'Industrial Conglomerates', 
'20106010': 'Construction Machinery & Heavy Trucks', 
'20106015': 'Agricultural & Farm Machinery', 
'20106020': 'Industrial Machinery', 
'20107010': 'Trading Companies & Distributors', 
'20201010': 'Commercial Printing', 
'20201020': 'Data Processing Services  (discontinued effective close of April 30, 2003)', 
'20201030': 'Diversified Commercial & Professional Services (discontinued effective close of August 31, 2008)', 
'20201040': 'Human Resource & Employment Services (discontinued effective close of August 31, 2008)', 
'20201050': 'Environmental & Facilities Services', 
'20201060': 'Office Services & Supplies', 
'20201070': 'Diversified Support Services', 
'20201080': 'Security & Alarm Services', 
'20202010': 'Human Resource & Employment Services', 
'20202020': 'Research & Consulting Services', 
'20301010': 'Air Freight & Logistics', 
'20302010': 'Airlines', 
'20303010': 'Marine', 
'20304010': 'Railroads', 
'20304020': 'Trucking', 
'20305010': 'Airport Services', 
'20305020': 'Highways & Railtracks', 
'20305030': 'Marine Ports & Services', 
'25101010': 'Auto Parts & Equipment', 
'25101020': 'Tires & Rubber', 
'25102010': 'Automobile Manufacturers', 
'25102020': 'Motorcycle Manufacturers', 
'25201010': 'Consumer Electronics', 
'25201020': 'Home Furnishings', 
'25201030': 'Homebuilding', 
'25201040': 'Household Appliances', 
'25201050': 'Housewares & Specialties', 
'25202010': 'Leisure Products', 
'25202020': 'Photographic Products (discontinued effective close of February 28, 2014)', 
'25203010': 'Apparel, Accessories & Luxury Goods', 
'25203020': 'Footwear', 
'25203030': 'Textiles', 
'25301010': 'Casinos & Gaming', 
'25301020': 'Hotels, Resorts & Cruise Lines', 
'25301030': 'Leisure Facilities', 
'25301040': 'Restaurants', 
'25302010': 'Education Services', 
'25302020': 'Specialized Consumer Services', 
'25401010': 'Advertising (discontinued effective close of September 28, 2018)', 
'25401020': 'Broadcasting (discontinued effective close of September 28, 2018)', 
'25401025': 'Cable & Satellite (discontinued effective close of September 28, 2018)', 
'25401030': 'Movies & Entertainment (discontinued effective close of September 28, 2018)', 
'25401040': 'Publishing (discontinued effective close of September 28, 2018)', 
'25501010': 'Distributors', 
'25502010': 'Catalog Retail (discontinued effective close of August 31, 2016)', 
'25502020': 'Internet & Direct Marketing Retail', 
'25503010': 'Department Stores', 
'25503020': 'General Merchandise Stores', 
'25504010': 'Apparel Retail', 
'25504020': 'Computer & Electronics Retail', 
'25504030': 'Home Improvement Retail', 
'25504040': 'Specialty Stores', 
'25504050': 'Automotive Retail', 
'25504060': 'Homefurnishing Retail', 
'30101010': 'Drug Retail', 
'30101020': 'Food Distributors', 
'30101030': 'Food Retail', 
'30101040': 'Hypermarkets & Super Centers', 
'30201010': 'Brewers', 
'30201020': 'Distillers & Vintners', 
'30201030': 'Soft Drinks', 
'30202010': 'Agricultural Products', 
'30202020': 'Meat, Poultry & Fish (discontinued effective close of March 28 2002)', 
'30202030': 'Packaged Foods & Meats', 
'30203010': 'Tobacco', 
'30301010': 'Household Products', 
'30302010': 'Personal Products', 
'35101010': 'Health Care Equipment', 
'35101020': 'Health Care Supplies', 
'35102010': 'Health Care Distributors', 
'35102015': 'Health Care  Services', 
'35102020': 'Health Care Facilities', 
'35102030': 'Managed Health Care', 
'35103010': 'Health Care Technology', 
'35201010': 'Biotechnology', 
'35202010': 'Pharmaceuticals', 
'35203010': 'Life Sciences Tools & Services', 
'40101010': 'Diversified Banks', 
'40101015': 'Regional Banks', 
'40102010': 'Thrifts & Mortgage Finance', 
'40201010': 'Consumer Finance (discontinued effective close of April 30, 2003)', 
'40201020': 'Other Diversified Financial Services', 
'40201030': 'Multi-Sector Holdings', 
'40201040': 'Specialized Finance', 
'40202010': 'Consumer Finance', 
'40203010': 'Asset Management & Custody Banks', 
'40203020': 'Investment Banking & Brokerage', 
'40203030': 'Diversified Capital Markets', 
'40203040': 'Financial Exchanges & Data', 
'40204010': 'Mortgage REITs', 
'40301010': 'Insurance Brokers', 
'40301020': 'Life & Health Insurance', 
'40301030': 'Multi-line Insurance', 
'40301040': 'Property & Casualty Insurance', 
'40301050': 'Reinsurance', 
'40401010': 'Real Estate Investment Trusts (discontinued effective close of April 28, 2006)', 
'40401020': 'Real Estate Management & Development (discontinued effective close of April 28, 2006)', 
'40402010': 'Diversified REITs (discontinued effective close of August 31, 2016)', 
'40402020': 'Industrial REITs (discontinued effective close of August 31, 2016)', 
'40402030': 'Mortgage REITs (discontinued effective close of August 31, 2016)', 
'40402035': 'Hotel & Resort REITs (discontinued effective close of August 31, 2016)', 
'40402040': 'Office REITs (discontinued effective close of August 31, 2016)', 
'40402045': 'Health Care REITs (discontinued effective close of August 31, 2016)', 
'40402050': 'Residential REITs (discontinued effective close of August 31, 2016)', 
'40402060': 'Retail REITs (discontinued effective close of August 31, 2016)', 
'40402070': 'Specialized REITs (discontinued effective close of August 31, 2016)', 
'40403010': 'Diversified Real Estate Activities (discontinued effective close of August 31, 2016)', 
'40403020': 'Real Estate Operating Companies (discontinued effective close of August 31, 2016)', 
'40403030': 'Real Estate Development (discontinued effective close of August 31, 2016)', 
'40403040': 'Real Estate Services (discontinued effective close of August 31, 2016)', 
'45101010': 'Internet Software & Services (discontinued effective close of September 28, 2018)', 
'45102010': 'IT Consulting & Other Services', 
'45102020': 'Data Processing & Outsourced Services', 
'45102030': 'Internet Services & Infrastructure', 
'45103010': 'Application Software', 
'45103020': 'Systems Software', 
'45103030': 'Home Entertainment Software (discontinued effective close of September 28, 2018)', 
'45201020': 'Communications Equipment', 
'45201010': 'Networking Equipment (discontinued effective close of April 30, 2003)', 
'45202010': 'Computer Hardware (discontinued effective close of February 28, 2014)', 
'45202020': 'Computer Storage & Peripherals (discontinued effective close of February 28, 2014)', 
'45202030': 'Technology Hardware, Storage & Peripherals', 
'45203010': 'Electronic Equipment & Instruments ', 
'45203015': 'Electronic Components', 
'45203020': 'Electronic Manufacturing Services', 
'45203030': 'Technology Distributors', 
'45204010': 'Office Electronics (discontinued effective close of February 28, 2014)', 
'45205010': 'Semiconductor Equipment (discontinued effective close of April 30, 2003)', 
'45205020': 'Semiconductors (discontinued effective close of April 30, 2003)', 
'45301010': 'Semiconductor Equipment ', 
'45301020': 'Semiconductors', 
'50101010': 'Alternative Carriers', 
'50101020': 'Integrated Telecommunication Services', 
'50102010': 'Wireless Telecommunication Services', 
'50201010': 'Advertising', 
'50201020': 'Broadcasting', 
'50201030': 'Cable & Satellite', 
'50201040': 'Publishing', 
'50202010': 'Movies & Entertainment', 
'50202020': 'Interactive Home Entertainment', 
'50203010': 'Interactive Media & Services', 
'55101010': 'Electric Utilities', 
'55102010': 'Gas Utilities', 
'55103010': 'Multi-Utilities', 
'55104010': 'Water Utilities', 
'55105010': 'Independent Power Producers & Energy Traders', 
'55105020': 'Renewable Electricity ', 
'60101010': 'Diversified REITs', 
'60101020': 'Industrial REITs', 
'60101030': 'Hotel & Resort REITs ', 
'60101040': 'Office REITs ', 
'60101050': 'Health Care REITs ', 
'60101060': 'Residential REITs', 
'60101070': 'Retail REITs', 
'60101080': 'Specialized REITs ', 
'60102010': 'Diversified Real Estate Activities ', 
'60102020': 'Real Estate Operating Companies', 
'60102030': 'Real Estate Development ', 
'60102040': 'Real Estate Services ',
'99': 'Missing or Unknown GICS',
'9999': 'Missing or Unknown GICS',
'999999': 'Missing or Unknown GICS',
'99999999': 'Missing or Unknown GICS',
}

# -*- coding: utf-8 -*-
"""
Created on Mon Mar  6 17:22:06 2017

@author: ub71894 (4e8e6d0b), CSG
"""

['Attachdefaults.py', 'CreateBenchmarkMatrix.py', 'CreateBenchmarkMatrix_old_openpyxl.py', 'MFA.py', 'PDModel.py', 'Process.py', 'SFA.py', '_info_data.py', '__init__.py']
[336, 758, 1157, 2058, 2166, 3643, 4602, 7120, 7127]
# -*- coding: utf-8 -*-
"""
Created on Tue Apr  4 10:21:43 2017

@author: ub64283 Liwen Zhang (original author: XU79799)
@author: ub71894 (4e8e6d0b), Add TMnotches(), withknotchesrate() and TMstats()

Version: 1.1: 20191101, modify code to fit new openpyxl function

"""
import pandas as pd
import numpy as np
import openpyxl
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.styles import Border, Side, Alignment, Font
from openpyxl.formatting.rule import IconSetRule
from openpyxl import  Workbook, load_workbook
import scipy.stats as stats
from openpyxl.utils.cell import coordinate_from_string, column_index_from_string


def range_to_index(range_in_str): # New Ver. 1.1. The best patch I have ever developed...

    delimiter = range_in_str.find(':')
    min_cell = range_in_str[:delimiter]
    max_cell = range_in_str[(delimiter+1):]

    xy = coordinate_from_string(min_cell) 
    min_col = column_index_from_string(xy[0]) 
    min_row = xy[1]
    xy = coordinate_from_string(max_cell) 
    max_col = column_index_from_string(xy[0]) 
    max_row = xy[1]

    return({'min_col':min_col, 'max_col':max_col, 'min_row':min_row, 'max_row':max_row})




#Count the number of obligors which have same rating with a margin of y    
def within(MM,y):
    """
    Summary:
    Counts the number of obligors which have the same rating within a margin of y

    Variables:
    MM: type = DataFrame. A dataframe which contains the transition matrix of two
    ratings systems.

    y: type = int. The margin of ratings that will contribute to the final sum.

    Returns:
    An integer counting the number of obligors with the same rating within a margin of y
    """
    x=np.array(MM)
    N=len(x)
    sum1=0
    if y == 0:
        return np.sum(np.diag(x))
    else:
        for i in range(N):
            ll=max(0,i-y) #checks the lower limit of the margin
            uu=min(N+1,i+y+1) #checks the upper limit of the margin
            sum1=sum1+np.sum(x[ll:uu,i]) #for each cell, adds the number that falls within the margin
        return(sum1)

def down_notch(MM):
    N=MM.shape[0]
    sum1 = 0
    for notch in range(1,N):
        temp = 0
        for i in range(notch):
            temp += MM[i,(N-notch+i)]
        sum1 += temp*(N-notch)
    return(sum1)
        
def worse_ratings(MM):
    x=np.array(MM)
    N=len(x)
    sum1=0
    for i in range(N):
        ll=0
        sum1=sum1+np.sum(x[ll:i,i])
    return(sum1)

def up_notch(MM):
    N=MM.shape[0]
    sum1=0
    for notch in range(1,N):
        temp=0
        for i in range(notch):
            temp += MM[(N-notch+i),i]
        sum1 += temp*(N-notch)
    return(sum1)
    
def better_ratings(MM):
    x=np.array(MM)
    N=len(x)
    sum1=0
    for i in range(N):
        uu=N+1
        sum1=sum1+np.sum(x[(i+1):uu,i])
    return(sum1)

def ComputeListOutstandingObligors(df):
    res=[]
    
#    for i in range(df.shape[0]):
#       if (abs(df['Old_Ratings'].iloc[i]-df['New_Ratings_oldms'].iloc[i])>=4):
#           res.append(df['CUSTOMERID'].iloc[i])
    return res

def CreateStyleSheet_TransitionMatrix(sheet, rowshift, colshift, PDRR):
    
    ### assume len(PDRR) <= 21 ###
    
    len_pdrr = len(PDRR)
    
    #### Work on style ###
    for i in range(rowshift-3,rowshift+len_pdrr):
        sheet.row_dimensions[i].height=12
    for j in ['AB','AC','AD']:
        sheet.column_dimensions[j].width=7
    sheet.column_dimensions['B'].width=3
    sheet.column_dimensions['AA'].width=9
    sheet.column_dimensions['AG'].width=9
    sheet.column_dimensions['AH'].width=5
    sheet.column_dimensions['AE'].width=8
    sheet.column_dimensions['AF'].width=8

    for _row in sheet.iter_rows(**range_to_index('B'+ str(rowshift-2) + ':' + chr(68+len_pdrr) + str(rowshift-1))):
        for _cell in _row:
            fill=PatternFill(start_color='FFFFFF', end_color='FFFFFF',fill_type='solid')
            _cell.fill=fill 
    for _row in sheet.iter_rows(**range_to_index('B'+ str(rowshift-2) + ':C' + str(rowshift+len_pdrr))):
        for _cell in _row:
            fill=PatternFill(start_color='FFFFFF', end_color='FFFFFF',fill_type='solid')
            _cell.fill=fill 
    for _row in sheet.iter_rows(**range_to_index(chr(68+len_pdrr) + str(rowshift-2)+ ':' + chr(68+len_pdrr) + str(rowshift+len_pdrr))):
        for _cell in _row:
            fill=PatternFill(start_color='FFFFFF', end_color='FFFFFF',fill_type='solid')
            _cell.fill=fill 
    
    for _row in sheet.iter_rows(**range_to_index('C' + str(rowshift-1)+ ':' + chr(68+len_pdrr) + str(rowshift+len_pdrr))):
        for _cell in _row:
            _cell.alignment=Alignment(horizontal='center',vertical='center')
            _cell.font=Font(size=9)
 
    #for colNum in range(15):
    for colNum in range(len_pdrr):
         col1=colshift+colNum
         sheet.column_dimensions[get_column_letter(col1)].width = 5   
   
   #Format of percentage/notch table
    for _row in sheet.iter_rows(**range_to_index('V'+ str(rowshift-1) + ':A' + chr(65+2) + str(rowshift+1))): 
       for _cell in _row:
           _cell.font=Font(size=9, bold=True)
           _cell.alignment=Alignment(vertical='center', horizontal='center')
           fill=PatternFill(start_color='FFFFFF', end_color='FFFFFF',fill_type='solid')
           _cell.fill=fill 
           
    #Add borders for percentage/notch table:
    side=Side(style='thin', color="FF000000")
    for _row in sheet.iter_rows(**range_to_index('V' + str(rowshift-1) + ':A' + chr(65+2) + str(rowshift-1))): #top border
       for _cell in _row:
           _cell.border=Border(top=side) 
    for _row in sheet.iter_rows(**range_to_index('V' + str(rowshift) + ':A' + chr(65+2) + str(rowshift+1))): #bottom border
       for _cell in _row:
           _cell.border=Border(bottom=side)
    for _row in sheet.iter_rows(**range_to_index('V' + str(rowshift+1) + ':A' + chr(65+2) + str(rowshift+1))): #another bottom border
       for _cell in _row:
           _cell.border=Border(bottom=side)
    sheet['V'+str(rowshift-1)].border=Border(left=side, right=side, top=side)
    sheet['V'+str(rowshift)].border=Border(left=side, right=side)         
    sheet['V'+str(rowshift+1)].border=Border(left=side, right=side, top=side, bottom=side)
    sheet['Z'+str(rowshift-1)].border=Border(right=side, top=side)
    sheet['Z'+str(rowshift)].border=Border(right=side, bottom=side)       
    sheet['Z'+str(rowshift+1)].border=Border(right=side, bottom=side, top=side)   
    sheet['A'+chr(65+2)+str(rowshift-1)].border=Border(right=side, left=side, top=side)
    sheet['A'+chr(65+2)+str(rowshift)].border=Border(right=side, left=side) 
    sheet['A'+chr(65+2)+str(rowshift+1)].border=Border(right=side, left=side, top=side, bottom=side)
    
    ### Do formatting ###
    for rowNum in PDRR:
        rowNum=rowNum-1
        for colNum in range(len(PDRR)): 
            row1=rowshift+rowNum
            col1=colshift+colNum
            if rowNum==colNum:
                fill=PatternFill(start_color='FFFF00', end_color='FFFF00',fill_type='solid')
                sheet.cell(row=row1,column=col1).fill=fill
            elif ((abs(rowNum-colNum) == 1) | (abs(rowNum-colNum)== 2)):
                fill=PatternFill(start_color='FFFFCC', end_color='FFFFCC',fill_type='solid')
                sheet.cell(row=row1,column=col1).fill=fill
            elif abs(rowNum-colNum) == 3:
                fill=PatternFill(start_color='FCD5B4', end_color='FCD5B4',fill_type='solid')
                sheet.cell(row=row1,column=col1).fill=fill
            elif ((abs(rowNum-colNum) > 3) & (not(sheet.cell(row=row1,column=col1).value==None))): 
                fill=PatternFill(start_color='E4DFEC', end_color='E4DFEC',fill_type='solid')
                sheet.cell(row=row1,column=col1).fill=fill 
            else:
                fill=PatternFill(start_color='FFFFFF', end_color='FFFFFF',fill_type='solid')
                sheet.cell(row=row1,column=col1).fill=fill     
        
    #Title format        
    sheet.merge_cells('B'+str(rowshift-3)+':'+chr(67+len_pdrr)+str(rowshift-3))
    sheet.merge_cells('D'+str(rowshift-2)+ ':'+chr(67+len_pdrr)+str(rowshift-2))
    sheet.merge_cells('B'+str(rowshift)+ ':B'+str(rowshift+len_pdrr-1))
    sheet['B'+str(rowshift)].alignment=Alignment(horizontal='center',text_rotation=90, vertical='center')
    sheet['B'+str(rowshift-3)].alignment=Alignment(horizontal='center', vertical='center')
    sheet['D'+str(rowshift-2)].alignment=Alignment(horizontal='center', vertical='center')
    sheet['B'+str(rowshift-3)].font=Font(bold=True)
    sheet['D'+str(rowshift-2)].font=Font(italic=True)
    sheet['B'+str(rowshift)].font=Font(italic=True)
    
    #Add borders
    for _row in sheet.iter_rows(**range_to_index('B'+ str(rowshift-3) + ':' + chr(67+len_pdrr) + str(rowshift-3))): #top border
       for _cell in _row:
           _cell.border=Border(top=side)
    for _row in sheet.iter_rows(**range_to_index('B'+ str(rowshift) + ':' +chr(67+len_pdrr) + str(rowshift))): #top border 2
       for _cell in _row:
           _cell.border=Border(top=side)
    for i in ['B'+ str(rowshift-2),'B' + str(rowshift-1)]: #left border
        a = sheet[i]
        a.border=Border(left=side)
    for _row in sheet.iter_rows(**range_to_index('B' + str(rowshift) + ':B' + str(rowshift+len_pdrr-1))): #left border
       for _cell in _row:
           _cell.border=Border(left=side, right=side)
    for _row in sheet.iter_rows(**range_to_index('C'+ str(rowshift+len_pdrr-1) + ':' + chr(67+len_pdrr) + str(rowshift+len_pdrr-1))): #bottom border
       for _cell in _row:
           _cell.border=Border(bottom=side)
    for _row in sheet.iter_rows(**range_to_index(chr(67+len_pdrr) + str(rowshift) + ':' + chr(67+len_pdrr) + str(rowshift+len_pdrr-1))): #right border
       for _cell in _row:
           _cell.border=Border(right=side)
    sheet['B'+str(rowshift-3)].border=Border(left=side, top=side)
    sheet['B'+str(rowshift+len_pdrr-1)].border=Border(left=side, bottom=side, right=side)
    sheet[chr(67+len_pdrr)+str(rowshift+len_pdrr-1)].border=Border(bottom=side, right=side)
    sheet[chr(67+len_pdrr)+str(rowshift-1)].border=Border(right=side, bottom=side)
    sheet[chr(67+len_pdrr)+str(rowshift-2)].border=Border(right=side)
    sheet[chr(67+len_pdrr)+str(rowshift-3)].border=Border(right=side, top=side)
    
#    for j in ['D','E','F','G','H','I','J','K','L','M','N','O','P','Q','R']:
#        sheet.column_dimensions[j].width=4   
                               
    for j in range(68, 67+len_pdrr):
        if j <= 90:
            sheet.column_dimensions[chr(j)].width=5
        if j >= 91:
            sheet.column_dimensions['A' + chr(j-26)].width=5
    
    sheet.column_dimensions[chr(68+len_pdrr)].width=6
    sheet.column_dimensions['C'].width=6


##### ##### ##### ##### ##### ##### #####
##### #####   Main Function   ##### #####
##### ##### ##### ##### ##### ##### #####

# This function Creates a sheet to compare the two different rating
# Output: will add a sheet in the output file 
def CreateBenchmarkMatrix(Data, output_filename, sheetname, xaxis, yaxis, PDRR):
     
    #add 'try' on 08/23/2017 to make it work on existing file
    try:
        wb = load_workbook(filename = output_filename)
        sheet = wb.create_sheet() 
        sheet.title = sheetname

    except FileNotFoundError:
        wb =  Workbook()
        sheet = wb.create_sheet() 
        sheet.title = sheetname
        wb.remove_sheet(wb.get_sheet_by_name('Sheet')) 

   
    colshift = 4 #not changeable
    rowshift = 5 #changeable
    len_pdrr = len(PDRR)
    
    #### Create table ### 
    
    #label the vertical axis
    for rowNum in range(len_pdrr):
        row1 = rowNum+rowshift
        sheet.cell(row=row1,column=3).value=PDRR[rowNum]
    
    #label the horizontal axis
    for colNum in range(len_pdrr):
        col1 = colNum+colshift
        sheet.cell(row=rowshift-1,column=col1).value=PDRR[colNum]
    
    sheet.cell(row=rowshift-3,column=colshift-2).value = 'Benchmarking Matrix'
    sheet.cell(row=rowshift-2,column=colshift).value = xaxis
    sheet.cell(row=rowshift,column=colshift-2).value = yaxis

    ### Fill up table ###
    aa = np.zeros((len_pdrr,len_pdrr))
    for rowNum in range(len_pdrr):
        for colNum in range(len_pdrr):
            row1=rowshift+rowNum
            col1=colshift+colNum
            temp = sum((Data[yaxis]==PDRR[rowNum]) & (Data[xaxis]==PDRR[colNum]))
            temp=int(temp)            
            if ((not(temp==0))):        
                sheet.cell(row=row1,column=col1).value= temp
                aa[rowNum,colNum]=temp                  
    
    ### Add list of oustanding obligors
#    sheet.cell(row=rowshift+6,column=27).value='Old Prelim PD'
#    sheet.cell(row=rowshift+5,column=27).value='New Prelim PD'        
#    sheet.cell(row=rowshift+4,column=27).value='Outstanding Obligors'
#    sheet.cell(row=rowshift+4,column=colshift+len_pdrr+2).font=Font(bold=True)
#    list_outstanding=ComputeListOutstandingObligors(Data)
#    for j in range(len(list_outstanding)):
#        sheet.cell(row=rowshift+5+(j % 7),column=colshift+len_pdrr+2+j//7).value=int(list_outstanding[j])  
    
    ### Create percentage Table ###
    sheet.cell(row=rowshift-1,column=len_pdrr+8).value='Match'
    sheet.cell(row=rowshift-1,column=len_pdrr+9).value='Within 1'
    sheet.cell(row=rowshift-1,column=len_pdrr+10).value='Within 2'
    sheet.cell(row=rowshift-1,column=len_pdrr+11).value='Outside 5'
    sheet.cell(row=rowshift-1,column=len_pdrr+12).value='Downgrade'
    sheet.cell(row=rowshift-1,column=len_pdrr+13).value='Upgrade'
    sheet.cell(row=rowshift-1,column=len_pdrr+14).value='Total'
    sheet.cell(row=rowshift,column=len_pdrr+7).value='Percentage'
    sheet.cell(row=rowshift+1,column=len_pdrr+7).value='Notch'
    sheet.cell(row=rowshift,column=len_pdrr+8).value=within(aa,0)/np.sum(aa) # Match
    sheet.cell(row=rowshift,column=len_pdrr+8).number_format='.0%'
    sheet.cell(row=rowshift,column=len_pdrr+9).value=within(aa,1)/np.sum(aa) # Within 1
    sheet.cell(row=rowshift,column=len_pdrr+9).number_format='.0%'
    sheet.cell(row=rowshift,column=len_pdrr+10).value=within(aa,2)/np.sum(aa) # Within 2
    sheet.cell(row=rowshift,column=len_pdrr+10).number_format='.0%'
    sheet.cell(row=rowshift,column=len_pdrr+11).value=1-within(aa,4)/np.sum(aa) # Outside 5
    sheet.cell(row=rowshift,column=len_pdrr+11).number_format='.0%'
    sheet.cell(row=rowshift,column=len_pdrr+12).value=worse_ratings(aa)/np.sum(aa) # Downgrade
    sheet.cell(row=rowshift,column=len_pdrr+12).number_format='.0%'
    down_measure = down_notch(aa)
    sheet.cell(row=rowshift+1,column=len_pdrr+12).value=down_measure
    sheet.cell(row=rowshift,column=len_pdrr+13).value=better_ratings(aa)/np.sum(aa) # Upgrade
    sheet.cell(row=rowshift,column=len_pdrr+13).number_format='.0%'
    up_measure = up_notch(aa)
    sheet.cell(row=rowshift+1,column=len_pdrr+13).value=up_measure
    sheet.cell(row=rowshift,column=len_pdrr+14).value=np.sum(aa)
    sheet.cell(row=rowshift+1,column=len_pdrr+14).value=down_measure + up_measure
    
    CreateStyleSheet_TransitionMatrix(sheet, rowshift, colshift, PDRR)
    
    ### Add total per rows and columns
    sheet.cell(row=rowshift-1,column=colshift+len_pdrr).value = 'Total'
    for i in range(len_pdrr):
        row1=rowshift+i
        sheet.cell(row=row1,column=colshift+len_pdrr).value = sum(aa[i,:])
    sheet.cell(row=rowshift+len_pdrr,column=colshift-1).value='Total'
    for i in range(len_pdrr):
        col1=colshift+i
        sheet.cell(row=rowshift+len_pdrr,column=col1).value = sum(aa[:,i])
    sheet.cell(row=rowshift+len_pdrr,column=colshift+len_pdrr).value = np.sum(aa)    
        
    wb.save(output_filename)


def TMnotches(Data, xaxis, yaxis, PDRR):
     
    len_pdrr = len(PDRR)
    
    ### Fill up table ###
    aa = np.zeros((len_pdrr,len_pdrr))
    for rowNum in range(len_pdrr):
        for colNum in range(len_pdrr):
            temp = sum((Data[yaxis]==PDRR[rowNum]) & (Data[xaxis]==PDRR[colNum]))
            temp=int(temp)            
            if ((not(temp==0))):        
                aa[rowNum,colNum]=temp                  
    
    return((down_notch(aa),up_notch(aa)))


def withknotchesrate(Data, xaxis, yaxis, PDRR, k=2):
     
    len_pdrr = len(PDRR)
    
    ### Fill up table ###
    aa = np.zeros((len_pdrr,len_pdrr))
    for rowNum in range(len_pdrr):
        for colNum in range(len_pdrr):
            temp = sum((Data[yaxis]==PDRR[rowNum]) & (Data[xaxis]==PDRR[colNum]))
            temp=int(temp)            
            if ((not(temp==0))):        
                aa[rowNum,colNum]=temp                  
    
    return(within(aa,k)/np.sum(aa))


def TMstats(Data, xaxis, yaxis, PDRR):
     
    len_pdrr = len(PDRR)
    
    ### Fill up table ###
    aa = np.zeros((len_pdrr,len_pdrr))
    for rowNum in range(len_pdrr):
        for colNum in range(len_pdrr):
            temp = sum((Data[yaxis]==PDRR[rowNum]) & (Data[xaxis]==PDRR[colNum]))
            temp=int(temp)            
            if ((not(temp==0))):        
                aa[rowNum,colNum]=temp                  
    
    stats = {
    'Match':within(aa,0)/np.sum(aa),
    'Within_1':within(aa,1)/np.sum(aa),
    'Within_2':within(aa,2)/np.sum(aa),
    'Outside_5':1-within(aa,4)/np.sum(aa),
    'Downgrade':worse_ratings(aa)/np.sum(aa),
    'Upgrade':better_ratings(aa)/np.sum(aa),
    #'CC_3': aa[2,:].sum(),
    #'CC_3-4': aa[2:4,:].sum(),
    #'CC_3-5': aa[2:5,:].sum(),
    #'CC_3-6': aa[2:6,:].sum(),
    }
    
    return(stats)
    # -*- coding: utf-8 -*-
"""
Created on Tue Apr  4 10:21:43 2017

@author: ub64283 Liwen Zhang (original author: XU79799)
@author: ub71894 (4e8e6d0b), Add TMnotches(), withknotchesrate() and TMstats()
"""
import pandas as pd
import numpy as np
import openpyxl
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.styles import Border, Side, Alignment, Font
from openpyxl.formatting.rule import IconSetRule
from openpyxl import  Workbook, load_workbook
import scipy.stats as stats

#Count the number of obligors which have same rating with a margin of y    
def within(MM,y):
    """
    Summary:
    Counts the number of obligors which have the same rating within a margin of y

    Variables:
    MM: type = DataFrame. A dataframe which contains the transition matrix of two
    ratings systems.

    y: type = int. The margin of ratings that will contribute to the final sum.

    Returns:
    An integer counting the number of obligors with the same rating within a margin of y
    """
    x=np.array(MM)
    N=len(x)
    sum1=0
    if y == 0:
        return np.sum(np.diag(x))
    else:
        for i in range(N):
            ll=max(0,i-y) #checks the lower limit of the margin
            uu=min(N+1,i+y+1) #checks the upper limit of the margin
            sum1=sum1+np.sum(x[ll:uu,i]) #for each cell, adds the number that falls within the margin
        return(sum1)

def down_notch(MM):
    N=MM.shape[0]
    sum1 = 0
    for notch in range(1,N):
        temp = 0
        for i in range(notch):
            temp += MM[i,(N-notch+i)]
        sum1 += temp*(N-notch)
    return(sum1)
        
def worse_ratings(MM):
    x=np.array(MM)
    N=len(x)
    sum1=0
    for i in range(N):
        ll=0
        sum1=sum1+np.sum(x[ll:i,i])
    return(sum1)

def up_notch(MM):
    N=MM.shape[0]
    sum1=0
    for notch in range(1,N):
        temp=0
        for i in range(notch):
            temp += MM[(N-notch+i),i]
        sum1 += temp*(N-notch)
    return(sum1)
    
def better_ratings(MM):
    x=np.array(MM)
    N=len(x)
    sum1=0
    for i in range(N):
        uu=N+1
        sum1=sum1+np.sum(x[(i+1):uu,i])
    return(sum1)

def ComputeListOutstandingObligors(df):
    res=[]
    
#    for i in range(df.shape[0]):
#       if (abs(df['Old_Ratings'].iloc[i]-df['New_Ratings_oldms'].iloc[i])>=4):
#           res.append(df['CUSTOMERID'].iloc[i])
    return res

def CreateStyleSheet_TransitionMatrix(sheet, rowshift, colshift, PDRR):
    
    ### assume len(PDRR) <= 21 ###
    
    len_pdrr = len(PDRR)
    
    #### Work on style ###
    for i in range(rowshift-3,rowshift+len_pdrr):
        sheet.row_dimensions[i].height=12
    for j in ['AB','AC','AD']:
        sheet.column_dimensions[j].width=7
    sheet.column_dimensions['B'].width=3
    sheet.column_dimensions['AA'].width=9
    sheet.column_dimensions['AG'].width=9
    sheet.column_dimensions['AH'].width=5
    sheet.column_dimensions['AE'].width=8
    sheet.column_dimensions['AF'].width=8

    for _row in sheet.iter_rows('B'+ str(rowshift-2) + ':' + chr(68+len_pdrr) + str(rowshift-1)):
        for _cell in _row:
            fill=PatternFill(start_color='FFFFFF', end_color='FFFFFF',fill_type='solid')
            _cell.fill=fill 
    for _row in sheet.iter_rows('B'+ str(rowshift-2) + ':C' + str(rowshift+len_pdrr)):
        for _cell in _row:
            fill=PatternFill(start_color='FFFFFF', end_color='FFFFFF',fill_type='solid')
            _cell.fill=fill 
    for _row in sheet.iter_rows(chr(68+len_pdrr) + str(rowshift-2)+ ':' + chr(68+len_pdrr) + str(rowshift+len_pdrr)):
        for _cell in _row:
            fill=PatternFill(start_color='FFFFFF', end_color='FFFFFF',fill_type='solid')
            _cell.fill=fill 
    
    for _row in sheet.iter_rows('C' + str(rowshift-1)+ ':' + chr(68+len_pdrr) + str(rowshift+len_pdrr)):
        for _cell in _row:
            _cell.alignment=Alignment(horizontal='center',vertical='center')
            _cell.font=Font(size=9)
 
    #for colNum in range(15):
    for colNum in range(len_pdrr):
         col1=colshift+colNum
         sheet.column_dimensions[get_column_letter(col1)].width = 5   
   
   #Format of percentage/notch table
    for _row in sheet.iter_rows('V'+ str(rowshift-1) + ':A' + chr(65+2) + str(rowshift+1)): 
       for _cell in _row:
           _cell.font=Font(size=9, bold=True)
           _cell.alignment=Alignment(vertical='center', horizontal='center')
           fill=PatternFill(start_color='FFFFFF', end_color='FFFFFF',fill_type='solid')
           _cell.fill=fill 
           
    #Add borders for percentage/notch table:
    side=Side(style='thin', color="FF000000")
    for _row in sheet.iter_rows('V' + str(rowshift-1) + ':A' + chr(65+2) + str(rowshift-1)): #top border
       for _cell in _row:
           _cell.border=Border(top=side) 
    for _row in sheet.iter_rows('V' + str(rowshift) + ':A' + chr(65+2) + str(rowshift+1)): #bottom border
       for _cell in _row:
           _cell.border=Border(bottom=side)
    for _row in sheet.iter_rows('V' + str(rowshift+1) + ':A' + chr(65+2) + str(rowshift+1)): #another bottom border
       for _cell in _row:
           _cell.border=Border(bottom=side)
    sheet['V'+str(rowshift-1)].border=Border(left=side, right=side, top=side)
    sheet['V'+str(rowshift)].border=Border(left=side, right=side)         
    sheet['V'+str(rowshift+1)].border=Border(left=side, right=side, top=side, bottom=side)
    sheet['Z'+str(rowshift-1)].border=Border(right=side, top=side)
    sheet['Z'+str(rowshift)].border=Border(right=side, bottom=side)       
    sheet['Z'+str(rowshift+1)].border=Border(right=side, bottom=side, top=side)   
    sheet['A'+chr(65+2)+str(rowshift-1)].border=Border(right=side, left=side, top=side)
    sheet['A'+chr(65+2)+str(rowshift)].border=Border(right=side, left=side) 
    sheet['A'+chr(65+2)+str(rowshift+1)].border=Border(right=side, left=side, top=side, bottom=side)
    
    ### Do formatting ###
    for rowNum in PDRR:
        rowNum=rowNum-1
        for colNum in range(len(PDRR)): 
            row1=rowshift+rowNum
            col1=colshift+colNum
            if rowNum==colNum:
                fill=PatternFill(start_color='FFFF00', end_color='FFFF00',fill_type='solid')
                sheet.cell(row=row1,column=col1).fill=fill
            elif ((abs(rowNum-colNum) == 1) | (abs(rowNum-colNum)== 2)):
                fill=PatternFill(start_color='FFFFCC', end_color='FFFFCC',fill_type='solid')
                sheet.cell(row=row1,column=col1).fill=fill
            elif abs(rowNum-colNum) == 3:
                fill=PatternFill(start_color='FCD5B4', end_color='FCD5B4',fill_type='solid')
                sheet.cell(row=row1,column=col1).fill=fill
            elif ((abs(rowNum-colNum) > 3) & (not(sheet.cell(row=row1,column=col1).value==None))): 
                fill=PatternFill(start_color='E4DFEC', end_color='E4DFEC',fill_type='solid')
                sheet.cell(row=row1,column=col1).fill=fill 
            else:
                fill=PatternFill(start_color='FFFFFF', end_color='FFFFFF',fill_type='solid')
                sheet.cell(row=row1,column=col1).fill=fill     
        
    #Title format        
    sheet.merge_cells('B'+str(rowshift-3)+':'+chr(67+len_pdrr)+str(rowshift-3))
    sheet.merge_cells('D'+str(rowshift-2)+ ':'+chr(67+len_pdrr)+str(rowshift-2))
    sheet.merge_cells('B'+str(rowshift)+ ':B'+str(rowshift+len_pdrr-1))
    sheet['B'+str(rowshift)].alignment=Alignment(horizontal='center',text_rotation=90, vertical='center')
    sheet['B'+str(rowshift-3)].alignment=Alignment(horizontal='center', vertical='center')
    sheet['D'+str(rowshift-2)].alignment=Alignment(horizontal='center', vertical='center')
    sheet['B'+str(rowshift-3)].font=Font(bold=True)
    sheet['D'+str(rowshift-2)].font=Font(italic=True)
    sheet['B'+str(rowshift)].font=Font(italic=True)
    
    #Add borders
    for _row in sheet.iter_rows('B'+ str(rowshift-3) + ':' + chr(67+len_pdrr) + str(rowshift-3)): #top border
       for _cell in _row:
           _cell.border=Border(top=side)
    for _row in sheet.iter_rows('B'+ str(rowshift) + ':' +chr(67+len_pdrr) + str(rowshift)): #top border 2
       for _cell in _row:
           _cell.border=Border(top=side)
    for i in ['B'+ str(rowshift-2),'B' + str(rowshift-1)]: #left border
        a = sheet[i]
        a.border=Border(left=side)
    for _row in sheet.iter_rows('B' + str(rowshift) + ':B' + str(rowshift+len_pdrr-1)): #left border
       for _cell in _row:
           _cell.border=Border(left=side, right=side)
    for _row in sheet.iter_rows('C'+ str(rowshift+len_pdrr-1) + ':' + chr(67+len_pdrr) + str(rowshift+len_pdrr-1)): #bottom border
       for _cell in _row:
           _cell.border=Border(bottom=side)
    for _row in sheet.iter_rows(chr(67+len_pdrr) + str(rowshift) + ':' + chr(67+len_pdrr) + str(rowshift+len_pdrr-1)): #right border
       for _cell in _row:
           _cell.border=Border(right=side)
    sheet['B'+str(rowshift-3)].border=Border(left=side, top=side)
    sheet['B'+str(rowshift+len_pdrr-1)].border=Border(left=side, bottom=side, right=side)
    sheet[chr(67+len_pdrr)+str(rowshift+len_pdrr-1)].border=Border(bottom=side, right=side)
    sheet[chr(67+len_pdrr)+str(rowshift-1)].border=Border(right=side, bottom=side)
    sheet[chr(67+len_pdrr)+str(rowshift-2)].border=Border(right=side)
    sheet[chr(67+len_pdrr)+str(rowshift-3)].border=Border(right=side, top=side)
    
#    for j in ['D','E','F','G','H','I','J','K','L','M','N','O','P','Q','R']:
#        sheet.column_dimensions[j].width=4   
                               
    for j in range(68, 67+len_pdrr):
        if j <= 90:
            sheet.column_dimensions[chr(j)].width=5
        if j >= 91:
            sheet.column_dimensions['A' + chr(j-26)].width=5
    
    sheet.column_dimensions[chr(68+len_pdrr)].width=6
    sheet.column_dimensions['C'].width=6


##### ##### ##### ##### ##### ##### #####
##### #####   Main Function   ##### #####
##### ##### ##### ##### ##### ##### #####

# This function Creates a sheet to compare the two different rating
# Output: will add a sheet in the output file 
def CreateBenchmarkMatrix(Data, output_filename, sheetname, xaxis, yaxis, PDRR):
     
    #add 'try' on 08/23/2017 to make it work on existing file
    try:
        wb = load_workbook(filename = output_filename)
        sheet = wb.create_sheet() 
        sheet.title = sheetname

    except FileNotFoundError:
        wb =  Workbook()
        sheet = wb.create_sheet() 
        sheet.title = sheetname
        wb.remove_sheet(wb.get_sheet_by_name('Sheet')) 

   
    colshift = 4 #not changeable
    rowshift = 5 #changeable
    len_pdrr = len(PDRR)
    
    #### Create table ### 
    
    #label the vertical axis
    for rowNum in range(len_pdrr):
        row1 = rowNum+rowshift
        sheet.cell(row=row1,column=3).value=PDRR[rowNum]
    
    #label the horizontal axis
    for colNum in range(len_pdrr):
        col1 = colNum+colshift
        sheet.cell(row=rowshift-1,column=col1).value=PDRR[colNum]
    
    sheet.cell(row=rowshift-3,column=colshift-2).value = 'Benchmarking Matrix'
    sheet.cell(row=rowshift-2,column=colshift).value = xaxis
    sheet.cell(row=rowshift,column=colshift-2).value = yaxis

    ### Fill up table ###
    aa = np.zeros((len_pdrr,len_pdrr))
    for rowNum in range(len_pdrr):
        for colNum in range(len_pdrr):
            row1=rowshift+rowNum
            col1=colshift+colNum
            temp = sum((Data[yaxis]==PDRR[rowNum]) & (Data[xaxis]==PDRR[colNum]))
            temp=int(temp)            
            if ((not(temp==0))):        
                sheet.cell(row=row1,column=col1).value= temp
                aa[rowNum,colNum]=temp                  
    
    ### Add list of oustanding obligors
#    sheet.cell(row=rowshift+6,column=27).value='Old Prelim PD'
#    sheet.cell(row=rowshift+5,column=27).value='New Prelim PD'        
#    sheet.cell(row=rowshift+4,column=27).value='Outstanding Obligors'
#    sheet.cell(row=rowshift+4,column=colshift+len_pdrr+2).font=Font(bold=True)
#    list_outstanding=ComputeListOutstandingObligors(Data)
#    for j in range(len(list_outstanding)):
#        sheet.cell(row=rowshift+5+(j % 7),column=colshift+len_pdrr+2+j//7).value=int(list_outstanding[j])  
    
    ### Create percentage Table ###
    sheet.cell(row=rowshift-1,column=len_pdrr+8).value='Match'
    sheet.cell(row=rowshift-1,column=len_pdrr+9).value='Within 1'
    sheet.cell(row=rowshift-1,column=len_pdrr+10).value='Within 2'
    sheet.cell(row=rowshift-1,column=len_pdrr+11).value='Outside 5'
    sheet.cell(row=rowshift-1,column=len_pdrr+12).value='Downgrade'
    sheet.cell(row=rowshift-1,column=len_pdrr+13).value='Upgrade'
    sheet.cell(row=rowshift-1,column=len_pdrr+14).value='Total'
    sheet.cell(row=rowshift,column=len_pdrr+7).value='Percentage'
    sheet.cell(row=rowshift+1,column=len_pdrr+7).value='Notch'
    sheet.cell(row=rowshift,column=len_pdrr+8).value=within(aa,0)/np.sum(aa) # Match
    sheet.cell(row=rowshift,column=len_pdrr+8).number_format='.0%'
    sheet.cell(row=rowshift,column=len_pdrr+9).value=within(aa,1)/np.sum(aa) # Within 1
    sheet.cell(row=rowshift,column=len_pdrr+9).number_format='.0%'
    sheet.cell(row=rowshift,column=len_pdrr+10).value=within(aa,2)/np.sum(aa) # Within 2
    sheet.cell(row=rowshift,column=len_pdrr+10).number_format='.0%'
    sheet.cell(row=rowshift,column=len_pdrr+11).value=1-within(aa,4)/np.sum(aa) # Outside 5
    sheet.cell(row=rowshift,column=len_pdrr+11).number_format='.0%'
    sheet.cell(row=rowshift,column=len_pdrr+12).value=worse_ratings(aa)/np.sum(aa) # Downgrade
    sheet.cell(row=rowshift,column=len_pdrr+12).number_format='.0%'
    down_measure = down_notch(aa)
    sheet.cell(row=rowshift+1,column=len_pdrr+12).value=down_measure
    sheet.cell(row=rowshift,column=len_pdrr+13).value=better_ratings(aa)/np.sum(aa) # Upgrade
    sheet.cell(row=rowshift,column=len_pdrr+13).number_format='.0%'
    up_measure = up_notch(aa)
    sheet.cell(row=rowshift+1,column=len_pdrr+13).value=up_measure
    sheet.cell(row=rowshift,column=len_pdrr+14).value=np.sum(aa)
    sheet.cell(row=rowshift+1,column=len_pdrr+14).value=down_measure + up_measure
    
    CreateStyleSheet_TransitionMatrix(sheet, rowshift, colshift, PDRR)
    
    ### Add total per rows and columns
    sheet.cell(row=rowshift-1,column=colshift+len_pdrr).value = 'Total'
    for i in range(len_pdrr):
        row1=rowshift+i
        sheet.cell(row=row1,column=colshift+len_pdrr).value = sum(aa[i,:])
    sheet.cell(row=rowshift+len_pdrr,column=colshift-1).value='Total'
    for i in range(len_pdrr):
        col1=colshift+i
        sheet.cell(row=rowshift+len_pdrr,column=col1).value = sum(aa[:,i])
    sheet.cell(row=rowshift+len_pdrr,column=colshift+len_pdrr).value = np.sum(aa)    
        
    wb.save(output_filename)


def TMnotches(Data, xaxis, yaxis, PDRR):
     
    len_pdrr = len(PDRR)
    
    ### Fill up table ###
    aa = np.zeros((len_pdrr,len_pdrr))
    for rowNum in range(len_pdrr):
        for colNum in range(len_pdrr):
            temp = sum((Data[yaxis]==PDRR[rowNum]) & (Data[xaxis]==PDRR[colNum]))
            temp=int(temp)            
            if ((not(temp==0))):        
                aa[rowNum,colNum]=temp                  
    
    return((down_notch(aa),up_notch(aa)))


def withknotchesrate(Data, xaxis, yaxis, PDRR, k=2):
     
    len_pdrr = len(PDRR)
    
    ### Fill up table ###
    aa = np.zeros((len_pdrr,len_pdrr))
    for rowNum in range(len_pdrr):
        for colNum in range(len_pdrr):
            temp = sum((Data[yaxis]==PDRR[rowNum]) & (Data[xaxis]==PDRR[colNum]))
            temp=int(temp)            
            if ((not(temp==0))):        
                aa[rowNum,colNum]=temp                  
    
    return(within(aa,k)/np.sum(aa))


def TMstats(Data, xaxis, yaxis, PDRR):
     
    len_pdrr = len(PDRR)
    
    ### Fill up table ###
    aa = np.zeros((len_pdrr,len_pdrr))
    for rowNum in range(len_pdrr):
        for colNum in range(len_pdrr):
            temp = sum((Data[yaxis]==PDRR[rowNum]) & (Data[xaxis]==PDRR[colNum]))
            temp=int(temp)            
            if ((not(temp==0))):        
                aa[rowNum,colNum]=temp                  
    
    stats = {
    'Match':within(aa,0)/np.sum(aa),
    'Within_1':within(aa,1)/np.sum(aa),
    'Within_2':within(aa,2)/np.sum(aa),
    'Outside_5':1-within(aa,4)/np.sum(aa),
    'Downgrade':worse_ratings(aa)/np.sum(aa),
    'Upgrade':better_ratings(aa)/np.sum(aa),
    'CC_3': aa[2,:].sum(),
    'CC_3-4': aa[2:4,:].sum(),
    'CC_3-5': aa[2:5,:].sum(),
    'CC_3-6': aa[2:6,:].sum(),
    }
    
    return(stats)

    # -*- coding: utf-8 -*-
"""
Created on Mon Mar 20 14:11:42 2017

Version: 1.0: Initial build
Version: 1.1: 20170512, Modify the class to handle missing data. And add new method LinearReg(); Minor change in __SomersD()
Version: 1.2: 20180511, Add new private method __LinReg() and modify method modelselection(): add linear regression as one option.
@author: ub71894 (4e8e6d0b), CSG
"""
import os
import pandas as pd
import numpy as np
import warnings
import matplotlib.pyplot as plt
import statsmodels.api as sm
from sklearn.metrics import roc_auc_score
from itertools import combinations, product
from numba import jit
from PDScorecardTool.Process import SomersD

#%%
class MFA(object):
    
    '''    
    This class is used to realize all calculations in MFA procedure.
    
    MFA(data, model):

        data:       data for development. Make sure it has column 'def_flag' as default
                    indicator and the factors' names are the same as the ones in 'model'

        model:      PDModle class. It saves all parameters for the model
    

        quant_only: boolean, default 'False'. Added in Ver.1.6
                    set as 'True' if only normalize on quant factors.

        missing:    str, default 'median'. Added in Ver.1.1
                    statistics of the data used to fill missing data (NA data)



    Methods:

        describe()

        corr(factors='all', raw=False, output=True)

        VIF(factors='all', fullmodel='all')

        reweight(on='good', multiplier=1, random_state=None)

        recover()

        LogReg(factors='all', summary=False)

        LinearReg(factors='all', dependentvar='Final_PD_Risk_Rating', summary=False)

        modelAR(quant_weight=0, quali_weight=0, quantweight=0.5)
        
        plotAR(quant_weight=0, quali_weight=0, quantweight_range=[0,1], savefig=True)
        
        modelselection(factors, atleast_p=3, best_k=5, by='SomersD')
        
        ARgridsearch(**kw):  
        
    '''

    def __init__(self, data, model, quant_only=False, missing='median'):

        self.model = model
        self.rawdata = data.copy()
        self.rawdata.reset_index(drop=True, inplace=True)
        self.stats = self.__describe()
        self.normdata = self.__normalization(quant_only, missing)
        # save original normdata for later recover
        self.oridata = self.normdata.copy()


    def __describe(self):
        '''

        Privata method that calculate key statistics for quant factors on raw data

        '''
        warnings.filterwarnings('ignore', category=Warning)
        dat = self.rawdata[self.model.quant_factor]
        stats = dat.describe(percentiles=[0.01,0.05,0.1,0.25,0.5,0.75,0.9,0.95,0.99])  
        temp={}
        temp['Skew'] = pd.Series(dat.skew())
        temp['Kurtosis'] = dat.kurtosis()
        stats = pd.concat([stats, pd.DataFrame(temp).T])
        return stats


    def __normalization(self, quant_only, missing):
        '''

        Private method that apply normalization on financial factors.

        '''
        warnings.filterwarnings('ignore', category=Warning)
        # Invalid_Neg
        normdata = self.rawdata.copy()
        for i,neg_source in enumerate(self.model.Invalid_Neg):
            if neg_source:
                col=self.model.quant_factor[i]
                normdata[col][ (normdata[col]<0) & (normdata[neg_source]<0) ] = self.model.cap[i]
                # treat NA in 'neg_source' as negative value
                normdata[col][ (normdata[col]<0) & pd.isnull(normdata[neg_source]) ] = self.model.cap[i]

        # cap/floor for quant factors:
        for i, col in enumerate(self.model.quant_factor):
            normdata[col] = np.clip(normdata[col], self.model.floor[i], self.model.cap[i])        

        # quant factors transformation:
        for i, col in enumerate(self.model.quant_factor):
            if self.model.quant_log[i]:
                normdata[col] = np.log(normdata[col])
            
        # quant factors normalization:
        for i, col in enumerate(self.model.quant_factor):
            normdata[col]  = 50*(normdata[col] - self.model.doc_mean[i]) / self.model.doc_std[i]      
        
        # quant factors flip sign:  
        for i, col in enumerate(self.model.quant_factor):
            normdata[col] = normdata[col] * self.model.quant_multiplier[i]     


        if quant_only:
            exec('normdata[self.model.quant_factor] = normdata[self.model.quant_factor].\
            fillna(normdata[self.model.quant_factor].{missing}())'.format(missing=missing))
        else:
            # calibration for quali factors:
            for i, col in enumerate(self.model.quali_factor):
                normdata[col].replace(to_replace=self.model.qualimapping[i],inplace=True) 
                 
            # fill missing data. added in Ver. 1.1
            exec('normdata[self.model.quant_factor+self.model.quali_factor] = normdata[self.model.quant_factor+self.model.quali_factor].\
                fillna(normdata[self.model.quant_factor+self.model.quali_factor].{missing}())'.format(missing=missing))

        return normdata       


    def __AUROC(self, score, dependentvar, isthere_def=True) :
        '''

        Private method that calculate area under ROC.

        '''
        if isthere_def:
            score = np.array(score)
            y_true  = [int(x) for x in list(self.normdata[dependentvar])] 
            return roc_auc_score(y_true, score) 
        else:
            return 'NA'


    def __SomersD_byAUROC(self, score, dependentvar, isthere_def=True) :
        '''

        Private method that calculate SomersD by 2*AUROC-1.

        '''
        if isthere_def:
            score = np.array(score)
            y_true  = [int(x) for x in list(self.normdata[dependentvar])] 
            return 2*roc_auc_score(y_true, score)-1
        else:
            return 'NA'


    @jit
    def __SomersD(self, score, dependentvar, sign=1, unit_concordance=True): 
        '''
        Private method to calculate SomersD
        New version of SomersD function which leverages numba.jit to accelerate
        the calculation. 

        '''
        x = np.array(score)
        # added in Ver 1.1 to deal with missing PDRR sometimes.
        y_true = self.normdata[dependentvar].copy()
        y_true.dropna(inplace=True)
        y = np.array(y_true)
        n = len(x)
        C = 0
        D = 0
        if unit_concordance==True:
            for i in range(n):
                for j in range(i):
                    dx = x[i] - x[j]
                    dy = y[i] - y[j]
                    if (dx*dy) > 0:
                        cij = 1
                        dij = 1
                    elif  (dx*dy) < 0:
                        cij = -1
                        dij = 1
                    else:
                        cij = 0
                        if dy==0: 
                            dij = 0
                        else: 
                            dij=1
                    C = C + cij
                    D = D + dij
        else:
            for i in range(n):
                for j in range(i):
                    dx = x[i] - x[j]
                    dy = y[i] - y[j]
                    if  (dx*dy) > 0:
                        cij = abs(dy)
                        dij = abs(dy)
                    elif (dx*dy) < 0:
                        cij = -abs(dy)
                        dij = abs(dy)
                    else:
                        cij = 0
                        if dy==0: 
                            dij = 0
                        else: 
                            dij=1
                    C = C + cij
                    D = D + dij    
        
        return sign*C/D


    def describe(self):
        '''

        Public method that present key statistics for quant factors on raw data
        
        '''
        return self.stats


    def corr(self, factors='all', raw=False, output=True):
        '''

        Public method that calculate correlation matrix for multifactors

        Parameters: 

        factors:    str {'quant', 'quali', 'all'} or None, default 'all'
                    Multifactors for matrix calculations. 

        raw:        boolean, default False
                    whether using raw data or normalized data. Default setting is
                    False since usually we care about the corr of normalized factors.

        output:     boolean, default True
                    whether save the corr matrix to excel file.

        '''
        warnings.filterwarnings('ignore', category=Warning)
        if factors=='quant':
            names= self.model.quant_factor
        elif factors=='quali':
            names= self.model.quali_factor
        elif factors=='all':
            names = self.model.quant_factor + self.model.quali_factor
        else: # user input a customized list
            names = factors
            
        
        if raw:
            suffix = '_raw'
            corr = self.rawdata[names].corr()
            if output:
                corr.to_excel('CorrMat_'+factors+suffix+'.xlsx')
            return corr

        else:
            suffix = '_norm'
            corr = self.normdata[names].corr()
            if output:
                corr.to_excel('CorrMat_'+factors+suffix+'.xlsx')
            return corr    


    def VIF(self, factors='all', fullmodel='all'):
        '''

        Public method that calculate VIF for factors

        Parameters: 

        factors:    str {'quant','quali', 'all', factorname}, default 'all'
                    factors for VIF calculation. 

        fullmodel:  str {'quant','quali', 'all'}, default 'all'
                    factors busket for VIF calculation. 

        '''        

        if factors=='quant':
            names= self.model.quant_factor
        elif factors=='quali':
            names= self.model.quali_factor
        elif factors=='all':
            names= self.model.quant_factor + self.model.quali_factor 
        else: # user input a single factor's name
            names = [factors]   
        
        if fullmodel=='quant':
            fullmodelnames= self.model.quant_factor
        elif fullmodel=='quali':
            fullmodelnames= self.model.quali_factor
        else:
            fullmodelnames = self.model.quant_factor + self.model.quali_factor 

        result=[]
        for factor in names:
            ols_model = sm.OLS(endog=np.array(self.normdata[factor]),exog=\
                sm.add_constant(self.normdata[fullmodelnames].drop(factor,axis=1)), missing='drop')
            ols_result = ols_model.fit()
            result.append(1/(1-ols_result.rsquared))

        return pd.DataFrame({'VIF':result}, index=names)


    def reweight(self, on='good', multiplier=1, random_state=None):
        '''

        Public method that performs data 'sampling' for reweighting purpose.
        Parameters: 

        on:             str 'good' or 'bad', default 'good'
                        which event for reweight

        multiplier:     int or float, default 1

        random_state:   int or numpy.random.RandomState
                        Seed for the random number generator (if int), or 
                        numpy RandomState object.
    
        ''' 
        enum   = {'good':0,'bad':1}
        data   = self.normdata.query('def_flag=={on}'.format(on=enum[on])).copy()
        data_c = self.normdata.query('def_flag=={not_on}'.format(not_on=1-enum[on]))
        
        if multiplier<=1: # sample without replacement
            sample = int(multiplier*len(data))
            new = pd.concat([data.sample(n=sample, replace=False, random_state=random_state), data_c], ignore_index=True)   

        else:       # sample without replacement for fraction part, times sample for int part
            int_part = int(multiplier)
            frac_part = multiplier-int_part
            sample = int(frac_part*len(data))
            new = pd.concat([data]*int_part, ignore_index=True)
            if sample !=0:
                new = pd.concat([new, data.sample(n=sample, replace=False, random_state=random_state)], ignore_index=True)
            new = pd.concat([new, data_c], ignore_index=True)
        
        self.normdata = new.copy()


    def recover(self):
        '''

        Public method that recover reweighted normdata back to original normdata
    

        '''
        print('Current "normdata" has {n} rows.\n'.format(n=len(self.normdata)))
        self.normdata = self.oridata.copy()
        print('After recovery, "normdata" is back to {n} rows.\n'.format(n=len(self.normdata)))


    def LogReg(self, factors='all', summary=False):
        '''

        Public method that perform Logistic Regression with intercept

        Parameters: 

        factors:    str {'quant','quali', all'} or a list of strs, default 'all'
                    factors for LR. 

        summary:    boolean, default False
                    whether to print out the summary for LR.
        

        Return:     a tuple has parameters implied weights and corresponding AR


        '''      
        if factors=='quant':
            names = self.model.quant_factor
        elif factors=='quali':
            names = self.model.quali_factor
        elif factors=='all':
            names = self.model.quant_factor + self.model.quali_factor 
        else:
            names = factors


        x = sm.add_constant(self.normdata[names],prepend = True)
        logit = sm.Logit(self.normdata.def_flag, x)
        result = logit.fit(disp=0)
        if summary:
            print(result.summary())

        if factors!='all':
            coeff = result.params[1:]
            weight_coeff = [x/coeff.sum() for x in coeff]
            score = weight_coeff*self.normdata[names] 
            AUROC = roc_auc_score(self.normdata.def_flag,score.sum(axis=1))
            SomersD = 2*AUROC-1

            result1 = pd.DataFrame({'weights':weight_coeff}, index=names)
            result2 = pd.DataFrame({'AR':[AUROC, SomersD]}, index=['AUROC', 'SomersD'])
            return (result1, result2)

        else: # both quant and quali factors
            idx =len(self.model.quant_factor)+1
            coeff_quant = result.params[1:idx]
            coeff_quali = result.params[idx:]
            weight_coeff_quant = [x/coeff_quant.sum() for x in coeff_quant]
            weight_coeff_quali = [x/coeff_quali.sum() for x in coeff_quali]
            weight_quantmodule = coeff_quant.sum() / (coeff_quant.sum() + coeff_quali.sum())
            AUROC = roc_auc_score(self.normdata.def_flag, result.fittedvalues)
            SomersD = 2*AUROC-1

            result1 = pd.DataFrame({'weights':weight_coeff_quant+weight_coeff_quali}, index=names)
            result2 = pd.DataFrame({'AR':[AUROC, SomersD]}, index=['AUROC', 'SomersD'])
            result3 = pd.DataFrame({'quantmoduleweight':weight_quantmodule}, index=['byLogisticRegresion'])
            return (result1, result2, result3)


    def LinearReg(self, factors='all', dependentvar='Final_PD_Risk_Rating', summary=False):
        '''

        Public method that perform Logistic Regression with intercept

        Parameters: 

        factors:        str {'quant','quali', all'} or a list of strs, default 'all'
                        factors for LR. 

        dependentvar:   str, default 'Final_PD_Risk_Rating' 
                        column name of y which is dependent variable in the regression

        summary:        boolean, default False
                        whether to print out the summary for LR.
        

        Return:     a tuple has parameters implied weights and corresponding AR


        ''' 
        
        if factors=='quant':
            names = self.model.quant_factor
        elif factors=='quali':
            names = self.model.quali_factor
        elif factors=='all':
            names = self.model.quant_factor + self.model.quali_factor 
        else:
            names = factors

        #dataforLR = self.normdata.copy()
        #dataforLR.dropna(subset=[names+[dependentvar]], how='any', inplace=True)

        #x = sm.add_constant(dataforLR[names],prepend = True)
        #OLS = sm.OLS(dataforLR[dependentvar], x)
        x = sm.add_constant(self.normdata[names],prepend = True)
        OLS = sm.OLS(self.normdata[dependentvar], x, missing='drop')
        result = OLS.fit(disp=0)
        if summary:
            print(result.summary())

        if factors!='all':
            coeff = result.params[1:]
            weight_coeff = [x/coeff.sum() for x in coeff]
            SomersD = self.__SomersD(result.fittedvalues, dependentvar)

            result1 = pd.DataFrame({'weights':weight_coeff}, index=names)
            result2 = pd.DataFrame({'AR':[SomersD]}, index=['SomersD'])
            return (result1, result2)

        else: # both quant and quali factors
            idx =len(self.model.quant_factor)+1
            coeff_quant = result.params[1:idx]
            coeff_quali = result.params[idx:]
            weight_coeff_quant = [x/coeff_quant.sum() for x in coeff_quant]
            weight_coeff_quali = [x/coeff_quali.sum() for x in coeff_quali]
            weight_quantmodule = coeff_quant.sum() / (coeff_quant.sum() + coeff_quali.sum())
            SomersD = self.__SomersD(result.fittedvalues, dependentvar)

            result1 = pd.DataFrame({'weights':weight_coeff_quant+weight_coeff_quali}, index=names)
            result2 = pd.DataFrame({'AR':[SomersD]}, index=['SomersD'])
            result3 = pd.DataFrame({'quantmoduleweight':weight_quantmodule}, index=['byLogisticRegresion'])
            return (result1, result2, result3)

    
    def modelAR(self, quant_weight=0, quali_weight=0, quantweight=0.5, isthere_def=True, dependentvar='def_flag', use_msms=False, update_msms=False):
        '''

        Public method that calculate model's AR according to input model setting

        Parameters: 

        quant_weight:   list, default 0 
                        list of quant factors' weights

        quali_weight:   list, default 0 
                        list of quali factors' weights

        quantweight:    float, default 0.5
                        the weight of quant module in total score.

        isthere_def:    boolean, default True
                        whether the analysis is on the default event. If not, for 
                        example we want to check the accuracy of factor on PDRR,
                        then we should input  
                        isthere_def=False, dependentvar='PDRR'

        dependentvar:   str, default 'def_flag'
                        column name for dependent variable.          

        use_msms:       boolean, default False
                        'msms' means module statistics which saved in model setting
                        whether use module mean and std in model setting.
                        To use the mean and std of quant and quali module in class 'model' if True.
                        To use the mean and std of quant and quali module that calibrated from the data if False.
        update_msms:    boolean, default False
                        update model.quantmean, model.qualimean, model.quantstd and model.qualistd with calibrated value from data
        Return:   AR in a Dataframe 
        '''      
        if quant_weight and quali_weight: # means all factors

            temp_quant = quant_weight*self.normdata[self.model.quant_factor]
            temp_quali = quali_weight*self.normdata[self.model.quali_factor]

            if use_msms:
                quantmean =  self.model.quantmean
                quantstd =  self.model.quantstd
                qualimean =  self.model.qualimean
                qualistd =  self.model.qualistd
            else:
                quantmean =  temp_quant.sum(axis=1).mean()
                quantstd =  temp_quant.sum(axis=1).std()
                qualimean =  temp_quali.sum(axis=1).mean()
                qualistd =  temp_quali.sum(axis=1).std()

            if update_msms:
                self.model.quantmean = quantmean
                self.model.quantstd = quantstd
                self.model.qualimean = qualimean
                self.model.qualistd = qualistd
                self.model.quantweight = quantweight
                self.model.qualiweight = 1 - quantweight


            score = quantweight*50*((temp_quant).sum(axis=1)-quantmean) / quantstd+\
                    (1-quantweight)*50*((temp_quali).sum(axis=1)-qualimean) / qualistd

            AUROC = self.__AUROC(score, dependentvar, isthere_def)
            if isthere_def:
                SomersD = 2*AUROC-1
            else:
                SomersD = self.__SomersD(score, dependentvar)
            return(pd.DataFrame({'AR':[AUROC, SomersD]}, index=['AUROC', 'SomersD']))

        elif quant_weight:
            score =  (quant_weight*self.normdata[self.model.quant_factor]).sum(axis=1)
            AUROC = self.__AUROC(score, dependentvar, isthere_def)
            if isthere_def:
                SomersD = 2*AUROC-1
            else:
                SomersD = self.__SomersD(score, dependentvar)
            return(pd.DataFrame({'AR':[AUROC, SomersD]}, index=['AUROC', 'SomersD'])) 

        elif quali_weight:
            score =  (quali_weight*self.normdata[self.model.quali_factor]).sum(axis=1)
            AUROC = self.__AUROC(score, dependentvar, isthere_def)
            if isthere_def:
                SomersD = 2*AUROC-1
            else:
                SomersD = self.__SomersD(score, dependentvar)
            return(pd.DataFrame({'AR':[AUROC, SomersD]}, index=['AUROC', 'SomersD']))   
        
        else:
            return 0


    def plotAR(self, quant_weight=0, quali_weight=0, quantweight_range=[0,1], isthere_def=True, dependentvar='def_flag', savefig=True): 
        '''

        Public method that draws the relationship between model's AR and quantweight

        Parameters: 

        quant_weight:       list, default 0 
                            list of quant factors' weights

        quali_weight:       list, default 0 
                            list of quali factors' weights

        quantweight_range:  list, default [0,1]
                            starting and ending points of quantweight in this plot  

        isthere_def:        boolean, default True
                            whether the analysis is on the default event. If not, for 
                            example we want to check the accuracy of factor on PDRR,
                            then we should input  
                            isthere_def=False, dependentvar='PDRR'

        dependentvar:       str, default 'def_flag'
                            column name for dependent variable. 

        savefig:            boolean, default True
                            whether save png file in current working directory. 
        
        '''      
        weightlist=[];  ARlist=[]
        weight =  quantweight_range[0]  
        while weight<=quantweight_range[1]:
            weightlist.append(weight)
            ARlist.append(self.modelAR(quant_weight, quali_weight, quantweight=weight,\
                isthere_def=isthere_def, dependentvar=dependentvar).loc['SomersD','AR'])
            weight += 0.01  
        AR_max = max(ARlist)
        weigh_max = weightlist[ARlist.index(max(ARlist))]
        
        # plot
        f,ax = plt.subplots(1,1,figsize=(8,8))  
        plt.plot(weightlist, ARlist, 'b')
        ax.annotate('Max={ar:.3f} @ weight={wt:.2f}'.format(ar=AR_max, wt=weigh_max), \
            xy=(weigh_max, AR_max), xytext=(weigh_max-0.1, AR_max+0.002),\
            arrowprops=dict(facecolor='green', shrink=0.05))
        ax.set_ylim([min(ARlist),max(ARlist)+0.01])
        plt.xlabel('Weight of Quant')
        plt.ylabel('AccuracyRate by SomersD')

        if savefig:
            s=1;
            for file in os.listdir('.'):
                if file.startswith("ARplot"):
                    s+=1
            f.savefig('ARplot_{suffix}.png'.format(suffix=s))

        return(f)


    def __LogReg(self, names):
        '''
        Privata method that run LR with fixed setting in factors.
        '''
        x = sm.add_constant(self.normdata[names],prepend = True)
        logit = sm.Logit(self.normdata.def_flag, x)
        result = logit.fit(disp=0)
        BIC = result.bic
        AUROC = roc_auc_score(self.normdata.def_flag, result.fittedvalues)
        SomersD = 2*AUROC-1
        return (pd.DataFrame({'Measure':[SomersD, BIC]}, index=['SomersD','BIC']))

    def __LinReg(self, names, dependentvar):
        '''
        Privata method that run LR with fixed setting in factors.
        '''
        x = sm.add_constant(self.normdata[names],prepend = True)
        linear = sm.OLS(self.normdata[dependentvar], x, missing='drop')
        result = linear.fit(disp=0)
        R2 = result.rsquared
        SD = SomersD(self.normdata[dependentvar], result.fittedvalues)
        
        return (pd.DataFrame({'Measure':[SD, R2]}, index=['SomersD','R-squared']))

    def modelselection(self, factors, atleast_p=3, best_k=5, by='SomersD', regressor='logistic', dependentvar=None):
        '''

        Public method that performs model selection by Logistic Regression. 
        According to 'SomersD' or 'BIC', it returns the best k model settings.

        Parameters: 

        factors:        list of strings
                        full list of factor candidates

        atleast_p:      int, default 3
                        minimal requirement for the number of factors in the model.

        best_k:         int, default 5
                        how many best settings are kept.

        by:             str {'SomersD' or 'BIC'}, default 'SomersD'
                        the measure used as criterion in model selection.

        regressor:      str {'logistic' or 'linear'}. Add in Ver. 1.2
                        the regression model that user chooses

        dependentvar:   str, the name for dependent variable in linear regression.
                        Add in Ver. 1.2

        Return:         best_k settings and measure in a Dataframe 
        ''' 
        if regressor=='logistic':

            enum = {'SomersD':0, 'BIC':1}
            idx = enum.pop(by, 0)   

            setting=[];    measure=[]
            for p in range(atleast_p, len(factors)+1):
                for names in combinations(factors, p):
                    setting.append(names)
                    measure.append(self.__LogReg(list(names)).iloc[idx,0])
            result = pd.DataFrame({'Setting':setting,'Measure':measure}, columns=['Setting','Measure'])
            result.sort_values(by='Measure', ascending=idx, inplace=True)
            return (result.reset_index(drop=True).iloc[:best_k,])

        elif regressor=='linear':

            enum = {'SomersD':0, 'R-squared':1}
            idx = enum.pop(by, 0)   

            setting=[];    measure=[]
            for p in range(atleast_p, len(factors)+1):
                for names in combinations(factors, p):
                    setting.append(names)
                    measure.append(self.__LinReg(list(names), dependentvar).iloc[idx,0])
            result = pd.DataFrame({'Setting':setting,'Measure':measure}, columns=['Setting','Measure'])
            result.sort_values(by='Measure', ascending=idx, inplace=True)
            return (result.reset_index(drop=True).iloc[:best_k,])
        else:
            return (0)


    def ARgridsearch(self, **kw):   
        '''

        Public method that performs Grid Search for factors' best weights according to
        Somers'D.

        Parameters: 

        quant_factor:       list of strings(quant factors's names)
                            ex. model.quant_factor or ['quant1','quant2']

        quant_weight_range: list of tuples that indicate the beginning and ending 
                            points in searching for quant factors. 
                            ex. [(0.25,0.35),(0.1,0.2),(0.2,0.3),(0.1,0.2),(0.25,0.35)]

        quali_factor:       list of strings(quali factors's names)
                            ex. model.quali_factor or ['quali1','quali2']

        quali_weight_range: list of tuples that indicate the beginning and ending 
                            points in searching for quali factors. 
                            ex. [(0.25,0.35),(0.4,0.5),(0.2,0.3),(0.05,0.15)]

        delta_factor:       float, default 0.05
                            grid increment for factors in searching      

        quantweight_range:  tuple that indicates the beginning and ending 
                            points in searching for quantweight.    

        delta_quantweight:  float, default 0.05
                            grid increment for quantweight in searching   

        best_k:             int, default 20
                            how many best settings are kept.

        isthere_def:        boolean, default True
                            whether the analysis is on the default event. If not, for 
                            example we want to check the accuracy of factor on PDRR,
                            then we should input  
                            isthere_def=False, dependentvar='PDRR'

        dependentvar:       str, default 'def_flag'
                            column name for dependent variable.      

        use_msms:           boolean, default False
                            'msms' means module statistics which saved in model setting
                            whether use module mean and std in model setting.
                            To use the mean and std of quant and quali module in class 'model' if True.
                            To use the mean and std of quant and quali module that calibrated from the data if False.

        Return:   best_k weights settings and AR(SomersD) in a Dataframe 
        '''     
        quant_factor        = kw.pop('quant_factor',False)
        quant_weight_range  = kw.pop('quant_weight_range',False)    
        quali_factor        = kw.pop('quali_factor',False)
        quali_weight_range  = kw.pop('quali_weight_range',False)
        delta_factor        = kw.pop('delta_factor',0.05)
        quantweight_range   = kw.pop('quantweight_range',False)
        delta_quantweight   = kw.pop('delta_quantweight',0.05)
        best_k              = kw.pop('best_k',20)   
        isthere_def         = kw.pop('isthere_def',True)   
        dependentvar        = kw.pop('dependentvar','def_flag')   
        use_msms            = kw.pop('use_msms',False)

        if isthere_def:
            arfunc = self.__SomersD_byAUROC
        else:
            arfunc = self.__SomersD
            
        rows_list = [];     args = []       

        # case 1: both quant and quali:
        if quant_factor and quali_factor:
            # prepare data
            quant_weight_range  = dict(zip(quant_factor,quant_weight_range))
            quali_weight_range  = dict(zip(quali_factor,quali_weight_range))    

            # prepare grid for each quant and quali factors:
            for i in quant_factor:
                args.append(np.linspace(quant_weight_range[i][0],quant_weight_range[i][1], \
                    num=int(round((quant_weight_range[i][1]-quant_weight_range[i][0])/delta_factor+1))))
            for i in quali_factor:
                args.append(np.linspace(quali_weight_range[i][0],quali_weight_range[i][1], \
                    num=int(round((quali_weight_range[i][1]-quali_weight_range[i][0])/delta_factor+1))))        
            if len(quantweight_range)==1:
                args.append(np.array(quantweight_range))
            else:
                args.append(np.linspace(quantweight_range[0],quantweight_range[1],num=int(round((quantweight_range[1]-quantweight_range[0])/delta_quantweight+1))))
            
            lens_qq = len(quant_factor)
            weights = product(*args)   

            if use_msms:
                quantmean =  self.model.quantmean
                quantstd =  self.model.quantstd
                qualimean =  self.model.qualimean
                qualistd =  self.model.qualistd

            for all in weights:
                # check the sum of weights:
                if sum(all[:lens_qq])==1 and sum(all[lens_qq:-1])==1:
                    quant_weights = list(all)[:lens_qq]
                    quali_weights = list(all)[lens_qq:-1]

                    temp_quant = quant_weights*self.normdata[quant_factor]
                    temp_quali = quali_weights*self.normdata[quali_factor]
                    
                    if not use_msms:
                        quantmean =  temp_quant.sum(axis=1).mean()
                        quantstd =  temp_quant.sum(axis=1).std()
                        qualimean =  temp_quali.sum(axis=1).mean()
                        qualistd =  temp_quali.sum(axis=1).std()

                    score =  all[-1]*50*((temp_quant).sum(axis=1)-quantmean) / quantstd + \
                    (1-all[-1])*50*((temp_quali).sum(axis=1)-qualimean) / qualistd
                
                    accrat = arfunc(score, dependentvar)

                    dict1 = {}
                    dict1.update({'AR':accrat})
                    dict1.update(dict(zip(quant_factor,quant_weights)))
                    dict1.update(dict(zip(quali_factor,quali_weights)))
                    dict1.update({'quantweight':all[-1]})
                    rows_list.append(dict1) 
    

            result = pd.DataFrame(rows_list)
            result = result[quant_factor+quali_factor+['quantweight']+['AR']]
            result.sort_values(by='AR',ascending=False, inplace=True)
            result.reset_index(drop=True, inplace=True)
            return(result.head(best_k)) 
    
        elif bool(quant_factor) != bool(quali_factor):# means quant_factor xor quali_factor
            
            if quant_factor:
                quant_weight_range  = dict(zip(quant_factor,quant_weight_range))
                names = quant_factor
                for i in names:
                    args.append(np.linspace(quant_weight_range[i][0],quant_weight_range[i][1], \
                    num=int(round((quant_weight_range[i][1]-quant_weight_range[i][0])/delta_factor+1))))
            else:
                quali_weight_range  = dict(zip(quali_factor,quali_weight_range))
                names = quali_factor
                for i in names:
                    args.append(np.linspace(quali_weight_range[i][0],quali_weight_range[i][1], \
                    num=int(round((quali_weight_range[i][1]-quali_weight_range[i][0])/delta_factor+1))))    

            weights = product(*args)    
            for all in weights:    
                # check the sum of weights:   
                if sum(all)==1:
                    names_weights = list(all)                   
                    score = (self.normdata[names]*names_weights).sum(axis=1)        
                    accrat = arfunc(score, dependentvar)

                    dict1 = {}
                    dict1.update({'AR':accrat})
                    dict1.update(dict(zip(names,names_weights)))
                    rows_list.append(dict1) 

            result = pd.DataFrame(rows_list)
            result = result[names+['AR']]
            result.sort_values(by='AR',ascending=False, inplace=True)
            result.reset_index(drop=True, inplace=True)
            return(result.head(best_k))
           
        else:
            return(0)



# -*- coding: utf-8 -*-
"""
Created on Fri Oct  7 10:47:28 2016

Version: 1.0: Initial build
Version: 1.1, 20170327, Add method update()
Version: 1.2, 20170330, Add masterscale as new class attribute.
Version: 1.3, 20170502, Change masterscale as optional attribute.
Version: 1.4, 20180508, Add method reset()
Version: 1.5, 20180810, Add method save()

@author: ub71894 (4e8e6d0b), CSG

"""

import collections
import pandas as pd
import numpy as np
import warnings
import pickle

class PDModel(object):

    def __init__(self,  PDInfo_file, model_name, version, quant_factor, quali_factor, masterscale_file=False):

        self.model_name = model_name
        self.version = version
        self.quant_factor = quant_factor
        self.quali_factor = quali_factor
        self.masterscale_file = masterscale_file

        PDInfo = pd.read_excel(PDInfo_file,sheet_name=r'PD Models')

        Parameters=PDInfo.loc[(PDInfo['Model Name.1']==self.model_name) & (PDInfo['Version Number']==self.version)]
        # Quant parameters df
        quant_para=pd.DataFrame(Parameters.loc[:,'Quant_Name_1':'Quant_Wt_1'].values, 
            columns=['Factor', 'Lower', 'Upper', 'Invalid_Neg', 'Take_log','Mean', 'Std', 'Multiplier','Weight'])     
        for i in range((len(self.quant_factor)-1)):
            quant_para=pd.concat([quant_para, 
                pd.DataFrame(Parameters.loc[:,'Quant_Name_'+str(i+2):'Quant_Wt_'+str(i+2)].values,
                columns=['Factor', 'Lower', 'Upper', 'Invalid_Neg', 'Take_log','Mean', 'Std', 'Multiplier','Weight'])],  ignore_index=True)              
        # Quali parameters df
        quali_para=pd.DataFrame(Parameters.loc[:,'Qual_1_Question':'Qual_1_Wt'].values, 
            columns=['Factor', 'A', 'B', 'C', 'D','E', 'F', 'G','Weight'])
        for i in range((len(self.quali_factor)-1)):
            quali_para=pd.concat([quali_para, 
                pd.DataFrame(Parameters.loc[:,'Qual_'+str(i+2)+'_Question':'Qual_'+str(i+2)+'_Wt'].values, 
                columns=['Factor', 'A', 'B', 'C', 'D','E', 'F', 'G','Weight']) ], ignore_index=True)
            
        self.quant_weight=quant_para['Weight'].tolist()  
        self.floor=quant_para['Lower'].tolist()
        self.cap=quant_para['Upper'].tolist()
        self.doc_mean=quant_para['Mean'].tolist()
        self.doc_std=quant_para['Std'].tolist()
        self.quant_log = quant_para['Take_log'].fillna(value = False).tolist()
        self.quant_multiplier = quant_para['Multiplier'].tolist()        

        self.quali_weight=quali_para['Weight'].tolist()    
        qualimapping=[]
        for i in range(len(quali_para)):
           qualimapping.append(quali_para.loc[i,'A':'G'].dropna().to_dict()) 
        self.qualimapping = qualimapping           
           
        self.quantmean=Parameters['Quant_Mean_Module'].iat[0]
        self.quantstd=Parameters['Quant_StdDev_Module'].iat[0]
        self.quantweight=Parameters['Quant_Weight'].iat[0] 
        self.qualimean=Parameters['Qual_Mean_Module'].iat[0]
        self.qualistd=Parameters['Qual_StdDev_Module'].iat[0]
        self.qualiweight=1-self.quantweight
        self.Invalid_Neg=quant_para['Invalid_Neg'].tolist()  

        if Parameters['Less than or Equal to'].iat[0]>=999999:
            self.slope1=self.slope2=Parameters['Slope'].iat[0]
            self.intercept1=self.intercept2=Parameters['Intercept'].iat[0]
            self.cutoff=1e7
        else:
            self.slope1=Parameters['Slope'].iat[0]
            self.intercept1=Parameters['Intercept'].iat[0]
            self.slope2=Parameters['Slope.1'].iat[0]
            self.intercept2=Parameters['Intercept.1'].iat[0]
            self.cutoff=Parameters['Less than or Equal to'].iat[0]
        # new in Ver. 1.3
        if self.masterscale_file:
            self.MS = pd.read_excel(self.masterscale_file) # add in Ver. 1.2
        else:
            print("Waring: MasterScale file is not assigned to PDModel instance.\n")

    
    def reset(self):
       for key, value in self.__dict__.items():
            if key!='MS':
                exec('self.{key} = 0'.format(key=key)) 

    def update(self, kw):
        for key, value in kw.items():
            exec('self.{key} = value'.format(key=key))

    def __repr__(self):
        od = collections.OrderedDict(sorted(self.__dict__.items()))
        inner_lines = '\n'.join('%s = %s' % (k, v) for k, v in od.items())
        return '<\n%s\n>' % inner_lines

    def save(self, filename):
        filehandler = open(filename, 'wb')
        pickle.dump(self, filehandler)

    __str__=__repr__

# -*- coding: utf-8 -*-
"""
Created on Thu Mar 16 12:50:12 2017

Version: 1.0:   Initial build
Version: 1.1:   20170327, Add function quanttrans() which calculate cap/floor, 
                mean and std of quant factors.
Version: 1.2:   20170405, Add function normalization(), PD_frPDRR(), 
                logitPD_frPDRR(), getPDRR(), getTM(), and getTMnotches().
                Jack debugs method qualitrans() to make it work correctly in the
                model whose has no true default.
Version: 1.3:   20170424, Add function getwithknotchesrate(), SomersD() and 
                buildpseudodef()
Version: 1.4:   20170512, Modify function qualitrans() and quanttrans(), 
                normalization()
Version: 1.5:   20170802, Modify function getPDRR(). Add new function 
                RAreplica()
Version: 1.6:   20180508, Modify function normalization(), PD_frPDRR() and 
                logitPD_frPDRR()
Version: 1.7:   20180713, Add function buildpseudodef2()
Version: 1.8:   20180808, Add function MAUG_mapping(), NAICS_mapping()
Version: 1.9:   20180820, Add function ExtRating_mapping()
Version: 2.0:   20180906, Modify function getPDRR(), getTM(), getTMnotches(), 
                getwithknotchesrate() to make them have option for the version 
                of Masterscale
Version: 2.1:   20190604, Add function PD_frPDRR_autoMS() and 
                logitPD_frPDRR_autoMS()
Version: 2.2:   20191105, Add function gcar_converter()
Version: 2.3:   20200121, Modify a typo in MAUG_mapping()
Version: 2.4:   20200318, Add a tuple for americas office code: gcar_americas_office_code
                gcar_cif_cleaner()
Version: 2.5:   20200330, Modify MAUG_mapping(), Major Update for NAICS_mapping()           
Version: 2.6:   20200331, Add 3 functions: func_sd, func_rla and func_ovd 
                          All of them are designed for GroupBy.apply.
Version: 2.7:   20200605, Modify function SomersD()      
Version: 2.8:   20200930, Modify function gcar_converter()   
Version: 2.9:   20201113, Add function GICS_mapping()  
Version: 3.0:   20201119, Add function correlation_dist(), reorder_cols(), 
                          cluster_corr(), cluster_corr_old(), plot_cluster_corr(),
                          pca_n_comp()   
Version: 3.1:   20201228, Modify function SomersD() to fix the numba waring which is 
                          related to data type. No change in calculation algorithm.
                          Add function buildpseudodef_btmu()

@author: ub71894 (4e8e6d0b), CSG
"""

import pandas as pd
import numpy as np
import warnings
from numba import jit
from PDScorecardTool.CreateBenchmarkMatrix import CreateBenchmarkMatrix
from PDScorecardTool.CreateBenchmarkMatrix import TMnotches, withknotchesrate
from PDScorecardTool._info_data import naics_code, gics_code
from scipy.cluster import hierarchy
from scipy.stats import pearsonr
from sklearn import decomposition
from sklearn.preprocessing import StandardScaler
import matplotlib.pyplot as plt
import seaborn as sns
import locale
locale.setlocale(locale.LC_NUMERIC, '')


# for qualitative factors recalibration

def qualitrans(data, model, isthere_def=True, dependentvar='def_flag', output=True):
    """
    This function transforms raw qualitative factors (letters) into numeric value
    based on logit mean PD of each bucket.
    
    Parameters:

        data:           the input dataset in DataFrame. Make sure no NA in quali
                        factors

        model:          PDModel class
                        
        isthere_def:    boolean, default True
                        whether the calibration is based on true default event.

        dependentvar:   str, default 'def_flag'
                        column name for dependent variable. If 'isthere_def=False',
                        'dependentvar' should be some PD.  

        output:         boolean, default True
                        whether save the calibration to excel file.                           

    Return:
        a list of dictionaries that saves all mapping between letters to values
    
    """    

    calibration = {}
    for quali_col in model.quali_factor: #for each quali column
        dat = data[[quali_col,dependentvar]].copy() #takes quali_col and dependentvar columns from data (preserves index)
        dat.dropna(how='any', inplace=True)  # added in Ver. 1.4
        temp = dat.groupby(by=[quali_col]).mean() #for quali_col, find mean dependentvar
        temp.rename(columns={dependentvar:'mean_PD'},inplace=True) #renames dependentvar to mean_PD
        temp['count'] = dat.groupby(by=[quali_col]).count() #checks number of items with quali_col
        if isthere_def: #if there is a default
            temp['defaults'] = data[[quali_col,dependentvar]].groupby(by=[quali_col]).sum()#add to default column
            temp = temp[['count','defaults','mean_PD']] #sort into count, defaults, mean_PD
            # to set 1 as the minimal #def. Move here in Ver. 1.2
            temp.defaults.replace({0:1}, inplace=True)
            temp.loc[temp.defaults==1, 'mean_PD'] = 1/temp.loc[temp.defaults==1, 'count']
        else:
            temp = temp[['count','mean_PD']] #just sort into count, mean_pd
        
        temp['Rawscore'] = np.log(temp.mean_PD/(1-temp.mean_PD)) #calculates raw score        
        #new mean
        temp['mean'] = sum(temp['Rawscore'] * temp['count'])/temp['count'].sum() #takes the mean of raw score                
        #new std
        temp['std'] = np.sqrt(sum((temp['count'] * ((temp['Rawscore'] - temp['mean'])**2))/(temp['count'].sum()-1))) #takes the std of the raw score
        #new score
        temp['score'] = 50* (temp['Rawscore']-temp['mean'])/temp['std']#calculated new score
        calibration.update({quali_col:temp}) #adds tempscore to each qualitative column
    if output:
        writer_Def = pd.ExcelWriter('qualitrans_output.xlsx', engine='xlsxwriter')#creates a Pandas Excel Writer
        s=0
        for quali_col in model.quali_factor:
            calibration[quali_col].to_excel(writer_Def,startrow=(8*s+1), startcol=2)
            #Writes to excel
            s+=1
    #convert format to PDModel class's attribute
    result=[]
    for quali_col in model.quali_factor:
        result.append(dict(dict(calibration[quali_col])['score']))

    return({'qualimapping':result})


def quanttrans(data, model, floor=0.05, cap=0.95):
    """
    This function calculates the floor, cap, mean and std of the data. The output
    can be used as the updated parameters for quantitative factors normailization.
    Modified in Ver. 1.4

    Parameters:

        data:   the input dataset in DataFrame. Make sure no NA in quant
                factors

        model:  PDModel class

        floor:  float, default 0.05
                quantile for floor 

        cap:    float, default 0.95
                quantile for cap 

        
    Return:
        a dictionary that saves floor, cap, doc_mean, doc_std
    
    """    
    normdata = data.copy()
    warnings.filterwarnings('ignore', category=Warning)

    # get new cap and floor from valid observations:
    floor_list=[];  cap_list=[]

    for i, factor in enumerate(model.quant_factor):
        if not model.Invalid_Neg[i]:
            floor_list.append(normdata[factor].quantile(floor))
            cap_list.append(normdata[factor].quantile(cap))
        else:
            negative = model.Invalid_Neg[i]
            floor_list.append(normdata[factor][normdata[negative]>0].quantile(floor))
            cap_list.append(normdata[factor][normdata[negative]>0].quantile(cap)) 


    # Invalid_Neg
    for i,neg_source in enumerate(model.Invalid_Neg):
        if neg_source:
            col=model.quant_factor[i]
            normdata[col][ (normdata[col]<0) & (normdata[neg_source]<0) ] = cap_list[i]
            normdata[col][ (normdata[col]<0) & pd.isnull(normdata[neg_source]) ] = cap_list[i]

    # cap/floor for quant factors:
    for i, col in enumerate(model.quant_factor):
        normdata[col] = np.clip(normdata[col], floor_list[i], cap_list[i])        

    # quant factors transformation:
    for i, col in enumerate(model.quant_factor):
        if model.quant_log[i]:
            normdata[col] = np.log(normdata[col])

    # get new mean and std:
    doc_mean=[];    doc_std=[]

    for i, col in enumerate(model.quant_factor):
        doc_mean.append(normdata[col].mean())     
        doc_std.append(normdata[col].std())     

    dictionary = {'floor':floor_list, 'cap':cap_list, 'doc_mean':doc_mean, 'doc_std':doc_std}
    
    return(dictionary)


def normalization(data, model, quant_only=False, missing='median'):
    """
    This function performs normalization on rawdata with the setting in class 'model'
    
    Parameters:
    
        data:       the input dataset in DataFrame. Make sure no NA in quant
                    factors

        model:      PDModel class


        quant_only: boolean, default 'False'. Added in Ver.1.6
                    set as 'True' if only normalize on quant factors.


        missing:    str, default 'median'. Added in Ver.1.4
                    statistics of the data used to fill missing data (NA data)
       


    Return:
        a DataFrame that has normalized factors.
    
    """    
    warnings.filterwarnings('ignore', category=Warning)
    # Invalid_Neg
    normdata = data.copy()
    # modify the loop below in Ver 1.4:
    for i,neg_source in enumerate(model.Invalid_Neg):
        if neg_source:
            col=model.quant_factor[i]
            normdata[col][ (normdata[col]<0) & (normdata[neg_source]<0) ] = model.cap[i]
            # treat NA in 'neg_source' as negative value
            normdata[col][ (normdata[col]<0) & pd.isnull(normdata[neg_source]) ] = model.cap[i]

    # cap/floor for quant factors:
    for i, col in enumerate(model.quant_factor):
        normdata[col] = np.clip(normdata[col], model.floor[i], model.cap[i])        

    # quant factors transformation:
    for i, col in enumerate(model.quant_factor):
        if model.quant_log[i]:
            normdata[col] = np.log(normdata[col])
            
    # quant factors normalization:
    for i, col in enumerate(model.quant_factor):
        normdata[col]  = 50*(normdata[col] - model.doc_mean[i]) / model.doc_std[i]      
        
    # quant factors flip sign:  
    for i, col in enumerate(model.quant_factor):
        normdata[col] = normdata[col] * model.quant_multiplier[i]     

    if quant_only:
        exec('normdata[model.quant_factor] = normdata[model.quant_factor].\
            fillna(normdata[model.quant_factor].{missing}())'.format(missing=missing))
    else:
        # calibration for quali factors:
        for i, col in enumerate(model.quali_factor):
            normdata[col].replace(to_replace=model.qualimapping[i],inplace=True) 
        # fill missing data. added in Ver. 1.4
        exec('normdata[model.quant_factor+model.quali_factor] = normdata[model.quant_factor+model.quali_factor].\
            fillna(normdata[model.quant_factor+model.quali_factor].{missing}())'.format(missing=missing))

    return normdata


def PD_frPDRR(data, model, PDRR, ms_ver='new'):
    """
    This function calculates implied PD according to the PDRR which user assigned.
    The underling MasterScale is the new version one. 
    
    Parameters:

        data:   the input dataset in DataFrame. Make sure no NA in quant
                factors

        model:  PDModel class

        PDRR:   str
                column name for PDRR. ex. PDRR='Final_PD_Risk_Rating'
        
        ms_ver: str, default 'new'
                version of master default 'old' or 'new'
    Return:
        a DataFrame which is a copy of input data but has one more column called
        'PD_frPDRR'.
    
    """    
    dat = data.copy()
    if ms_ver=='old':
        mid_pd = model.MS['old_mid']
    else:
        mid_pd = model.MS['new_mid']

    MS_dict = dict(zip(model.MS['PDRR'], mid_pd))

    dat['PD_frPDRR'] = data[PDRR].transform(lambda x: np.nan if pd.isnull(x) else MS_dict[x]) 
    return (dat)


def logitPD_frPDRR(data, model, PDRR, ms_ver='new'):
    """
    This function calculates implied logitPD according to the PDRR which user assigned.
    The underling MasterScale is the new version one.
    
    Parameters:

        data:   the input dataset in DataFrame. Make sure no NA in quant
                factors

        model:  PDModel class

        PDRR:   str
                column name for PDRR. ex. PDRR='Final_PD_Risk_Rating'

        ms_ver: str, default 'new'
                version of master default 'old' or 'new'
    Return:
        a DataFrame which is a copy of input data but has one more column called
        'logitPD_frPDRR'.
    
    """        
    dat = data.copy()
    if ms_ver=='old':
        mid_pd = model.MS['old_mid']
    else:
        mid_pd = model.MS['new_mid']

    MS_dict = dict(zip(model.MS['PDRR'], mid_pd))
    pl_pd = data[PDRR].transform(lambda x: np.nan if pd.isnull(x) else MS_dict[x]).tolist()

    # convert PD to logit PD
    dat['logitPD_frPDRR'] = [np.log(x/(1-x)) for x in pl_pd]
    return (dat)


def PD_frPDRR_autoMS(data, model, PDRR, timestamp='archive_date'):  # add in ver. 2.1     

    """
    This function calculates implied PD according to the PDRR which user assigned.
    The underling MasterScale is automatically assigned by archive_date
    (New MS is in production from 05/01/2016)
    
    Parameters:

        data:       the input dataset in DataFrame. Make sure no NA in quant
                    factors
    
        model:      PDModel class
    
        PDRR:       str
                    column name for PDRR. ex. PDRR='Final_PD_Risk_Rating'
        
        timestamp:  str, default 'archive_date'
                    new MS apply after 20160501.
    Return:
        a DataFrame which is a copy of input data but has one more column called
        'PD_frPDRR'.
    
    """    

    dat_old = data.query('{}<20160501'.format(timestamp))
    dat_new = data.query('{}>=20160501'.format(timestamp))
    old_mid = model.MS['old_mid']
    new_mid = model.MS['new_mid']
    old_MS_dict = dict(zip(model.MS['PDRR'], old_mid))
    new_MS_dict = dict(zip(model.MS['PDRR'], new_mid))

    dat_old['PD_frPDRR'] = dat_old[PDRR].transform(lambda x: np.nan if pd.isnull(x) else old_MS_dict[x]) 
    dat_new['PD_frPDRR'] = dat_new[PDRR].transform(lambda x: np.nan if pd.isnull(x) else new_MS_dict[x]) 

    return (pd.concat([dat_old, dat_new], axis=0).sort_index())


def logitPD_frPDRR_autoMS(data, model, PDRR, timestamp='archive_date'): # add in ver. 2.1

    """
    This function calculates implied logitPD according to the PDRR which user assigned.
    The underling MasterScale is automatically assigned by archive_date
    (New MS is in production from 05/01/2016)
    
    Parameters:

        data:       the input dataset in DataFrame. Make sure no NA in quant
                    factors
    
        model:      PDModel class
    
        PDRR:       str
                    column name for PDRR. ex. PDRR='Final_PD_Risk_Rating'
        
        timestamp:  str, default 'archive_date'
                    new MS apply after 20160501.
    Return:
        a DataFrame which is a copy of input data but has one more column called
        'logitPD_frPDRR'.
    
    """    

    dat_old = data.query('{}<20160501'.format(timestamp))
    dat_new = data.query('{}>=20160501'.format(timestamp))
    old_mid = model.MS['old_mid']
    new_mid = model.MS['new_mid']
    old_MS_dict = dict(zip(model.MS['PDRR'], old_mid))
    new_MS_dict = dict(zip(model.MS['PDRR'], new_mid))
    pl_pd_old = dat_old[PDRR].transform(lambda x: np.nan if pd.isnull(x) else old_MS_dict[x]) 
    pl_pd_new = dat_new[PDRR].transform(lambda x: np.nan if pd.isnull(x) else new_MS_dict[x]) 

    # convert PD to logit PD
    dat_old['logitPD_frPDRR'] = [np.log(x/(1-x)) for x in pl_pd_old]
    dat_new['logitPD_frPDRR'] = [np.log(x/(1-x)) for x in pl_pd_new]

    return (pd.concat([dat_old, dat_new], axis=0).sort_index())


def getTotalscore(data, model):
    """
    This function applies 'model' on the 'data': Goes through normalization, weighted
    summation to get model's total score for each obligor.
    
    Parameters:

        data:   the input dataset in DataFrame. Make sure no NA in quant
                factors

        model:  PDModel class


    Return:
        a DataFrame which is a copy of input data but has more columns. The result
        is in column 'Ratings'.
    
    """    
    dat = normalization(data,model)

    if (model.quant_factor and model.quali_factor): # That's mean the model have both quant and quali factors
        dat['quantscore'] = (model.quant_weight * dat[model.quant_factor].values).sum(axis=1)
        dat['quantscore'] = 50*( dat['quantscore'] - model.quantmean) / model.quantstd
        dat['qualiscore'] = (model.quali_weight * dat[model.quali_factor].values).sum(axis=1)
        dat['qualiscore'] = 50*( dat['qualiscore'] - model.qualimean) / model.qualistd
        dat['Totalscore'] = dat['quantscore']*model.quantweight + dat['qualiscore'] *model.qualiweight
    else: # quant factors only. New in Version 1.5
        dat['Totalscore'] =  (model.quant_weight * dat[model.quant_factor].values).sum(axis=1)

    return (dat)


def getPDRR(data, model, ms_ver='new'):
    """
    This function applies 'model' on the 'data': Goes through normalization, weighted
    summation, calibration to get model PD for each obligor and uses the 'MasterScale'
    in 'model' to get the model PDRR (or 'Prelim_PD_Risk_Rating_Uncap')
    
    Parameters:

        data:   the input dataset in DataFrame. Make sure no NA in quant
                factors

        model:  PDModel class

        ms_ver: str, default 'new' New in Version 2.0
                version of master default 'old' or 'new'
    Return:
        a DataFrame which is a copy of input data but has more columns. The result
        is in column 'Ratings'.
    
    """    
    dat = normalization(data,model)

    if (model.quant_factor and model.quali_factor): # That's mean the model have both quant and quali factors
        dat['quantscore'] = (model.quant_weight * dat[model.quant_factor].values).sum(axis=1)
        dat['quantscore'] = 50*( dat['quantscore'] - model.quantmean) / model.quantstd
        dat['qualiscore'] = (model.quali_weight * dat[model.quali_factor].values).sum(axis=1)
        dat['qualiscore'] = 50*( dat['qualiscore'] - model.qualimean) / model.qualistd
        dat['score'] = dat['quantscore']*model.quantweight + dat['qualiscore'] *model.qualiweight
    else: # quant factors only. New in Version 1.5
        dat['score'] =  (model.quant_weight * dat[model.quant_factor].values).sum(axis=1)

    logitPD = []
    for obs in dat.iterrows():
        if obs[1].score < model.cutoff:
            logitPD.append(model.intercept1 + model.slope1*obs[1]['score'])
        else:
            logitPD.append(model.intercept2 + model.slope2*obs[1]['score'])
    dat['logitPD'] = logitPD

    dat['PD'] = dat['logitPD'].apply(lambda x: 100*np.exp(x)/(1+np.exp(x)))
    Ratings = []
    if ms_ver=='old':
        low_pd = model.MS['old_low']
    else:
        low_pd = model.MS['new_low']

    for i in dat.iterrows():
        Ratings.append(sum(low_pd<=(i[1].PD/100)))
    dat['Ratings'] = Ratings
    return (dat)


def getTM(data, model1, model2, ms_ver='new', PDRR_range=(1,20)):
    """
    This function generates Transition Matrix in an Excel file. The TM is coming
    from the same data but different model settings.
    
    Parameters:

        data:       the input dataset in DataFrame. Make sure no NA in quant
                    factors

        model1:     PDModel class
                    Model 1's setting

        model2:     PDModel class
                    Model 2's setting  

        ms_ver:     str, default 'new' New in Version 2.0
                    version of master default 'old' or 'new'

        PDRR_range: tuple, default (1,20)
                    the range of PDRR, (min rating, max rating)
        
    Return:
        An excel file called 'Matrix_Output.xlsx' 
    
    """    

    dat1 = getPDRR(data, model1, ms_ver)
    dat2 = getPDRR(data, model2, ms_ver)
    dat = pd.concat([dat1.Ratings, dat2.Ratings], axis=1)
    dat.columns = pd.Index(['PDRR_model1','PDRR_model2'])
    CreateBenchmarkMatrix(dat, 'Matrix_Output.xlsx', 'Benchmarking Matrix', 'PDRR_model1', 'PDRR_model2', PDRR=range(PDRR_range[0],PDRR_range[1]+1))

    
def getTMnotches(data, model1, model2, ms_ver='new', PDRR_range=(1,20)):
    """
    This function calculates two statistics from function getTM(): 
    'down_notches' and 'up_notches'. They're used to measure the mismatch.
    
    Parameters:

        data:       the input dataset in DataFrame. Make sure no NA in quant
                    factors

        model1:     PDModel class
                    Model 1's setting

        model2:     PDModel class
                    Model 2's setting  

        ms_ver:     str, default 'new' New in Version 2.0
                    version of master default 'old' or 'new'
                          
        PDRR_range: tuple, default (1,20)
                    the range of PDRR, (min rating, max rating)
        
    Return:
        tuple: (down_notches, up_notches)
    
    """
    dat1 = getPDRR(data, model1, ms_ver)
    dat2 = getPDRR(data, model2, ms_ver)
    dat = pd.concat([dat1.Ratings, dat2.Ratings], axis=1)
    dat.columns = pd.Index(['PDRR_model1','PDRR_model2'])
    return(TMnotches(dat,'PDRR_model1', 'PDRR_model2', PDRR=range(PDRR_range[0],PDRR_range[1]+1)))


def getwithknotchesrate(data, model1, model2, ms_ver='new', PDRR_range=(1,20), k=2):
    """
    This function calculates the rate that with k notches matches.
    
    Parameters:

        data:       the input dataset in DataFrame. Make sure no NA in quant
                    factors

        model1:     PDModel class
                    Model 1's setting

        model2:     PDModel class
                    Model 2's setting  
        
        ms_ver:     str, default 'new' New in Version 2.0
                    version of master default 'old' or 'new'
                                              
        PDRR_range: tuple, default (1,20)
                    the range of PDRR, (min rating, max rating)

        k:          int, default 2
                    the rate is within k notches
    Return:
        tuple: (down_notches, up_notches)
    
    """
    dat1 = getPDRR(data, model1, ms_ver)
    dat2 = getPDRR(data, model2, ms_ver)
    dat = pd.concat([dat1.Ratings, dat2.Ratings], axis=1)
    dat.columns = pd.Index(['PDRR_model1','PDRR_model2'])
    return(withknotchesrate(dat,'PDRR_model1', 'PDRR_model2', PDRR=range(PDRR_range[0],PDRR_range[1]+1),k=k))
 

def buildpseudodef(data, def_flag='def_flag', PDRR='Final_PD_Risk_Rating', pseudodef_PDRR=[13,14,15]):
    """
    This function constructs pseudo defaults according the raings
    
    Parameters:

        data:           the input dataset in DataFrame. Make sure no NA in quant
                        factors

        def_flag:       str, default 'def_flag'
                        column name for default flag

        PDRR:           str, default 'Final_PD_Risk_Rating'
                        column name for PDRR
                                  
        pseudodef_PDRR: list, default [13,14,15]
                        the range of PDRR that treated as defaulted
        
    Return:
        DataFrame: Data with pseudo defaults.

    """
    warnings.filterwarnings('ignore', category=Warning)
    dat = data.copy()
    n_truedef = int(dat[def_flag].sum())
    print ("There're {n} true defaults in the input dateset.".format(n=n_truedef))
    dat2 = dat[dat[PDRR].isin(pseudodef_PDRR)]
    dat2[def_flag]=1
    dat[dat[PDRR].isin(pseudodef_PDRR)] = dat2
    n_totaldef = int(dat[def_flag].sum())
    n_pseudodef = n_totaldef - n_truedef
    print ("There're {n} defaults in the output dateset. Among them, {m} are pseudo defaults.".\
        format(n=n_totaldef, m=n_pseudodef))

    return(dat)



def buildpseudodef2(data, **kw):  # add in ver. 1.7
    """
    This function constructs pseudo defaults based on the multiple conditions:
    1. current (at time t) PDRR is at 'pseudodef_PDRR'
    2. the downgrade gap between time t and time t-1 is equal or larger than 'gap_PDRR'
    
    Parameters:

        data:           the input dataset in DataFrame. Make sure no NA in quant
                        factors

        def_flag:       str, default 'def_flag'
                        column name for default flag

        PDRR:           str, default 'Final_PD_Risk_Rating'
                        column name for PDRR

        pseudodef_PDRR: list, default [13,14]
                        the range of PDRR that treated as defaulted

        gap_PDRR:       int, default 2
                        the gap between PDRR at t-1 and PDRR at t 

        timestamp:      str, default 'archive_date'
                        column name for timestamp
        
        idn:            str, default 'CUSTOMERID'
                        column name for customer identifier

    Return:
        DataFrame: Data with pseudo defaults.

    """
    def_flag =          kw.pop('def_flag','def_flag')
    PDRR =              kw.pop('PDRR','Final_PD_Risk_Rating')
    pseudodef_PDRR =    kw.pop('pseudodef_PDRR',[13,14])
    gap_PDRR =          kw.pop('gap_PDRR',2 )
    timestamp =         kw.pop('timestamp','archive_date')
    idn =               kw.pop('idn','CUSTOMERID')


    #warnings.filterwarnings('ignore', category=Warning)
    dat = data.copy()
    n_truedef = int(dat[def_flag].sum())
    print ("There're {n} true defaults in the input dateset.".format(n=n_truedef))

    dat.sort_values(by=[idn,timestamp], inplace=True)
    dat.reset_index(drop=True, inplace=True)
    idlist = dat[idn].unique().tolist()

    id_pseudo=[] # stores all indices of 'dat' that match our criterion
    for id in idlist:
        temp = dat[dat[idn]==id]
        for i in range(len(temp)-1):      
            diff = temp[PDRR].iloc[i+1]-temp[PDRR].iloc[i]
            if (diff>=gap_PDRR) & (temp[PDRR].iloc[i+1] in pseudodef_PDRR):
                id_pseudo.append(temp.index[i+1])
            else:
                continue

    for id in id_pseudo:
        temp = dat.loc[id]
        temp.def_flag = True
        dat.loc[id] = temp

    n_totaldef = int(dat[def_flag].sum())
    n_pseudodef = n_totaldef - n_truedef
    print ("There're {n} defaults in the output dateset. Among them, {m} are pseudo defaults.".\
        format(n=n_totaldef, m=n_pseudodef))

    return(dat)



def SomersD(y_true, y_score, sign=1, unit_concordance=True): 
    '''

    New version of SomersD function which leverages numba.jit to accelerate
    the calculation. In production from Ver. 1.6.

    '''

    # mofified in Ver 2.7. To make it auto drop missing value before calculation
    dat = pd.DataFrame()                        #new lines in Ver 2.7                 
    dat['true'] = y_true                        #new lines in Ver 2.7                 
    dat['score'] = y_score                      #new lines in Ver 2.7                
    dat.dropna(how='any', inplace=True)         #new lines in Ver 2.7                   
    x = np.array(dat.score, dtype=np.float32)                   #new lines in Ver 2.7                 
    y = np.array(dat.true, dtype=np.float32)                    #new lines in Ver 2.7     

    return(_SD(x, y, sign, unit_concordance))   #new lines in Ver 3.1 

@jit
def _SD(x, y, sign, unit_concordance): #new private function in in Ver 3.1
    n = len(x)
    C = 0
    D = 0
    if unit_concordance==True:
        for i in range(n):
            for j in range(i):
                dx = x[i] - x[j]
                dy = y[i] - y[j]
                if (dx*dy) > 0:
                    cij = 1
                    dij = 1
                elif  (dx*dy) < 0:
                    cij = -1
                    dij = 1
                else:
                    cij = 0
                    if dy==0: 
                        dij = 0
                    else: 
                        dij=1
                C = C + cij
                D = D + dij
    else:
        for i in range(n):
            for j in range(i):
                dx = x[i] - x[j]
                dy = y[i] - y[j]
                if  (dx*dy) > 0:
                    cij = abs(dy)
                    dij = abs(dy)
                elif (dx*dy) < 0:
                    cij = -abs(dy)
                    dij = abs(dy)
                else:
                    cij = 0
                    if dy==0: 
                        dij = 0
                    else: 
                        dij=1
                C = C + cij
                D = D + dij            
    return sign*C/D



def RAreplica(data, model, ms_ver='new'):
    """
    This function applies 'model' on the 'data': Goes through normalization, weighted
    summation, calibration to get model PD for each obligor and uses the 'MasterScale'
    in 'model' to get the PDRR and all intermediate results. New in Version 1.5.
    
    Parameters:

        data:   the input dataset in DataFrame. Make sure no NA in quant
                factors

        model:  PDModel class

        ms_ver: str, default 'new'
                the masterscale version to use.  {‘old’, ‘new’, timecutoff}
                'old':      old version ms which was used before April,2016
                'new':      new version ms which was used after April,2016
                timecutoff: e.x. '2016/05/01'
                            the function will apply old version ms for all observations 
                            if its 'archive_date' is before this date; and apply new 
                            version ms for all observations if its 'archive_date' is on 
                            or after this date. Note that: user MUST make sure all observations
                            have valid 'archive_date'. 

    Return:
        a DataFrame which is a copy of input data but has more columns. And the function will
        export the DataFrame into an Excel File 'RAreplica.xlsx' into current working directory. 
    
    """    
    dat = normalization(data,model)

    if (model.quant_factor and model.quali_factor): # That's mean the model have both quant and quali factors
        dat['quantscore'] = (model.quant_weight * dat[model.quant_factor].values).sum(axis=1)
        dat['quantscore'] = 50*( dat['quantscore'] - model.quantmean) / model.quantstd
        dat['qualiscore'] = (model.quali_weight * dat[model.quali_factor].values).sum(axis=1)
        dat['qualiscore'] = 50*( dat['qualiscore'] - model.qualimean) / model.qualistd
        dat['finalscore'] = dat['quantscore']*model.quantweight + dat['qualiscore'] *model.qualiweight
    else: # quant factors only
        dat['finalscore'] =  (model.quant_weight * dat[model.quant_factor].values).sum(axis=1)

    logitPD = []
    for obs in dat.iterrows():
        if obs[1].finalscore < model.cutoff:
            logitPD.append(model.intercept1 + model.slope1*obs[1]['finalscore'])
        else:
            logitPD.append(model.intercept2 + model.slope2*obs[1]['finalscore'])
    dat['logitPD'] = logitPD
    dat['PD'] = dat['logitPD'].apply(lambda x: 100*np.exp(x)/(1+np.exp(x)))

    Ratings = []
    if ms_ver=='new':
        ms_used = model.MS['new_low']
        for i in dat.iterrows():
            Ratings.append(sum(ms_used<=(i[1].PD/100)))
        dat['Ratings'] = Ratings
    elif ms_ver=='old':
        ms_used = model.MS['old_low']
        for i in dat.iterrows():
            Ratings.append(sum(ms_used<=(i[1].PD/100)))
        dat['Ratings'] = Ratings
    else:
        for i in dat.iterrows():
            if i[1].archive_date >= pd.to_datetime(ms_ver): #use new masterscale
                Ratings.append(sum(model.MS['new_low']<=(i[1].PD/100)))
            else: # use old masterscale
                Ratings.append(sum(model.MS['old_low']<=(i[1].PD/100)))

    dat['Ratings'] = Ratings
    dat.to_excel('RAreplica.xlsx')

    return (dat)


def MAUG_mapping(data, col_guidance='Underwriter_Guideline'):
    """
    This function is to get MAUG code based on 'col_guidance'.(usually, its name is 'Underwriter_Guideline'
    in RA) And then mapping MAUG to industry based on doc 'MAUG-0100 UG to MAUG Mapping.pdf'. New in Version 1.8

    Parameters:

        data:           the input dataset in DataFrame. 

        col_guidance:   str, default 'Underwriter_Guideline'
                        column name

    Return:
        a DataFrame which is a copy of input data but has 2 more columns. 'MAUG' and 'Industry_by_MAUG'
    
    """    
    processed_data = data.copy()
    processed_data['MAUG'] = ['MAUG_'+str(x) for x in processed_data[col_guidance]] # modified on Ver.2.5
        
    # 440 is old UG code and it still exist after 20160723
    processed_data['MAUG'] = processed_data['MAUG'].replace(to_replace={'MAUG_440':'MAUG_125'})
    # 490 is old UG code and it still exist after 20160723
    processed_data['MAUG'] = processed_data['MAUG'].replace(to_replace={'MAUG_490':'MAUG_305'})
    # 705 is old UG code and it still exist after 20160723
    processed_data['MAUG'] = processed_data['MAUG'].replace(to_replace={'MAUG_705':'MAUG_340'})
    # 720 is old UG code and it still exist after 20160723
    processed_data['MAUG'] = processed_data['MAUG'].replace(to_replace={'MAUG_720':'MAUG_350'})
    # 835 is old UG code and it still exist after 20160723
    processed_data['MAUG'] = processed_data['MAUG'].replace(to_replace={'MAUG_835':'MAUG_325'})
    # 150-1 is old UG code and it still exist after 20160723
    processed_data['MAUG'] = processed_data['MAUG'].replace(to_replace={'MAUG_150-1':'MAUG_105'})

    processed_data['MAUG'] = processed_data['MAUG'].replace(to_replace={'MAUG_227':'MAUG_115'})

    processed_data['MAUG'] = processed_data['MAUG'].replace(to_replace={'MAUG_240':'MAUG_115'})

    processed_data['MAUG'] = processed_data['MAUG'].replace(to_replace={'MAUG_245':'MAUG_115'})

    processed_data['MAUG'] = processed_data['MAUG'].replace(to_replace={'MAUG_340-1':'MAUG_105'})

    processed_data['MAUG'] = processed_data['MAUG'].replace(to_replace={'MAUG_408':'MAUG_160'})

    processed_data['MAUG'] = processed_data['MAUG'].replace(to_replace={'MAUG_450':'MAUG_105'})

    processed_data['MAUG'] = processed_data['MAUG'].replace(to_replace={'MAUG_455':'MAUG_205'})

    processed_data['MAUG'] = processed_data['MAUG'].replace(to_replace={'MAUG_470':'MAUG_180'})

    processed_data['MAUG'] = processed_data['MAUG'].replace(to_replace={'MAUG_480':'MAUG_165'})

    processed_data['MAUG'] = processed_data['MAUG'].replace(to_replace={'MAUG_810':np.nan})

    processed_data['MAUG'] = processed_data['MAUG'].replace(to_replace={'MAUG_820':np.nan})

    processed_data['MAUG'] = processed_data['MAUG'].replace(to_replace={'MAUG_840':np.nan})

    processed_data['MAUG'] = processed_data['MAUG'].replace(to_replace={'MAUG_230':np.nan})

    processed_data['MAUG'] = processed_data['MAUG'].replace(to_replace={'MAUG_235':np.nan})

    processed_data['MAUG'] = processed_data['MAUG'].replace(to_replace={'MAUG_750-01':np.nan})

    processed_data['MAUG'] = processed_data['MAUG'].replace(to_replace={'MAUG_465':np.nan})

    processed_data['MAUG'] = processed_data['MAUG'].replace(to_replace={'MAUG_CBUG-185':'MAUG_180'})

    processed_data['MAUG'] = processed_data['MAUG'].replace(to_replace={'MAUG_481':'MAUG_105'})

    processed_data['MAUG'] = processed_data['MAUG'].replace(to_replace={'MAUG_806':'MAUG_395'})

    processed_data['MAUG'] = processed_data['MAUG'].replace(to_replace={'MAUG_730':'MAUG_350'})

    processed_data['MAUG'] = processed_data['MAUG'].replace(to_replace={'MAUG_445':'MAUG_105'})

    processed_data['MAUG'] = processed_data['MAUG'].replace(to_replace={'MAUG_405-01':'MAUG_155'})

    processed_data['MAUG'] = processed_data['MAUG'].replace(to_replace={'MAUG_405-02':'MAUG_155'})

    processed_data['MAUG'] = processed_data['MAUG'].replace(to_replace={'MAUG_405-03':'MAUG_155'})

    processed_data['MAUG'] = processed_data['MAUG'].replace(to_replace={'MAUG_910':'MAUG_435'})

    processed_data['MAUG'] = processed_data['MAUG'].replace(to_replace={'MAUG_905':'MAUG_105'})

    processed_data['MAUG'] = processed_data['MAUG'].replace(to_replace={'MAUG_850':'MAUG_105'})

    processed_data['MAUG'] = processed_data['MAUG'].replace(to_replace={'MAUG_225':'MAUG_115'})

    processed_data['MAUG'] = processed_data['MAUG'].replace(to_replace={'MAUG_837':'MAUG_320'})

    processed_data['MAUG'] = processed_data['MAUG'].replace(to_replace={'MAUG_855':'MAUG_105'})

    processed_data['MAUG'] = processed_data['MAUG'].replace(to_replace={'MAUG_626':'MAUG_355'})

    processed_data['MAUG'] = processed_data['MAUG'].replace(to_replace={'MAUG_250':'MAUG_115'})

    processed_data['MAUG'] = processed_data['MAUG'].replace(to_replace={'MAUG_805':'MAUG_415'})

    processed_data['MAUG'] = processed_data['MAUG'].replace(to_replace={'MAUG_845':'MAUG_390'})

    processed_data['MAUG'] = processed_data['MAUG'].replace(to_replace={'MAUG_485':'MAUG_310'})

    processed_data['MAUG'] = processed_data['MAUG'].replace(to_replace={'MAUG_CBUG-135':'MAUG_105'})

    processed_data['MAUG'] = processed_data['MAUG'].replace(to_replace={'MAUG_425-1':'MAUG_110'})

    processed_data['MAUG'] = processed_data['MAUG'].replace(to_replace={'MAUG_CBUG-190':'MAUG_165'})

    processed_data['MAUG'] = processed_data['MAUG'].replace(to_replace={'MAUG_CBUG-125':'MAUG_175'})

    processed_data['MAUG'] = processed_data['MAUG'].replace(to_replace={'MAUG_CBUG-130':'MAUG_120'})

    processed_data['MAUG'] = processed_data['MAUG'].replace(to_replace={'MAUG_340-2':'MAUG_330'})


    sofar_mapping = {
    'MAUG_105': 'General Industries',
    'MAUG_110': 'Entertainment',
    'MAUG_115': 'Media and Telecom',
    'MAUG_120': 'Food and Beverage',
    'MAUG_125': 'Health Care',
    'MAUG_130': 'Insurance',
    'MAUG_135': 'Power & Utilities',
    'MAUG_140': 'Independent Exploration and Production',
    'MAUG_143': 'Midstream Energy',
    'MAUG_145': 'Independent Refining',
    'MAUG_147': 'Drilling and Oilfield Services',
    'MAUG_150': 'Auto and Auto Parts',
    'MAUG_155': 'Agribusiness',
    'MAUG_160': 'Wine Industry',
    'MAUG_165': 'Technology',
    'MAUG_170': 'Leasing Companies',
    'MAUG_175': 'Engineering and Construction',
    'MAUG_180': 'Retail',
    'MAUG_205': 'Asian Corporate Banking',
    'MAUG_210': 'Japanese Corporate Banking in Canada',
    'MAUG_215': 'Asian Corporate Leasing & Finance',
    'MAUG_305': 'Business Banking',
    'MAUG_307': 'Small Business Banking',
    'MAUG_310': 'Business Diversity Lending',
    'MAUG_320': 'Corporate Leasing Transactions',
    'MAUG_325': 'Leasing and Asset Finance Division',
    'MAUG_330': 'Project Finance',
    'MAUG_333': 'Commodity Finance',
    'MAUG_335': 'Mezzanine Finance',
    'MAUG_340': 'Commercial Finance Loans',
    'MAUG_345': 'Commodity and Structured Trade Finance',
    'MAUG_350': 'Trading Asset Reliant',
    'MAUG_355': 'Capital Call Bridge / Subscription Lending',
    'MAUG_360': 'Securities Broker / Dealers',
    'MAUG_365': 'Asset Managers',
    'MAUG_367': 'Investment Funds (Mutual Funds)',
    'MAUG_370': 'Clearinghouses and Exchanges',
    'MAUG_375': 'Industrial Development Bonds',
    'MAUG_380': 'Insured Domestic Institutions',
    'MAUG_381': 'Foreign Banks',
    'MAUG_383': 'Global Financial Solutions',
    'MAUG_385': 'Securitization',
    'MAUG_390': 'Debt Obligation and Passive Equity Investments',
    'MAUG_395': 'Individuals',
    'MAUG_405': 'Real Estate Industries',
    'MAUG_410': 'Institutional Real Estate Lending',
    'MAUG_415': 'Homeowner Association Loans',
    'MAUG_420': 'Retail Equity Equivalent Loan Program',
    'MAUG_425': 'Debtor-in-Possession Loans',
    'MAUG_430': 'Finance and Mortgage Companies',
    'MAUG_435': 'Investor Perm Loans',
    'MAUG_505': 'Community Development Financing',
    'MAUG_510': 'Public Finance',
    'MAUG_515': 'Non-Profit Organizations',
    'MAUG_605': 'Brazil Agribusiness',
    'MAUG_610': 'Placeholder for pending General Corporate Latin America',
    'MAUG_615': 'Placeholder for pending Retail Latin America',
    'MAUG_620': 'Placeholder for pending Utilities Latin America',
    'MAUG_625': 'Placeholder for pending Oil & Gas Latin America',
    'MAUG_630': 'Placeholder for pending Metals and Mining Latin America',
    'MAUG_635': 'Placeholder for pending Food and Beverage Latin America'
    }

    processed_data['Industry_by_MAUG'] = processed_data['MAUG'].replace(sofar_mapping)

    return (processed_data)


def NAICS_mapping(data, col_guidance='NAICS_Cd', first_digits=2):
    """
    This function is to convert NAICS code to industry. New in Version 1.8. Major
    update in Version 2.5 to make it work for different granular level mapping

    Parameters:

        data:           the input dataset in DataFrame. 

        col_guidance:   str, default 'NAICS_Cd'
                        column name

        first_digits:   integer from 2 to 6, default 2. 
                        First digits used for the industry mapping

    Return:
        a DataFrame which is a copy of input data but has one more column: 'Industry_by_NAICS'
    
    """    
    processed_data = data.copy()
    processed_data['Industry_by_NAICS'] = processed_data[col_guidance].fillna(999999)
    processed_data['Industry_by_NAICS'] = [str(int(x))[:first_digits] for x in processed_data['Industry_by_NAICS']]
    processed_data['Industry_by_NAICS'] = processed_data['Industry_by_NAICS'].replace(naics_code)
    
    # to deal with 'Others':
    pl_code = [str(x) for x in range(0,10**first_digits) if str(x) not in list(naics_code.keys()) ]
    pd_others = dict(zip(pl_code,['Unknow_NAICS_Code',]*len(pl_code)))
    processed_data['Industry_by_NAICS'] = processed_data['Industry_by_NAICS'].replace(pd_others)

    return (processed_data)


def ExtRating_mapping(data, col_ext_rating='SPRating'):
    """
    This function is to convert external S&P rating to internal PDRR. New in Version 1.9

    Parameters:

        data:               the input dataset in DataFrame. 

        col_ext_rating:     str, default 'SPRating'
                            column name for S&P rating

    Return:
        a DataFrame which is a copy of input data but has one more column: 'ExternalRating'
    
    """   
    processed_data = data.copy()

    ext_ratings_mapping={
    'AAA':2,
    'AA+':2,
    'AA':2,
    'AA-':2, # modified on 1/31/2019, previous is '3'
    'A+':3,
    'A':3,
    'A-':4,
    'BBB+':5,
    'BBB':6,
    'BBB-':7,
    'BB+':8,
    'BB':9,  
    'BB-':10,
    'B+':11,
    'B':12,
    'B-':13,
    'CCC+':14, # modified on 10/314/2020, previous is '15'
    'CCC':15,
    'CCC-':15,
    'CC':15,


    }

    processed_data['ExternalRating'] = processed_data[col_ext_rating]
    processed_data['ExternalRating'] = processed_data['ExternalRating'].replace(ext_ratings_mapping)
    return (processed_data)


def gcar_converter(data, cols, inplace=False, ignore_pct=True): # Added in Ver. 2.2
    '''
    This function is to convert GCAR data's financial statement into float. New in Version 2.2

    Parameters:

        data:           the input dataset in DataFrame. 
    
        cols:           str, or list of str
                        column name for the conversion
    
        inplace:        bool, default False
                        If True, do operation inplace and return None.

        ignore_pct:     bool, default True. New in Version 2.8
                        If True, convert '123.45%' to 123.45        
                        If False, convert '123.45%' to 1.2345                

    Return:
        DataFrame with converted columns.
    '''
    def _mycast(s):

        if pd.isnull(s):
            return(np.nan)

        elif isinstance(s, float) or isinstance(s, int):
            return (s)

        elif s[-1]=='%' and ignore_pct:  # to convert '123.45%' to 123.45
            if ',' in s:
                return (locale.atof(s[:-1]))
            else:
                return (float(s[:-1]))

        elif s[-1]=='%' and not ignore_pct:  # to convert '123.45%' to 1.2345
            if ',' in s:
                return (locale.atof(s[:-1])/100)
            else:
                return (float(s[:-1])/100)

        elif ',' in s:  # to convert '123,345.000' to 123456.0
            return (locale.atof(s))

        else:
            try: 
                return (float(s))
            except ValueError:
                return(np.nan)

    if isinstance(cols, str):
        cols=[cols,]

    if inplace:
        data[cols] = data[cols].applymap(_mycast)
    else:
        dat = data.copy()
        dat[cols] = dat[cols].applymap(_mycast)
        return (dat)



def gcar_cif_cleaner(data, col='Borrower CIF', inplace=False): # Added in Ver. 2.4
    '''
    This function is to convert GCAR data's CIF into correct string format. New in Version 2.4

    Parameters:

        data:       the input dataset in DataFrame. 

        col:        str, default 'Borrower CIF'
                    column name for the CIF

        inplace:    bool, default False
                    If True, do operation inplace and return None.

    Return:
        DataFrame with converted columns.
    '''
    def _my_cast(cif):
        if cif[0]!='R':
            return (f'{int(cif):08d}')
        else:
            return (cif)

    if inplace:
        data[col] = data[col].astype(str)
        data[col] = data[col].apply(_my_cast)
    else:
        tmp_series = data[col].astype(str)
        return (tmp_series.apply(_my_cast))


# Added in Ver. 2.4
gcar_Americas_office_code = (
3103,3104,3108,3110,3116,3120,3134,3138,3147,3149,3153,3157,3158,3160,
3161,3165,3170,3191,3220,3500,3770,
3281,3282,3286,
3250,3260) # add on 06/10/2020


gcar_EMEA_office_code = (
3330, 3332, 3461, 6277, 3591, 6276, 3337, 3340, 3341, 3342, 3343, 33422, 
3345, 6412, 3219, 3348, 3349, 6415, 3351, 6418, 3225, 3226, 3354, 3356, 
3357, 6430, 3231, 3232, 3361, 3234, 3233, 3364, 3237, 3238, 3239, 3235, 
3236, 3240, 3371, 6444, 6447, 3377, 6457, 6423, 6334, 6335, 6207, 6337, 
6338, 3523, 3524, 3525, 3526, 7232, 7233, 6990, 6227, 6359, 6360, 6362, 
6363, 3550, 3424, 6431, 3301, 6246, 3304, 3561, 3305, 6249, 6762, 3309, 
3310, 6434, 3321)  # add on 09/01/2020


def func_sd(data, target_col='Final_PD_Risk_Rating', score_col='Prelim_PD_Risk_Rating', min_num=6):
    '''
    This function is to calculate SomersD between target_col and score_col.
    And it's designed for 'GroupBy.apply' and applied to group-wise object.
    New in Version 2.6

    Parameters:

        data:           the input dataset in DataFrame. 

        target_col:     str, default 'Final_PD_Risk_Rating'
                        column name for the targeted column.

        score_col:      str, default 'Prelim_PD_Risk_Rating'
                        column name for the score column.

        min_num:        int, default 6
                        the threshold. If the number of obs is less than it, the result is NA 
                        when the number of obs less than 

    Return:
        combined results is used in GroupBy.apply
    '''
    if len(data) < min_num:
        return np.nan
    else:
        return SomersD(data[target_col], data[score_col])




def func_rla(data, rla_col='RLA_Notches'):
    '''
    This function is to calculate RLA rate.
    And it's designed for 'GroupBy.apply' and applied to group-wise object.
    New in Version 2.6

    Parameters:

        data:       the input dataset in DataFrame. 

        rla_col:    str, default 'RLA_Notches'
                    column name for RLA column.

    Return:
        combined results is used in GroupBy.apply
    '''
    N = len(data)
    n_RLA = len(data.query(f'{rla_col}!=0'))    
    return (n_RLA/N)


def func_ovd(data, ovd_col='Override_Action'):
    '''
    This function is to calculate Override rate.
    And it's designed for 'GroupBy.apply' and applied to group-wise object.
    New in Version 2.6

    Parameters:

        data:       the input dataset in DataFrame. 

        ovd_col:    str, default 'Override_Action'
                    column name for Override.

    Return:
        combined results is used in GroupBy.apply
    '''

    N = len(data)   
    n_Override = len(data.query(f'{ovd_col}!=0'))
    return (n_Override/N)



#  Blueberry, Cinnabar, Selective Yellow, Sea Green
google_color = ['#4285F4','#EA4335','#FBBC05','#34A853']



def GICS_mapping(data, col_guidance, first_digits=2):
    """
    This function is to convert GICS code to industry. New in Version 2.9. 


    Parameters:

        data:           the input dataset in DataFrame. 

        col_guidance:   str
                        column name for GICS code

        first_digits:   integer [2,4,6,8], default 2. 
                        First digits used for the industry mapping

    Return:
        a DataFrame which is a copy of input data but has one more column: 'Industry_by_GICS'
    
    """    
    processed_data = data.copy()
    processed_data['Industry_by_GICS'] = processed_data[col_guidance].fillna(99999999)
    processed_data['Industry_by_GICS'] = [str(int(x))[:first_digits] for x in processed_data['Industry_by_GICS']]
    processed_data['Industry_by_GICS'] = processed_data['Industry_by_GICS'].replace(gics_code)
    
    return (processed_data)


def cluster_corr_old(dat, pl_cols):

    pl_cat1 = [pl_cols[0],]
    pl_cat2 = []

    for name in pl_cols[1:]:
        if np.abs(dat[[name]+pl_cat1].corr()).iloc[0,1:].mean() > 0.75:
            pl_cat1.append(name)
        else:
            pl_cat2.append(name)
    if len(pl_cat2)==0:
        return((pl_cat1,))
    else:
        return((pl_cat1,) + cluster_corr(dat, pl_cat2))


def correlation_dist(m, n):
    """
    This function is to calculate the distance of two variables based on correlation
    New in Version 3.0

    Parameters:

        m,n:   2 variables

    Return:
        distance, scalar
    
    """   

    df = pd.DataFrame({'m':m,'n':n})
    df.dropna(inplace=True)
    return(1-np.abs(pearsonr(df.m, df.n)[0]))


def reorder_cols(dat, cols):
    """
    This function is to re-rank the columns based on their correlation distance
    New in Version 3.0

    Parameters:

        dat:    the input dataset in DataFrame. 

        cols:   list of str
                column names


    Return:

        a list which is a re-ranked column names
    
    """    

    m = len(cols)
    X = dat[cols].T.to_numpy()
    k = 0
    dm = np.empty((m * (m - 1)) // 2, dtype=np.double)
    for i in range(0, m - 1):
        for j in range(i + 1, m):
            dm[k] = correlation_dist(X[i], X[j])
            k = k + 1

    yy = hierarchy.linkage(dm, method='average', optimal_ordering=False)
    Z = hierarchy.dendrogram(yy, no_plot=True, color_threshold=-np.inf)
    list_reorder = [cols[i] for i in Z['leaves']]

    return(list_reorder)


def cluster_corr(dat, list_cols):
    """
    This function is to group columns based on re-ranked columns names.
    New in Version 3.0
    
    Parameters:

        dat:        the input dataset in DataFrame. 

        list_cols:  list of str
                    column names


    Return:

        a tuple (grouped_names , re-ranked names)
    
    """    


    pl_cols = reorder_cols(dat, list_cols)
    pl_cat = []
    X = np.abs(dat[pl_cols].corr()).to_numpy()
    m = len(pl_cols)
    start=0
    for i in range(m-1):
        tmp = X[i+1,i]
        if tmp>=0.7:
            continue
        else:
            end=i
            pl_cat.append(pl_cols[start:(end+1)])
            start=end+1
            continue
    pl_cat.append(pl_cols[start:])
    return((pl_cat, pl_cols))


def plot_cluster_corr(dat, cat_list):
    """
    This function is to generate heatmap plot for the re-ranked cols
    New in Version 3.0
    
    Parameters:

        dat:        the input dataset in DataFrame. 

        cat_list:   tuple, the first output of function cluster_corr()

    """        
    cols = []
    for cat in cat_list:
        cols = cols+cat

    f, ax = plt.subplots(figsize=(10*len(cols)/25, 6*len(cols)/25))
    sns.heatmap(np.abs(dat[cols].corr()), linewidths=.3, cmap='Blues', ax=ax)


def pca_n_comp(dat, cols, threshold=0.9):
    df = dat[cols].dropna(subset=cols)
    scaler = StandardScaler()
    scaler.fit(df)
    df_pca = scaler.transform(df)

    for i in range(1,len(cols)+1):
        pca = decomposition.PCA(n_components=i)
        pca.fit(df_pca)
        if pca.explained_variance_ratio_.sum()>threshold:
            n = i
            explained_variance = pca.explained_variance_ratio_.sum()
            break
    return((n, explained_variance))


# Added in Ver. 3.0
compustat_Americas_country_code = ('USA','CAN','BMU','BRA','MEX','CHL','ARG',
    'CYM','COL','PER','BHS','BRB','DOM','HND','PAN')

# Added in Ver. 3.0
compustat_EMEA_country_code = ('GBR','FRA','NLD','DEU','CHE','IRL','SWE','LUX',
    'ESP','RUS','GRC','ITA','FIN','NOR','MCO','BEL','DNK','PRT','ZAF','HUN',
    'ISR','AUT','CYP','POL','TUR','UKR')





def buildpseudodef_btmu(data, BTMU='Primary Evaluation', threshold=81):
    """
    This function constructs pseudo defaults according the raings
    New in Version 3.1
    Parameters:

        data:       the input dataset in DataFrame. 

        BTMU:       str, default 'Primary Evaluation'
                    column name for BTMU Rating
                                  
        threshold:  integer, default 81
                    numerical value of pseudo default threshold
                    all observations who have equal or worse rating will be 
                    tagged as 'default'
        
    Return:
        DataFrame: Data with pseudo defaults.

    """
    #warnings.filterwarnings('ignore', category=Warning)
    dat = data.copy()
    dat['def_flag'] = np.where(dat[BTMU]>=threshold,1,0)
    n_pseudodef = dat['def_flag'].sum()
    print (f"There're {dat['def_flag'].sum()} pseudo defaults in the output dateset.")
    return(dat)# -*- coding: utf-8 -*-
"""
Created on Fri Feb 24 10:57:30 2017

Version: 1.0: Initial build
Version: 1.1, 20170317, Modify method barplot() to avoid negative y tick for mean PD
Version: 1.2, 20170320, Debug. Delete 'plt.show()' in qqplot() to avoid saving blank png.
Version: 1.3, 20170321, Mofify method barplot() by Jack to make it compatible to 
                        grade combinations, such as 'B/C'.
Version: 1.4, 20170327, Modify method __normalization() to correctly treat 'Invalid Negative'
                        in some quant factors in normalization procedure.
Version: 1.5, 20170329, Modify method barplot() to correctly plot mean PD point when its default
                        comes from floor(at least 1 default for each category)
Version: 1.6, 20170404, Modify method __Somersd() to accelerate the calculation and add one new 
                        method 'logitPDplot()' to plot the relationship between quant factor and 
                        obligor's logitPD.
Version: 1.7, 20170404, Modify method 'logitPDplot()' and add new method 'PDplot()'
Version: 1.8: 20170510, Modify the class to handle missing data. Minor change in __SomersD().

@author: ub71894 (4e8e6d0b), CSG
"""

import pandas as pd
import numpy as np
import warnings
import seaborn as sns
import matplotlib.pyplot as plt
import statsmodels.api as sm
from sklearn.metrics import roc_auc_score
import itertools
from numba import jit
from PDScorecardTool.Process import logitPD_frPDRR
#%%
class SFA(object):
    
    '''    
    This class is used to generate plots, tables and statistics related to SFA
    
    SFA(data, model):

        data:       data for development. Make sure it has column 'def_flag' as default
                    indicator and the factors' names are the same as the ones in 'model'
        model:      PDModle class. It saves all parameters for the model
    
        missing:    str, default 'median'. Added in Ver.1.8
                    statistics of the data used to fill missing data (NA data)

    Methods:

        describe()

        distplot(factors='quant', raw=True, savefig=True, clip=[0.01,0.99])

        barplot(factors='quali', savefig=True)

        qqplot(factors='quant', raw=True, savefig=True, clip=[0.01,0.99])

        violinplot(factors='quant', raw=True, savefig=True, clip=[0.01,0.99])

        boxplot(factors='quant', raw=True, savefig=True, clip=[0.01,0.99])

        logitPDplot(factors='quant', raw=True, PDsource='def_flag', bin_on='default', numofbins=10, order=1, savefig=True)
        
        corr(factors='all', raw=False, output=True)

        ARanalysis(factors='all', isthere_def=True, dependentvar='def_flag', output=True)
    

    Applicability:
                            |  Only quant   |  Only quali   |     ALL              
    ---------------------------------------------------------------------
    Only raw data:          |  describe()   |  barplot()    |
    ---------------------------------------------------------------------
                            |               |               |
    Only normalized data:   |               |               | ARanalysis()
    ---------------------------------------------------------------------
                            |               |               |
    Both:                   |  distplot()   |               | corr()
                            |  violinplot() |               |
                            |  boxplot()    |               |
                            |  qqplot()     |               | 
                            |  logitPDplot()|               |
                            |  PDplot()     |               |
    =====================================================================
    '''

    def __init__(self, data, model, missing='median'):

        self.model = model
        self.rawdata = data.copy()
        self.rawdata.reset_index(drop=True, inplace=True)
        self.stats = self.__describe()
        self.normdata = self.__normalization(missing)
        

    def __describe(self):
        '''

        Privata method that calculate key statistics for quant factors on raw data

        '''
        warnings.filterwarnings('ignore', category=Warning)
        dat = self.rawdata[self.model.quant_factor]
        stats = dat.describe(percentiles=[0.01,0.05,0.1,0.25,0.5,0.75,0.9,0.95,0.99])  
        temp={}
        temp['Skew'] = pd.Series(dat.skew())
        temp['Kurtosis'] = dat.kurtosis()
        stats = pd.concat([stats, pd.DataFrame(temp).T])
        return stats


    def __normalization(self, missing):
        '''

        Private method that apply normalization on financial factors.

        '''
        warnings.filterwarnings('ignore', category=Warning)
        # Invalid_Neg
        normdata = self.rawdata.copy()
        # modify the loop below in Ver 1.4:
        for i,neg_source in enumerate(self.model.Invalid_Neg):
            if neg_source:
                col=self.model.quant_factor[i]
                normdata[col][ (normdata[col]<0) & (normdata[neg_source]<0) ] = self.model.cap[i]
                # treat NA in 'neg_source' as negative value
                normdata[col][ (normdata[col]<0) & pd.isnull(normdata[neg_source]) ] = self.model.cap[i]

        # cap/floor for quant factors:
        for i, col in enumerate(self.model.quant_factor):
            normdata[col] = np.clip(normdata[col], self.model.floor[i], self.model.cap[i])        

        # calibration for quali factors:
        for i, col in enumerate(self.model.quali_factor):
            normdata[col].replace(to_replace=self.model.qualimapping[i],inplace=True) 

        # quant factors transformation:
        for i, col in enumerate(self.model.quant_factor):
            if self.model.quant_log[i]:
                normdata[col] = np.log(normdata[col])
            
        # quant factors normalization:
        for i, col in enumerate(self.model.quant_factor):
            normdata[col]  = 50*(normdata[col] - self.model.doc_mean[i]) / self.model.doc_std[i]      
        
        # quant factors flip sign:  
        for i, col in enumerate(self.model.quant_factor):
            normdata[col] = normdata[col] * self.model.quant_multiplier[i]     
        
        # fill missing data. added in Ver. 1.8
        exec('normdata[self.model.quant_factor+self.model.quali_factor] = normdata[self.model.quant_factor+self.model.quali_factor].\
            fillna(normdata[self.model.quant_factor+self.model.quali_factor].{missing}(), inplace=True)'.format(missing=missing))
        
        return normdata
    

    def __Spearman(self, factor, dependentvar) :
        '''

        Private method that calculate Spearman correlation.

        '''
        rho = self.ardata.corr(method='spearman')[factor][dependentvar]
        return rho


    def __Pearson(self, factor, dependentvar) :
        '''

        Private method that calculate Pearson correlation.

        '''       
        rho = self.ardata.corr(method='pearson')[factor][dependentvar]
        return rho        


    def __AUROC(self, factor, dependentvar, isthere_def) :
        '''

        Private method that calculate area under ROC.

        '''
        if isthere_def:
            y_score = self.ardata[factor].tolist()
            y_true  = [int(x) for x in list(self.ardata[dependentvar])] 
            return roc_auc_score(y_true, y_score) 
        else:
            return 'NA'


    def __SomersD_old(self, factor, dependentvar, sign=1, unit_concordance=True): # debugged on 20160817
        '''

        Private method that calculate SomersD using Josh's code. It's debugged 
        by me on 20160817. when dependentvar='def_flag', SomersD = 2*AUROC-1

        '''
        x = np.array(self.ardata[factor])
        y = np.array(self.ardata[dependentvar])
        n = len(x)
        C = 0
        D = 0
        if unit_concordance==True:
            for i, j in itertools.combinations(np.arange(n), 2):
                dx = x[i] - x[j]
                dy = y[i] - y[j]
                if (((dx<0) & (dy<0)) | ((dx>0) & (dy>0))):
                    cij = 1
                    dij = 1
                elif (((dx<0) & (dy>0)) | ((dx>0) & (dy<0))):
                    cij = -1
                    dij = 1
                else:
                    cij = 0
                    if dy==0: 
                        dij = 0
                    else: 
                        dij=1
                C = C + cij
                D = D + dij
        else:
            for i, j in itertools.combinations(np.arange(n), 2):
                dx = x[i] - x[j]
                dy = y[i] - y[j]
                if (((dx<0) & (dy<0)) | ((dx>0) & (dy>0))):
                    cij = abs(dy)
                    dij = abs(dy)
                elif (((dx<0) & (dy>0)) | ((dx>0) & (dy<0))):
                    cij = -abs(dy)
                    dij = abs(dy)
                else:
                    cij = 0
                    if dy==0: 
                        dij = 0
                    else: 
                        dij=1
                C = C + cij
                D = D + dij    
        
        return sign*C/D



    @jit
    def __SomersD(self, factor, dependentvar, sign=1, unit_concordance=True): 
        '''

        New version of SomersD function which leverages numba.jit to accelerate
        the calculation. In production from Ver. 1.6.

        '''
        x = np.array(self.ardata[factor])
        # added in Ver 1.8 to deal with missing PDRR sometimes.
        y_true = self.normdata[dependentvar].copy()
        y_true.dropna(inplace=True)
        y = np.array(y_true)
        n = len(x)
        C = 0
        D = 0
        if unit_concordance==True:
            for i in range(n):
                for j in range(i):
                    dx = x[i] - x[j]
                    dy = y[i] - y[j]
                    if (dx*dy) > 0:
                        cij = 1
                        dij = 1
                    elif  (dx*dy) < 0:
                        cij = -1
                        dij = 1
                    else:
                        cij = 0
                        if dy==0: 
                            dij = 0
                        else: 
                            dij=1
                    C = C + cij
                    D = D + dij
        else:
            for i in range(n):
                for j in range(i):
                    dx = x[i] - x[j]
                    dy = y[i] - y[j]
                    if  (dx*dy) > 0:
                        cij = abs(dy)
                        dij = abs(dy)
                    elif (dx*dy) < 0:
                        cij = -abs(dy)
                        dij = abs(dy)
                    else:
                        cij = 0
                        if dy==0: 
                            dij = 0
                        else: 
                            dij=1
                    C = C + cij
                    D = D + dij    
        
        return sign*C/D


    def describe(self):
        '''

        Public method that present key statistics for quant factors on raw data
        
        '''
        return self.stats
        

    def distplot(self, factors='quant', raw=True, savefig=True, clip=[0.01,0.99]):
        '''

        Public method that present distribution plot for quant factors
        It will ignore +-inf and NA data
        Note that, for raw quant factors, it makes plot based on clipped data 
        and exclude outliers.
        
        Parameters: 

        factors:    str {'quant','all', factorname} or None, default 'quant'
                    quant factors for plotting. 

        raw:        boolean, default True
                    whether using raw data or normalized data

        savefig:    boolean, default True
                    whether save png file in current working directory.  

        clip:       list of 2 scalars. default [0.01,0.99]
                    Lower and upper bounds percentiles for datapoints used to fit
                    ONLY work for raw data
        '''
        warnings.filterwarnings('ignore', category=Warning)
        if factors=='all' or factors=='quant':
            factors= self.model.quant_factor 
        else: # user input a single factor's name
            factors = [factors]
        if raw:
            suffix = '_raw' + str(int(clip[0]*100))+'-'+str(int(clip[1]*100))
            for factor in factors:
                # only use clipped data and exclude outliers:
                floor, cap = self.rawdata[factor].quantile(clip)
                dataforplot = self.rawdata.query('{floor}<{factor}<{cap}'.format(factor=factor,floor=floor,cap=cap))            
                # 3 plots: hist, pdf and cdf
                fig,ax=plt.subplots(1,3,figsize=(18,6))
                fig.suptitle(factor+suffix)
                sns.distplot(dataforplot[factor], kde=False , ax=ax[0], color='g')
                ax[0].set_xlabel('Histogram')
                sns.distplot(dataforplot[factor], hist=False, kde_kws=dict(shade=True), ax=ax[1], color='g')
                ax[1].set_xlabel('Probability density function')
                sns.distplot(dataforplot[factor], hist=False, kde_kws=dict(cumulative=True,shade=True), ax=ax[2], color='g')
                ax[2].set_xlabel('Cumulative density function')
                if savefig:
                    fig.savefig(factor+suffix+'_quantdist.png')
        else:
            suffix = '_norm'
            for factor in factors:     
                # 3 plots: hist, pdf and cdf    
                fig,ax=plt.subplots(1,3,figsize=(18,6))
                fig.suptitle(factor+suffix)
                sns.distplot(self.normdata[factor], kde=False , ax=ax[0], color='g')
                ax[0].set_xlabel('Histogram')
                sns.distplot(self.normdata[factor], hist=False, kde_kws=dict(shade=True), ax=ax[1], color='g')
                ax[1].set_xlabel('Probability density function')
                sns.distplot(self.normdata[factor], hist=False, kde_kws=dict(cumulative=True,shade=True), ax=ax[2], color='g')
                ax[2].set_xlabel('Cumulative density function')
                if savefig:
                    fig.savefig(factor+suffix+'_quantdist.png')


    def barplot(self, factors='quali', savefig=True):
        '''

        Public method that present distribution plot for quali factors
        It will ignore NA data

        Parameters: 

        factors:    str {'quali','all', factorname} or None, default 'quali'
                    quali factors for plotting. 

        savefig:    boolean, default True
                    whether save png file in current working directory.  
        
        '''

        
        # add this adjusted mean function in Ver 1.5 for plotting
        def _adjmean(x):
            if x.sum()==0:
                return (1/x.size)
            else:
                return x.mean()

        warnings.filterwarnings('ignore', category=Warning)
        if factors=='all' or factors=='quali':
            factors= self.model.quali_factor
        else: # user input a single factor's name
            factors = [factors]

        for factor in factors:
            letter = list(set(self.rawdata[factor])) # add in Ver. 1.3
            letter.sort()
            fig = plt.figure()
            fig.suptitle(factor)
            ax1 = fig.add_subplot(111)
            sns.countplot(x=factor, data=self.rawdata, ax=ax1,order=letter, alpha=0.7, palette='Blues_d')
            ax1.set_xlabel('Answers')
            ax2 = ax1.twinx()
            sns.pointplot(x=factor, y='def_flag', data=self.rawdata, ax=ax2, order=letter, ci=0, estimator=_adjmean) # modified in Ver 1.5
            ax2.grid(None)
            ax2.set_ylabel('mean of PD')
            ax2.set_ylim(bottom=0) # add this line in Ver 1.1
            if savefig:
                fig.savefig(factor+'_qualibar.png')


    def qqplot(self, factors='quant', raw=True, savefig=True, clip=[0.01,0.99]):
        '''

        Public method that make qq plot on quant factors
        It uses statsmodels.api.qqplot to create graph.
        Note that:  1. for raw quant factors, it makes plot based on clipped data 
                      and exclude outliers.
                    2. Usually this plot is trival for our modeling work since we 
                        don't expect the quant factor has normal distribution.
        Parameters: 

        factors:    str {'quant','all', factorname} or None, default 'quant'
                    quant factors for plotting. 

        raw:        boolean, default True
                    whether using raw data or normalized data.

        savefig:    boolean, default True
                    whether save png file in current working directory. 

        clip:       list of 2 scalars. default [0.01,0.99]
                    Lower and upper bounds percentiles for datapoints used to fit
                    ONLY work for raw data
        '''
        warnings.filterwarnings('ignore', category=Warning)
        if factors=='quant' or factors=='all':
            factors= self.model.quant_factor
        else: # user input a single factor's name
            factors = [factors]
        if raw:
            suffix = '_raw' + str(int(clip[0]*100))+'-'+str(int(clip[1]*100))
            for factor in factors:
                # only use clipped data and exclude outliers:
                floor, cap = self.rawdata[factor].quantile(clip)
                dataforplot = self.rawdata.query('{floor}<{factor}<{cap}'.format(factor=factor,floor=floor,cap=cap))  
                dataforplot = dataforplot[factor]
                sm.qqplot(dataforplot,fit=True,line='45')
                plt.title(dataforplot.name)
                if savefig:
                    plt.savefig(factor+suffix+'_quantqqplot.png')
        else:
            suffix = '_norm'
            for factor in factors:
                dataforplot= self.normdata[factor]
                sm.qqplot(dataforplot,fit=True,line='45')
                plt.title(dataforplot.name)
                if savefig:
                    plt.savefig(factor+suffix+'_quantqqplot.png')


    def violinplot(self, factors='quant', raw=True, savefig=True, clip=[0.01,0.99]):
        '''

        Public method that make violin plot to show distribution and quartile of
        the quant factors. It splits data into default(bad) and non-default(good)
        and seaborn.violinplot() to create the graph.

        Note that, for raw quant factors, it makes plot based on clipped data 
        and exclude outliers.

        Parameters: 

        factors:    str {'quant','all', factorname} or None, default 'quant'
                    quant factors for plotting. 

        raw:        boolean, default True
                    whether using raw data or normalized data.

        savefig:    boolean, default True
                    whether save png file in current working directory. 

        clip:       list of 2 scalars. default [0.01,0.99]
                    Lower and upper bounds percentiles for datapoints used to fit
                    ONLY work for raw data
        '''
        warnings.filterwarnings('ignore', category=Warning)
        if factors=='quant' or factors=='all':
            factors= self.model.quant_factor
        else: # user input a single factor's name
            factors = [factors]
        if raw:
            suffix = '_raw' + str(int(clip[0]*100))+'-'+str(int(clip[1]*100))
            for factor in factors:
                # only use clipped data and exclude outliers:
                floor, cap = self.rawdata[factor].quantile(clip)
                dataforplot = self.rawdata.query('{floor}<{factor}<{cap}'.format(factor=factor,floor=floor,cap=cap))  
                dataforplot["all"] = ""
                dataforplot['event'] = dataforplot['def_flag'].replace(to_replace={1:'Default',0:'Non-Default'})
                fig,ax=plt.subplots(1,1,figsize=(8,8))
                ax = sns.violinplot(x='all', y=factor, hue='event', inner="quartile", data=dataforplot, split=True,  cut=0, palette="Paired")
                if savefig:
                    fig.savefig(factor+suffix+'_quantviolinplot.png')
        else:
            suffix = '_norm'
            dataforplot= self.normdata.copy()
            for factor in factors:
                dataforplot["all"] = ""
                dataforplot['event'] = dataforplot['def_flag'].replace(to_replace={1:'Default',0:'Non-Default'})
                fig,ax=plt.subplots(1,1,figsize=(8,8))
                ax = sns.violinplot(x='all', y=factor, hue='event', inner="quartile", data=dataforplot, split=True,  cut=0, palette="Paired")
                if savefig:
                    fig.savefig(factor+suffix+'_quantviolinplot.png')  


    def boxplot(self, factors='quant', raw=True, savefig=True, clip=[0.01,0.99]):
        '''

        Public method that make box plot to show distribution and quartile of
        the quant factors. It splits data into default(bad) and non-default(good)
        and seaborn.boxplot() to create the graph.

        Note that, for raw quant factors, it makes plot based on clipped data 
        and exclude outliers.

        Parameters: 

        factors:    str {'quant','all', factorname} or None, default 'quant'
                    quant factors for plotting. 

        raw:        boolean, default True
                    whether using raw data or normalized data.

        savefig:    boolean, default True
                    whether save png file in current working directory. 

        clip:       list of 2 scalars. default [0.01,0.99]
                    Lower and upper bounds percentiles for datapoints used to fit
                    ONLY work for raw data
        '''
        warnings.filterwarnings('ignore', category=Warning)
        if factors=='quant' or factors=='all':
            factors= self.model.quant_factor
        else: # user input a single factor's name
            factors = [factors]
        if raw:
            suffix = '_raw_clip' + str(int(clip[0]*100))+'-'+str(int(clip[1]*100))
            for factor in factors:
                # only use clipped data and exclude outliers:
                floor, cap = self.rawdata[factor].quantile(clip)
                dataforplot = self.rawdata.query('{floor}<{factor}<{cap}'.format(factor=factor,floor=floor,cap=cap))  
                dataforplot["all"] = ""
                dataforplot['event'] = dataforplot['def_flag'].replace(to_replace={1:'Default',0:'Non-Default'})
                fig,ax=plt.subplots(1,1,figsize=(8,8))
                ax = sns.boxplot(x="all", y=factor, hue='event', data=dataforplot, palette="Paired")
                if savefig:
                    fig.savefig(factor+suffix+'_quantboxplot.png')
        else:
            suffix = '_norm'
            dataforplot = self.normdata.copy()
            for factor in factors:
                dataforplot["all"] = ""
                dataforplot['event'] = dataforplot['def_flag'].replace(to_replace={1:'Default',0:'Non-Default'})
                fig,ax=plt.subplots(1,1,figsize=(8,8))
                ax = sns.boxplot(x="all", y=factor, hue='event', data=dataforplot, palette="Paired")
                if savefig:
                    fig.savefig(factor+suffix+'_quantboxplot.png')       

    
    def logitPDplot(self, factors='quant', raw=True, PDsource='def_flag', bin_on='default', numofbins=10, \
        plot_on_binnum=False, order=1,  lowess=False, savefig=True):
        '''

        Public method that plots the relationship between logitPD to factor value 


        Parameters: 

        factors:        str {'quant','all', factorname} or None, default 'quant'
                        quant factors for plotting. 

        raw:            boolean, default True
                        whether using raw data or normalized data

        PDsource:       str, default 'def_flag'
                        the name of the column which is the source of PD. For true default
                        empirical PD, PDsource='def_flag'; For non-true default PD, user can
                        assign a PDRR name. ex. PDsource='Final_PD_Risk_Rating'

        bin_on:         {'default','obligor'}, default 'default'
                        when PDsource='def_flag', user needs to choose the even binning is on
                        the number of default or the number of obligor. In other words, user 
                        needs to choose "even obligors" or "even defaults" in each bin. 

        numofbins:      int, default 10
                        the number of bins in binning procedure.

        plot_on_binnum: boolean, default False, Added in Ver. 1.7
                        plot logitPD vs. factor mean in each bin, if False
                        plot logitPD vs. the sequence number of each bin, if True

        order:          int, default 1
                        the plot will estimate a linear regression in default setting.
                        if order>1, the plot will estimate a polynomial regression.

        lowess:         boolean, default False, Added in Ver. 1.7
                        if True, estimate a nonparametric lowess model (locally weighted linear regression)
        
        savefig:        boolean, default True
                        whether save png file in current working directory. 

        '''

        if factors=='quant' or factors=='all':
            factors= self.model.quant_factor
        else: # user input a single factor's name
            factors = [factors]

        if raw:
            data = self.rawdata.copy()
            suffix = '_raw'
        else:
            data = self.normdata.copy()
            suffix = '_norm'


        if (PDsource=='def_flag') & (bin_on=='default'):
            suffix += '_evendef_(#bins={nbins}_order={order})'.format(nbins=numofbins,order=order)
            for factor in factors:
                data_plot = data[[factor]+['def_flag']].copy()
                # get the bins from default data:
                temp = data_plot.query("def_flag>0")                
                try:
                    _,bins = pd.qcut(temp[factor], numofbins, retbins=True)
                    bins[0] = -np.inf; bins[-1] = np.inf    # Added in Ver. 1.7                
                except ValueError:
                    print('Not enought defaults for the binning. Try smaller number of bins or Even-Obligors binning method')
                    return 0
                # apply the bins to all data
                data_plot['bin'] = pd.cut(data_plot[factor], bins, labels=list(range(1,numofbins+1))) # Added in Ver. 1.7
                bucket = data_plot.groupby(by='bin').mean()
                bucket['count'] = data_plot.groupby(by='bin').count().def_flag
                bucket['Ndef'] = data_plot.groupby(by='bin').sum().def_flag
                bucket['logit_PD'] = np.array([np.log(x/(1-x)) for x in bucket.def_flag.tolist()])
                bucket.reset_index(drop=False, inplace=True) # Added in Ver. 1.7
                # kick out -inf in logit PD due to 0 default in some bin
                bucket = bucket.query('logit_PD > {invalid}'.format(invalid=-np.inf))                
                # scatter plot with regression 
                if plot_on_binnum: # Added in Ver. 1.7
                    if lowess:
                        g = sns.lmplot(x='bin', y="logit_PD", data=bucket, lowess=lowess, ci=None, scatter_kws={"s": 20})
                    else:
                        g = sns.lmplot(x='bin', y="logit_PD", data=bucket, order=order, ci=None, scatter_kws={"s": 20})
                else:
                    if lowess:
                        g = sns.lmplot(x=factor, y="logit_PD", data=bucket, lowess=lowess, ci=None, scatter_kws={"s": 20})
                    else:
                        g = sns.lmplot(x=factor, y="logit_PD", data=bucket, order=order, ci=None, scatter_kws={"s": 20})
                
                if savefig:
                    g.savefig(factor+suffix+'_logitPDplot.png')


        elif (PDsource=='def_flag') & (bin_on=='obligor'):
            suffix += '_evenobligor_(#bins={nbins}_order={order})'.format(nbins=numofbins,order=order)
            for factor in factors:
                data_plot = data[[factor]+['def_flag']].copy()  
                try:
                    data_plot['bin'] = pd.qcut(data_plot[factor], numofbins, labels=list(range(1,numofbins+1))) # Added in Ver. 1.7
                except ValueError:
                    print('Too many obligors have the same value due to cap/floor. Try smaller number of bins.')
                    return 0
                bucket = data_plot.groupby(by='bin').mean()
                bucket['count'] = data_plot.groupby(by='bin').count().def_flag
                bucket['Ndef'] = data_plot.groupby(by='bin').sum().def_flag
                bucket['logit_PD'] = np.array([np.log(x/(1-x)) for x in bucket.def_flag.tolist()])
                bucket.reset_index(drop=False, inplace=True) # Added in Ver. 1.7
                # kick out -inf in logit PD due to 0 default in some bin
                bucket = bucket.query('logit_PD > {invalid}'.format(invalid=-np.inf))
                # scatter plot with regression 
                if plot_on_binnum: # Added in Ver. 1.7
                    if lowess:
                        g = sns.lmplot(x='bin', y="logit_PD", data=bucket, lowess=lowess, ci=None, scatter_kws={"s": 20})
                    else:
                        g = sns.lmplot(x='bin', y="logit_PD", data=bucket, order=order, ci=None, scatter_kws={"s": 20})
                else:
                    if lowess:
                        g = sns.lmplot(x=factor, y="logit_PD", data=bucket, lowess=lowess, ci=None, scatter_kws={"s": 20})
                    else:
                        g = sns.lmplot(x=factor, y="logit_PD", data=bucket, order=order, ci=None, scatter_kws={"s": 20})

                if savefig:
                    g.savefig(factor+suffix+'_logitPDplot.png')
        

        elif PDsource: # means PDsource has been assigned
            try:
                data[PDsource]
            except KeyError:
                print('No column in dataset is called "{PDsource}".'.format(PDsource=PDsource))
                return 0
            suffix += '_{PDRR}_(order={order})'.format(PDRR=PDsource,order=order)
            # prepare PDRR implied logitPD of each obligor:
            data = logitPD_frPDRR(data, self.model, PDsource)
            for factor in factors:
                # scatter plot with regression 
                g = sns.lmplot(x=factor, y="logitPD_frPDRR", data=data, order=order, ci=None, scatter_kws={"s": 20})
                if savefig:    
                    g.savefig(factor+suffix+'_logitPDplot.png')    


        else:
            return 0


    def PDplot(self, factors='quant', raw=True, PDsource='def_flag', bin_on='default', numofbins=10, \
        plot_on_binnum=False, order=1,  lowess=False, savefig=True):
        '''

        Public method that plots the relationship between PD to factor value 


        Parameters: 

        factors:        str {'quant','all', factorname} or None, default 'quant'
                        quant factors for plotting. 

        raw:            boolean, default True
                        whether using raw data or normalized data

        PDsource:       str, default 'def_flag'
                        the name of the column which is the source of PD. For true default
                        empirical PD, PDsource='def_flag'; For non-true default PD, user can
                        assign a PDRR name. ex. PDsource='Final_PD_Risk_Rating'

        bin_on:         {'default','obligor'}, default 'default'
                        when PDsource='def_flag', user needs to choose the even binning is on
                        the number of default or the number of obligor. In other words, user 
                        needs to choose "even obligors" or "even defaults" in each bin. 

        numofbins:      int, default 10
                        the number of bins in binning procedure.

        plot_on_binnum: boolean, default False, Added in Ver. 1.7
                        plot logitPD vs. factor mean in each bin, if False
                        plot logitPD vs. the sequence number of each bin, if True

        order:          int, default 1
                        the plot will estimate a linear regression in default setting.
                        if order>1, the plot will estimate a polynomial regression.
        
        lowess:         boolean, default False, Added in Ver. 1.7
                        if True, estimate a nonparametric lowess model (locally weighted linear regression)
        
        savefig:        boolean, default True
                        whether save png file in current working directory. 

        '''

        if factors=='quant' or factors=='all':
            factors= self.model.quant_factor
        else: # user input a single factor's name
            factors = [factors]

        if raw:
            data = self.rawdata.copy()
            suffix = '_raw'
        else:
            data = self.normdata.copy()
            suffix = '_norm'


        if (PDsource=='def_flag') & (bin_on=='default'):
            suffix += '_evendef_(#bins={nbins}_order={order})'.format(nbins=numofbins,order=order)
            for factor in factors:
                data_plot = data[[factor]+['def_flag']].copy()
                # get the bins from default data:
                temp = data_plot.query("def_flag>0")                
                try:
                    _,bins = pd.qcut(temp[factor], numofbins, retbins=True)
                    bins[0] = -np.inf; bins[-1] = np.inf    # Added in Ver. 1.7                
                except ValueError:
                    print('Not enought defaults for the binning. Try smaller number of bins or Even-Obligors binning method')
                    return 0
                # apply the bins to all data
                data_plot['bin'] = pd.cut(data_plot[factor], bins, labels=list(range(1,numofbins+1))) # Added in Ver. 1.7
                bucket = data_plot.groupby(by='bin').mean()
                bucket['count'] = data_plot.groupby(by='bin').count().def_flag
                bucket['Ndef'] = data_plot.groupby(by='bin').sum().def_flag
                bucket['logit_PD'] = np.array([np.log(x/(1-x)) for x in bucket.def_flag.tolist()])
                bucket.reset_index(drop=False, inplace=True) # Added in Ver. 1.7
                # kick out -inf in logit PD due to 0 default in some bin
                #bucket = bucket.query('logit_PD > {invalid}'.format(invalid=-np.inf))                
                # scatter plot with regression 
                if plot_on_binnum: # Added in Ver. 1.7
                    if lowess:
                        g = sns.lmplot(x='bin', y="def_flag", data=bucket, lowess=lowess, ci=None, scatter_kws={"s": 20})
                    else:
                        g = sns.lmplot(x='bin', y="def_flag", data=bucket, order=order, ci=None, scatter_kws={"s": 20})
                else:
                    if lowess:
                        g = sns.lmplot(x=factor, y="def_flag", data=bucket, lowess=lowess, ci=None, scatter_kws={"s": 20})
                    else:
                        g = sns.lmplot(x=factor, y="def_flag", data=bucket, order=order, ci=None, scatter_kws={"s": 20})

                if savefig:
                    g.savefig(factor+suffix+'_PDplot.png')


        elif (PDsource=='def_flag') & (bin_on=='obligor'):
            suffix += '_evenobligor_(#bins={nbins}_order={order})'.format(nbins=numofbins,order=order)
            for factor in factors:
                data_plot = data[[factor]+['def_flag']].copy()  
                try:
                    data_plot['bin'] = pd.qcut(data_plot[factor], numofbins, labels=list(range(1,numofbins+1))) # Added in Ver. 1.7
                except ValueError:
                    print('Too many obligors have the same value due to cap/floor. Try smaller number of bins.')
                    return 0
                bucket = data_plot.groupby(by='bin').mean()
                bucket['count'] = data_plot.groupby(by='bin').count().def_flag
                bucket['Ndef'] = data_plot.groupby(by='bin').sum().def_flag
                bucket['logit_PD'] = np.array([np.log(x/(1-x)) for x in bucket.def_flag.tolist()])
                bucket.reset_index(drop=False, inplace=True) # Added in Ver. 1.7
                # kick out -inf in logit PD due to 0 default in some bin
                #bucket = bucket.query('logit_PD > {invalid}'.format(invalid=-np.inf))
                # scatter plot with regression 
                if plot_on_binnum: # Added in Ver. 1.7
                    if lowess:
                        g = sns.lmplot(x='bin', y="def_flag", data=bucket, lowess=lowess, ci=None, scatter_kws={"s": 20})
                    else:
                        g = sns.lmplot(x='bin', y="def_flag", data=bucket, order=order, ci=None, scatter_kws={"s": 20})
                else:
                    if lowess:
                        g = sns.lmplot(x=factor, y="def_flag", data=bucket, lowess=lowess, ci=None, scatter_kws={"s": 20})
                    else:
                        g = sns.lmplot(x=factor, y="def_flag", data=bucket, order=order, ci=None, scatter_kws={"s": 20})
                    
                if savefig:
                    g.savefig(factor+suffix+'_PDplot.png')
        

        elif PDsource: # means PDsource has been assigned
            try:
                data[PDsource]
            except KeyError:
                print('No column in dataset is called "{PDsource}".'.format(PDsource=PDsource))
                return 0
            suffix += '_{PDRR}_(order={order})'.format(PDRR=PDsource,order=order)
            # prepare PDRR implied logitPD of each obligor:
            data = PD_frPDRR(data, self.model, PDsource)
            for factor in factors:
                # scatter plot with regression 
                g = sns.lmplot(x=factor, y="PD_frPDRR", data=data, order=order, ci=None, scatter_kws={"s": 20})
                if savefig:    
                    g.savefig(factor+suffix+'_PDplot.png')    


        else:
            return 0


    def corr(self, factors='all', raw=False, output=True):
        '''

        Public method that calculate correlation matrix for multifactors

        Parameters: 

        factors:    str {'quant', 'quali', 'all'} or None, default 'all'
                    Multifactors for matrix calculations. 

        raw:        boolean, default False
                    whether using raw data or normalized data. Default setting is
                    False since usually we care about the corr of normalized factors. 

        output:     boolean, default True
                    whether save the corr matrix to excel file.

        '''
        warnings.filterwarnings('ignore', category=Warning)
        if factors=='quant':
            names= self.model.quant_factor
        elif factors=='quali':
            names= self.model.quali_factor
        else: # all factors
            names = self.model.quant_factor + self.model.quali_factor
        
        if raw:
            suffix = '_raw'
            corr = self.rawdata[names].corr()
            if output:
                corr.to_excel('CorrMat_'+factors+suffix+'.xlsx')
            return corr

        else:
            suffix = '_norm'
            corr = self.normdata[names].corr()
            if output:
                corr.to_excel('CorrMat_'+factors+suffix+'.xlsx')
            return corr    


    def ARanalysis(self, factors='all', isthere_def=True, dependentvar='def_flag', output=True):
        '''
        
        Public method that calculate correlation matrix for multifactors

        Parameters: 

        factors:        str {'quant','quali', all', factorname}, default 'all'
                        Multifactors for matrix calculations. 

        isthere_def:    boolean, default True
                        whether the analysis is on the default event. If not, for 
                        example we want to check the accuracy of factor on PDRR,
                        then we should input  
                        isthere_def=False, dependentvar='PDRR'

        dependentvar:   str, default 'def_flag'
                        column name for dependent variable.  
                                    
        output:         boolean, default True
                        whether save the corr matrix to excel file.

        '''
        warnings.filterwarnings('ignore', category=Warning)
        if factors=='quant':
            names= self.model.quant_factor
        elif factors=='quali':
            names= self.model.quali_factor
        elif factors=='all':
            names= self.model.quant_factor + self.model.quali_factor 
        else: # user input a single factor's name
            names = [factors]   

        self.ardata = self.normdata[names+[dependentvar]].copy()
        
        d={}
        for factor in names:
            rho = self.__Pearson(factor, dependentvar)
            spearmanrho = self.__Spearman(factor, dependentvar)
            AUROC = self.__AUROC(factor, dependentvar, isthere_def)
            if isthere_def:
                SomersD = 2*AUROC-1
            else:
                SomersD = self.__SomersD(factor, dependentvar)
            d.update({factor:pd.Series([rho, spearmanrho, AUROC, SomersD], index= ['Correlation','Spearman','AUROC','SomersD'])})   
        result = pd.DataFrame(d, columns=[names])
        if output:
            result.to_excel('AR_'+factors+'.xlsx')
        return result            
        
naics_code ={
"11":"Agriculture, Forestry, Fishing and Hunting",
"111":"Crop Production",
"1111":"Oilseed and Grain Farming",
"11111":"Soybean Farming",
"111110":"Soybean Farming",
"11112":"Oilseed (except Soybean) Farming",
"111120":"Oilseed (except Soybean) Farming ",
"11113":"Dry Pea and Bean Farming",
"111130":"Dry Pea and Bean Farming ",
"11114":"Wheat Farming",
"111140":"Wheat Farming",
"11115":"Corn Farming",
"111150":"Corn Farming ",
"11116":"Rice Farming",
"111160":"Rice Farming",
"11119":"Other Grain Farming",
"111191":"Oilseed and Grain Combination Farming ",
"111199":"All Other Grain Farming ",
"1112":"Vegetable and Melon Farming",
"11121":"Vegetable and Melon Farming",
"111211":"Potato Farming ",
"111219":"Other Vegetable (except Potato) and Melon Farming ",
"1113":"Fruit and Tree Nut Farming",
"11131":"Orange Groves",
"111310":"Orange Groves",
"11132":"Citrus (except Orange) Groves",
"111320":"Citrus (except Orange) Groves ",
"11133":"Noncitrus Fruit and Tree Nut Farming",
"111331":"Apple Orchards ",
"111332":"Grape Vineyards ",
"111333":"Strawberry Farming ",
"111334":"Berry (except Strawberry) Farming ",
"111335":"Tree Nut Farming ",
"111336":"Fruit and Tree Nut Combination Farming ",
"111339":"Other Noncitrus Fruit Farming ",
"1114":"Greenhouse, Nursery, and Floriculture Production",
"11141":"Food Crops Grown Under Cover",
"111411":"Mushroom Production ",
"111419":"Other Food Crops Grown Under Cover ",
"11142":"Nursery and Floriculture Production",
"111421":"Nursery and Tree Production ",
"111422":"Floriculture Production ",
"1119":"Other Crop Farming",
"11191":"Tobacco Farming",
"111910":"Tobacco Farming",
"11192":"Cotton Farming",
"111920":"Cotton Farming",
"11193":"Sugarcane Farming",
"111930":"Sugarcane Farming",
"11194":"Hay Farming",
"111940":"Hay Farming ",
"11199":"All Other Crop Farming",
"111991":"Sugar Beet Farming ",
"111992":"Peanut Farming ",
"111998":"All Other Miscellaneous Crop Farming ",
"112":"Animal Production and Aquaculture",
"1121":"Cattle Ranching and Farming",
"11211":"Beef Cattle Ranching and Farming, including Feedlots",
"112111":"Beef Cattle Ranching and Farming ",
"112112":"Cattle Feedlots ",
"11212":"Dairy Cattle and Milk Production",
"112120":"Dairy Cattle and Milk Production",
"11213":"Dual-Purpose Cattle Ranching and Farming",
"112130":"Dual-Purpose Cattle Ranching and Farming ",
"1122":"Hog and Pig Farming",
"11221":"Hog and Pig Farming",
"112210":"Hog and Pig Farming ",
"1123":"Poultry and Egg Production",
"11231":"Chicken Egg Production",
"112310":"Chicken Egg Production ",
"11232":"Broilers and Other Meat Type Chicken Production",
"112320":"Broilers and Other Meat Type Chicken Production ",
"11233":"Turkey Production",
"112330":"Turkey Production",
"11234":"Poultry Hatcheries",
"112340":"Poultry Hatcheries",
"11239":"Other Poultry Production",
"112390":"Other Poultry Production ",
"1124":"Sheep and Goat Farming",
"11241":"Sheep Farming",
"112410":"Sheep Farming",
"11242":"Goat Farming",
"112420":"Goat Farming",
"1125":"Aquaculture",
"11251":"Aquaculture",
"112511":"Finfish Farming and Fish Hatcheries ",
"112512":"Shellfish Farming ",
"112519":"Other Aquaculture ",
"1129":"Other Animal Production",
"11291":"Apiculture",
"112910":"Apiculture",
"11292":"Horses and Other Equine Production",
"112920":"Horses and Other Equine Production",
"11293":"Fur-Bearing Animal and Rabbit Production",
"112930":"Fur-Bearing Animal and Rabbit Production",
"11299":"All Other Animal Production",
"112990":"All Other Animal Production ",
"113":"Forestry and Logging",
"1131":"Timber Tract Operations",
"11311":"Timber Tract Operations",
"113110":"Timber Tract Operations",
"1132":"Forest Nurseries and Gathering of Forest Products",
"11321":"Forest Nurseries and Gathering of Forest Products",
"113210":"Forest Nurseries and Gathering of Forest Products ",
"1133":"Logging",
"11331":"Logging",
"113310":"Logging ",
"114":"Fishing, Hunting and Trapping",
"1141":"Fishing",
"11411":"Fishing",
"114111":"Finfish Fishing ",
"114112":"Shellfish Fishing ",
"114119":"Other Marine Fishing ",
"1142":"Hunting and Trapping",
"11421":"Hunting and Trapping",
"114210":"Hunting and Trapping",
"115":"Support Activities for Agriculture and Forestry",
"1151":"Support Activities for Crop Production",
"11511":"Support Activities for Crop Production",
"115111":"Cotton Ginning ",
"115112":"Soil Preparation, Planting, and Cultivating ",
"115113":"Crop Harvesting, Primarily by Machine ",
"115114":"Postharvest Crop Activities (except Cotton Ginning) ",
"115115":"Farm Labor Contractors and Crew Leaders ",
"115116":"Farm Management Services ",
"1152":"Support Activities for Animal Production",
"11521":"Support Activities for Animal Production",
"115210":"Support Activities for Animal Production",
"1153":"Support Activities for Forestry",
"11531":"Support Activities for Forestry",
"115310":"Support Activities for Forestry",
"21":"Mining, Quarrying, and Oil and Gas Extraction",
"211":"Oil and Gas Extraction",
"2111":"Oil and Gas Extraction",
"21112":"Crude Petroleum Extraction ",
"211120":"Crude Petroleum Extraction ",
"21113":"Natural Gas Extraction ",
"211130":"Natural Gas Extraction ",
"212":"Mining (except Oil and Gas)",
"2121":"Coal Mining",
"21211":"Coal Mining",
"212111":"Bituminous Coal and Lignite Surface Mining ",
"212112":"Bituminous Coal Underground Mining ",
"212113":"Anthracite Mining ",
"2122":"Metal Ore Mining",
"21221":"Iron Ore Mining",
"212210":"Iron Ore Mining",
"21222":"Gold Ore and Silver Ore Mining",
"212221":"Gold Ore Mining ",
"212222":"Silver Ore Mining ",
"21223":"Copper, Nickel, Lead, and Zinc Mining",
"212230":"Copper, Nickel, Lead, and Zinc Mining ",
"21229":"Other Metal Ore Mining",
"212291":"Uranium-Radium-Vanadium Ore Mining ",
"212299":"All Other Metal Ore Mining ",
"2123":"Nonmetallic Mineral Mining and Quarrying",
"21231":"Stone Mining and Quarrying",
"212311":"Dimension Stone Mining and Quarrying ",
"212312":"Crushed and Broken Limestone Mining and Quarrying ",
"212313":"Crushed and Broken Granite Mining and Quarrying ",
"212319":"Other Crushed and Broken Stone Mining and Quarrying ",
"21232":"Sand, Gravel, Clay, and Ceramic and Refractory Minerals Mining and Quarrying",
"212321":"Construction Sand and Gravel Mining ",
"212322":"Industrial Sand Mining ",
"212324":"Kaolin and Ball Clay Mining ",
"212325":"Clay and Ceramic and Refractory Minerals Mining ",
"21239":"Other Nonmetallic Mineral Mining and Quarrying",
"212391":"Potash, Soda, and Borate Mineral Mining ",
"212392":"Phosphate Rock Mining ",
"212393":"Other Chemical and Fertilizer Mineral Mining ",
"212399":"All Other Nonmetallic Mineral Mining ",
"213":"Support Activities for Mining",
"2131":"Support Activities for Mining",
"21311":"Support Activities for Mining",
"213111":"Drilling Oil and Gas Wells",
"213112":"Support Activities for Oil and Gas Operations ",
"213113":"Support Activities for Coal Mining ",
"213114":"Support Activities for Metal Mining ",
"213115":"Support Activities for Nonmetallic Minerals (except Fuels) Mining ",
"22":"Utilities",
"221":"Utilities ",
"2211":"Electric Power Generation, Transmission and Distribution",
"22111":"Electric Power Generation ",
"221111":"Hydroelectric Power Generation ",
"221112":"Fossil Fuel Electric Power Generation ",
"221113":"Nuclear Electric Power Generation ",
"221114":"Solar Electric Power Generation ",
"221115":"Wind Electric Power Generation ",
"221116":"Geothermal Electric Power Generation ",
"221117":"Biomass Electric Power Generation ",
"221118":"Other Electric Power Generation ",
"22112":"Electric Power Transmission, Control, and Distribution ",
"221121":"Electric Bulk Power Transmission and Control ",
"221122":"Electric Power Distribution ",
"2212":"Natural Gas Distribution ",
"22121":"Natural Gas Distribution ",
"221210":"Natural Gas Distribution ",
"2213":"Water, Sewage and Other Systems ",
"22131":"Water Supply and Irrigation Systems ",
"221310":"Water Supply and Irrigation Systems ",
"22132":"Sewage Treatment Facilities ",
"221320":"Sewage Treatment Facilities ",
"22133":"Steam and Air-Conditioning Supply ",
"221330":"Steam and Air-Conditioning Supply ",
"23":"Construction",
"236":"Construction of Buildings",
"2361":"Residential Building Construction",
"23611":"Residential Building Construction",
"236115":"New Single-Family Housing Construction (except For-Sale Builders) ",
"236116":"New Multifamily Housing Construction (except For-Sale Builders) ",
"236117":"New Housing For-Sale Builders ",
"236118":"Residential Remodelers ",
"2362":"Nonresidential Building Construction",
"23621":"Industrial Building Construction",
"236210":"Industrial Building Construction ",
"23622":"Commercial and Institutional Building Construction",
"236220":"Commercial and Institutional Building Construction ",
"237":"Heavy and Civil Engineering Construction",
"2371":"Utility System Construction",
"23711":"Water and Sewer Line and Related Structures Construction",
"237110":"Water and Sewer Line and Related Structures Construction ",
"23712":"Oil and Gas Pipeline and Related Structures Construction",
"237120":"Oil and Gas Pipeline and Related Structures Construction ",
"23713":"Power and Communication Line and Related Structures Construction",
"237130":"Power and Communication Line and Related Structures Construction ",
"2372":"Land Subdivision",
"23721":"Land Subdivision",
"237210":"Land Subdivision ",
"2373":"Highway, Street, and Bridge Construction",
"23731":"Highway, Street, and Bridge Construction",
"237310":"Highway, Street, and Bridge Construction ",
"2379":"Other Heavy and Civil Engineering Construction",
"23799":"Other Heavy and Civil Engineering Construction",
"237990":"Other Heavy and Civil Engineering Construction ",
"238":"Specialty Trade Contractors",
"2381":"Foundation, Structure, and Building Exterior Contractors",
"23811":"Poured Concrete Foundation and Structure Contractors ",
"238110":"Poured Concrete Foundation and Structure Contractors ",
"23812":"Structural Steel and Precast Concrete Contractors ",
"238120":"Structural Steel and Precast Concrete Contractors ",
"23813":"Framing Contractors ",
"238130":"Framing Contractors ",
"23814":"Masonry Contractors ",
"238140":"Masonry Contractors ",
"23815":"Glass and Glazing Contractors ",
"238150":"Glass and Glazing Contractors ",
"23816":"Roofing Contractors ",
"238160":"Roofing Contractors ",
"23817":"Siding Contractors ",
"238170":"Siding Contractors ",
"23819":"Other Foundation, Structure, and Building Exterior Contractors ",
"238190":"Other Foundation, Structure, and Building Exterior Contractors ",
"2382":"Building Equipment Contractors",
"23821":"Electrical Contractors and Other Wiring Installation Contractors",
"238210":"Electrical Contractors and Other Wiring Installation Contractors",
"23822":"Plumbing, Heating, and Air-Conditioning Contractors",
"238220":"Plumbing, Heating, and Air-Conditioning Contractors ",
"23829":"Other Building Equipment Contractors",
"238290":"Other Building Equipment Contractors ",
"2383":"Building Finishing Contractors",
"23831":"Drywall and Insulation Contractors",
"238310":"Drywall and Insulation Contractors ",
"23832":"Painting and Wall Covering Contractors",
"238320":"Painting and Wall Covering Contractors",
"23833":"Flooring Contractors",
"238330":"Flooring Contractors",
"23834":"Tile and Terrazzo Contractors",
"238340":"Tile and Terrazzo Contractors",
"23835":"Finish Carpentry Contractors",
"238350":"Finish Carpentry Contractors",
"23839":"Other Building Finishing Contractors",
"238390":"Other Building Finishing Contractors",
"2389":"Other Specialty Trade Contractors",
"23891":"Site Preparation Contractors",
"238910":"Site Preparation Contractors",
"23899":"All Other Specialty Trade Contractors",
"238990":"All Other Specialty Trade Contractors",
"31":"Manufacturing",
"32":"Manufacturing",
"33":"Manufacturing",
"311":"Food Manufacturing",
"3111":"Animal Food Manufacturing",
"31111":"Animal Food Manufacturing",
"311111":"Dog and Cat Food Manufacturing ",
"311119":"Other Animal Food Manufacturing ",
"3112":"Grain and Oilseed Milling",
"31121":"Flour Milling and Malt Manufacturing",
"311211":"Flour Milling ",
"311212":"Rice Milling ",
"311213":"Malt Manufacturing ",
"31122":"Starch and Vegetable Fats and Oils Manufacturing",
"311221":"Wet Corn Milling ",
"311224":"Soybean and Other Oilseed Processing ",
"311225":"Fats and Oils Refining and Blending ",
"31123":"Breakfast Cereal Manufacturing",
"311230":"Breakfast Cereal Manufacturing",
"3113":"Sugar and Confectionery Product Manufacturing",
"31131":"Sugar Manufacturing",
"311313":"Beet Sugar Manufacturing ",
"311314":"Cane Sugar Manufacturing ",
"31134":"Nonchocolate Confectionery Manufacturing",
"311340":"Nonchocolate Confectionery Manufacturing",
"31135":"Chocolate and Confectionery Manufacturing",
"311351":"Chocolate and Confectionery Manufacturing from Cacao Beans ",
"311352":"Confectionery Manufacturing from Purchased Chocolate ",
"3114":"Fruit and Vegetable Preserving and Specialty Food Manufacturing",
"31141":"Frozen Food Manufacturing",
"311411":"Frozen Fruit, Juice, and Vegetable Manufacturing ",
"311412":"Frozen Specialty Food Manufacturing ",
"31142":"Fruit and Vegetable Canning, Pickling, and Drying",
"311421":"Fruit and Vegetable Canning ",
"311422":"Specialty Canning ",
"311423":"Dried and Dehydrated Food Manufacturing ",
"3115":"Dairy Product Manufacturing",
"31151":"Dairy Product (except Frozen) Manufacturing",
"311511":"Fluid Milk Manufacturing ",
"311512":"Creamery Butter Manufacturing ",
"311513":"Cheese Manufacturing ",
"311514":"Dry, Condensed, and Evaporated Dairy Product Manufacturing ",
"31152":"Ice Cream and Frozen Dessert Manufacturing",
"311520":"Ice Cream and Frozen Dessert Manufacturing",
"3116":"Animal Slaughtering and Processing",
"31161":"Animal Slaughtering and Processing",
"311611":"Animal (except Poultry) Slaughtering ",
"311612":"Meat Processed from Carcasses ",
"311613":"Rendering and Meat Byproduct Processing ",
"311615":"Poultry Processing ",
"3117":"Seafood Product Preparation and Packaging",
"31171":"Seafood Product Preparation and Packaging",
"311710":"Seafood Product Preparation and Packaging",
"3118":"Bakeries and Tortilla Manufacturing",
"31181":"Bread and Bakery Product Manufacturing",
"311811":"Retail Bakeries ",
"311812":"Commercial Bakeries ",
"311813":"Frozen Cakes, Pies, and Other Pastries Manufacturing ",
"31182":"Cookie, Cracker, and Pasta Manufacturing",
"311821":"Cookie and Cracker Manufacturing ",
"311824":"Dry Pasta, Dough, and Flour Mixes Manufacturing from Purchased Flour ",
"31183":"Tortilla Manufacturing",
"311830":"Tortilla Manufacturing",
"3119":"Other Food Manufacturing",
"31191":"Snack Food Manufacturing",
"311911":"Roasted Nuts and Peanut Butter Manufacturing ",
"311919":"Other Snack Food Manufacturing ",
"31192":"Coffee and Tea Manufacturing",
"311920":"Coffee and Tea Manufacturing ",
"31193":"Flavoring Syrup and Concentrate Manufacturing",
"311930":"Flavoring Syrup and Concentrate Manufacturing",
"31194":"Seasoning and Dressing Manufacturing",
"311941":"Mayonnaise, Dressing, and Other Prepared Sauce Manufacturing ",
"311942":"Spice and Extract Manufacturing ",
"31199":"All Other Food Manufacturing",
"311991":"Perishable Prepared Food Manufacturing ",
"311999":"All Other Miscellaneous Food Manufacturing ",
"312":"Beverage and Tobacco Product Manufacturing",
"3121":"Beverage Manufacturing",
"31211":"Soft Drink and Ice Manufacturing",
"312111":"Soft Drink Manufacturing ",
"312112":"Bottled Water Manufacturing ",
"312113":"Ice Manufacturing ",
"31212":"Breweries",
"312120":"Breweries",
"31213":"Wineries",
"312130":"Wineries ",
"31214":"Distilleries",
"312140":"Distilleries ",
"3122":"Tobacco Manufacturing",
"31223":"Tobacco Manufacturing",
"312230":"Tobacco Manufacturing ",
"313":"Textile Mills",
"3131":"Fiber, Yarn, and Thread Mills",
"31311":"Fiber, Yarn, and Thread Mills",
"313110":"Fiber, Yarn, and Thread Mills ",
"3132":"Fabric Mills",
"31321":"Broadwoven Fabric Mills",
"313210":"Broadwoven Fabric Mills",
"31322":"Narrow Fabric Mills and Schiffli Machine Embroidery",
"313220":"Narrow Fabric Mills and Schiffli Machine Embroidery",
"31323":"Nonwoven Fabric Mills",
"313230":"Nonwoven Fabric Mills",
"31324":"Knit Fabric Mills",
"313240":"Knit Fabric Mills",
"3133":"Textile and Fabric Finishing and Fabric Coating Mills",
"31331":"Textile and Fabric Finishing Mills",
"313310":"Textile and Fabric Finishing Mills ",
"31332":"Fabric Coating Mills",
"313320":"Fabric Coating Mills",
"314":"Textile Product Mills",
"3141":"Textile Furnishings Mills",
"31411":"Carpet and Rug Mills",
"314110":"Carpet and Rug Mills",
"31412":"Curtain and Linen Mills",
"314120":"Curtain and Linen Mills",
"3149":"Other Textile Product Mills",
"31491":"Textile Bag and Canvas Mills",
"314910":"Textile Bag and Canvas Mills ",
"31499":"All Other Textile Product Mills",
"314994":"Rope, Cordage, Twine, Tire Cord, and Tire Fabric Mills ",
"314999":"All Other Miscellaneous Textile Product Mills ",
"315":"Apparel Manufacturing",
"3151":"Apparel Knitting Mills",
"31511":"Hosiery and Sock Mills",
"315110":"Hosiery and Sock Mills",
"31519":"Other Apparel Knitting Mills",
"315190":"Other Apparel Knitting Mills ",
"3152":"Cut and Sew Apparel Manufacturing",
"31521":"Cut and Sew Apparel Contractors ",
"315210":"Cut and Sew Apparel Contractors ",
"31522":"Men’s and Boys’ Cut and Sew Apparel Manufacturing ",
"315220":"Men’s and Boys’ Cut and Sew Apparel Manufacturing ",
"31524":"Women’s, Girls’, and Infants’ Cut and Sew Apparel Manufacturing",
"315240":"Women’s, Girls’, and Infants’ Cut and Sew Apparel Manufacturing ",
"31528":"Other Cut and Sew Apparel Manufacturing ",
"315280":"Other Cut and Sew Apparel Manufacturing ",
"3159":"Apparel Accessories and Other Apparel Manufacturing",
"31599":"Apparel Accessories and Other Apparel Manufacturing",
"315990":"Apparel Accessories and Other Apparel Manufacturing ",
"316":"Leather and Allied Product Manufacturing",
"3161":"Leather and Hide Tanning and Finishing",
"31611":"Leather and Hide Tanning and Finishing",
"316110":"Leather and Hide Tanning and Finishing",
"3162":"Footwear Manufacturing",
"31621":"Footwear Manufacturing",
"316210":"Footwear Manufacturing ",
"3169":"Other Leather and Allied Product Manufacturing",
"31699":"Other Leather and Allied Product Manufacturing",
"316992":"Women's Handbag and Purse Manufacturing ",
"316998":"All Other Leather Good and Allied Product Manufacturing ",
"321":"Wood Product Manufacturing",
"3211":"Sawmills and Wood Preservation",
"32111":"Sawmills and Wood Preservation",
"321113":"Sawmills ",
"321114":"Wood Preservation ",
"3212":"Veneer, Plywood, and Engineered Wood Product Manufacturing",
"32121":"Veneer, Plywood, and Engineered Wood Product Manufacturing",
"321211":"Hardwood Veneer and Plywood Manufacturing ",
"321212":"Softwood Veneer and Plywood Manufacturing ",
"321213":"Engineered Wood Member (except Truss) Manufacturing ",
"321214":"Truss Manufacturing ",
"321219":"Reconstituted Wood Product Manufacturing ",
"3219":"Other Wood Product Manufacturing",
"32191":"Millwork",
"321911":"Wood Window and Door Manufacturing ",
"321912":"Cut Stock, Resawing Lumber, and Planing ",
"321918":"Other Millwork (including Flooring) ",
"32192":"Wood Container and Pallet Manufacturing",
"321920":"Wood Container and Pallet Manufacturing",
"32199":"All Other Wood Product Manufacturing",
"321991":"Manufactured Home (Mobile Home) Manufacturing ",
"321992":"Prefabricated Wood Building Manufacturing ",
"321999":"All Other Miscellaneous Wood Product Manufacturing ",
"322":"Paper Manufacturing",
"3221":"Pulp, Paper, and Paperboard Mills",
"32211":"Pulp Mills",
"322110":"Pulp Mills ",
"32212":"Paper Mills",
"322121":"Paper (except Newsprint) Mills ",
"322122":"Newsprint Mills ",
"32213":"Paperboard Mills",
"322130":"Paperboard Mills ",
"3222":"Converted Paper Product Manufacturing",
"32221":"Paperboard Container Manufacturing",
"322211":"Corrugated and Solid Fiber Box Manufacturing ",
"322212":"Folding Paperboard Box Manufacturing ",
"322219":"Other Paperboard Container Manufacturing ",
"32222":"Paper Bag and Coated and Treated Paper Manufacturing",
"322220":"Paper Bag and Coated and Treated Paper Manufacturing",
"32223":"Stationery Product Manufacturing",
"322230":"Stationery Product Manufacturing",
"32229":"Other Converted Paper Product Manufacturing",
"322291":"Sanitary Paper Product Manufacturing ",
"322299":"All Other Converted Paper Product Manufacturing ",
"323":"Printing and Related Support Activities",
"3231":"Printing and Related Support Activities",
"32311":"Printing",
"323111":"Commercial Printing (except Screen and Books) ",
"323113":"Commercial Screen Printing ",
"323117":"Books Printing ",
"32312":"Support Activities for Printing",
"323120":"Support Activities for Printing",
"324":"Petroleum and Coal Products Manufacturing",
"3241":"Petroleum and Coal Products Manufacturing",
"32411":"Petroleum Refineries",
"324110":"Petroleum Refineries",
"32412":"Asphalt Paving, Roofing, and Saturated Materials Manufacturing",
"324121":"Asphalt Paving Mixture and Block Manufacturing ",
"324122":"Asphalt Shingle and Coating Materials Manufacturing ",
"32419":"Other Petroleum and Coal Products Manufacturing",
"324191":"Petroleum Lubricating Oil and Grease Manufacturing ",
"324199":"All Other Petroleum and Coal Products Manufacturing ",
"325":"Chemical Manufacturing",
"3251":"Basic Chemical Manufacturing",
"32511":"Petrochemical Manufacturing",
"325110":"Petrochemical Manufacturing",
"32512":"Industrial Gas Manufacturing",
"325120":"Industrial Gas Manufacturing",
"32513":"Synthetic Dye and Pigment Manufacturing",
"325130":"Synthetic Dye and Pigment Manufacturing",
"32518":"Other Basic Inorganic Chemical Manufacturing",
"325180":"Other Basic Inorganic Chemical Manufacturing ",
"32519":"Other Basic Organic Chemical Manufacturing",
"325193":"Ethyl Alcohol Manufacturing ",
"325194":"Cyclic Crude, Intermediate, and Gum and Wood Chemical Manufacturing ",
"325199":"All Other Basic Organic Chemical Manufacturing ",
"3252":"Resin, Synthetic Rubber, and Artificial and Synthetic Fibers and Filaments Manufacturing",
"32521":"Resin and Synthetic Rubber Manufacturing",
"325211":"Plastics Material and Resin Manufacturing ",
"325212":"Synthetic Rubber Manufacturing ",
"32522":"Artificial and Synthetic Fibers and Filaments Manufacturing",
"325220":"Artificial and Synthetic Fibers and Filaments Manufacturing",
"3253":"Pesticide, Fertilizer, and Other Agricultural Chemical Manufacturing",
"32531":"Fertilizer Manufacturing",
"325311":"Nitrogenous Fertilizer Manufacturing ",
"325312":"Phosphatic Fertilizer Manufacturing ",
"325314":"Fertilizer (Mixing Only) Manufacturing ",
"32532":"Pesticide and Other Agricultural Chemical Manufacturing",
"325320":"Pesticide and Other Agricultural Chemical Manufacturing",
"3254":"Pharmaceutical and Medicine Manufacturing",
"32541":"Pharmaceutical and Medicine Manufacturing",
"325411":"Medicinal and Botanical Manufacturing ",
"325412":"Pharmaceutical Preparation Manufacturing ",
"325413":"In-Vitro Diagnostic Substance Manufacturing ",
"325414":"Biological Product (except Diagnostic) Manufacturing ",
"3255":"Paint, Coating, and Adhesive Manufacturing",
"32551":"Paint and Coating Manufacturing",
"325510":"Paint and Coating Manufacturing",
"32552":"Adhesive Manufacturing",
"325520":"Adhesive Manufacturing",
"3256":"Soap, Cleaning Compound, and Toilet Preparation Manufacturing",
"32561":"Soap and Cleaning Compound Manufacturing",
"325611":"Soap and Other Detergent Manufacturing ",
"325612":"Polish and Other Sanitation Good Manufacturing ",
"325613":"Surface Active Agent Manufacturing ",
"32562":"Toilet Preparation Manufacturing",
"325620":"Toilet Preparation Manufacturing",
"3259":"Other Chemical Product and Preparation Manufacturing",
"32591":"Printing Ink Manufacturing",
"325910":"Printing Ink Manufacturing",
"32592":"Explosives Manufacturing",
"325920":"Explosives Manufacturing",
"32599":"All Other Chemical Product and Preparation Manufacturing",
"325991":"Custom Compounding of Purchased Resins ",
"325992":"Photographic Film, Paper, Plate, and Chemical Manufacturing ",
"325998":"All Other Miscellaneous Chemical Product and Preparation Manufacturing ",
"326":"Plastics and Rubber Products Manufacturing",
"3261":"Plastics Product Manufacturing",
"32611":"Plastics Packaging Materials and Unlaminated Film and Sheet Manufacturing",
"326111":"Plastics Bag and Pouch Manufacturing ",
"326112":"Plastics Packaging Film and Sheet (including Laminated) Manufacturing ",
"326113":"Unlaminated Plastics Film and Sheet (except Packaging) Manufacturing ",
"32612":"Plastics Pipe, Pipe Fitting, and Unlaminated Profile Shape Manufacturing",
"326121":"Unlaminated Plastics Profile Shape Manufacturing ",
"326122":"Plastics Pipe and Pipe Fitting Manufacturing ",
"32613":"Laminated Plastics Plate, Sheet (except Packaging), and Shape Manufacturing",
"326130":"Laminated Plastics Plate, Sheet (except Packaging), and Shape Manufacturing",
"32614":"Polystyrene Foam Product Manufacturing",
"326140":"Polystyrene Foam Product Manufacturing",
"32615":"Urethane and Other Foam Product (except Polystyrene) Manufacturing",
"326150":"Urethane and Other Foam Product (except Polystyrene) Manufacturing",
"32616":"Plastics Bottle Manufacturing",
"326160":"Plastics Bottle Manufacturing",
"32619":"Other Plastics Product Manufacturing",
"326191":"Plastics Plumbing Fixture Manufacturing ",
"326199":"All Other Plastics Product Manufacturing ",
"3262":"Rubber Product Manufacturing",
"32621":"Tire Manufacturing",
"326211":"Tire Manufacturing (except Retreading) ",
"326212":"Tire Retreading ",
"32622":"Rubber and Plastics Hoses and Belting Manufacturing",
"326220":"Rubber and Plastics Hoses and Belting Manufacturing",
"32629":"Other Rubber Product Manufacturing",
"326291":"Rubber Product Manufacturing for Mechanical Use ",
"326299":"All Other Rubber Product Manufacturing ",
"327":"Nonmetallic Mineral Product Manufacturing",
"3271":"Clay Product and Refractory Manufacturing",
"32711":"Pottery, Ceramics, and Plumbing Fixture Manufacturing",
"327110":"Pottery, Ceramics, and Plumbing Fixture Manufacturing ",
"32712":"Clay Building Material and Refractories Manufacturing",
"327120":"Clay Building Material and Refractories Manufacturing ",
"3272":"Glass and Glass Product Manufacturing",
"32721":"Glass and Glass Product Manufacturing",
"327211":"Flat Glass Manufacturing ",
"327212":"Other Pressed and Blown Glass and Glassware Manufacturing ",
"327213":"Glass Container Manufacturing ",
"327215":"Glass Product Manufacturing Made of Purchased Glass ",
"3273":"Cement and Concrete Product Manufacturing",
"32731":"Cement Manufacturing",
"327310":"Cement Manufacturing",
"32732":"Ready-Mix Concrete Manufacturing",
"327320":"Ready-Mix Concrete Manufacturing",
"32733":"Concrete Pipe, Brick, and Block Manufacturing",
"327331":"Concrete Block and Brick Manufacturing ",
"327332":"Concrete Pipe Manufacturing ",
"32739":"Other Concrete Product Manufacturing",
"327390":"Other Concrete Product Manufacturing ",
"3274":"Lime and Gypsum Product Manufacturing",
"32741":"Lime Manufacturing",
"327410":"Lime Manufacturing",
"32742":"Gypsum Product Manufacturing",
"327420":"Gypsum Product Manufacturing",
"3279":"Other Nonmetallic Mineral Product Manufacturing",
"32791":"Abrasive Product Manufacturing",
"327910":"Abrasive Product Manufacturing",
"32799":"All Other Nonmetallic Mineral Product Manufacturing",
"327991":"Cut Stone and Stone Product Manufacturing ",
"327992":"Ground or Treated Mineral and Earth Manufacturing ",
"327993":"Mineral Wool Manufacturing ",
"327999":"All Other Miscellaneous Nonmetallic Mineral Product Manufacturing ",
"331":"Primary Metal Manufacturing",
"3311":"Iron and Steel Mills and Ferroalloy Manufacturing",
"33111":"Iron and Steel Mills and Ferroalloy Manufacturing",
"331110":"Iron and Steel Mills and Ferroalloy Manufacturing ",
"3312":"Steel Product Manufacturing from Purchased Steel",
"33121":"Iron and Steel Pipe and Tube Manufacturing from Purchased Steel",
"331210":"Iron and Steel Pipe and Tube Manufacturing from Purchased Steel",
"33122":"Rolling and Drawing of Purchased Steel",
"331221":"Rolled Steel Shape Manufacturing ",
"331222":"Steel Wire Drawing ",
"3313":"Alumina and Aluminum Production and Processing",
"33131":"Alumina and Aluminum Production and Processing",
"331313":"Alumina Refining and Primary Aluminum Production ",
"331314":"Secondary Smelting and Alloying of Aluminum ",
"331315":"Aluminum Sheet, Plate, and Foil Manufacturing ",
"331318":"Other Aluminum Rolling, Drawing, and Extruding ",
"3314":"Nonferrous Metal (except Aluminum) Production and Processing",
"33141":"Nonferrous Metal (except Aluminum) Smelting and Refining",
"331410":"Nonferrous Metal (except Aluminum) Smelting and Refining ",
"33142":"Copper Rolling, Drawing, Extruding, and Alloying",
"331420":"Copper Rolling, Drawing, Extruding, and Alloying",
"33149":"Nonferrous Metal (except Copper and Aluminum) Rolling, Drawing, Extruding, and Alloying",
"331491":"Nonferrous Metal (except Copper and Aluminum) Rolling, Drawing, and Extruding ",
"331492":"Secondary Smelting, Refining, and Alloying of Nonferrous Metal (except Copper and Aluminum) ",
"3315":"Foundries",
"33151":"Ferrous Metal Foundries",
"331511":"Iron Foundries ",
"331512":"Steel Investment Foundries ",
"331513":"Steel Foundries (except Investment) ",
"33152":"Nonferrous Metal Foundries",
"331523":"Nonferrous Metal Die-Casting Foundries ",
"331524":"Aluminum Foundries (except Die-Casting) ",
"331529":"Other Nonferrous Metal Foundries (except Die-Casting) ",
"332":"Fabricated Metal Product Manufacturing",
"3321":"Forging and Stamping",
"33211":"Forging and Stamping",
"332111":"Iron and Steel Forging ",
"332112":"Nonferrous Forging ",
"332114":"Custom Roll Forming ",
"332117":"Powder Metallurgy Part Manufacturing ",
"332119":"Metal Crown, Closure, and Other Metal Stamping (except Automotive) ",
"3322":"Cutlery and Handtool Manufacturing",
"33221":"Cutlery and Handtool Manufacturing",
"332215":"Metal Kitchen Cookware, Utensil, Cutlery, and Flatware (except Precious) Manufacturing ",
"332216":"Saw Blade and Handtool Manufacturing ",
"3323":"Architectural and Structural Metals Manufacturing",
"33231":"Plate Work and Fabricated Structural Product Manufacturing",
"332311":"Prefabricated Metal Building and Component Manufacturing ",
"332312":"Fabricated Structural Metal Manufacturing ",
"332313":"Plate Work Manufacturing ",
"33232":"Ornamental and Architectural Metal Products Manufacturing",
"332321":"Metal Window and Door Manufacturing ",
"332322":"Sheet Metal Work Manufacturing ",
"332323":"Ornamental and Architectural Metal Work Manufacturing ",
"3324":"Boiler, Tank, and Shipping Container Manufacturing",
"33241":"Power Boiler and Heat Exchanger Manufacturing",
"332410":"Power Boiler and Heat Exchanger Manufacturing",
"33242":"Metal Tank (Heavy Gauge) Manufacturing",
"332420":"Metal Tank (Heavy Gauge) Manufacturing",
"33243":"Metal Can, Box, and Other Metal Container (Light Gauge) Manufacturing",
"332431":"Metal Can Manufacturing ",
"332439":"Other Metal Container Manufacturing ",
"3325":"Hardware Manufacturing",
"33251":"Hardware Manufacturing",
"332510":"Hardware Manufacturing",
"3326":"Spring and Wire Product Manufacturing",
"33261":"Spring and Wire Product Manufacturing",
"332613":"Spring Manufacturing ",
"332618":"Other Fabricated Wire Product Manufacturing ",
"3327":"Machine Shops; Turned Product; and Screw, Nut, and Bolt Manufacturing",
"33271":"Machine Shops",
"332710":"Machine Shops",
"33272":"Turned Product and Screw, Nut, and Bolt Manufacturing",
"332721":"Precision Turned Product Manufacturing ",
"332722":"Bolt, Nut, Screw, Rivet, and Washer Manufacturing ",
"3328":"Coating, Engraving, Heat Treating, and Allied Activities",
"33281":"Coating, Engraving, Heat Treating, and Allied Activities",
"332811":"Metal Heat Treating ",
"332812":"Metal Coating, Engraving (except Jewelry and Silverware), and Allied Services to Manufacturers ",
"332813":"Electroplating, Plating, Polishing, Anodizing, and Coloring ",
"3329":"Other Fabricated Metal Product Manufacturing",
"33291":"Metal Valve Manufacturing",
"332911":"Industrial Valve Manufacturing ",
"332912":"Fluid Power Valve and Hose Fitting Manufacturing ",
"332913":"Plumbing Fixture Fitting and Trim Manufacturing ",
"332919":"Other Metal Valve and Pipe Fitting Manufacturing ",
"33299":"All Other Fabricated Metal Product Manufacturing",
"332991":"Ball and Roller Bearing Manufacturing",
"332992":"Small Arms Ammunition Manufacturing ",
"332993":"Ammunition (except Small Arms) Manufacturing ",
"332994":"Small Arms, Ordnance, and Ordnance Accessories Manufacturing ",
"332996":"Fabricated Pipe and Pipe Fitting Manufacturing ",
"332999":"All Other Miscellaneous Fabricated Metal Product Manufacturing ",
"333":"Machinery Manufacturing",
"3331":"Agriculture, Construction, and Mining Machinery Manufacturing",
"33311":"Agricultural Implement Manufacturing",
"333111":"Farm Machinery and Equipment Manufacturing ",
"333112":"Lawn and Garden Tractor and Home Lawn and Garden Equipment Manufacturing ",
"33312":"Construction Machinery Manufacturing",
"333120":"Construction Machinery Manufacturing",
"33313":"Mining and Oil and Gas Field Machinery Manufacturing",
"333131":"Mining Machinery and Equipment Manufacturing ",
"333132":"Oil and Gas Field Machinery and Equipment Manufacturing ",
"3332":"Industrial Machinery Manufacturing",
"33324":"Industrial Machinery Manufacturing",
"333241":"Food Product Machinery Manufacturing ",
"333242":"Semiconductor Machinery Manufacturing ",
"333243":"Sawmill, Woodworking, and Paper Machinery Manufacturing ",
"333244":"Printing Machinery and Equipment Manufacturing ",
"333249":"Other Industrial Machinery Manufacturing ",
"3333":"Commercial and Service Industry Machinery Manufacturing",
"33331":"Commercial and Service Industry Machinery Manufacturing",
"333314":"Optical Instrument and Lens Manufacturing ",
"333316":"Photographic and Photocopying Equipment Manufacturing ",
"333318":"Other Commercial and Service Industry Machinery Manufacturing ",
"3334":"Ventilation, Heating, Air-Conditioning, and Commercial Refrigeration Equipment Manufacturing",
"33341":"Ventilation, Heating, Air-Conditioning, and Commercial Refrigeration Equipment Manufacturing",
"333413":"Industrial and Commercial Fan and Blower and Air Purification Equipment Manufacturing ",
"333414":"Heating Equipment (except Warm Air Furnaces) Manufacturing ",
"333415":"Air-Conditioning and Warm Air Heating Equipment and Commercial and Industrial Refrigeration Equipment Manufacturing ",
"3335":"Metalworking Machinery Manufacturing",
"33351":"Metalworking Machinery Manufacturing",
"333511":"Industrial Mold Manufacturing ",
"333514":"Special Die and Tool, Die Set, Jig, and Fixture Manufacturing ",
"333515":"Cutting Tool and Machine Tool Accessory Manufacturing ",
"333517":"Machine Tool Manufacturing ",
"333519":"Rolling Mill and Other Metalworking Machinery Manufacturing ",
"3336":"Engine, Turbine, and Power Transmission Equipment Manufacturing",
"33361":"Engine, Turbine, and Power Transmission Equipment Manufacturing",
"333611":"Turbine and Turbine Generator Set Units Manufacturing ",
"333612":"Speed Changer, Industrial High-Speed Drive, and Gear Manufacturing ",
"333613":"Mechanical Power Transmission Equipment Manufacturing ",
"333618":"Other Engine Equipment Manufacturing ",
"3339":"Other General Purpose Machinery Manufacturing",
"33391":"Pump and Compressor Manufacturing",
"333912":"Air and Gas Compressor Manufacturing ",
"333914":"Measuring, Dispensing, and Other Pumping Equipment Manufacturing ",
"33392":"Material Handling Equipment Manufacturing",
"333921":"Elevator and Moving Stairway Manufacturing ",
"333922":"Conveyor and Conveying Equipment Manufacturing ",
"333923":"Overhead Traveling Crane, Hoist, and Monorail System Manufacturing ",
"333924":"Industrial Truck, Tractor, Trailer, and Stacker Machinery Manufacturing ",
"33399":"All Other General Purpose Machinery Manufacturing",
"333991":"Power-Driven Handtool Manufacturing ",
"333992":"Welding and Soldering Equipment Manufacturing ",
"333993":"Packaging Machinery Manufacturing ",
"333994":"Industrial Process Furnace and Oven Manufacturing ",
"333995":"Fluid Power Cylinder and Actuator Manufacturing ",
"333996":"Fluid Power Pump and Motor Manufacturing ",
"333997":"Scale and Balance Manufacturing ",
"333999":"All Other Miscellaneous General Purpose Machinery Manufacturing ",
"334":"Computer and Electronic Product Manufacturing",
"3341":"Computer and Peripheral Equipment Manufacturing",
"33411":"Computer and Peripheral Equipment Manufacturing",
"334111":"Electronic Computer Manufacturing ",
"334112":"Computer Storage Device Manufacturing ",
"334118":"Computer Terminal and Other Computer Peripheral Equipment Manufacturing ",
"3342":"Communications Equipment Manufacturing",
"33421":"Telephone Apparatus Manufacturing",
"334210":"Telephone Apparatus Manufacturing",
"33422":"Radio and Television Broadcasting and Wireless Communications Equipment Manufacturing",
"334220":"Radio and Television Broadcasting and Wireless Communications Equipment Manufacturing",
"33429":"Other Communications Equipment Manufacturing",
"334290":"Other Communications Equipment Manufacturing",
"3343":"Audio and Video Equipment Manufacturing",
"33431":"Audio and Video Equipment Manufacturing",
"334310":"Audio and Video Equipment Manufacturing",
"3344":"Semiconductor and Other Electronic Component Manufacturing",
"33441":"Semiconductor and Other Electronic Component Manufacturing",
"334412":"Bare Printed Circuit Board Manufacturing  ",
"334413":"Semiconductor and Related Device Manufacturing ",
"334416":"Capacitor, Resistor, Coil, Transformer, and Other Inductor Manufacturing ",
"334417":"Electronic Connector Manufacturing ",
"334418":"Printed Circuit Assembly (Electronic Assembly) Manufacturing ",
"334419":"Other Electronic Component Manufacturing ",
"3345":"Navigational, Measuring, Electromedical, and Control Instruments Manufacturing",
"33451":"Navigational, Measuring, Electromedical, and Control Instruments Manufacturing",
"334510":"Electromedical and Electrotherapeutic Apparatus Manufacturing ",
"334511":"Search, Detection, Navigation, Guidance, Aeronautical, and Nautical System and Instrument Manufacturing ",
"334512":"Automatic Environmental Control Manufacturing for Residential, Commercial, and Appliance Use ",
"334513":"Instruments and Related Products Manufacturing for Measuring, Displaying, and Controlling Industrial Process Variables ",
"334514":"Totalizing Fluid Meter and Counting Device Manufacturing ",
"334515":"Instrument Manufacturing for Measuring and Testing Electricity and Electrical Signals ",
"334516":"Analytical Laboratory Instrument Manufacturing ",
"334517":"Irradiation Apparatus Manufacturing ",
"334519":"Other Measuring and Controlling Device Manufacturing ",
"3346":"Manufacturing and Reproducing Magnetic and Optical Media",
"33461":"Manufacturing and Reproducing Magnetic and Optical Media",
"334613":"Blank Magnetic and Optical Recording Media Manufacturing ",
"334614":"Software and Other Prerecorded Compact Disc, Tape, and Record Reproducing ",
"335":"Electrical Equipment, Appliance, and Component Manufacturing",
"3351":"Electric Lighting Equipment Manufacturing",
"33511":"Electric Lamp Bulb and Part Manufacturing",
"335110":"Electric Lamp Bulb and Part Manufacturing",
"33512":"Lighting Fixture Manufacturing",
"335121":"Residential Electric Lighting Fixture Manufacturing ",
"335122":"Commercial, Industrial, and Institutional Electric Lighting Fixture Manufacturing ",
"335129":"Other Lighting Equipment Manufacturing ",
"3352":"Household Appliance Manufacturing",
"33521":"Small Electrical Appliance Manufacturing",
"335210":"Small Electrical Appliance Manufacturing",
"33522":"Major Household Appliance Manufacturing ",
"335220":"Major Household Appliance Manufacturing ",
"3353":"Electrical Equipment Manufacturing",
"33531":"Electrical Equipment Manufacturing",
"335311":"Power, Distribution, and Specialty Transformer Manufacturing ",
"335312":"Motor and Generator Manufacturing ",
"335313":"Switchgear and Switchboard Apparatus Manufacturing ",
"335314":"Relay and Industrial Control Manufacturing ",
"3359":"Other Electrical Equipment and Component Manufacturing",
"33591":"Battery Manufacturing",
"335911":"Storage Battery Manufacturing ",
"335912":"Primary Battery Manufacturing ",
"33592":"Communication and Energy Wire and Cable Manufacturing",
"335921":"Fiber Optic Cable Manufacturing ",
"335929":"Other Communication and Energy Wire Manufacturing ",
"33593":"Wiring Device Manufacturing",
"335931":"Current-Carrying Wiring Device Manufacturing ",
"335932":"Noncurrent-Carrying Wiring Device Manufacturing ",
"33599":"All Other Electrical Equipment and Component Manufacturing",
"335991":"Carbon and Graphite Product Manufacturing ",
"335999":"All Other Miscellaneous Electrical Equipment and Component Manufacturing ",
"336":"Transportation Equipment Manufacturing",
"3361":"Motor Vehicle Manufacturing",
"33611":"Automobile and Light Duty Motor Vehicle Manufacturing",
"336111":"Automobile Manufacturing ",
"336112":"Light Truck and Utility Vehicle Manufacturing ",
"33612":"Heavy Duty Truck Manufacturing",
"336120":"Heavy Duty Truck Manufacturing",
"3362":"Motor Vehicle Body and Trailer Manufacturing",
"33621":"Motor Vehicle Body and Trailer Manufacturing",
"336211":"Motor Vehicle Body Manufacturing ",
"336212":"Truck Trailer Manufacturing ",
"336213":"Motor Home Manufacturing ",
"336214":"Travel Trailer and Camper Manufacturing ",
"3363":"Motor Vehicle Parts Manufacturing",
"33631":"Motor Vehicle Gasoline Engine and Engine Parts Manufacturing",
"336310":"Motor Vehicle Gasoline Engine and Engine Parts Manufacturing",
"33632":"Motor Vehicle Electrical and Electronic Equipment Manufacturing",
"336320":"Motor Vehicle Electrical and Electronic Equipment Manufacturing",
"33633":"Motor Vehicle Steering and Suspension Components (except Spring) Manufacturing",
"336330":"Motor Vehicle Steering and Suspension Components (except Spring) Manufacturing",
"33634":"Motor Vehicle Brake System Manufacturing",
"336340":"Motor Vehicle Brake System Manufacturing",
"33635":"Motor Vehicle Transmission and Power Train Parts Manufacturing",
"336350":"Motor Vehicle Transmission and Power Train Parts Manufacturing",
"33636":"Motor Vehicle Seating and Interior Trim Manufacturing",
"336360":"Motor Vehicle Seating and Interior Trim Manufacturing",
"33637":"Motor Vehicle Metal Stamping",
"336370":"Motor Vehicle Metal Stamping",
"33639":"Other Motor Vehicle Parts Manufacturing",
"336390":"Other Motor Vehicle Parts Manufacturing",
"3364":"Aerospace Product and Parts Manufacturing",
"33641":"Aerospace Product and Parts Manufacturing",
"336411":"Aircraft Manufacturing ",
"336412":"Aircraft Engine and Engine Parts Manufacturing ",
"336413":"Other Aircraft Parts and Auxiliary Equipment Manufacturing ",
"336414":"Guided Missile and Space Vehicle Manufacturing ",
"336415":"Guided Missile and Space Vehicle Propulsion Unit and Propulsion Unit Parts Manufacturing ",
"336419":"Other Guided Missile and Space Vehicle Parts and Auxiliary Equipment Manufacturing ",
"3365":"Railroad Rolling Stock Manufacturing",
"33651":"Railroad Rolling Stock Manufacturing",
"336510":"Railroad Rolling Stock Manufacturing",
"3366":"Ship and Boat Building",
"33661":"Ship and Boat Building",
"336611":"Ship Building and Repairing ",
"336612":"Boat Building ",
"3369":"Other Transportation Equipment Manufacturing",
"33699":"Other Transportation Equipment Manufacturing",
"336991":"Motorcycle, Bicycle, and Parts Manufacturing ",
"336992":"Military Armored Vehicle, Tank, and Tank Component Manufacturing ",
"336999":"All Other Transportation Equipment Manufacturing ",
"337":"Furniture and Related Product Manufacturing",
"3371":"Household and Institutional Furniture and Kitchen Cabinet Manufacturing",
"33711":"Wood Kitchen Cabinet and Countertop Manufacturing",
"337110":"Wood Kitchen Cabinet and Countertop Manufacturing",
"33712":"Household and Institutional Furniture Manufacturing",
"337121":"Upholstered Household Furniture Manufacturing ",
"337122":"Nonupholstered Wood Household Furniture Manufacturing ",
"337124":"Metal Household Furniture Manufacturing ",
"337125":"Household Furniture (except Wood and Metal) Manufacturing ",
"337127":"Institutional Furniture Manufacturing ",
"3372":"Office Furniture (including Fixtures) Manufacturing",
"33721":"Office Furniture (including Fixtures) Manufacturing",
"337211":"Wood Office Furniture Manufacturing ",
"337212":"Custom Architectural Woodwork and Millwork Manufacturing ",
"337214":"Office Furniture (except Wood) Manufacturing ",
"337215":"Showcase, Partition, Shelving, and Locker Manufacturing ",
"3379":"Other Furniture Related Product Manufacturing",
"33791":"Mattress Manufacturing",
"337910":"Mattress Manufacturing",
"33792":"Blind and Shade Manufacturing",
"337920":"Blind and Shade Manufacturing",
"339":"Miscellaneous Manufacturing",
"3391":"Medical Equipment and Supplies Manufacturing",
"33911":"Medical Equipment and Supplies Manufacturing",
"339112":"Surgical and Medical Instrument Manufacturing ",
"339113":"Surgical Appliance and Supplies Manufacturing ",
"339114":"Dental Equipment and Supplies Manufacturing ",
"339115":"Ophthalmic Goods Manufacturing ",
"339116":"Dental Laboratories ",
"3399":"Other Miscellaneous Manufacturing",
"33991":"Jewelry and Silverware Manufacturing",
"339910":"Jewelry and Silverware Manufacturing ",
"33992":"Sporting and Athletic Goods Manufacturing",
"339920":"Sporting and Athletic Goods Manufacturing",
"33993":"Doll, Toy, and Game Manufacturing",
"339930":"Doll, Toy, and Game Manufacturing",
"33994":"Office Supplies (except Paper) Manufacturing",
"339940":"Office Supplies (except Paper) Manufacturing",
"33995":"Sign Manufacturing",
"339950":"Sign Manufacturing",
"33999":"All Other Miscellaneous Manufacturing",
"339991":"Gasket, Packing, and Sealing Device Manufacturing ",
"339992":"Musical Instrument Manufacturing ",
"339993":"Fastener, Button, Needle, and Pin Manufacturing ",
"339994":"Broom, Brush, and Mop Manufacturing ",
"339995":"Burial Casket Manufacturing ",
"339999":"All Other Miscellaneous Manufacturing ",
"42":"Wholesale Trade",
"423":"Merchant Wholesalers, Durable Goods ",
"4231":"Motor Vehicle and Motor Vehicle Parts and Supplies Merchant Wholesalers ",
"42311":"Automobile and Other Motor Vehicle Merchant Wholesalers ",
"423110":"Automobile and Other Motor Vehicle Merchant Wholesalers ",
"42312":"Motor Vehicle Supplies and New Parts Merchant Wholesalers ",
"423120":"Motor Vehicle Supplies and New Parts Merchant Wholesalers ",
"42313":"Tire and Tube Merchant Wholesalers ",
"423130":"Tire and Tube Merchant Wholesalers ",
"42314":"Motor Vehicle Parts (Used) Merchant Wholesalers ",
"423140":"Motor Vehicle Parts (Used) Merchant Wholesalers ",
"4232":"Furniture and Home Furnishing Merchant Wholesalers ",
"42321":"Furniture Merchant Wholesalers ",
"423210":"Furniture Merchant Wholesalers ",
"42322":"Home Furnishing Merchant Wholesalers ",
"423220":"Home Furnishing Merchant Wholesalers ",
"4233":"Lumber and Other Construction Materials Merchant Wholesalers ",
"42331":"Lumber, Plywood, Millwork, and Wood Panel Merchant Wholesalers ",
"423310":"Lumber, Plywood, Millwork, and Wood Panel Merchant Wholesalers ",
"42332":"Brick, Stone, and Related Construction Material Merchant Wholesalers ",
"423320":"Brick, Stone, and Related Construction Material Merchant Wholesalers ",
"42333":"Roofing, Siding, and Insulation Material Merchant Wholesalers ",
"423330":"Roofing, Siding, and Insulation Material Merchant Wholesalers ",
"42339":"Other Construction Material Merchant Wholesalers ",
"423390":"Other Construction Material Merchant Wholesalers ",
"4234":"Professional and Commercial Equipment and Supplies Merchant Wholesalers ",
"42341":"Photographic Equipment and Supplies Merchant Wholesalers ",
"423410":"Photographic Equipment and Supplies Merchant Wholesalers ",
"42342":"Office Equipment Merchant Wholesalers ",
"423420":"Office Equipment Merchant Wholesalers ",
"42343":"Computer and Computer Peripheral Equipment and Software Merchant Wholesalers ",
"423430":"Computer and Computer Peripheral Equipment and Software Merchant Wholesalers ",
"42344":"Other Commercial Equipment Merchant Wholesalers ",
"423440":"Other Commercial Equipment Merchant Wholesalers ",
"42345":"Medical, Dental, and Hospital Equipment and Supplies Merchant Wholesalers ",
"423450":"Medical, Dental, and Hospital Equipment and Supplies Merchant Wholesalers ",
"42346":"Ophthalmic Goods Merchant Wholesalers ",
"423460":"Ophthalmic Goods Merchant Wholesalers ",
"42349":"Other Professional Equipment and Supplies Merchant Wholesalers ",
"423490":"Other Professional Equipment and Supplies Merchant Wholesalers ",
"4235":"Metal and Mineral (except Petroleum) Merchant Wholesalers ",
"42351":"Metal Service Centers and Other Metal Merchant Wholesalers ",
"423510":"Metal Service Centers and Other Metal Merchant Wholesalers ",
"42352":"Coal and Other Mineral and Ore Merchant Wholesalers ",
"423520":"Coal and Other Mineral and Ore Merchant Wholesalers ",
"4236":"Household Appliances and Electrical and Electronic Goods Merchant Wholesalers ",
"42361":"Electrical Apparatus and Equipment, Wiring Supplies, and Related Equipment Merchant Wholesalers ",
"423610":"Electrical Apparatus and Equipment, Wiring Supplies, and Related Equipment Merchant Wholesalers ",
"42362":"Household Appliances, Electric Housewares, and Consumer Electronics Merchant Wholesalers ",
"423620":"Household Appliances, Electric Housewares, and Consumer Electronics Merchant Wholesalers ",
"42369":"Other Electronic Parts and Equipment Merchant Wholesalers ",
"423690":"Other Electronic Parts and Equipment Merchant Wholesalers ",
"4237":"Hardware, and Plumbing and Heating Equipment and Supplies Merchant Wholesalers ",
"42371":"Hardware Merchant Wholesalers ",
"423710":"Hardware Merchant Wholesalers ",
"42372":"Plumbing and Heating Equipment and Supplies (Hydronics) Merchant Wholesalers ",
"423720":"Plumbing and Heating Equipment and Supplies (Hydronics) Merchant Wholesalers ",
"42373":"Warm Air Heating and Air-Conditioning Equipment and Supplies Merchant Wholesalers ",
"423730":"Warm Air Heating and Air-Conditioning Equipment and Supplies Merchant Wholesalers ",
"42374":"Refrigeration Equipment and Supplies Merchant Wholesalers ",
"423740":"Refrigeration Equipment and Supplies Merchant Wholesalers ",
"4238":"Machinery, Equipment, and Supplies Merchant Wholesalers ",
"42381":"Construction and Mining (except Oil Well) Machinery and Equipment Merchant Wholesalers ",
"423810":"Construction and Mining (except Oil Well) Machinery and Equipment Merchant Wholesalers ",
"42382":"Farm and Garden Machinery and Equipment Merchant Wholesalers ",
"423820":"Farm and Garden Machinery and Equipment Merchant Wholesalers ",
"42383":"Industrial Machinery and Equipment Merchant Wholesalers ",
"423830":"Industrial Machinery and Equipment Merchant Wholesalers ",
"42384":"Industrial Supplies Merchant Wholesalers ",
"423840":"Industrial Supplies Merchant Wholesalers",
"42385":"Service Establishment Equipment and Supplies Merchant Wholesalers ",
"423850":"Service Establishment Equipment and Supplies Merchant Wholesalers ",
"42386":"Transportation Equipment and Supplies (except Motor Vehicle) Merchant Wholesalers ",
"423860":"Transportation Equipment and Supplies (except Motor Vehicle) Merchant Wholesalers ",
"4239":"Miscellaneous Durable Goods Merchant Wholesalers ",
"42391":"Sporting and Recreational Goods and Supplies Merchant Wholesalers",
"423910":"Sporting and Recreational Goods and Supplies Merchant Wholesalers ",
"42392":"Toy and Hobby Goods and Supplies Merchant Wholesalers ",
"423920":"Toy and Hobby Goods and Supplies Merchant Wholesalers ",
"42393":"Recyclable Material Merchant Wholesalers ",
"423930":"Recyclable Material Merchant Wholesalers ",
"42394":"Jewelry, Watch, Precious Stone, and Precious Metal Merchant Wholesalers ",
"423940":"Jewelry, Watch, Precious Stone, and Precious Metal Merchant Wholesalers ",
"42399":"Other Miscellaneous Durable Goods Merchant Wholesalers ",
"423990":"Other Miscellaneous Durable Goods Merchant Wholesalers ",
"424":"Merchant Wholesalers, Nondurable Goods ",
"4241":"Paper and Paper Product Merchant Wholesalers ",
"42411":"Printing and Writing Paper Merchant Wholesalers ",
"424110":"Printing and Writing Paper Merchant Wholesalers ",
"42412":"Stationery and Office Supplies Merchant Wholesalers ",
"424120":"Stationery and Office Supplies Merchant Wholesalers ",
"42413":"Industrial and Personal Service Paper Merchant Wholesalers ",
"424130":"Industrial and Personal Service Paper Merchant Wholesalers ",
"4242":"Drugs and Druggists' Sundries Merchant Wholesalers ",
"42421":"Drugs and Druggists' Sundries Merchant Wholesalers ",
"424210":"Drugs and Druggists' Sundries Merchant Wholesalers ",
"4243":"Apparel, Piece Goods, and Notions Merchant Wholesalers ",
"42431":"Piece Goods, Notions, and Other Dry Goods Merchant Wholesalers ",
"424310":"Piece Goods, Notions, and Other Dry Goods Merchant Wholesalers ",
"42432":"Men's and Boys' Clothing and Furnishings Merchant Wholesalers ",
"424320":"Men's and Boys' Clothing and Furnishings Merchant Wholesalers ",
"42433":"Women's, Children's, and Infants' Clothing and Accessories Merchant Wholesalers ",
"424330":"Women's, Children's, and Infants' Clothing and Accessories Merchant Wholesalers ",
"42434":"Footwear Merchant Wholesalers ",
"424340":"Footwear Merchant Wholesalers ",
"4244":"Grocery and Related Product Merchant Wholesalers ",
"42441":"General Line Grocery Merchant Wholesalers ",
"424410":"General Line Grocery Merchant Wholesalers ",
"42442":"Packaged Frozen Food Merchant Wholesalers ",
"424420":"Packaged Frozen Food Merchant Wholesalers ",
"42443":"Dairy Product (except Dried or Canned) Merchant Wholesalers ",
"424430":"Dairy Product (except Dried or Canned) Merchant Wholesalers ",
"42444":"Poultry and Poultry Product Merchant Wholesalers ",
"424440":"Poultry and Poultry Product Merchant Wholesalers ",
"42445":"Confectionery Merchant Wholesalers ",
"424450":"Confectionery Merchant Wholesalers ",
"42446":"Fish and Seafood Merchant Wholesalers ",
"424460":"Fish and Seafood Merchant Wholesalers ",
"42447":"Meat and Meat Product Merchant Wholesalers ",
"424470":"Meat and Meat Product Merchant Wholesalers ",
"42448":"Fresh Fruit and Vegetable Merchant Wholesalers ",
"424480":"Fresh Fruit and Vegetable Merchant Wholesalers ",
"42449":"Other Grocery and Related Products Merchant Wholesalers ",
"424490":"Other Grocery and Related Products Merchant Wholesalers ",
"4245":"Farm Product Raw Material Merchant Wholesalers ",
"42451":"Grain and Field Bean Merchant Wholesalers ",
"424510":"Grain and Field Bean Merchant Wholesalers ",
"42452":"Livestock Merchant Wholesalers ",
"424520":"Livestock Merchant Wholesalers ",
"42459":"Other Farm Product Raw Material Merchant Wholesalers ",
"424590":"Other Farm Product Raw Material Merchant Wholesalers ",
"4246":"Chemical and Allied Products Merchant Wholesalers ",
"42461":"Plastics Materials and Basic Forms and Shapes Merchant Wholesalers ",
"424610":"Plastics Materials and Basic Forms and Shapes Merchant Wholesalers ",
"42469":"Other Chemical and Allied Products Merchant Wholesalers ",
"424690":"Other Chemical and Allied Products Merchant Wholesalers ",
"4247":"Petroleum and Petroleum Products Merchant Wholesalers ",
"42471":"Petroleum Bulk Stations and Terminals ",
"424710":"Petroleum Bulk Stations and Terminals ",
"42472":"Petroleum and Petroleum Products Merchant Wholesalers (except Bulk Stations and Terminals) ",
"424720":"Petroleum and Petroleum Products Merchant Wholesalers (except Bulk Stations and Terminals) ",
"4248":"Beer, Wine, and Distilled Alcoholic Beverage Merchant Wholesalers ",
"42481":"Beer and Ale Merchant Wholesalers ",
"424810":"Beer and Ale Merchant Wholesalers ",
"42482":"Wine and Distilled Alcoholic Beverage Merchant Wholesalers ",
"424820":"Wine and Distilled Alcoholic Beverage Merchant Wholesalers ",
"4249":"Miscellaneous Nondurable Goods Merchant Wholesalers ",
"42491":"Farm Supplies Merchant Wholesalers ",
"424910":"Farm Supplies Merchant Wholesalers ",
"42492":"Book, Periodical, and Newspaper Merchant Wholesalers ",
"424920":"Book, Periodical, and Newspaper Merchant Wholesalers ",
"42493":"Flower, Nursery Stock, and Florists' Supplies Merchant Wholesalers ",
"424930":"Flower, Nursery Stock, and Florists' Supplies Merchant Wholesalers ",
"42494":"Tobacco and Tobacco Product Merchant Wholesalers ",
"424940":"Tobacco and Tobacco Product Merchant Wholesalers ",
"42495":"Paint, Varnish, and Supplies Merchant Wholesalers ",
"424950":"Paint, Varnish, and Supplies Merchant Wholesalers ",
"42499":"Other Miscellaneous Nondurable Goods Merchant Wholesalers ",
"424990":"Other Miscellaneous Nondurable Goods Merchant Wholesalers ",
"425":"Wholesale Electronic Markets and Agents and Brokers ",
"4251":"Wholesale Electronic Markets and Agents and Brokers ",
"42511":"Business to Business Electronic Markets ",
"425110":"Business to Business Electronic Markets ",
"42512":"Wholesale Trade Agents and Brokers ",
"425120":"Wholesale Trade Agents and Brokers ",
"44":"Retail Trade",
"45":"Retail Trade",
"441":"Motor Vehicle and Parts Dealers ",
"4411":"Automobile Dealers ",
"44111":"New Car Dealers ",
"441110":"New Car Dealers ",
"44112":"Used Car Dealers ",
"441120":"Used Car Dealers ",
"4412":"Other Motor Vehicle Dealers ",
"44121":"Recreational Vehicle Dealers ",
"441210":"Recreational Vehicle Dealers ",
"44122":"Motorcycle, Boat, and Other Motor Vehicle Dealers ",
"441222":"Boat Dealers ",
"441228":"Motorcycle, ATV, and All Other Motor Vehicle Dealers ",
"4413":"Automotive Parts, Accessories, and Tire Stores ",
"44131":"Automotive Parts and Accessories Stores ",
"441310":"Automotive Parts and Accessories Stores ",
"44132":"Tire Dealers ",
"441320":"Tire Dealers ",
"442":"Furniture and Home Furnishings Stores ",
"4421":"Furniture Stores ",
"44211":"Furniture Stores ",
"442110":"Furniture Stores ",
"4422":"Home Furnishings Stores ",
"44221":"Floor Covering Stores ",
"442210":"Floor Covering Stores ",
"44229":"Other Home Furnishings Stores ",
"442291":"Window Treatment Stores ",
"442299":"All Other Home Furnishings Stores ",
"443":"Electronics and Appliance Stores ",
"4431":"Electronics and Appliance Stores ",
"44314":"Electronics and Appliance Stores ",
"443141":"Household Appliance Stores ",
"443142":"Electronics Stores ",
"444":"Building Material and Garden Equipment and Supplies Dealers ",
"4441":"Building Material and Supplies Dealers ",
"44411":"Home Centers ",
"444110":"Home Centers ",
"44412":"Paint and Wallpaper Stores ",
"444120":"Paint and Wallpaper Stores ",
"44413":"Hardware Stores ",
"444130":"Hardware Stores ",
"44419":"Other Building Material Dealers ",
"444190":"Other Building Material Dealers ",
"4442":"Lawn and Garden Equipment and Supplies Stores ",
"44421":"Outdoor Power Equipment Stores ",
"444210":"Outdoor Power Equipment Stores ",
"44422":"Nursery, Garden Center, and Farm Supply Stores ",
"444220":"Nursery, Garden Center, and Farm Supply Stores ",
"445":"Food and Beverage Stores ",
"4451":"Grocery Stores ",
"44511":"Supermarkets and Other Grocery (except Convenience) Stores ",
"445110":"Supermarkets and Other Grocery (except Convenience) Stores ",
"44512":"Convenience Stores ",
"445120":"Convenience Stores ",
"4452":"Specialty Food Stores ",
"44521":"Meat Markets ",
"445210":"Meat Markets ",
"44522":"Fish and Seafood Markets ",
"445220":"Fish and Seafood Markets ",
"44523":"Fruit and Vegetable Markets ",
"445230":"Fruit and Vegetable Markets ",
"44529":"Other Specialty Food Stores ",
"445291":"Baked Goods Stores ",
"445292":"Confectionery and Nut Stores ",
"445299":"All Other Specialty Food Stores ",
"4453":"Beer, Wine, and Liquor Stores ",
"44531":"Beer, Wine, and Liquor Stores ",
"445310":"Beer, Wine, and Liquor Stores ",
"446":"Health and Personal Care Stores ",
"4461":"Health and Personal Care Stores ",
"44611":"Pharmacies and Drug Stores ",
"446110":"Pharmacies and Drug Stores ",
"44612":"Cosmetics, Beauty Supplies, and Perfume Stores ",
"446120":"Cosmetics, Beauty Supplies, and Perfume Stores ",
"44613":"Optical Goods Stores ",
"446130":"Optical Goods Stores ",
"44619":"Other Health and Personal Care Stores ",
"446191":"Food (Health) Supplement Stores ",
"446199":"All Other Health and Personal Care Stores ",
"447":"Gasoline Stations ",
"4471":"Gasoline Stations ",
"44711":"Gasoline Stations with Convenience Stores ",
"447110":"Gasoline Stations with Convenience Stores ",
"44719":"Other Gasoline Stations ",
"447190":"Other Gasoline Stations ",
"448":"Clothing and Clothing Accessories Stores ",
"4481":"Clothing Stores ",
"44811":"Men's Clothing Stores ",
"448110":"Men's Clothing Stores ",
"44812":"Women's Clothing Stores ",
"448120":"Women's Clothing Stores ",
"44813":"Children's and Infants' Clothing Stores ",
"448130":"Children's and Infants' Clothing Stores ",
"44814":"Family Clothing Stores ",
"448140":"Family Clothing Stores ",
"44815":"Clothing Accessories Stores ",
"448150":"Clothing Accessories Stores ",
"44819":"Other Clothing Stores ",
"448190":"Other Clothing Stores ",
"4482":"Shoe Stores ",
"44821":"Shoe Stores ",
"448210":"Shoe Stores ",
"4483":"Jewelry, Luggage, and Leather Goods Stores ",
"44831":"Jewelry Stores ",
"448310":"Jewelry Stores ",
"44832":"Luggage and Leather Goods Stores ",
"448320":"Luggage and Leather Goods Stores ",
"451":"Sporting Goods, Hobby, Musical Instrument, and Book Stores ",
"4511":"Sporting Goods, Hobby, and Musical Instrument Stores ",
"45111":"Sporting Goods Stores ",
"451110":"Sporting Goods Stores ",
"45112":"Hobby, Toy, and Game Stores ",
"451120":"Hobby, Toy, and Game Stores ",
"45113":"Sewing, Needlework, and Piece Goods Stores ",
"451130":"Sewing, Needlework, and Piece Goods Stores ",
"45114":"Musical Instrument and Supplies Stores ",
"451140":"Musical Instrument and Supplies Stores ",
"4512":"Book Stores and News Dealers ",
"45121":"Book Stores and News Dealers ",
"451211":"Book Stores ",
"451212":"News Dealers and Newsstands ",
"452":"General Merchandise Stores ",
"4522":"Department Stores ",
"45221":"Department Stores ",
"452210":"Department Stores ",
"4523":"General Merchandise Stores, including Warehouse Clubs and Supercenters ",
"45231":"General Merchandise Stores, including Warehouse Clubs and Supercenters ",
"452311":"Warehouse Clubs and Supercenters ",
"452319":"All Other General Merchandise Stores ",
"453":"Miscellaneous Store Retailers ",
"4531":"Florists ",
"45311":"Florists ",
"453110":"Florists ",
"4532":"Office Supplies, Stationery, and Gift Stores ",
"45321":"Office Supplies and Stationery Stores ",
"453210":"Office Supplies and Stationery Stores ",
"45322":"Gift, Novelty, and Souvenir Stores ",
"453220":"Gift, Novelty, and Souvenir Stores ",
"4533":"Used Merchandise Stores ",
"45331":"Used Merchandise Stores ",
"453310":"Used Merchandise Stores ",
"4539":"Other Miscellaneous Store Retailers ",
"45391":"Pet and Pet Supplies Stores ",
"453910":"Pet and Pet Supplies Stores ",
"45392":"Art Dealers ",
"453920":"Art Dealers ",
"45393":"Manufactured (Mobile) Home Dealers ",
"453930":"Manufactured (Mobile) Home Dealers ",
"45399":"All Other Miscellaneous Store Retailers ",
"453991":"Tobacco Stores ",
"453998":"All Other Miscellaneous Store Retailers (except Tobacco Stores) ",
"454":"Nonstore Retailers ",
"4541":"Electronic Shopping and Mail-Order Houses ",
"45411":"Electronic Shopping and Mail-Order Houses ",
"454110":"Electronic Shopping and Mail-Order Houses ",
"4542":"Vending Machine Operators ",
"45421":"Vending Machine Operators ",
"454210":"Vending Machine Operators ",
"4543":"Direct Selling Establishments ",
"45431":"Fuel Dealers ",
"454310":"Fuel Dealers ",
"45439":"Other Direct Selling Establishments ",
"454390":"Other Direct Selling Establishments ",
"48":"Transportation and Warehousing",
"49":"Transportation and Warehousing",
"481":"Air Transportation",
"4811":"Scheduled Air Transportation",
"48111":"Scheduled Air Transportation",
"481111":"Scheduled Passenger Air Transportation ",
"481112":"Scheduled Freight Air Transportation ",
"4812":"Nonscheduled Air Transportation",
"48121":"Nonscheduled Air Transportation",
"481211":"Nonscheduled Chartered Passenger Air Transportation ",
"481212":"Nonscheduled Chartered Freight Air Transportation ",
"481219":"Other Nonscheduled Air Transportation ",
"482":"Rail Transportation",
"4821":"Rail Transportation",
"48211":"Rail Transportation",
"482111":"Line-Haul Railroads ",
"482112":"Short Line Railroads ",
"483":"Water Transportation",
"4831":"Deep Sea, Coastal, and Great Lakes Water Transportation",
"48311":"Deep Sea, Coastal, and Great Lakes Water Transportation",
"483111":"Deep Sea Freight Transportation ",
"483112":"Deep Sea Passenger Transportation ",
"483113":"Coastal and Great Lakes Freight Transportation ",
"483114":"Coastal and Great Lakes Passenger Transportation ",
"4832":"Inland Water Transportation",
"48321":"Inland Water Transportation",
"483211":"Inland Water Freight Transportation ",
"483212":"Inland Water Passenger Transportation ",
"484":"Truck Transportation",
"4841":"General Freight Trucking",
"48411":"General Freight Trucking, Local",
"484110":"General Freight Trucking, Local ",
"48412":"General Freight Trucking, Long-Distance",
"484121":"General Freight Trucking, Long-Distance, Truckload ",
"484122":"General Freight Trucking, Long-Distance, Less Than Truckload ",
"4842":"Specialized Freight Trucking",
"48421":"Used Household and Office Goods Moving",
"484210":"Used Household and Office Goods Moving",
"48422":"Specialized Freight (except Used Goods) Trucking, Local",
"484220":"Specialized Freight (except Used Goods) Trucking, Local ",
"48423":"Specialized Freight (except Used Goods) Trucking, Long-Distance",
"484230":"Specialized Freight (except Used Goods) Trucking, Long-Distance ",
"485":"Transit and Ground Passenger Transportation",
"4851":"Urban Transit Systems",
"48511":"Urban Transit Systems",
"485111":"Mixed Mode Transit Systems ",
"485112":"Commuter Rail Systems ",
"485113":"Bus and Other Motor Vehicle Transit Systems ",
"485119":"Other Urban Transit Systems ",
"4852":"Interurban and Rural Bus Transportation",
"48521":"Interurban and Rural Bus Transportation",
"485210":"Interurban and Rural Bus Transportation",
"4853":"Taxi and Limousine Service",
"48531":"Taxi Service",
"485310":"Taxi Service ",
"48532":"Limousine Service",
"485320":"Limousine Service",
"4854":"School and Employee Bus Transportation",
"48541":"School and Employee Bus Transportation",
"485410":"School and Employee Bus Transportation",
"4855":"Charter Bus Industry",
"48551":"Charter Bus Industry",
"485510":"Charter Bus Industry",
"4859":"Other Transit and Ground Passenger Transportation",
"48599":"Other Transit and Ground Passenger Transportation",
"485991":"Special Needs Transportation ",
"485999":"All Other Transit and Ground Passenger Transportation ",
"486":"Pipeline Transportation",
"4861":"Pipeline Transportation of Crude Oil",
"48611":"Pipeline Transportation of Crude Oil",
"486110":"Pipeline Transportation of Crude Oil",
"4862":"Pipeline Transportation of Natural Gas",
"48621":"Pipeline Transportation of Natural Gas",
"486210":"Pipeline Transportation of Natural Gas",
"4869":"Other Pipeline Transportation",
"48691":"Pipeline Transportation of Refined Petroleum Products",
"486910":"Pipeline Transportation of Refined Petroleum Products",
"48699":"All Other Pipeline Transportation",
"486990":"All Other Pipeline Transportation",
"487":"Scenic and Sightseeing Transportation",
"4871":"Scenic and Sightseeing Transportation, Land",
"48711":"Scenic and Sightseeing Transportation, Land",
"487110":"Scenic and Sightseeing Transportation, Land",
"4872":"Scenic and Sightseeing Transportation, Water",
"48721":"Scenic and Sightseeing Transportation, Water",
"487210":"Scenic and Sightseeing Transportation, Water",
"4879":"Scenic and Sightseeing Transportation, Other",
"48799":"Scenic and Sightseeing Transportation, Other",
"487990":"Scenic and Sightseeing Transportation, Other",
"488":"Support Activities for Transportation",
"4881":"Support Activities for Air Transportation",
"48811":"Airport Operations",
"488111":"Air Traffic Control",
"488119":"Other Airport Operations ",
"48819":"Other Support Activities for Air Transportation",
"488190":"Other Support Activities for Air Transportation",
"4882":"Support Activities for Rail Transportation",
"48821":"Support Activities for Rail Transportation",
"488210":"Support Activities for Rail Transportation",
"4883":"Support Activities for Water Transportation",
"48831":"Port and Harbor Operations",
"488310":"Port and Harbor Operations",
"48832":"Marine Cargo Handling",
"488320":"Marine Cargo Handling",
"48833":"Navigational Services to Shipping",
"488330":"Navigational Services to Shipping ",
"48839":"Other Support Activities for Water Transportation",
"488390":"Other Support Activities for Water Transportation",
"4884":"Support Activities for Road Transportation",
"48841":"Motor Vehicle Towing",
"488410":"Motor Vehicle Towing",
"48849":"Other Support Activities for Road Transportation",
"488490":"Other Support Activities for Road Transportation ",
"4885":"Freight Transportation Arrangement",
"48851":"Freight Transportation Arrangement",
"488510":"Freight Transportation Arrangement ",
"4889":"Other Support Activities for Transportation",
"48899":"Other Support Activities for Transportation",
"488991":"Packing and Crating ",
"488999":"All Other Support Activities for Transportation ",
"491":"Postal Service",
"4911":"Postal Service",
"49111":"Postal Service",
"491110":"Postal Service",
"492":"Couriers and Messengers",
"4921":"Couriers and Express Delivery Services",
"49211":"Couriers and Express Delivery Services",
"492110":"Couriers and Express Delivery Services",
"4922":"Local Messengers and Local Delivery",
"49221":"Local Messengers and Local Delivery",
"492210":"Local Messengers and Local Delivery",
"493":"Warehousing and Storage",
"4931":"Warehousing and Storage",
"49311":"General Warehousing and Storage",
"493110":"General Warehousing and Storage ",
"49312":"Refrigerated Warehousing and Storage",
"493120":"Refrigerated Warehousing and Storage",
"49313":"Farm Product Warehousing and Storage",
"493130":"Farm Product Warehousing and Storage",
"49319":"Other Warehousing and Storage",
"493190":"Other Warehousing and Storage",
"51":"Information",
"511":"Publishing Industries (except Internet)",
"5111":"Newspaper, Periodical, Book, and Directory Publishers",
"51111":"Newspaper Publishers",
"511110":"Newspaper Publishers ",
"51112":"Periodical Publishers",
"511120":"Periodical Publishers ",
"51113":"Book Publishers",
"511130":"Book Publishers ",
"51114":"Directory and Mailing List Publishers",
"511140":"Directory and Mailing List Publishers ",
"51119":"Other Publishers",
"511191":"Greeting Card Publishers ",
"511199":"All Other Publishers ",
"5112":"Software Publishers",
"51121":"Software Publishers",
"511210":"Software Publishers",
"512":"Motion Picture and Sound Recording Industries",
"5121":"Motion Picture and Video Industries",
"51211":"Motion Picture and Video Production",
"512110":"Motion Picture and Video Production ",
"51212":"Motion Picture and Video Distribution",
"512120":"Motion Picture and Video Distribution",
"51213":"Motion Picture and Video Exhibition",
"512131":"Motion Picture Theaters (except Drive-Ins) ",
"512132":"Drive-In Motion Picture Theaters ",
"51219":"Postproduction Services and Other Motion Picture and Video Industries",
"512191":"Teleproduction and Other Postproduction Services ",
"512199":"Other Motion Picture and Video Industries ",
"5122":"Sound Recording Industries",
"51223":"Music Publishers",
"512230":"Music Publishers",
"51224":"Sound Recording Studios",
"512240":"Sound Recording Studios",
"51225":"Record Production and Distribution",
"512250":"Record Production and Distribution",
"51229":"Other Sound Recording Industries",
"512290":"Other Sound Recording Industries",
"515":"Broadcasting (except Internet)",
"5151":"Radio and Television Broadcasting",
"51511":"Radio Broadcasting",
"515111":"Radio Networks ",
"515112":"Radio Stations ",
"51512":"Television Broadcasting",
"515120":"Television Broadcasting",
"5152":"Cable and Other Subscription Programming",
"51521":"Cable and Other Subscription Programming",
"515210":"Cable and Other Subscription Programming",
"517":"Telecommunications",
"5173":"Wired and Wireless Telecommunications Carriers",
"51731":"Wired and Wireless Telecommunications Carriers",
"517311":"Wired Telecommunications Carriers ",
"517312":"Wireless Telecommunications Carriers (except Satellite)",
"5174":"Satellite Telecommunications",
"51741":"Satellite Telecommunications",
"517410":"Satellite Telecommunications",
"5179":"Other Telecommunications",
"51791":"Other Telecommunications",
"517911":"Telecommunications Resellers ",
"517919":"All Other Telecommunications ",
"518":"Data Processing, Hosting, and Related Services",
"5182":"Data Processing, Hosting, and Related Services",
"51821":"Data Processing, Hosting, and Related Services",
"518210":"Data Processing, Hosting, and Related Services",
"519":"Other Information Services",
"5191":"Other Information Services",
"51911":"News Syndicates",
"519110":"News Syndicates",
"51912":"Libraries and Archives",
"519120":"Libraries and Archives ",
"51913":"Internet Publishing and Broadcasting and Web Search Portals",
"519130":"Internet Publishing and Broadcasting and Web Search Portals",
"51919":"All Other Information Services",
"519190":"All Other Information Services",
"52":"Finance and Insurance",
"521":"Monetary Authorities-Central Bank",
"5211":"Monetary Authorities-Central Bank",
"52111":"Monetary Authorities-Central Bank",
"521110":"Monetary Authorities-Central Bank",
"522":"Credit Intermediation and Related Activities",
"5221":"Depository Credit Intermediation ",
"52211":"Commercial Banking ",
"522110":"Commercial Banking ",
"52212":"Savings Institutions ",
"522120":"Savings Institutions ",
"52213":"Credit Unions ",
"522130":"Credit Unions ",
"52219":"Other Depository Credit Intermediation ",
"522190":"Other Depository Credit Intermediation ",
"5222":"Nondepository Credit Intermediation ",
"52221":"Credit Card Issuing ",
"522210":"Credit Card Issuing ",
"52222":"Sales Financing ",
"522220":"Sales Financing ",
"52229":"Other Nondepository Credit Intermediation ",
"522291":"Consumer Lending ",
"522292":"Real Estate Credit ",
"522293":"International Trade Financing ",
"522294":"Secondary Market Financing ",
"522298":"All Other Nondepository Credit Intermediation ",
"5223":"Activities Related to Credit Intermediation ",
"52231":"Mortgage and Nonmortgage Loan Brokers ",
"522310":"Mortgage and Nonmortgage Loan Brokers ",
"52232":"Financial Transactions Processing, Reserve, and Clearinghouse Activities ",
"522320":"Financial Transactions Processing, Reserve, and Clearinghouse Activities ",
"52239":"Other Activities Related to Credit Intermediation ",
"522390":"Other Activities Related to Credit Intermediation ",
"523":"Securities, Commodity Contracts, and Other Financial Investments and Related Activities",
"5231":"Securities and Commodity Contracts Intermediation and Brokerage",
"52311":"Investment Banking and Securities Dealing ",
"523110":"Investment Banking and Securities Dealing ",
"52312":"Securities Brokerage ",
"523120":"Securities Brokerage ",
"52313":"Commodity Contracts Dealing ",
"523130":"Commodity Contracts Dealing ",
"52314":"Commodity Contracts Brokerage ",
"523140":"Commodity Contracts Brokerage ",
"5232":"Securities and Commodity Exchanges",
"52321":"Securities and Commodity Exchanges",
"523210":"Securities and Commodity Exchanges",
"5239":"Other Financial Investment Activities",
"52391":"Miscellaneous Intermediation ",
"523910":"Miscellaneous Intermediation ",
"52392":"Portfolio Management ",
"523920":"Portfolio Management ",
"52393":"Investment Advice ",
"523930":"Investment Advice ",
"52399":"All Other Financial Investment Activities ",
"523991":"Trust, Fiduciary, and Custody Activities ",
"523999":"Miscellaneous Financial Investment Activities ",
"524":"Insurance Carriers and Related Activities",
"5241":"Insurance Carriers",
"52411":"Direct Life, Health, and Medical Insurance Carriers ",
"524113":"Direct Life Insurance Carriers ",
"524114":"Direct Health and Medical Insurance Carriers ",
"52412":"Direct Insurance (except Life, Health, and Medical) Carriers ",
"524126":"Direct Property and Casualty Insurance Carriers ",
"524127":"Direct Title Insurance Carriers ",
"524128":"Other Direct Insurance (except Life, Health, and Medical) Carriers ",
"52413":"Reinsurance Carriers ",
"524130":"Reinsurance Carriers ",
"5242":"Agencies, Brokerages, and Other Insurance Related Activities",
"52421":"Insurance Agencies and Brokerages ",
"524210":"Insurance Agencies and Brokerages ",
"52429":"Other Insurance Related Activities ",
"524291":"Claims Adjusting ",
"524292":"Third Party Administration of Insurance and Pension Funds ",
"524298":"All Other Insurance Related Activities ",
"525":"Funds, Trusts, and Other Financial Vehicles ",
"5251":"Insurance and Employee Benefit Funds ",
"52511":"Pension Funds ",
"525110":"Pension Funds ",
"52512":"Health and Welfare Funds ",
"525120":"Health and Welfare Funds ",
"52519":"Other Insurance Funds ",
"525190":"Other Insurance Funds ",
"5259":"Other Investment Pools and Funds",
"52591":"Open-End Investment Funds ",
"525910":"Open-End Investment Funds ",
"52592":"Trusts, Estates, and Agency Accounts ",
"525920":"Trusts, Estates, and Agency Accounts ",
"52599":"Other Financial Vehicles ",
"525990":"Other Financial Vehicles ",
"53":"Real Estate and Rental and Leasing",
"531":"Real Estate",
"5311":"Lessors of Real Estate",
"53111":"Lessors of Residential Buildings and Dwellings ",
"531110":"Lessors of Residential Buildings and Dwellings ",
"53112":"Lessors of Nonresidential Buildings (except Miniwarehouses) ",
"531120":"Lessors of Nonresidential Buildings (except Miniwarehouses) ",
"53113":"Lessors of Miniwarehouses and Self-Storage Units ",
"531130":"Lessors of Miniwarehouses and Self-Storage Units ",
"53119":"Lessors of Other Real Estate Property ",
"531190":"Lessors of Other Real Estate Property ",
"5312":"Offices of Real Estate Agents and Brokers",
"53121":"Offices of Real Estate Agents and Brokers",
"531210":"Offices of Real Estate Agents and Brokers",
"5313":"Activities Related to Real Estate",
"53131":"Real Estate Property Managers ",
"531311":"Residential Property Managers ",
"531312":"Nonresidential Property Managers ",
"53132":"Offices of Real Estate Appraisers ",
"531320":"Offices of Real Estate Appraisers ",
"53139":"Other Activities Related to Real Estate ",
"531390":"Other Activities Related to Real Estate ",
"532":"Rental and Leasing Services",
"5321":"Automotive Equipment Rental and Leasing",
"53211":"Passenger Car Rental and Leasing",
"532111":"Passenger Car Rental ",
"532112":"Passenger Car Leasing ",
"53212":"Truck, Utility Trailer, and RV (Recreational Vehicle) Rental and Leasing",
"532120":"Truck, Utility Trailer, and RV (Recreational Vehicle) Rental and Leasing ",
"5322":"Consumer Goods Rental",
"53221":"Consumer Electronics and Appliances Rental",
"532210":"Consumer Electronics and Appliances Rental",
"53228":"Other Consumer Goods Rental ",
"532281":"Formal Wear and Costume Rental",
"532282":"Video Tape and Disc Rental",
"532283":"Home Health Equipment Rental ",
"532284":"Recreational Goods Rental ",
"532289":"All Other Consumer Goods Rental ",
"5323":"General Rental Centers",
"53231":"General Rental Centers",
"532310":"General Rental Centers",
"5324":"Commercial and Industrial Machinery and Equipment Rental and Leasing",
"53241":"Construction, Transportation, Mining, and Forestry Machinery and Equipment Rental and Leasing",
"532411":"Commercial Air, Rail, and Water Transportation Equipment Rental and Leasing ",
"532412":"Construction, Mining, and Forestry Machinery and Equipment Rental and Leasing ",
"53242":"Office Machinery and Equipment Rental and Leasing",
"532420":"Office Machinery and Equipment Rental and Leasing",
"53249":"Other Commercial and Industrial Machinery and Equipment Rental and Leasing",
"532490":"Other Commercial and Industrial Machinery and Equipment Rental and Leasing ",
"533":"Lessors of Nonfinancial Intangible Assets (except Copyrighted Works)",
"5331":"Lessors of Nonfinancial Intangible Assets (except Copyrighted Works)",
"53311":"Lessors of Nonfinancial Intangible Assets (except Copyrighted Works)",
"533110":"Lessors of Nonfinancial Intangible Assets (except Copyrighted Works)",
"54":"Professional, Scientific, and Technical Services",
"541":"Professional, Scientific, and Technical Services",
"5411":"Legal Services",
"54111":"Offices of Lawyers",
"541110":"Offices of Lawyers",
"54112":"Offices of Notaries",
"541120":"Offices of Notaries",
"54119":"Other Legal Services",
"541191":"Title Abstract and Settlement Offices ",
"541199":"All Other Legal Services ",
"5412":"Accounting, Tax Preparation, Bookkeeping, and Payroll Services",
"54121":"Accounting, Tax Preparation, Bookkeeping, and Payroll Services",
"541211":"Offices of Certified Public Accountants ",
"541213":"Tax Preparation Services ",
"541214":"Payroll Services ",
"541219":"Other Accounting Services ",
"5413":"Architectural, Engineering, and Related Services",
"54131":"Architectural Services",
"541310":"Architectural Services",
"54132":"Landscape Architectural Services",
"541320":"Landscape Architectural Services",
"54133":"Engineering Services",
"541330":"Engineering Services",
"54134":"Drafting Services",
"541340":"Drafting Services",
"54135":"Building Inspection Services",
"541350":"Building Inspection Services",
"54136":"Geophysical Surveying and Mapping Services",
"541360":"Geophysical Surveying and Mapping Services",
"54137":"Surveying and Mapping (except Geophysical) Services",
"541370":"Surveying and Mapping (except Geophysical) Services",
"54138":"Testing Laboratories",
"541380":"Testing Laboratories",
"5414":"Specialized Design Services",
"54141":"Interior Design Services",
"541410":"Interior Design Services",
"54142":"Industrial Design Services",
"541420":"Industrial Design Services",
"54143":"Graphic Design Services",
"541430":"Graphic Design Services",
"54149":"Other Specialized Design Services",
"541490":"Other Specialized Design Services",
"5415":"Computer Systems Design and Related Services",
"54151":"Computer Systems Design and Related Services",
"541511":"Custom Computer Programming Services ",
"541512":"Computer Systems Design Services ",
"541513":"Computer Facilities Management Services ",
"541519":"Other Computer Related Services",
"5416":"Management, Scientific, and Technical Consulting Services",
"54161":"Management Consulting Services",
"541611":"Administrative Management and General Management Consulting Services ",
"541612":"Human Resources Consulting Services ",
"541613":"Marketing Consulting Services ",
"541614":"Process, Physical Distribution, and Logistics Consulting Services ",
"541618":"Other Management Consulting Services ",
"54162":"Environmental Consulting Services",
"541620":"Environmental Consulting Services",
"54169":"Other Scientific and Technical Consulting Services",
"541690":"Other Scientific and Technical Consulting Services",
"5417":"Scientific Research and Development Services",
"54171":"Research and Development in the Physical, Engineering, and Life Sciences",
"541713":"Research and Development in Nanotechnology ",
"541714":"Research and Development in Biotechnology (except Nanobiotechnology)",
"541715":"Research and Development in the Physical, Engineering, and Life Sciences (except Nanotechnology and Biotechnology) ",
"54172":"Research and Development in the Social Sciences and Humanities",
"541720":"Research and Development in the Social Sciences and Humanities ",
"5418":"Advertising, Public Relations, and Related Services",
"54181":"Advertising Agencies",
"541810":"Advertising Agencies",
"54182":"Public Relations Agencies",
"541820":"Public Relations Agencies",
"54183":"Media Buying Agencies",
"541830":"Media Buying Agencies",
"54184":"Media Representatives",
"541840":"Media Representatives",
"54185":"Outdoor Advertising",
"541850":"Outdoor Advertising",
"54186":"Direct Mail Advertising",
"541860":"Direct Mail Advertising",
"54187":"Advertising Material Distribution Services",
"541870":"Advertising Material Distribution Services",
"54189":"Other Services Related to Advertising",
"541890":"Other Services Related to Advertising ",
"5419":"Other Professional, Scientific, and Technical Services",
"54191":"Marketing Research and Public Opinion Polling",
"541910":"Marketing Research and Public Opinion Polling",
"54192":"Photographic Services",
"541921":"Photography Studios, Portrait ",
"541922":"Commercial Photography ",
"54193":"Translation and Interpretation Services",
"541930":"Translation and Interpretation Services",
"54194":"Veterinary Services",
"541940":"Veterinary Services ",
"54199":"All Other Professional, Scientific, and Technical Services",
"541990":"All Other Professional, Scientific, and Technical Services",
"55":"Management of Companies and Enterprises",
"551":"Management of Companies and Enterprises",
"5511":"Management of Companies and Enterprises",
"55111":"Management of Companies and Enterprises",
"551111":"Offices of Bank Holding Companies ",
"551112":"Offices of Other Holding Companies ",
"551114":"Corporate, Subsidiary, and Regional Managing Offices ",
"56":"Administrative and Support and Waste Management and Remediation Services",
"561":"Administrative and Support Services",
"5611":"Office Administrative Services",
"56111":"Office Administrative Services",
"561110":"Office Administrative Services",
"5612":"Facilities Support Services",
"56121":"Facilities Support Services",
"561210":"Facilities Support Services",
"5613":"Employment Services",
"56131":"Employment Placement Agencies and Executive Search Services",
"561311":"Employment Placement Agencies ",
"561312":"Executive Search Services ",
"56132":"Temporary Help Services",
"561320":"Temporary Help Services",
"56133":"Professional Employer Organizations",
"561330":"Professional Employer Organizations",
"5614":"Business Support Services",
"56141":"Document Preparation Services",
"561410":"Document Preparation Services",
"56142":"Telephone Call Centers",
"561421":"Telephone Answering Services ",
"561422":"Telemarketing Bureaus and Other Contact Centers ",
"56143":"Business Service Centers",
"561431":"Private Mail Centers ",
"561439":"Other Business Service Centers (including Copy Shops) ",
"56144":"Collection Agencies",
"561440":"Collection Agencies",
"56145":"Credit Bureaus",
"561450":"Credit Bureaus",
"56149":"Other Business Support Services",
"561491":"Repossession Services ",
"561492":"Court Reporting and Stenotype Services ",
"561499":"All Other Business Support Services ",
"5615":"Travel Arrangement and Reservation Services",
"56151":"Travel Agencies",
"561510":"Travel Agencies",
"56152":"Tour Operators",
"561520":"Tour Operators",
"56159":"Other Travel Arrangement and Reservation Services",
"561591":"Convention and Visitors Bureaus ",
"561599":"All Other Travel Arrangement and Reservation Services ",
"5616":"Investigation and Security Services",
"56161":"Investigation, Guard, and Armored Car Services",
"561611":"Investigation Services ",
"561612":"Security Guards and Patrol Services ",
"561613":"Armored Car Services ",
"56162":"Security Systems Services",
"561621":"Security Systems Services (except Locksmiths) ",
"561622":"Locksmiths ",
"5617":"Services to Buildings and Dwellings",
"56171":"Exterminating and Pest Control Services",
"561710":"Exterminating and Pest Control Services",
"56172":"Janitorial Services",
"561720":"Janitorial Services ",
"56173":"Landscaping Services",
"561730":"Landscaping Services",
"56174":"Carpet and Upholstery Cleaning Services",
"561740":"Carpet and Upholstery Cleaning Services",
"56179":"Other Services to Buildings and Dwellings",
"561790":"Other Services to Buildings and Dwellings ",
"5619":"Other Support Services",
"56191":"Packaging and Labeling Services",
"561910":"Packaging and Labeling Services",
"56192":"Convention and Trade Show Organizers",
"561920":"Convention and Trade Show Organizers",
"56199":"All Other Support Services",
"561990":"All Other Support Services",
"562":"Waste Management and Remediation Services",
"5621":"Waste Collection ",
"56211":"Waste Collection ",
"562111":"Solid Waste Collection ",
"562112":"Hazardous Waste Collection ",
"562119":"Other Waste Collection ",
"5622":"Waste Treatment and Disposal ",
"56221":"Waste Treatment and Disposal ",
"562211":"Hazardous Waste Treatment and Disposal ",
"562212":"Solid Waste Landfill ",
"562213":"Solid Waste Combustors and Incinerators ",
"562219":"Other Nonhazardous Waste Treatment and Disposal ",
"5629":"Remediation and Other Waste Management Services ",
"56291":"Remediation Services ",
"562910":"Remediation Services ",
"56292":"Materials Recovery Facilities ",
"562920":"Materials Recovery Facilities ",
"56299":"All Other Waste Management Services ",
"562991":"Septic Tank and Related Services ",
"562998":"All Other Miscellaneous Waste Management Services ",
"61":"Educational Services",
"611":"Educational Services",
"6111":"Elementary and Secondary Schools",
"61111":"Elementary and Secondary Schools ",
"611110":"Elementary and Secondary Schools ",
"6112":"Junior Colleges",
"61121":"Junior Colleges",
"611210":"Junior Colleges ",
"6113":"Colleges, Universities, and Professional Schools",
"61131":"Colleges, Universities, and Professional Schools",
"611310":"Colleges, Universities, and Professional Schools ",
"6114":"Business Schools and Computer and Management Training",
"61141":"Business and Secretarial Schools",
"611410":"Business and Secretarial Schools ",
"61142":"Computer Training",
"611420":"Computer Training ",
"61143":"Professional and Management Development Training",
"611430":"Professional and Management Development Training ",
"6115":"Technical and Trade Schools ",
"61151":"Technical and Trade Schools",
"611511":"Cosmetology and Barber Schools ",
"611512":"Flight Training ",
"611513":"Apprenticeship Training ",
"611519":"Other Technical and Trade Schools ",
"6116":"Other Schools and Instruction",
"61161":"Fine Arts Schools",
"611610":"Fine Arts Schools ",
"61162":"Sports and Recreation Instruction",
"611620":"Sports and Recreation Instruction ",
"61163":"Language Schools",
"611630":"Language Schools ",
"61169":"All Other Schools and Instruction",
"611691":"Exam Preparation and Tutoring ",
"611692":"Automobile Driving Schools ",
"611699":"All Other Miscellaneous Schools and Instruction ",
"6117":"Educational Support Services",
"61171":"Educational Support Services",
"611710":"Educational Support Services",
"62":"Health Care and Social Assistance",
"621":"Ambulatory Health Care Services",
"6211":"Offices of Physicians",
"62111":"Offices of Physicians",
"621111":"Offices of Physicians (except Mental Health Specialists) ",
"621112":"Offices of Physicians, Mental Health Specialists ",
"6212":"Offices of Dentists",
"62121":"Offices of Dentists",
"621210":"Offices of Dentists ",
"6213":"Offices of Other Health Practitioners",
"62131":"Offices of Chiropractors",
"621310":"Offices of Chiropractors ",
"62132":"Offices of Optometrists",
"621320":"Offices of Optometrists",
"62133":"Offices of Mental Health Practitioners (except Physicians)",
"621330":"Offices of Mental Health Practitioners (except Physicians) ",
"62134":"Offices of Physical, Occupational and Speech Therapists, and Audiologists",
"621340":"Offices of Physical, Occupational and Speech Therapists, and Audiologists ",
"62139":"Offices of All Other Health Practitioners",
"621391":"Offices of Podiatrists ",
"621399":"Offices of All Other Miscellaneous Health Practitioners ",
"6214":"Outpatient Care Centers",
"62141":"Family Planning Centers",
"621410":"Family Planning Centers ",
"62142":"Outpatient Mental Health and Substance Abuse Centers",
"621420":"Outpatient Mental Health and Substance Abuse Centers ",
"62149":"Other Outpatient Care Centers",
"621491":"HMO Medical Centers ",
"621492":"Kidney Dialysis Centers ",
"621493":"Freestanding Ambulatory Surgical and Emergency Centers ",
"621498":"All Other Outpatient Care Centers ",
"6215":"Medical and Diagnostic Laboratories",
"62151":"Medical and Diagnostic Laboratories",
"621511":"Medical Laboratories ",
"621512":"Diagnostic Imaging Centers ",
"6216":"Home Health Care Services",
"62161":"Home Health Care Services",
"621610":"Home Health Care Services",
"6219":"Other Ambulatory Health Care Services",
"62191":"Ambulance Services",
"621910":"Ambulance Services ",
"62199":"All Other Ambulatory Health Care Services",
"621991":"Blood and Organ Banks ",
"621999":"All Other Miscellaneous Ambulatory Health Care Services ",
"622":"Hospitals",
"6221":"General Medical and Surgical Hospitals",
"62211":"General Medical and Surgical Hospitals",
"622110":"General Medical and Surgical Hospitals ",
"6222":"Psychiatric and Substance Abuse Hospitals",
"62221":"Psychiatric and Substance Abuse Hospitals",
"622210":"Psychiatric and Substance Abuse Hospitals ",
"6223":"Specialty (except Psychiatric and Substance Abuse) Hospitals",
"62231":"Specialty (except Psychiatric and Substance Abuse) Hospitals",
"622310":"Specialty (except Psychiatric and Substance Abuse) Hospitals ",
"623":"Nursing and Residential Care Facilities",
"6231":"Nursing Care Facilities (Skilled Nursing Facilities)",
"62311":"Nursing Care Facilities (Skilled Nursing Facilities)",
"623110":"Nursing Care Facilities (Skilled Nursing Facilities) ",
"6232":"Residential Intellectual and Developmental Disability, Mental Health, and Substance Abuse Facilities",
"62321":"Residential Intellectual and Developmental Disability Facilities",
"623210":"Residential Intellectual and Developmental Disability Facilities ",
"62322":"Residential Mental Health and Substance Abuse Facilities",
"623220":"Residential Mental Health and Substance Abuse Facilities ",
"6233":"Continuing Care Retirement Communities and Assisted Living Facilities for the Elderly",
"62331":"Continuing Care Retirement Communities and Assisted Living Facilities for the Elderly",
"623311":"Continuing Care Retirement Communities ",
"623312":"Assisted Living Facilities for the Elderly ",
"6239":"Other Residential Care Facilities",
"62399":"Other Residential Care Facilities",
"623990":"Other Residential Care Facilities ",
"624":"Social Assistance",
"6241":"Individual and Family Services",
"62411":"Child and Youth Services",
"624110":"Child and Youth Services ",
"62412":"Services for the Elderly and Persons with Disabilities",
"624120":"Services for the Elderly and Persons with Disabilities ",
"62419":"Other Individual and Family Services",
"624190":"Other Individual and Family Services ",
"6242":"Community Food and Housing, and Emergency and Other Relief Services",
"62421":"Community Food Services",
"624210":"Community Food Services ",
"62422":"Community Housing Services",
"624221":"Temporary Shelters ",
"624229":"Other Community Housing Services ",
"62423":"Emergency and Other Relief Services",
"624230":"Emergency and Other Relief Services ",
"6243":"Vocational Rehabilitation Services",
"62431":"Vocational Rehabilitation Services",
"624310":"Vocational Rehabilitation Services ",
"6244":"Child Day Care Services",
"62441":"Child Day Care Services",
"624410":"Child Day Care Services ",
"71":"Arts, Entertainment, and Recreation",
"711":"Performing Arts, Spectator Sports, and Related Industries",
"7111":"Performing Arts Companies",
"71111":"Theater Companies and Dinner Theaters",
"711110":"Theater Companies and Dinner Theaters ",
"71112":"Dance Companies",
"711120":"Dance Companies ",
"71113":"Musical Groups and Artists",
"711130":"Musical Groups and Artists ",
"71119":"Other Performing Arts Companies",
"711190":"Other Performing Arts Companies ",
"7112":"Spectator Sports",
"71121":"Spectator Sports",
"711211":"Sports Teams and Clubs ",
"711212":"Racetracks ",
"711219":"Other Spectator Sports ",
"7113":"Promoters of Performing Arts, Sports, and Similar Events",
"71131":"Promoters of Performing Arts, Sports, and Similar Events with Facilities",
"711310":"Promoters of Performing Arts, Sports, and Similar Events with Facilities ",
"71132":"Promoters of Performing Arts, Sports, and Similar Events without Facilities",
"711320":"Promoters of Performing Arts, Sports, and Similar Events without Facilities ",
"7114":"Agents and Managers for Artists, Athletes, Entertainers, and Other Public Figures",
"71141":"Agents and Managers for Artists, Athletes, Entertainers, and Other Public Figures",
"711410":"Agents and Managers for Artists, Athletes, Entertainers, and Other Public Figures",
"7115":"Independent Artists, Writers, and Performers",
"71151":"Independent Artists, Writers, and Performers",
"711510":"Independent Artists, Writers, and Performers ",
"712":"Museums, Historical Sites, and Similar Institutions",
"7121":"Museums, Historical Sites, and Similar Institutions",
"71211":"Museums",
"712110":"Museums ",
"71212":"Historical Sites",
"712120":"Historical Sites",
"71213":"Zoos and Botanical Gardens",
"712130":"Zoos and Botanical Gardens ",
"71219":"Nature Parks and Other Similar Institutions",
"712190":"Nature Parks and Other Similar Institutions",
"713":"Amusement, Gambling, and Recreation Industries",
"7131":"Amusement Parks and Arcades",
"71311":"Amusement and Theme Parks",
"713110":"Amusement and Theme Parks ",
"71312":"Amusement Arcades",
"713120":"Amusement Arcades",
"7132":"Gambling Industries",
"71321":"Casinos (except Casino Hotels)",
"713210":"Casinos (except Casino Hotels)",
"71329":"Other Gambling Industries",
"713290":"Other Gambling Industries ",
"7139":"Other Amusement and Recreation Industries",
"71391":"Golf Courses and Country Clubs",
"713910":"Golf Courses and Country Clubs",
"71392":"Skiing Facilities",
"713920":"Skiing Facilities",
"71393":"Marinas",
"713930":"Marinas",
"71394":"Fitness and Recreational Sports Centers",
"713940":"Fitness and Recreational Sports Centers ",
"71395":"Bowling Centers",
"713950":"Bowling Centers",
"71399":"All Other Amusement and Recreation Industries",
"713990":"All Other Amusement and Recreation Industries ",
"72":"Accommodation and Food Services",
"721":"Accommodation",
"7211":"Traveler Accommodation",
"72111":"Hotels (except Casino Hotels) and Motels",
"721110":"Hotels (except Casino Hotels) and Motels ",
"72112":"Casino Hotels",
"721120":"Casino Hotels",
"72119":"Other Traveler Accommodation",
"721191":"Bed-and-Breakfast Inns ",
"721199":"All Other Traveler Accommodation ",
"7212":"RV (Recreational Vehicle) Parks and Recreational Camps",
"72121":"RV (Recreational Vehicle) Parks and Recreational Camps",
"721211":"RV (Recreational Vehicle) Parks and Campgrounds ",
"721214":"Recreational and Vacation Camps (except Campgrounds) ",
"7213":"Rooming and Boarding Houses, Dormitories, and Workers' Camps",
"72131":"Rooming and Boarding Houses, Dormitories, and Workers' Camps",
"721310":"Rooming and Boarding Houses, Dormitories, and Workers' Camps ",
"722":"Food Services and Drinking Places",
"7223":"Special Food Services",
"72231":"Food Service Contractors",
"722310":"Food Service Contractors",
"72232":"Caterers",
"722320":"Caterers",
"72233":"Mobile Food Services",
"722330":"Mobile Food Services",
"7224":"Drinking Places (Alcoholic Beverages)",
"72241":"Drinking Places (Alcoholic Beverages)",
"722410":"Drinking Places (Alcoholic Beverages) ",
"7225":"Restaurants and Other Eating Places",
"72251":"Restaurants and Other Eating Places",
"722511":"Full-Service Restaurants ",
"722513":"Limited-Service Restaurants ",
"722514":"Cafeterias, Grill Buffets, and Buffets ",
"722515":"Snack and Nonalcoholic Beverage Bars ",
"81":"Other Services (except Public Administration)",
"811":"Repair and Maintenance",
"8111":"Automotive Repair and Maintenance",
"81111":"Automotive Mechanical and Electrical Repair and Maintenance",
"811111":"General Automotive Repair ",
"811112":"Automotive Exhaust System Repair ",
"811113":"Automotive Transmission Repair ",
"811118":"Other Automotive Mechanical and Electrical Repair and Maintenance ",
"81112":"Automotive Body, Paint, Interior, and Glass Repair",
"811121":"Automotive Body, Paint, and Interior Repair and Maintenance ",
"811122":"Automotive Glass Replacement Shops ",
"81119":"Other Automotive Repair and Maintenance",
"811191":"Automotive Oil Change and Lubrication Shops ",
"811192":"Car Washes ",
"811198":"All Other Automotive Repair and Maintenance ",
"8112":"Electronic and Precision Equipment Repair and Maintenance",
"81121":"Electronic and Precision Equipment Repair and Maintenance",
"811211":"Consumer Electronics Repair and Maintenance ",
"811212":"Computer and Office Machine Repair and Maintenance ",
"811213":"Communication Equipment Repair and Maintenance ",
"811219":"Other Electronic and Precision Equipment Repair and Maintenance ",
"8113":"Commercial and Industrial Machinery and Equipment (except Automotive and Electronic) Repair and Maintenance",
"81131":"Commercial and Industrial Machinery and Equipment (except Automotive and Electronic) Repair and Maintenance",
"811310":"Commercial and Industrial Machinery and Equipment (except Automotive and Electronic) Repair and Maintenance ",
"8114":"Personal and Household Goods Repair and Maintenance",
"81141":"Home and Garden Equipment and Appliance Repair and Maintenance",
"811411":"Home and Garden Equipment Repair and Maintenance ",
"811412":"Appliance Repair and Maintenance ",
"81142":"Reupholstery and Furniture Repair",
"811420":"Reupholstery and Furniture Repair",
"81143":"Footwear and Leather Goods Repair",
"811430":"Footwear and Leather Goods Repair",
"81149":"Other Personal and Household Goods Repair and Maintenance",
"811490":"Other Personal and Household Goods Repair and Maintenance ",
"812":"Personal and Laundry Services",
"8121":"Personal Care Services ",
"81211":"Hair, Nail, and Skin Care Services ",
"812111":"Barber Shops ",
"812112":"Beauty Salons ",
"812113":"Nail Salons ",
"81219":"Other Personal Care Services ",
"812191":"Diet and Weight Reducing Centers ",
"812199":"Other Personal Care Services ",
"8122":"Death Care Services ",
"81221":"Funeral Homes and Funeral Services ",
"812210":"Funeral Homes and Funeral Services ",
"81222":"Cemeteries and Crematories ",
"812220":"Cemeteries and Crematories ",
"8123":"Drycleaning and Laundry Services ",
"81231":"Coin-Operated Laundries and Drycleaners ",
"812310":"Coin-Operated Laundries and Drycleaners ",
"81232":"Drycleaning and Laundry Services (except Coin-Operated) ",
"812320":"Drycleaning and Laundry Services (except Coin-Operated) ",
"81233":"Linen and Uniform Supply ",
"812331":"Linen Supply ",
"812332":"Industrial Launderers ",
"8129":"Other Personal Services ",
"81291":"Pet Care (except Veterinary) Services ",
"812910":"Pet Care (except Veterinary) Services ",
"81292":"Photofinishing ",
"812921":"Photofinishing Laboratories (except One-Hour) ",
"812922":"One-Hour Photofinishing ",
"81293":"Parking Lots and Garages ",
"812930":"Parking Lots and Garages ",
"81299":"All Other Personal Services ",
"812990":"All Other Personal Services ",
"813":"Religious, Grantmaking, Civic, Professional, and Similar Organizations",
"8131":"Religious Organizations ",
"81311":"Religious Organizations ",
"813110":"Religious Organizations ",
"8132":"Grantmaking and Giving Services ",
"81321":"Grantmaking and Giving Services ",
"813211":"Grantmaking Foundations ",
"813212":"Voluntary Health Organizations ",
"813219":"Other Grantmaking and Giving Services ",
"8133":"Social Advocacy Organizations ",
"81331":"Social Advocacy Organizations ",
"813311":"Human Rights Organizations ",
"813312":"Environment, Conservation and Wildlife Organizations ",
"813319":"Other Social Advocacy Organizations ",
"8134":"Civic and Social Organizations ",
"81341":"Civic and Social Organizations ",
"813410":"Civic and Social Organizations ",
"8139":"Business, Professional, Labor, Political, and Similar Organizations ",
"81391":"Business Associations ",
"813910":"Business Associations ",
"81392":"Professional Organizations ",
"813920":"Professional Organizations ",
"81393":"Labor Unions and Similar Labor Organizations ",
"813930":"Labor Unions and Similar Labor Organizations ",
"81394":"Political Organizations ",
"813940":"Political Organizations ",
"81399":"Other Similar Organizations (except Business, Professional, Labor, and Political Organizations) ",
"813990":"Other Similar Organizations (except Business, Professional, Labor, and Political Organizations) ",
"814":"Private Households",
"8141":"Private Households",
"81411":"Private Households",
"814110":"Private Households",
"92":"Public Administration",
"921":"Executive, Legislative, and Other General Government Support ",
"9211":"Executive, Legislative, and Other General Government Support ",
"92111":"Executive Offices ",
"921110":"Executive Offices ",
"92112":"Legislative Bodies ",
"921120":"Legislative Bodies ",
"92113":"Public Finance Activities ",
"921130":"Public Finance Activities ",
"92114":"Executive and Legislative Offices, Combined ",
"921140":"Executive and Legislative Offices, Combined ",
"92115":"American Indian and Alaska Native Tribal Governments ",
"921150":"American Indian and Alaska Native Tribal Governments ",
"92119":"Other General Government Support ",
"921190":"Other General Government Support ",
"922":"Justice, Public Order, and Safety Activities ",
"9221":"Justice, Public Order, and Safety Activities ",
"92211":"Courts ",
"922110":"Courts ",
"92212":"Police Protection ",
"922120":"Police Protection ",
"92213":"Legal Counsel and Prosecution ",
"922130":"Legal Counsel and Prosecution ",
"92214":"Correctional Institutions ",
"922140":"Correctional Institutions ",
"92215":"Parole Offices and Probation Offices ",
"922150":"Parole Offices and Probation Offices ",
"92216":"Fire Protection ",
"922160":"Fire Protection ",
"92219":"Other Justice, Public Order, and Safety Activities ",
"922190":"Other Justice, Public Order, and Safety Activities ",
"923":"Administration of Human Resource Programs ",
"9231":"Administration of Human Resource Programs ",
"92311":"Administration of Education Programs ",
"923110":"Administration of Education Programs ",
"92312":"Administration of Public Health Programs ",
"923120":"Administration of Public Health Programs ",
"92313":"Administration of Human Resource Programs (except Education, Public Health, and Veterans' Affairs Programs) ",
"923130":"Administration of Human Resource Programs (except Education, Public Health, and Veterans' Affairs Programs) ",
"92314":"Administration of Veterans' Affairs ",
"923140":"Administration of Veterans' Affairs ",
"924":"Administration of Environmental Quality Programs ",
"9241":"Administration of Environmental Quality Programs ",
"92411":"Administration of Air and Water Resource and Solid Waste Management Programs ",
"924110":"Administration of Air and Water Resource and Solid Waste Management Programs ",
"92412":"Administration of Conservation Programs ",
"924120":"Administration of Conservation Programs ",
"925":"Administration of Housing Programs, Urban Planning, and Community Development ",
"9251":"Administration of Housing Programs, Urban Planning, and Community Development ",
"92511":"Administration of Housing Programs ",
"925110":"Administration of Housing Programs ",
"92512":"Administration of Urban Planning and Community and Rural Development ",
"925120":"Administration of Urban Planning and Community and Rural Development ",
"926":"Administration of Economic Programs ",
"9261":"Administration of Economic Programs ",
"92611":"Administration of General Economic Programs ",
"926110":"Administration of General Economic Programs ",
"92612":"Regulation and Administration of Transportation Programs ",
"926120":"Regulation and Administration of Transportation Programs ",
"92613":"Regulation and Administration of Communications, Electric, Gas, and Other Utilities ",
"926130":"Regulation and Administration of Communications, Electric, Gas, and Other Utilities ",
"92614":"Regulation of Agricultural Marketing and Commodities ",
"926140":"Regulation of Agricultural Marketing and Commodities ",
"92615":"Regulation, Licensing, and Inspection of Miscellaneous Commercial Sectors ",
"926150":"Regulation, Licensing, and Inspection of Miscellaneous Commercial Sectors ",
"927":"Space Research and Technology ",
"9271":"Space Research and Technology ",
"92711":"Space Research and Technology ",
"927110":"Space Research and Technology ",
"928":"National Security and International Affairs ",
"9281":"National Security and International Affairs ",
"92811":"National Security ",
"928110":"National Security ",
"92812":"International Affairs ",
"928120":"International Affairs ",
}



gics_code = {
'10': 'Energy', 
'15': 'Materials', 
'20': 'Industrials', 
'25': 'Consumer Discretionary', 
'30': 'Consumer Staples', 
'35': 'Health Care', 
'40': 'Financials', 
'45': 'Information Technology', 
'50': 'Communication Services', 
'55': 'Utilities', 
'60': 'Real Estate', 
'1010': 'Energy', 
'1510': 'Materials', 
'2010': 'Capital Goods', 
'2020': 'Commercial  & Professional Services', 
'2030': 'Transportation', 
'2510': 'Automobiles & Components', 
'2520': 'Consumer Durables & Apparel', 
'2530': 'Consumer Services', 
'2540': 'Media (discontinued effective close of September 28, 2018)', 
'2550': 'Retailing', 
'3010': 'Food & Staples Retailing', 
'3020': 'Food, Beverage & Tobacco', 
'3030': 'Household & Personal Products', 
'3510': 'Health Care Equipment & Services', 
'3520': 'Pharmaceuticals, Biotechnology & Life Sciences', 
'4010': 'Banks', 
'4020': 'Diversified Financials', 
'4030': 'Insurance', 
'4040': 'Real Estate - - discontinued effective close of Aug 31, 2016', 
'4510': 'Software & Services', 
'4520': 'Technology Hardware & Equipment', 
'4530': 'Semiconductors & Semiconductor Equipment', 
'5010': 'Telecommunication Services', 
'5020': 'Media & Entertainment', 
'5510': 'Utilities', 
'6010': 'Real Estate', 
'101010': 'Energy Equipment & Services', 
'101020': 'Oil, Gas & Consumable Fuels', 
'151010': 'Chemicals', 
'151020': 'Construction Materials', 
'151030': 'Containers & Packaging', 
'151040': 'Metals & Mining', 
'151050': 'Paper & Forest Products', 
'201010': 'Aerospace & Defense', 
'201020': 'Building Products', 
'201030': 'Construction & Engineering', 
'201040': 'Electrical Equipment', 
'201050': 'Industrial Conglomerates', 
'201060': 'Machinery', 
'201070': 'Trading Companies & Distributors', 
'202010': 'Commercial Services & Supplies', 
'202020': 'Professional Services', 
'203010': 'Air Freight & Logistics', 
'203020': 'Airlines', 
'203030': 'Marine', 
'203040': 'Road & Rail', 
'203050': 'Transportation Infrastructure', 
'251010': 'Auto Components', 
'251020': 'Automobiles', 
'252010': 'Household Durables', 
'252020': 'Leisure Products', 
'252030': 'Textiles, Apparel & Luxury Goods', 
'253010': 'Hotels, Restaurants & Leisure', 
'253020': 'Diversified Consumer Services', 
'254010': 'Media (discontinued effective close of September 28, 2018)', 
'255010': 'Distributors', 
'255020': 'Internet & Direct Marketing Retail', 
'255030': 'Multiline Retail', 
'255040': 'Specialty Retail', 
'301010': 'Food & Staples Retailing', 
'302010': 'Beverages', 
'302020': 'Food Products', 
'302030': 'Tobacco', 
'303010': 'Household Products', 
'303020': 'Personal Products', 
'351010': 'Health Care Equipment & Supplies', 
'351020': 'Health Care Providers & Services', 
'351030': 'Health Care Technology', 
'352010': 'Biotechnology', 
'352020': 'Pharmaceuticals', 
'352030': 'Life Sciences Tools & Services', 
'401010': 'Banks', 
'401020': 'Thrifts & Mortgage Finance', 
'402010': 'Diversified Financial Services', 
'402020': 'Consumer Finance', 
'402030': 'Capital Markets', 
'402040': 'Mortgage Real Estate Investment \nTrusts (REITs)', 
'403010': 'Insurance', 
'404010': 'Real Estate -- Discontinued effective 04/28/2006', 
'404020': 'Real Estate Investment Trusts (REITs) - discontinued effective close of Aug 31, 2016', 
'404030': 'Real Estate Management & Development (discontinued effective close of August 31, 2016)', 
'451010': 'Internet Software & Services (discontinued effective close of September 28, 2018)', 
'451020': 'IT Services', 
'451030': 'Software', 
'452010': 'Communications Equipment', 
'452020': 'Technology Hardware, Storage & Peripherals', 
'452030': 'Electronic Equipment, Instruments & Components', 
'452040': 'Office Electronics - Discontinued effective 02/28/2014', 
'452050': 'Semiconductor Equipment & Products -- Discontinued effective 04/30/2003.', 
'453010': 'Semiconductors & Semiconductor Equipment', 
'501010': 'Diversified Telecommunication Services', 
'501020': 'Wireless Telecommunication Services', 
'502010': 'Media', 
'502020': 'Entertainment', 
'502030': 'Interactive Media & Services', 
'551010': 'Electric Utilities', 
'551020': 'Gas Utilities', 
'551030': 'Multi-Utilities', 
'551040': 'Water Utilities', 
'551050': 'Independent Power and Renewable Electricity Producers', 
'601010': 'Equity Real Estate \nInvestment Trusts \n(REITs)', 
'601020': 'Real Estate Management & Development', 
'10101010': 'Oil & Gas Drilling', 
'10101020': 'Oil & Gas Equipment & Services', 
'10102010': 'Integrated Oil & Gas', 
'10102020': 'Oil & Gas Exploration & Production', 
'10102030': 'Oil & Gas Refining & Marketing', 
'10102040': 'Oil & Gas Storage & Transportation', 
'10102050': 'Coal & Consumable Fuels', 
'15101010': 'Commodity Chemicals', 
'15101020': 'Diversified Chemicals', 
'15101030': 'Fertilizers & Agricultural Chemicals', 
'15101040': 'Industrial Gases', 
'15101050': 'Specialty Chemicals', 
'15102010': 'Construction Materials', 
'15103010': 'Metal & Glass Containers', 
'15103020': 'Paper Packaging', 
'15104010': 'Aluminum', 
'15104020': 'Diversified Metals & Mining', 
'15104025': 'Copper', 
'15104030': 'Gold', 
'15104040': 'Precious Metals & Minerals', 
'15104045': 'Silver', 
'15104050': 'Steel', 
'15105010': 'Forest Products', 
'15105020': 'Paper Products', 
'20101010': 'Aerospace & Defense', 
'20102010': 'Building Products', 
'20103010': 'Construction & Engineering', 
'20104010': 'Electrical Components & Equipment', 
'20104020': 'Heavy Electrical Equipment', 
'20105010': 'Industrial Conglomerates', 
'20106010': 'Construction Machinery & Heavy Trucks', 
'20106015': 'Agricultural & Farm Machinery', 
'20106020': 'Industrial Machinery', 
'20107010': 'Trading Companies & Distributors', 
'20201010': 'Commercial Printing', 
'20201020': 'Data Processing Services  (discontinued effective close of April 30, 2003)', 
'20201030': 'Diversified Commercial & Professional Services (discontinued effective close of August 31, 2008)', 
'20201040': 'Human Resource & Employment Services (discontinued effective close of August 31, 2008)', 
'20201050': 'Environmental & Facilities Services', 
'20201060': 'Office Services & Supplies', 
'20201070': 'Diversified Support Services', 
'20201080': 'Security & Alarm Services', 
'20202010': 'Human Resource & Employment Services', 
'20202020': 'Research & Consulting Services', 
'20301010': 'Air Freight & Logistics', 
'20302010': 'Airlines', 
'20303010': 'Marine', 
'20304010': 'Railroads', 
'20304020': 'Trucking', 
'20305010': 'Airport Services', 
'20305020': 'Highways & Railtracks', 
'20305030': 'Marine Ports & Services', 
'25101010': 'Auto Parts & Equipment', 
'25101020': 'Tires & Rubber', 
'25102010': 'Automobile Manufacturers', 
'25102020': 'Motorcycle Manufacturers', 
'25201010': 'Consumer Electronics', 
'25201020': 'Home Furnishings', 
'25201030': 'Homebuilding', 
'25201040': 'Household Appliances', 
'25201050': 'Housewares & Specialties', 
'25202010': 'Leisure Products', 
'25202020': 'Photographic Products (discontinued effective close of February 28, 2014)', 
'25203010': 'Apparel, Accessories & Luxury Goods', 
'25203020': 'Footwear', 
'25203030': 'Textiles', 
'25301010': 'Casinos & Gaming', 
'25301020': 'Hotels, Resorts & Cruise Lines', 
'25301030': 'Leisure Facilities', 
'25301040': 'Restaurants', 
'25302010': 'Education Services', 
'25302020': 'Specialized Consumer Services', 
'25401010': 'Advertising (discontinued effective close of September 28, 2018)', 
'25401020': 'Broadcasting (discontinued effective close of September 28, 2018)', 
'25401025': 'Cable & Satellite (discontinued effective close of September 28, 2018)', 
'25401030': 'Movies & Entertainment (discontinued effective close of September 28, 2018)', 
'25401040': 'Publishing (discontinued effective close of September 28, 2018)', 
'25501010': 'Distributors', 
'25502010': 'Catalog Retail (discontinued effective close of August 31, 2016)', 
'25502020': 'Internet & Direct Marketing Retail', 
'25503010': 'Department Stores', 
'25503020': 'General Merchandise Stores', 
'25504010': 'Apparel Retail', 
'25504020': 'Computer & Electronics Retail', 
'25504030': 'Home Improvement Retail', 
'25504040': 'Specialty Stores', 
'25504050': 'Automotive Retail', 
'25504060': 'Homefurnishing Retail', 
'30101010': 'Drug Retail', 
'30101020': 'Food Distributors', 
'30101030': 'Food Retail', 
'30101040': 'Hypermarkets & Super Centers', 
'30201010': 'Brewers', 
'30201020': 'Distillers & Vintners', 
'30201030': 'Soft Drinks', 
'30202010': 'Agricultural Products', 
'30202020': 'Meat, Poultry & Fish (discontinued effective close of March 28 2002)', 
'30202030': 'Packaged Foods & Meats', 
'30203010': 'Tobacco', 
'30301010': 'Household Products', 
'30302010': 'Personal Products', 
'35101010': 'Health Care Equipment', 
'35101020': 'Health Care Supplies', 
'35102010': 'Health Care Distributors', 
'35102015': 'Health Care  Services', 
'35102020': 'Health Care Facilities', 
'35102030': 'Managed Health Care', 
'35103010': 'Health Care Technology', 
'35201010': 'Biotechnology', 
'35202010': 'Pharmaceuticals', 
'35203010': 'Life Sciences Tools & Services', 
'40101010': 'Diversified Banks', 
'40101015': 'Regional Banks', 
'40102010': 'Thrifts & Mortgage Finance', 
'40201010': 'Consumer Finance (discontinued effective close of April 30, 2003)', 
'40201020': 'Other Diversified Financial Services', 
'40201030': 'Multi-Sector Holdings', 
'40201040': 'Specialized Finance', 
'40202010': 'Consumer Finance', 
'40203010': 'Asset Management & Custody Banks', 
'40203020': 'Investment Banking & Brokerage', 
'40203030': 'Diversified Capital Markets', 
'40203040': 'Financial Exchanges & Data', 
'40204010': 'Mortgage REITs', 
'40301010': 'Insurance Brokers', 
'40301020': 'Life & Health Insurance', 
'40301030': 'Multi-line Insurance', 
'40301040': 'Property & Casualty Insurance', 
'40301050': 'Reinsurance', 
'40401010': 'Real Estate Investment Trusts (discontinued effective close of April 28, 2006)', 
'40401020': 'Real Estate Management & Development (discontinued effective close of April 28, 2006)', 
'40402010': 'Diversified REITs (discontinued effective close of August 31, 2016)', 
'40402020': 'Industrial REITs (discontinued effective close of August 31, 2016)', 
'40402030': 'Mortgage REITs (discontinued effective close of August 31, 2016)', 
'40402035': 'Hotel & Resort REITs (discontinued effective close of August 31, 2016)', 
'40402040': 'Office REITs (discontinued effective close of August 31, 2016)', 
'40402045': 'Health Care REITs (discontinued effective close of August 31, 2016)', 
'40402050': 'Residential REITs (discontinued effective close of August 31, 2016)', 
'40402060': 'Retail REITs (discontinued effective close of August 31, 2016)', 
'40402070': 'Specialized REITs (discontinued effective close of August 31, 2016)', 
'40403010': 'Diversified Real Estate Activities (discontinued effective close of August 31, 2016)', 
'40403020': 'Real Estate Operating Companies (discontinued effective close of August 31, 2016)', 
'40403030': 'Real Estate Development (discontinued effective close of August 31, 2016)', 
'40403040': 'Real Estate Services (discontinued effective close of August 31, 2016)', 
'45101010': 'Internet Software & Services (discontinued effective close of September 28, 2018)', 
'45102010': 'IT Consulting & Other Services', 
'45102020': 'Data Processing & Outsourced Services', 
'45102030': 'Internet Services & Infrastructure', 
'45103010': 'Application Software', 
'45103020': 'Systems Software', 
'45103030': 'Home Entertainment Software (discontinued effective close of September 28, 2018)', 
'45201020': 'Communications Equipment', 
'45201010': 'Networking Equipment (discontinued effective close of April 30, 2003)', 
'45202010': 'Computer Hardware (discontinued effective close of February 28, 2014)', 
'45202020': 'Computer Storage & Peripherals (discontinued effective close of February 28, 2014)', 
'45202030': 'Technology Hardware, Storage & Peripherals', 
'45203010': 'Electronic Equipment & Instruments ', 
'45203015': 'Electronic Components', 
'45203020': 'Electronic Manufacturing Services', 
'45203030': 'Technology Distributors', 
'45204010': 'Office Electronics (discontinued effective close of February 28, 2014)', 
'45205010': 'Semiconductor Equipment (discontinued effective close of April 30, 2003)', 
'45205020': 'Semiconductors (discontinued effective close of April 30, 2003)', 
'45301010': 'Semiconductor Equipment ', 
'45301020': 'Semiconductors', 
'50101010': 'Alternative Carriers', 
'50101020': 'Integrated Telecommunication Services', 
'50102010': 'Wireless Telecommunication Services', 
'50201010': 'Advertising', 
'50201020': 'Broadcasting', 
'50201030': 'Cable & Satellite', 
'50201040': 'Publishing', 
'50202010': 'Movies & Entertainment', 
'50202020': 'Interactive Home Entertainment', 
'50203010': 'Interactive Media & Services', 
'55101010': 'Electric Utilities', 
'55102010': 'Gas Utilities', 
'55103010': 'Multi-Utilities', 
'55104010': 'Water Utilities', 
'55105010': 'Independent Power Producers & Energy Traders', 
'55105020': 'Renewable Electricity ', 
'60101010': 'Diversified REITs', 
'60101020': 'Industrial REITs', 
'60101030': 'Hotel & Resort REITs ', 
'60101040': 'Office REITs ', 
'60101050': 'Health Care REITs ', 
'60101060': 'Residential REITs', 
'60101070': 'Retail REITs', 
'60101080': 'Specialized REITs ', 
'60102010': 'Diversified Real Estate Activities ', 
'60102020': 'Real Estate Operating Companies', 
'60102030': 'Real Estate Development ', 
'60102040': 'Real Estate Services ',
'99': 'Missing or Unknown GICS',
'9999': 'Missing or Unknown GICS',
'999999': 'Missing or Unknown GICS',
'99999999': 'Missing or Unknown GICS',
}

# -*- coding: utf-8 -*-
"""
Created on Mon Mar  6 17:22:06 2017

@author: ub71894 (4e8e6d0b), CSG
"""

['Attachdefaults.py', 'combo.py', 'CreateBenchmarkMatrix.py', 'CreateBenchmarkMatrix_old_openpyxl.py', 'MFA.py', 'PDModel.py', 'Process.py', 'SFA.py', '_info_data.py', '__init__.py']
[336, 7465, 7887, 8286, 9187, 9295, 10810, 11768, 14286, 14293]
